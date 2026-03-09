from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, g
from functools import wraps
import json
import os
import hashlib
import socket
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime, timedelta
from dotenv import load_dotenv
load_dotenv()
from database import init_db, get_db, build_snapshot, save_snapshot, restore_snapshot, get_job_data
from chatbot_engine import generate_bot_response
from claude_chatbot import generate_claude_response
from duplicate_detector import check_duplicate
from tax_rates import lookup_tax
from werkzeug.security import check_password_hash, generate_password_hash
import subprocess, tempfile
try:
    import weasyprint
except Exception:
    weasyprint = None

app = Flask(__name__)
app.config['JSON_SORT_KEYS'] = False
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.secret_key = os.environ.get('SECRET_KEY', 'construction-mgmt-secret-key-change-in-prod')

# Initialize DB on import (needed for gunicorn)
init_db()

# One-time backfill: set tax rates from zip codes for existing projects
def _backfill_tax_rates():
    conn = get_db()
    jobs = conn.execute("SELECT id, zip_code FROM jobs WHERE (tax_rate IS NULL OR tax_rate = 0) AND zip_code IS NOT NULL AND zip_code != ''").fetchall()
    updated = 0
    for job in jobs:
        try:
            info = lookup_tax(job['zip_code'])
            if info.get('tax_rate'):
                conn.execute('UPDATE jobs SET tax_rate = ? WHERE id = ?', (info['tax_rate'], job['id']))
                updated += 1
        except Exception:
            pass
    if updated:
        conn.commit()
    conn.close()
_backfill_tax_rates()

@app.after_request
def add_no_cache_headers(response):
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# Cache-busting version for static files — changes on each server restart
import time as _time
_static_version = str(int(_time.time()))

@app.context_processor
def inject_static_version():
    return {'sv': _static_version}

# ─── Auth Helpers ────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('must_change_password') and request.path != '/change-password':
            return render_template('change_password.html')
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            if session.get('must_change_password') and request.path != '/change-password':
                return render_template('change_password.html')
            if session.get('role') not in roles:
                return 'Access denied', 403
            return f(*args, **kwargs)
        return decorated
    return decorator

def api_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401
        return f(*args, **kwargs)
    return decorated

def api_role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return jsonify({'error': 'Not authenticated'}), 401
            if session.get('role') not in roles:
                return jsonify({'error': 'Access denied'}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator

@app.context_processor
def inject_user():
    if 'user_id' in session:
        return {
            'current_user': {
                'id': session['user_id'],
                'username': session.get('username', ''),
                'display_name': session.get('display_name', ''),
                'role': session.get('role', 'employee'),
            }
        }
    return {'current_user': None}

# ─── Auto Activity Logging Hook ──────────────────────────────────

_ROUTE_ENTITY_MAP = {
    '/api/jobs': 'job', '/api/change-orders': 'change_order', '/api/payapps': 'pay_app',
    '/api/bids': 'bid', '/api/rfis': 'rfi', '/api/submittals': 'submittal',
    '/api/accounting': 'accounting', '/api/payroll': 'payroll', '/api/warranties': 'warranty',
    '/api/service-calls': 'service_call', '/api/admin/users': 'user',
    '/api/howtos': 'howto', '/api/manuals': 'manual', '/api/codebooks': 'codebook',
    '/api/expenses': 'expense', '/api/payments': 'payment', '/api/invoices': 'invoice',
    '/api/contracts': 'contract', '/api/documents': 'document', '/api/transmittals': 'transmittal',
    '/api/licenses': 'license', '/api/recurring-expenses': 'recurring_expense',
    '/api/schedule': 'schedule', '/api/material-requests': 'material_request',
    '/api/inventory': 'inventory', '/api/plans': 'plan', '/api/permits': 'permit',
    '/api/photos': 'photo', '/api/lien-waivers': 'lien_waiver',
    '/api/supplier-quotes': 'supplier_quote', '/api/coi': 'coi',
    '/api/customers': 'customer', '/api/vendors': 'vendor',
    '/api/material-shipments': 'material_shipment', '/api/receiving': 'receiving',
    '/api/feedback': 'feedback', '/api/team-pay': 'team_pay',
    '/api/team-chat': 'team_chat',
}

_SKIP_PREFIXES = ('/api/heartbeat', '/api/notifications', '/static', '/api/chat', '/login', '/logout', '/change-password')

@app.after_request
def _auto_log_activity(response):
    if request.method not in ('POST', 'PUT', 'DELETE'):
        return response
    if not 200 <= response.status_code < 300:
        return response
    if 'user_id' not in session:
        return response
    path = request.path
    if any(path.startswith(p) for p in _SKIP_PREFIXES):
        return response
    # Determine action
    action = {'POST': 'create', 'PUT': 'update', 'DELETE': 'delete'}.get(request.method, 'action')
    parts = path.rstrip('/').split('/')
    for suffix, act in (('approve', 'approve'), ('generate-pdf', 'generate_pdf'), ('email', 'email'),
                        ('submit', 'submit'), ('reject', 'reject')):
        if suffix in parts:
            action = act
            break
    # Determine entity_type and entity_id
    entity_type, entity_id = '', None
    for prefix, etype in _ROUTE_ENTITY_MAP.items():
        if path.startswith(prefix):
            entity_type = etype
            remainder = path[len(prefix):].strip('/')
            if remainder:
                first_segment = remainder.split('/')[0]
                if first_segment.isdigit():
                    entity_id = int(first_segment)
            break
    desc = f"{action.replace('_', ' ').title()} {entity_type.replace('_', ' ')}"
    if entity_id:
        desc += f" #{entity_id}"
    try:
        from database import get_db as _gdb
        conn = _gdb()
        conn.execute(
            '''INSERT INTO activity_logs (user_id, action, entity_type, entity_id, description, ip_address)
               VALUES (?,?,?,?,?,?)''',
            (session['user_id'], action, entity_type, entity_id, desc, request.remote_addr or '')
        )
        conn.commit()
        conn.close()
    except Exception:
        pass
    return response

# ─── PWA / Icon Routes ────────────────────────────────────────────

@app.route('/apple-touch-icon.png')
@app.route('/apple-touch-icon-precomposed.png')
def apple_touch_icon():
    return send_file(os.path.join(app.static_folder, 'icons', 'apple-touch-icon.png'), mimetype='image/png')

# ─── Auth Routes ─────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE username = ? AND is_active = 1', (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['display_name'] = user['display_name']
            session['role'] = user['role']
            session['must_change_password'] = bool(user['must_change_password'])
            log_activity(user['id'], 'login', 'session', None, f"{user['display_name'] or user['username']} logged in")
            if user['must_change_password']:
                return render_template('change_password.html')
            return redirect(url_for('index'))
        return render_template('login.html', error='Invalid username or password')
    return render_template('login.html')

@app.route('/logout')
def logout():
    if 'user_id' in session:
        log_activity(session['user_id'], 'logout', 'session', None,
                     f"{session.get('display_name') or session.get('username', '')} logged out")
    session.clear()
    return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST' and 'new_password' in request.form:
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')
        if not new_pw or len(new_pw) < 6:
            return render_template('change_password.html', error='Password must be at least 6 characters')
        if new_pw != confirm_pw:
            return render_template('change_password.html', error='Passwords do not match')
        conn = get_db()
        conn.execute('UPDATE users SET password_hash = ?, must_change_password = 0 WHERE id = ?',
                     (generate_password_hash(new_pw), session['user_id']))
        conn.commit()
        conn.close()
        session['must_change_password'] = False
        return redirect(url_for('index'))
    return render_template('change_password.html')

# ─── Index / Role Redirect ──────────────────────────────────────

@app.route('/')
@login_required
def index():
    role = session.get('role', 'employee')
    if role in ('owner', 'admin'):
        return redirect(url_for('dashboard'))
    elif role == 'project_manager':
        return redirect(url_for('materials_list'))
    elif role == 'warehouse':
        return redirect(url_for('materials_list'))
    elif role == 'supplier':
        return redirect(url_for('materials_list'))
    else:
        return redirect(url_for('time_entry'))

# ─── Dashboard ───────────────────────────────────────────────────

@app.route('/dashboard')
@role_required('owner', 'admin')
def dashboard():
    return render_template('dashboard.html')

@app.route('/api/dashboard')
@api_role_required('owner', 'admin')
def api_dashboard():
    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    user_id = session.get('user_id')
    jobs = conn.execute('SELECT * FROM jobs ORDER BY name').fetchall()

    total_jobs = len(jobs)
    active_jobs = sum(1 for j in jobs if j['status'] == 'In Progress')
    completed_jobs = sum(1 for j in jobs if j['status'] == 'Complete')
    needs_bid = sum(1 for j in jobs if j['status'] == 'Needs Bid')

    # Financial summary
    total_expenses = conn.execute('SELECT COALESCE(SUM(amount), 0) FROM expenses').fetchone()[0]
    total_payments = conn.execute('SELECT COALESCE(SUM(amount), 0) FROM payments').fetchone()[0]
    total_invoiced = conn.execute("SELECT COALESCE(SUM(amount), 0) FROM client_invoices WHERE status = 'Paid'").fetchone()[0]
    total_outstanding = conn.execute("SELECT COALESCE(SUM(amount), 0) FROM client_invoices WHERE status IN ('Sent','Overdue','Partial')").fetchone()[0]

    # Estimated material costs (from bids/line_items)
    estimated_material_cost = 0
    for job in jobs:
        items = conn.execute(
            'SELECT total_net_price, qty_ordered, price_per FROM line_items WHERE job_id = ?',
            (job['id'],)
        ).fetchall()
        estimated_material_cost += sum(
            (row['total_net_price'] or 0) if (row['total_net_price'] or 0)
            else (row['qty_ordered'] or 0) * (row['price_per'] or 0)
            for row in items
        )

    # Actual material costs (from supplier invoices)
    actual_material_cost = conn.execute(
        "SELECT COALESCE(SUM(total), 0) FROM supplier_invoices WHERE is_duplicate = 0"
    ).fetchone()[0]

    # Open service calls
    open_calls_count = conn.execute("SELECT COUNT(*) FROM service_calls WHERE status NOT IN ('Resolved','Closed')").fetchone()[0]

    # Pending time entries
    pending_hours = conn.execute("SELECT COALESCE(SUM(hours), 0) FROM time_entries WHERE approved = 0").fetchone()[0]

    # Stage breakdown
    stages = ['Needs Bid', 'Takeoff Sent', 'Bid Complete', 'Awarded', 'In Progress', 'Complete']
    stage_counts = {}
    for s in stages:
        stage_counts[s] = sum(1 for j in jobs if j['status'] == s)

    # ─── Action Items ───────────────────────────────────────────
    # Reminders due soon (next 7 days)
    reminders_due = conn.execute(
        "SELECT * FROM reminders WHERE user_id = ? AND status = 'Active' AND due_date != '' AND due_date <= date(?, '+7 days') ORDER BY due_date LIMIT 8",
        (user_id, today)
    ).fetchall()

    # Open service calls (top 5 by priority)
    open_calls = conn.execute('''
        SELECT sc.*, j.name as job_name, u.display_name as assigned_name
        FROM service_calls sc
        LEFT JOIN jobs j ON sc.job_id = j.id
        LEFT JOIN users u ON sc.assigned_to = u.id
        WHERE sc.status NOT IN ('Resolved','Closed')
        ORDER BY CASE sc.priority WHEN 'Urgent' THEN 0 WHEN 'High' THEN 1 WHEN 'Normal' THEN 2 ELSE 3 END, sc.created_at DESC
        LIMIT 5
    ''').fetchall()

    # Overdue invoices
    overdue_invoices = conn.execute('''
        SELECT ci.*, j.name as job_name FROM client_invoices ci
        LEFT JOIN jobs j ON ci.job_id = j.id
        WHERE ci.status IN ('Sent','Overdue') AND ci.due_date != '' AND ci.due_date < ?
        ORDER BY ci.due_date LIMIT 5
    ''', (today,)).fetchall()

    # Pending submittals (rejected/resubmit)
    pending_submittals = conn.execute('''
        SELECT s.*, j.name as job_name FROM submittals s
        LEFT JOIN jobs j ON s.job_id = j.id
        WHERE s.status IN ('Rejected','Resubmit')
        ORDER BY s.updated_at DESC LIMIT 5
    ''').fetchall()

    # Expiring licenses (next 60 days)
    expiring_licenses = conn.execute(
        "SELECT * FROM licenses WHERE expiration_date != '' AND expiration_date <= date(?, '+60 days') AND expiration_date >= ? ORDER BY expiration_date LIMIT 5",
        (today, today)
    ).fetchall()

    # ─── Active Projects with progress ──────────────────────────
    active_project_list = []
    for j in jobs:
        if j['status'] != 'In Progress':
            continue
        jid = j['id']
        # Billing progress from pay apps
        contract_row = conn.execute(
            'SELECT COALESCE(SUM(original_contract_sum), 0) as total_contract FROM pay_app_contracts WHERE job_id = ?',
            (jid,)
        ).fetchone()
        billed_row = conn.execute('''
            SELECT COALESCE(SUM(ple.work_this_period + ple.materials_stored), 0) as total_billed
            FROM pay_app_line_entries ple
            JOIN pay_applications pa ON ple.pay_app_id = pa.id
            JOIN pay_app_contracts pac ON pa.contract_id = pac.id
            WHERE pac.job_id = ?
        ''', (jid,)).fetchone()
        contract_val = contract_row['total_contract'] or 0
        billed_val = billed_row['total_billed'] or 0
        pct = round(billed_val / contract_val * 100) if contract_val > 0 else 0
        # Completion % from schedule phases
        comp_row = conn.execute(
            "SELECT AVG(pct_complete) as avg_pct FROM job_schedule_events WHERE job_id = ? AND status != 'Cancelled'",
            (jid,)
        ).fetchone()
        pct_complete = round(comp_row['avg_pct'] or 0) if comp_row and comp_row['avg_pct'] else 0

        active_project_list.append({
            'id': jid, 'name': j['name'],
            'contract': round(contract_val, 2),
            'billed': round(billed_val, 2),
            'pct_billed': min(pct, 100),
            'pct_complete': min(pct_complete, 100),
        })
    active_project_list.sort(key=lambda p: p['name'])

    # ─── Recent Activity (last 10) ─────────────────────────────
    recent_activity = conn.execute('''
        SELECT al.*, u.display_name FROM activity_logs al
        LEFT JOIN users u ON al.user_id = u.id
        ORDER BY al.created_at DESC LIMIT 10
    ''').fetchall()

    # ─── Upcoming Schedule (next 14 days) ──────────────────────
    upcoming_schedule = conn.execute('''
        SELECT e.*, j.name as job_name, u.display_name as assigned_name
        FROM job_schedule_events e
        LEFT JOIN jobs j ON e.job_id = j.id
        LEFT JOIN users u ON e.assigned_to = u.id
        WHERE e.status != 'Complete' AND e.start_date != '' AND e.start_date <= date(?, '+14 days')
        ORDER BY e.start_date LIMIT 8
    ''', (today,)).fetchall()

    conn.close()
    return jsonify({
        'total_jobs': total_jobs,
        'active_jobs': active_jobs,
        'completed_jobs': completed_jobs,
        'needs_bid': needs_bid,
        'estimated_material_cost': round(estimated_material_cost, 2),
        'actual_material_cost': round(actual_material_cost, 2),
        'material_cost': round(estimated_material_cost, 2),
        'total_expenses': round(total_expenses, 2),
        'total_payments': round(total_payments, 2),
        'total_invoiced': round(total_invoiced, 2),
        'total_outstanding': round(total_outstanding, 2),
        'open_service_calls': open_calls_count,
        'pending_hours': round(pending_hours, 1),
        'stage_counts': stage_counts,
        'reminders_due': [dict(r) for r in reminders_due],
        'open_calls_list': [dict(c) for c in open_calls],
        'overdue_invoices': [dict(i) for i in overdue_invoices],
        'pending_submittals': [dict(s) for s in pending_submittals],
        'expiring_licenses': [dict(l) for l in expiring_licenses],
        'active_projects': active_project_list,
        'recent_activity': [dict(a) for a in recent_activity],
        'upcoming_schedule': [dict(e) for e in upcoming_schedule],
    })

# ─── Materials (relocated from old / and /job routes) ────────────

@app.route('/api/materials/import', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_materials_import():
    """Import an Excel workbook into line_items (and tracking entries) for a job."""
    from openpyxl import load_workbook

    file = request.files.get('file')
    job_id = request.form.get('job_id')
    if not file or not file.filename:
        return jsonify({'error': 'No file provided'}), 400
    if not job_id:
        return jsonify({'error': 'No job selected'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('xlsx', 'xls'):
        return jsonify({'error': 'Please upload an Excel file (.xlsx)'}), 400

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'error': f'Could not read Excel file: {str(e)[:200]}'}), 400

    conn = get_db()

    # Verify job exists
    job = conn.execute('SELECT id, name FROM jobs WHERE id = ?', (job_id,)).fetchone()
    if not job:
        conn.close()
        return jsonify({'error': 'Job not found'}), 404

    # Find the master list sheet (try common names, then fall back to first non-tracking sheet)
    master_sheet = None
    tracking_names_lower = ['received', 'shipped', 'shipped to site', 'invoiced']
    for name in ['Master List', 'master list', 'Sheet1', 'Materials', 'Line Items']:
        if name in wb.sheetnames:
            master_sheet = wb[name]
            break
    if not master_sheet:
        # Use first sheet that isn't a tracking sheet
        for name in wb.sheetnames:
            if name.lower().strip() not in tracking_names_lower:
                master_sheet = wb[name]
                break
    if not master_sheet:
        master_sheet = wb.active

    # Map header names to DB fields
    header_aliases = {
        'line_number': ['line #', 'line', 'line number', 'line no', '#', 'ln', 'no', 'no.'],
        'stock_ns': ['stock/ns', 'stock', 'stock ns', 'type', 's/ns', 'non-stock', 'nonstock', 'ns'],
        'sku': ['sku', 'part number', 'part #', 'part no', 'product code', 'item', 'item #',
                'item no', 'item number', 'catalog', 'catalog #', 'cat #', 'cat no', 'code',
                'material', 'mat', 'product #', 'product no', 'product number', 'part',
                'product'],
        'description': ['description', 'desc', 'item description', 'product', 'product description',
                        'material description', 'name', 'item name', 'material name',
                        'desc.', 'descriptions', 'comments'],
        'quote_qty': ['quote qty', 'quote quantity', 'qty quoted', 'est qty', 'estimated qty',
                      'est quantity', 'bid qty'],
        'qty_ordered': ['qty ordered', 'quantity ordered', 'ordered', 'qty', 'quantity',
                        'order qty', 'order quantity', 'qty.', 'qty on order', 'ship qty',
                        'qty to ship', 'pieces', 'pcs', 'count', 'units', 'order'],
        'price_per': ['price per', 'unit price', 'price', 'cost', 'unit cost', 'rate',
                      'price each', 'each', 'price/unit', 'cost/unit', 'unit', 'sell price',
                      'sell', 'net price each', 'net each', 'per'],
        'total_net_price': ['total net price', 'total price', 'order total', 'net price', 'extended',
                            'ext price', 'amount', 'ext.', 'ext', 'extension', 'line total',
                            'extended price', 'total cost', 'net total', 'net amount', 'net',
                            'total', 'quote total'],
        'notes': ['notes', 'note', 'remarks', 'comments', 'memo'],
    }

    # Scan first 10 rows for a header row (some files have title rows before headers)
    header_row = 1
    headers = {}
    col_map = {}
    for try_row in range(1, min(11, master_sheet.max_row + 1)):
        trial_headers = {}
        for col in range(1, master_sheet.max_column + 1):
            val = master_sheet.cell(row=try_row, column=col).value
            if val:
                trial_headers[str(val).strip().lower()] = col

        # Check if this row has at least description or sku
        trial_map = {}
        for field, aliases in header_aliases.items():
            for alias in aliases:
                if alias in trial_headers:
                    trial_map[field] = trial_headers[alias]
                    break
        if 'description' in trial_map or 'sku' in trial_map:
            headers = trial_headers
            col_map = trial_map
            header_row = try_row
            break

    # Must have at least description or sku to be useful
    if 'description' not in col_map and 'sku' not in col_map:
        # Show what headers we found to help debug
        row1_headers = []
        for r in range(1, min(4, master_sheet.max_row + 1)):
            vals = [str(master_sheet.cell(row=r, column=c).value or '') for c in range(1, min(master_sheet.max_column + 1, 15))]
            row1_headers.append(f'Row {r}: {", ".join(v for v in vals if v)}')
        found_str = ' | '.join(row1_headers)
        conn.close()
        return jsonify({'error': f'Could not find a Description or SKU column. '
                        f'Found in your file: {found_str}. '
                        f'Expected headers like: Line #, SKU, Description, Qty, Price Per, Total'}), 400

    # Clear existing line_items for this job (full re-import)
    existing = conn.execute('SELECT id FROM line_items WHERE job_id = ?', (job_id,)).fetchall()
    if existing:
        ids = [str(r['id']) for r in existing]
        id_list = ','.join(ids)
        conn.execute(f'DELETE FROM received_entries WHERE line_item_id IN ({id_list})')
        conn.execute(f'DELETE FROM shipped_entries WHERE line_item_id IN ({id_list})')
        conn.execute(f'DELETE FROM invoiced_entries WHERE line_item_id IN ({id_list})')
        conn.execute('DELETE FROM line_items WHERE job_id = ?', (job_id,))

    def safe_num(val):
        if val is None:
            return 0
        if isinstance(val, (int, float)):
            return val
        try:
            return float(str(val).replace('$', '').replace(',', ''))
        except (ValueError, TypeError):
            return 0

    # Import line items (start after the header row we found)
    imported = 0
    line_id_map = {}  # line_number -> new line_item id
    for row in range(header_row + 1, master_sheet.max_row + 1):
        desc = master_sheet.cell(row=row, column=col_map.get('description', 0)).value if 'description' in col_map else ''
        sku = master_sheet.cell(row=row, column=col_map.get('sku', 0)).value if 'sku' in col_map else ''

        # Skip empty rows
        if not desc and not sku:
            continue

        line_num = safe_num(master_sheet.cell(row=row, column=col_map.get('line_number', 0)).value) if 'line_number' in col_map else (imported + 1)
        if line_num == 0:
            line_num = imported + 1

        stock_ns = str(master_sheet.cell(row=row, column=col_map.get('stock_ns', 0)).value or '') if 'stock_ns' in col_map else ''
        quote_qty = safe_num(master_sheet.cell(row=row, column=col_map.get('quote_qty', 0)).value) if 'quote_qty' in col_map else 0
        qty_ordered = safe_num(master_sheet.cell(row=row, column=col_map.get('qty_ordered', 0)).value) if 'qty_ordered' in col_map else 0
        price_per = safe_num(master_sheet.cell(row=row, column=col_map.get('price_per', 0)).value) if 'price_per' in col_map else 0
        total_net = safe_num(master_sheet.cell(row=row, column=col_map.get('total_net_price', 0)).value) if 'total_net_price' in col_map else 0
        notes = str(master_sheet.cell(row=row, column=col_map.get('notes', 0)).value or '') if 'notes' in col_map else ''

        # Detect per-C or per-M pricing by comparing total to qty * price
        pricing_type = 'each'
        qty_for_calc = qty_ordered if qty_ordered > 0 else quote_qty
        if total_net and price_per and qty_for_calc:
            simple = qty_for_calc * price_per
            if simple > 0:
                ratio = total_net / simple
                if abs(ratio - 0.01) < 0.003:
                    pricing_type = 'per_c'
                elif abs(ratio - 0.001) < 0.0003:
                    pricing_type = 'per_m'

        # Auto-calc total if missing, using detected pricing type
        if total_net == 0 and qty_for_calc > 0 and price_per > 0:
            if pricing_type == 'per_c':
                total_net = qty_for_calc * price_per / 100
            elif pricing_type == 'per_m':
                total_net = qty_for_calc * price_per / 1000
            else:
                total_net = qty_for_calc * price_per

        cursor = conn.execute(
            '''INSERT INTO line_items (job_id, line_number, stock_ns, sku, description,
               quote_qty, qty_ordered, price_per, total_net_price, pricing_type, notes)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
            (job_id, int(line_num), stock_ns, str(sku or ''), str(desc or ''),
             quote_qty, qty_ordered, price_per, total_net, pricing_type, notes)
        )
        line_id_map[int(line_num)] = cursor.lastrowid
        # Also map by description (normalized) for tracking sheets that don't have line numbers
        desc_key = str(desc or '').strip().lower()
        if desc_key:
            line_id_map['desc:' + desc_key] = cursor.lastrowid
        imported += 1

    # Import tracking entries from Received/Shipped/Invoiced sheets
    tracking_sheets = {
        'Received': 'received_entries',
        'Shipped to Site': 'shipped_entries',
        'Shipped': 'shipped_entries',
        'Invoiced': 'invoiced_entries',
    }
    for sheet_name, table_name in tracking_sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Find the header row (scan first 10 rows for one with known column names)
        t_header_row = 1
        t_data_start = 2
        known_headers = ['description', 'desc', 'line #', 'line', 'sku', 'product',
                         'qty', 'quantity', 'total', 'item', 'name', 'material']
        for try_r in range(1, min(11, ws.max_row + 1)):
            row_vals = [str(ws.cell(row=try_r, column=c).value or '').strip().lower()
                       for c in range(1, min(ws.max_column + 1, 20))]
            matches = sum(1 for v in row_vals if v in known_headers or
                         any(kw in v for kw in ['description', 'qty', 'received', 'shipped', 'invoiced']))
            if matches >= 2:
                t_header_row = try_r
                t_data_start = try_r + 1
                break

        # Find which column has line numbers and which has descriptions
        t_headers = {}
        for c in range(1, min(ws.max_column + 1, 50)):
            val = ws.cell(row=t_header_row, column=c).value
            if val:
                t_headers[str(val).strip().lower()] = c

        line_col = None
        desc_col = None
        for alias in ['line #', 'line', 'line number', '#']:
            if alias in t_headers:
                line_col = t_headers[alias]
                break
        for alias in ['description', 'desc', 'item description', 'product', 'product description', 'name']:
            if alias in t_headers:
                desc_col = t_headers[alias]
                break

        # Find where quantity columns start (after description/totals)
        qty_start_col = None
        for c in range(1, min(ws.max_column + 1, 50)):
            hdr = str(ws.cell(row=t_header_row, column=c).value or '').strip().lower()
            # Skip "total" summary columns — we want the first actual entry column
            if 'total' in hdr:
                continue
            if any(kw in hdr for kw in ['qty received', 'qty shipped', 'invoice', 'qty 1']):
                qty_start_col = c
                break
        if qty_start_col is None:
            # Fall back: first column after "total" columns
            for c in range(3, min(ws.max_column + 1, 50)):
                hdr = str(ws.cell(row=t_header_row, column=c).value or '').strip().lower()
                if hdr and ('1' in hdr or 'date' in hdr.lower()):
                    qty_start_col = c
                    break
        if qty_start_col is None:
            qty_start_col = 3  # default fallback

        # Check row above header for dates (some files put dates in a separate row)
        date_headers = {}
        date_row = t_header_row - 1 if t_header_row > 1 else None
        for col in range(qty_start_col, min(ws.max_column + 1, qty_start_col + 15)):
            # Check header row first, then row above
            val = ws.cell(row=t_header_row, column=col).value
            if date_row and not val:
                val = ws.cell(row=date_row, column=col).value
            # Also check row below header for dates
            if not val and t_data_start <= ws.max_row:
                check_val = ws.cell(row=t_data_start, column=col).value
                if hasattr(check_val, 'strftime'):
                    val = check_val
                    t_data_start = t_data_start + 1  # dates row, data starts after
            if val:
                if hasattr(val, 'strftime'):
                    date_headers[col] = val.strftime('%Y-%m-%d')
                else:
                    date_headers[col] = str(val)

        for row in range(t_data_start, ws.max_row + 1):
            # Try to match by line number first, then by description
            line_item_id = None
            if line_col:
                line_num_val = safe_num(ws.cell(row=row, column=line_col).value)
                if line_num_val > 0:
                    line_item_id = line_id_map.get(int(line_num_val))

            if not line_item_id and desc_col:
                desc_val = str(ws.cell(row=row, column=desc_col).value or '').strip().lower()
                if desc_val:
                    line_item_id = line_id_map.get('desc:' + desc_val)

            # Also try column 1 as description if no line_col/desc_col matched
            if not line_item_id:
                desc_val = str(ws.cell(row=row, column=1).value or '').strip().lower()
                if desc_val:
                    line_item_id = line_id_map.get('desc:' + desc_val)

            if not line_item_id:
                continue

            for col in range(qty_start_col, min(ws.max_column + 1, qty_start_col + 15)):
                qty = safe_num(ws.cell(row=row, column=col).value)
                if qty == 0:
                    continue
                col_number = col - qty_start_col + 1
                if col_number < 1 or col_number > 15:
                    continue
                entry_date = date_headers.get(col, '')
                conn.execute(
                    f'''INSERT OR REPLACE INTO {table_name}
                        (line_item_id, column_number, quantity, entry_date)
                        VALUES (?,?,?,?)''',
                    (line_item_id, col_number, qty, entry_date)
                )

    conn.commit()
    conn.close()

    return jsonify({
        'ok': True,
        'imported': imported,
        'message': f'{imported} line items imported for {job["name"]}'
    }), 201


@app.route('/materials')
@role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def materials_list():
    conn = get_db()
    jobs = conn.execute('SELECT * FROM jobs ORDER BY name ASC').fetchall()
    conn.close()
    return render_template('materials/list.html', jobs=jobs)

@app.route('/materials/job/<int:job_id>')
@role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def materials_job(job_id):
    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    conn.close()
    if not job:
        return 'Job not found', 404
    return render_template('materials/job.html', job=job)

@app.route('/materials/job/<int:job_id>/history')
@role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def materials_history(job_id):
    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    conn.close()
    if not job:
        return 'Job not found', 404
    return render_template('materials/history.html', job=job)

# ─── Existing Job API Routes (unchanged functionality) ───────────

@app.route('/api/jobs', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def create_job():
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Job name is required'}), 400

    address = (data.get('address') or '').strip()
    city = (data.get('city') or '').strip()
    state = (data.get('state') or '').strip()
    zip_code = (data.get('zip_code') or '').strip()
    tax_rate = data.get('tax_rate')

    if tax_rate is None and zip_code:
        tax_info = lookup_tax(zip_code)
        tax_rate = tax_info['tax_rate']
        if not city:
            city = tax_info['city']
        if not state:
            state = tax_info['state']
    tax_rate = tax_rate or 0
    supplier_account = (data.get('supplier_account') or '').strip()
    customer_id = data.get('customer_id')
    if customer_id is not None:
        customer_id = int(customer_id) if customer_id else None

    conn = get_db()
    cursor = conn.execute(
        'INSERT INTO jobs (name, status, address, city, state, zip_code, tax_rate, supplier_account, customer_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
        (name, 'Needs Bid', address, city, state, zip_code, tax_rate, supplier_account, customer_id)
    )
    job_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': job_id, 'name': name}), 201

@app.route('/api/job/<int:job_id>')
@api_login_required
def get_job(job_id):
    conn = get_db()
    data = get_job_data(conn, job_id)
    conn.close()
    if not data:
        return jsonify({'error': 'Job not found'}), 404
    return jsonify(data)

@app.route('/api/job/<int:job_id>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def update_job(job_id):
    data = request.get_json()
    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    if not job:
        conn.close()
        return jsonify({'error': 'Job not found'}), 404

    name = (data.get('name') or '').strip()
    status = (data.get('status') or '').strip()
    if name:
        conn.execute('UPDATE jobs SET name = ?, updated_at = datetime("now","localtime") WHERE id = ?', (name, job_id))
    if status:
        conn.execute('UPDATE jobs SET status = ?, updated_at = datetime("now","localtime") WHERE id = ?', (status, job_id))

    for field in ('address', 'city', 'state', 'zip_code', 'supplier_account'):
        if field in data:
            val = (data[field] or '').strip()
            conn.execute(f'UPDATE jobs SET {field} = ?, updated_at = datetime("now","localtime") WHERE id = ?', (val, job_id))
    if 'project_manager_id' in data:
        pm_id = data['project_manager_id']
        conn.execute('UPDATE jobs SET project_manager_id = ?, updated_at = datetime("now","localtime") WHERE id = ?',
                     (int(pm_id) if pm_id else None, job_id))
    if 'tax_rate' in data:
        conn.execute('UPDATE jobs SET tax_rate = ?, updated_at = datetime("now","localtime") WHERE id = ?',
                     (float(data['tax_rate'] or 0), job_id))
    if 'customer_id' in data:
        cid = data['customer_id']
        conn.execute('UPDATE jobs SET customer_id = ?, updated_at = datetime("now","localtime") WHERE id = ?',
                     (int(cid) if cid else None, job_id))

    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/job/<int:job_id>', methods=['DELETE'])
@api_role_required('owner', 'admin')
def delete_job(job_id):
    conn = get_db()
    conn.execute('DELETE FROM jobs WHERE id = ?', (job_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/analytics')
@api_role_required('owner', 'admin')
def get_analytics():
    conn = get_db()
    jobs = conn.execute('SELECT * FROM jobs ORDER BY name').fetchall()

    stages = ['Needs Bid', 'Bid Complete', 'In Progress', 'Complete']
    result = {}
    grand_subtotal = 0
    grand_tax = 0
    grand_shipping = 0

    for stage in stages:
        result[stage] = {'count': 0, 'jobs': [], 'subtotal': 0, 'tax': 0, 'shipping': 0, 'total': 0}

    for job in jobs:
        stage = job['status'] if job['status'] in stages else 'Needs Bid'
        items = conn.execute(
            'SELECT total_net_price, qty_ordered, price_per FROM line_items WHERE job_id = ?',
            (job['id'],)
        ).fetchall()
        subtotal = sum((row['total_net_price'] or 0) if (row['total_net_price'] or 0) else (row['qty_ordered'] or 0) * (row['price_per'] or 0) for row in items)
        tax_rate = job['tax_rate'] or 0
        tax_amount = round(subtotal * tax_rate / 100, 2)

        is_out_of_state = (job['state'] or '').strip().upper() not in ('', 'OK', 'OKLAHOMA')
        shipping = 10000.0 if is_out_of_state else 0.0

        job_total = round(subtotal + tax_amount + shipping, 2)

        result[stage]['count'] += 1
        result[stage]['subtotal'] += subtotal
        result[stage]['tax'] += tax_amount
        result[stage]['shipping'] += shipping
        result[stage]['total'] += job_total
        result[stage]['jobs'].append({
            'id': job['id'],
            'name': job['name'],
            'location': f"{job['city'] or ''} {job['state'] or ''}".strip() or '—',
            'subtotal': round(subtotal, 2),
            'tax': tax_amount,
            'shipping': shipping,
            'total': job_total,
        })

        grand_subtotal += subtotal
        grand_tax += tax_amount
        grand_shipping += shipping

    for stage in stages:
        result[stage]['subtotal'] = round(result[stage]['subtotal'], 2)
        result[stage]['tax'] = round(result[stage]['tax'], 2)
        result[stage]['shipping'] = round(result[stage]['shipping'], 2)
        result[stage]['total'] = round(result[stage]['total'], 2)

    conn.close()
    return jsonify({
        'stages': result,
        'stage_order': stages,
        'grand_subtotal': round(grand_subtotal, 2),
        'grand_tax': round(grand_tax, 2),
        'grand_shipping': round(grand_shipping, 2),
        'grand_total': round(grand_subtotal + grand_tax + grand_shipping, 2),
    })

@app.route('/api/job/<int:job_id>/line-items', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def save_line_items(job_id):
    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    if not job:
        conn.close()
        return jsonify({'error': 'Job not found'}), 404

    save_snapshot(conn, job_id, 'Before master list update')

    data = request.get_json()
    items = data.get('line_items', [])

    existing = conn.execute('SELECT id, line_number FROM line_items WHERE job_id = ?', (job_id,)).fetchall()
    existing_ids = {row['id'] for row in existing}

    incoming_ids = set()
    for item in items:
        li_id = item.get('id')
        line_number = item.get('line_number', 0)
        stock_ns = item.get('stock_ns', '')
        sku = item.get('sku', '')
        description = item.get('description', '')
        quote_qty = item.get('quote_qty', 0) or 0
        qty_ordered = item.get('qty_ordered', 0) or 0
        price_per = item.get('price_per', 0) or 0
        total_net_price = item.get('total_net_price', 0) or 0
        pricing_type = item.get('pricing_type', 'each') or 'each'
        notes = item.get('notes', '') or ''

        if li_id and li_id in existing_ids:
            conn.execute(
                '''UPDATE line_items SET line_number=?, stock_ns=?, sku=?, description=?,
                   quote_qty=?, qty_ordered=?, price_per=?, total_net_price=?, pricing_type=?, notes=? WHERE id=? AND job_id=?''',
                (line_number, stock_ns, sku, description, quote_qty, qty_ordered, price_per, total_net_price, pricing_type, notes, li_id, job_id)
            )
            incoming_ids.add(li_id)
        else:
            cursor = conn.execute(
                '''INSERT INTO line_items (job_id, line_number, stock_ns, sku, description, quote_qty, qty_ordered, price_per, total_net_price, pricing_type, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (job_id, line_number, stock_ns, sku, description, quote_qty, qty_ordered, price_per, total_net_price, pricing_type, notes)
            )
            incoming_ids.add(cursor.lastrowid)

    removed = existing_ids - incoming_ids
    for rid in removed:
        conn.execute('DELETE FROM line_items WHERE id = ?', (rid,))

    conn.execute('UPDATE jobs SET updated_at = datetime("now","localtime") WHERE id = ?', (job_id,))
    conn.commit()

    result = get_job_data(conn, job_id)
    result['ok'] = True
    conn.close()
    return jsonify(result)

@app.route('/api/job/<int:job_id>/detect-pricing', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def detect_pricing_types(job_id):
    """Auto-detect per-C/per-M pricing on existing line items by comparing stored totals to qty*price."""
    conn = get_db()
    items = conn.execute('SELECT id, qty_ordered, quote_qty, price_per, total_net_price, pricing_type FROM line_items WHERE job_id = ?', (job_id,)).fetchall()
    updated = 0
    for item in items:
        total = item['total_net_price'] or 0
        price = item['price_per'] or 0
        qty = item['qty_ordered'] or item['quote_qty'] or 0
        if not (total and price and qty):
            continue
        simple = qty * price
        if simple == 0:
            continue
        ratio = total / simple
        new_type = 'each'
        if abs(ratio - 0.01) < 0.003:
            new_type = 'per_c'
        elif abs(ratio - 0.001) < 0.0003:
            new_type = 'per_m'
        if new_type != (item['pricing_type'] or 'each'):
            conn.execute('UPDATE line_items SET pricing_type = ? WHERE id = ?', (new_type, item['id']))
            updated += 1
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'updated': updated})

def _save_entries(job_id, table_name, data):
    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    if not job:
        conn.close()
        return jsonify({'error': 'Job not found'}), 404

    tab_label = table_name.replace('_entries', '')
    save_snapshot(conn, job_id, f'Before {tab_label} update')

    entries = data.get('entries', [])

    valid_items = conn.execute('SELECT id FROM line_items WHERE job_id = ?', (job_id,)).fetchall()
    valid_ids = {row['id'] for row in valid_items}

    for entry in entries:
        line_item_id = entry.get('line_item_id')
        if line_item_id not in valid_ids:
            continue
        column_number = entry.get('column_number')
        quantity = entry.get('quantity', 0) or 0

        if quantity == 0:
            conn.execute(
                f'DELETE FROM {table_name} WHERE line_item_id = ? AND column_number = ?',
                (line_item_id, column_number)
            )
        else:
            existing = conn.execute(
                f'SELECT id, entry_date FROM {table_name} WHERE line_item_id = ? AND column_number = ?',
                (line_item_id, column_number)
            ).fetchone()

            if existing:
                conn.execute(
                    f'UPDATE {table_name} SET quantity = ? WHERE id = ?',
                    (quantity, existing['id'])
                )
            else:
                entry_date = datetime.now().strftime('%Y-%m-%d')
                conn.execute(
                    f'INSERT INTO {table_name} (line_item_id, column_number, quantity, entry_date) VALUES (?, ?, ?, ?)',
                    (line_item_id, column_number, quantity, entry_date)
                )

    # Save custom column headers if provided
    col_headers = data.get('column_headers', {})
    if col_headers:
        tab = table_name.replace('_entries', '')
        for col_num_str, header_name in col_headers.items():
            col_num = int(col_num_str)
            if header_name:
                conn.execute(
                    '''INSERT INTO column_headers (job_id, tab_type, column_number, header_name)
                       VALUES (?, ?, ?, ?)
                       ON CONFLICT(job_id, tab_type, column_number) DO UPDATE SET header_name = ?''',
                    (job_id, tab, col_num, header_name, header_name)
                )
            else:
                conn.execute(
                    'DELETE FROM column_headers WHERE job_id = ? AND tab_type = ? AND column_number = ?',
                    (job_id, tab, col_num)
                )

    conn.execute('UPDATE jobs SET updated_at = datetime("now","localtime") WHERE id = ?', (job_id,))
    conn.commit()

    result = get_job_data(conn, job_id)
    result['ok'] = True
    conn.close()
    return jsonify(result)

@app.route('/api/job/<int:job_id>/received', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def save_received(job_id):
    return _save_entries(job_id, 'received_entries', request.get_json())

@app.route('/api/job/<int:job_id>/shipped', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def save_shipped(job_id):
    return _save_entries(job_id, 'shipped_entries', request.get_json())

@app.route('/api/job/<int:job_id>/invoiced', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def save_invoiced(job_id):
    return _save_entries(job_id, 'invoiced_entries', request.get_json())

# ─── Tax Lookup ─────────────────────────────────────────────────

@app.route('/api/tax-lookup/<zip_code>')
@api_login_required
def tax_lookup(zip_code):
    result = lookup_tax(zip_code)
    return jsonify(result)

# ─── Version History ────────────────────────────────────────────

@app.route('/api/job/<int:job_id>/versions')
@api_login_required
def list_versions(job_id):
    conn = get_db()
    versions = conn.execute(
        'SELECT id, description, created_at FROM versions WHERE job_id = ? ORDER BY created_at DESC',
        (job_id,)
    ).fetchall()
    conn.close()
    return jsonify([dict(v) for v in versions])

@app.route('/api/job/<int:job_id>/versions/<int:vid>')
@api_login_required
def get_version(job_id, vid):
    conn = get_db()
    version = conn.execute(
        'SELECT * FROM versions WHERE id = ? AND job_id = ?', (vid, job_id)
    ).fetchone()
    conn.close()
    if not version:
        return jsonify({'error': 'Version not found'}), 404
    return jsonify({
        'id': version['id'],
        'description': version['description'],
        'created_at': version['created_at'],
        'snapshot': json.loads(version['snapshot']),
    })

@app.route('/api/job/<int:job_id>/versions/<int:vid>/revert', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def revert_version(job_id, vid):
    conn = get_db()
    version = conn.execute(
        'SELECT * FROM versions WHERE id = ? AND job_id = ?', (vid, job_id)
    ).fetchone()
    if not version:
        conn.close()
        return jsonify({'error': 'Version not found'}), 404

    save_snapshot(conn, job_id, f'Before revert to version {vid}')

    snapshot_data = json.loads(version['snapshot'])
    restore_snapshot(conn, job_id, snapshot_data)
    conn.commit()

    result = get_job_data(conn, job_id)
    result['ok'] = True
    conn.close()
    return jsonify(result)

# ─── PDF Import ─────────────────────────────────────────────────

@app.route('/api/job/<int:job_id>/import-pdf', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def import_pdf(job_id):
    import re
    import pdfplumber

    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    conn.close()
    if not job:
        return jsonify({'error': 'Job not found'}), 404

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'File must be a PDF'}), 400

    try:
        pdf = pdfplumber.open(file)
        all_text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
        pdf.close()
    except Exception as e:
        return jsonify({'error': f'Could not read PDF: {str(e)}'}), 400

    pattern = re.compile(
        r'^(\d+)\s+(.+)\s+([\d,]+)\s+(\d+)\s+(\d+)\s+(\w+)\s+([\d,]+\.\d+)\s+([\d,]+\.\d+)$'
    )

    lines = all_text.split('\n')
    items = []
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        m = pattern.match(line)
        if m:
            qty_ordered = float(m.group(3).replace(',', ''))
            price_per = float(m.group(7).replace(',', ''))
            amount_net = float(m.group(8).replace(',', ''))

            desc = ''
            if i + 1 < len(lines):
                desc = lines[i + 1].strip()

            items.append({
                'line_number': int(m.group(1)),
                'sku': m.group(2).strip(),
                'description': desc,
                'qty_ordered': qty_ordered,
                'quote_qty': qty_ordered,
                'price_per': price_per,
                'total_net_price': amount_net,
                'stock_ns': '',
            })
            i += 2
        else:
            i += 1

    if not items:
        return jsonify({'error': 'No line items found in PDF. Check that the format matches a supplier quote.'}), 400

    return jsonify({'items': items, 'count': len(items)})

# ─── Excel Export ───────────────────────────────────────────────

@app.route('/api/job/<int:job_id>/export')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def export_excel(job_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn = get_db()
    data = get_job_data(conn, job_id)
    conn.close()
    if not data:
        return 'Job not found', 404

    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    editable_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    computed_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    amber_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    money_format = '#,##0.00'

    items = data['line_items']

    ws1 = wb.active
    ws1.title = 'Master List'
    headers1 = ['Line #', 'Non-Stock', 'Product', 'Description', 'Quote QTY',
                 'Order QTY', 'Price', 'Unit', 'Quote Total', 'Order Total',
                 'QTY Received', 'Missing', 'QTY Shipped', 'QTY Invoiced', 'Notes']
    for c, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for r, item in enumerate(items, 2):
        pricing_type = item.get('pricing_type', 'each')
        unit_label = 'Each' if pricing_type == 'each' else ('Per C' if pricing_type == 'per_c' else 'Per M')
        vals = [
            item['line_number'], item['stock_ns'], item['sku'], item['description'],
            item['quote_qty'], item['qty_ordered'], item['price_per'], unit_label,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.fill = editable_fill
            cell.border = thin_border
            if c in (5, 6):
                cell.number_format = '#,##0'
            elif c == 7:
                cell.number_format = money_format

        # Quote Total (col 9)
        cell = ws1.cell(row=r, column=9)
        cell.value = item.get('quote_total', 0)
        cell.fill = computed_fill
        cell.border = thin_border
        cell.number_format = money_format

        # Order Total (col 10)
        cell = ws1.cell(row=r, column=10)
        cell.value = item['total_net_price']
        cell.fill = computed_fill
        cell.border = thin_border
        cell.number_format = money_format

        # Received (col 11)
        cell = ws1.cell(row=r, column=11)
        cell.value = f"='Received'!D{r}"
        cell.fill = green_fill
        cell.border = thin_border
        cell.number_format = '#,##0'

        # Missing (col 12) = Order QTY - Received
        cell = ws1.cell(row=r, column=12)
        missing = max(0, (item['qty_ordered'] or 0) - (item['total_received'] or 0))
        cell.value = missing
        cell.fill = amber_fill if missing > 0 else green_fill
        cell.border = thin_border
        cell.number_format = '#,##0'

        # Shipped (col 13)
        cell = ws1.cell(row=r, column=13)
        cell.value = f"='Shipped to Site'!D{r}"
        cell.fill = green_fill
        cell.border = thin_border
        cell.number_format = '#,##0'

        # Invoiced (col 14)
        cell = ws1.cell(row=r, column=14)
        cell.value = f"='Invoiced'!D{r}"
        cell.fill = green_fill
        cell.border = thin_border
        cell.number_format = '#,##0'

        # Notes (col 15)
        cell = ws1.cell(row=r, column=15)
        cell.value = item.get('notes', '')
        cell.border = thin_border

    widths1 = [8, 10, 15, 35, 12, 12, 12, 8, 14, 14, 12, 10, 12, 12, 20]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    tab_configs = [
        ('Received', 'received_entries'),
        ('Shipped to Site', 'shipped_entries'),
        ('Invoiced', 'invoiced_entries'),
    ]

    for tab_name, entry_key in tab_configs:
        ws = wb.create_sheet(title=tab_name)
        headers = ['Line #', 'Description', 'Total Ordered', 'Total']
        date_headers = {}
        for item in items:
            entries = item.get(entry_key, {})
            for col_num, entry in entries.items():
                if entry.get('entry_date') and col_num not in date_headers:
                    date_headers[col_num] = entry['entry_date']

        for c in range(1, 16):
            headers.append(date_headers.get(str(c), f'Col {c}'))

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for r, item in enumerate(items, 2):
            cell = ws.cell(row=r, column=1, value=item['line_number'])
            cell.fill = computed_fill
            cell.border = thin_border

            cell = ws.cell(row=r, column=2, value=item['description'])
            cell.fill = computed_fill
            cell.border = thin_border

            cell = ws.cell(row=r, column=3)
            cell.value = f"='Master List'!F{r}"
            cell.fill = computed_fill
            cell.border = thin_border

            cell = ws.cell(row=r, column=4)
            cell.value = f'=SUM(E{r}:S{r})'
            cell.border = thin_border
            cell.number_format = '#,##0'

            total_val = sum(
                (item.get(entry_key, {}).get(str(c), {}).get('quantity', 0) or 0)
                for c in range(1, 16)
            )
            ordered_val = item['qty_ordered']
            if ordered_val > 0 and total_val >= ordered_val:
                cell.fill = green_fill
            elif total_val > 0:
                cell.fill = amber_fill
            else:
                cell.fill = computed_fill

            entries = item.get(entry_key, {})
            for col in range(1, 16):
                cell = ws.cell(row=r, column=col + 4)
                entry = entries.get(str(col), {})
                qty = entry.get('quantity', 0) or 0
                if qty:
                    cell.value = qty
                cell.fill = editable_fill
                cell.border = thin_border

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 10
        for col in range(5, 20):
            ws.column_dimensions[get_column_letter(col)].width = 12

    export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
    os.makedirs(export_dir, exist_ok=True)
    safe_name = "".join(c if c.isalnum() or c in ' -_' else '' for c in data['job']['name'])
    filepath = os.path.join(export_dir, f'{safe_name}.xlsx')
    wb.save(filepath)

    return send_file(filepath, as_attachment=True,
                     download_name=f"{safe_name}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─── Accounting ─────────────────────────────────────────────────

@app.route('/accounting')
@role_required('owner', 'admin')
def accounting_overview():
    conn = get_db()
    jobs = conn.execute('SELECT * FROM jobs ORDER BY name').fetchall()
    conn.close()
    return render_template('accounting/overview.html', jobs=jobs)

@app.route('/accounting/job/<int:job_id>')
@role_required('owner', 'admin')
def accounting_job(job_id):
    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    conn.close()
    if not job:
        return 'Job not found', 404
    return render_template('accounting/job.html', job=job)

@app.route('/api/accounting/job/<int:job_id>')
@api_role_required('owner', 'admin')
def api_accounting_job(job_id):
    conn = get_db()
    expenses = conn.execute('SELECT * FROM expenses WHERE job_id = ? ORDER BY expense_date DESC', (job_id,)).fetchall()
    payments = conn.execute('SELECT * FROM payments WHERE job_id = ? ORDER BY payment_date DESC', (job_id,)).fetchall()
    invoices = conn.execute('SELECT * FROM client_invoices WHERE job_id = ? ORDER BY issue_date DESC', (job_id,)).fetchall()

    # Material costs
    items = conn.execute(
        'SELECT total_net_price, qty_ordered, price_per FROM line_items WHERE job_id = ?', (job_id,)
    ).fetchall()
    material_cost = sum(
        (row['total_net_price'] or 0) if (row['total_net_price'] or 0)
        else (row['qty_ordered'] or 0) * (row['price_per'] or 0)
        for row in items
    )

    conn.close()
    return jsonify({
        'expenses': [dict(e) for e in expenses],
        'payments': [dict(p) for p in payments],
        'invoices': [dict(i) for i in invoices],
        'material_cost': round(material_cost, 2),
        'total_expenses': round(sum(e['amount'] or 0 for e in expenses), 2),
        'total_payments': round(sum(p['amount'] or 0 for p in payments), 2),
        'total_invoiced': round(sum(i['amount'] or 0 for i in invoices), 2),
    })

@app.route('/api/accounting/expenses', methods=['POST'])
@api_role_required('owner', 'admin')
def create_expense():
    data = request.get_json()
    conn = get_db()
    conn.execute(
        'INSERT INTO expenses (job_id, category, vendor, description, amount, expense_date, created_by) VALUES (?,?,?,?,?,?,?)',
        (data['job_id'], data.get('category',''), data.get('vendor',''), data.get('description',''),
         float(data.get('amount', 0)), data.get('expense_date', datetime.now().strftime('%Y-%m-%d')),
         session.get('user_id'))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/accounting/expenses/<int:eid>', methods=['DELETE'])
@api_role_required('owner', 'admin')
def delete_expense(eid):
    conn = get_db()
    conn.execute('DELETE FROM expenses WHERE id = ?', (eid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/accounting/payments', methods=['POST'])
@api_role_required('owner', 'admin')
def create_payment():
    data = request.get_json()
    conn = get_db()
    conn.execute(
        'INSERT INTO payments (job_id, amount, payment_method, reference_number, description, payment_date, created_by) VALUES (?,?,?,?,?,?,?)',
        (data['job_id'], float(data.get('amount', 0)), data.get('payment_method',''),
         data.get('reference_number',''), data.get('description',''),
         data.get('payment_date', datetime.now().strftime('%Y-%m-%d')), session.get('user_id'))
    )
    conn.commit()
    # Lien waiver auto-prompt: check if lien waiver exists for this job
    job_id_val = data.get('job_id')
    if job_id_val:
        conn2 = get_db()
        lw = conn2.execute("SELECT id FROM lien_waivers WHERE job_id = ? AND status IN ('Draft','Sent','Executed')", (job_id_val,)).fetchone()
        if not lw:
            job_row = conn2.execute('SELECT name FROM jobs WHERE id = ?', (job_id_val,)).fetchone()
            job_label = job_row['name'] if job_row else f'Job #{job_id_val}'
            users_to_notify = conn2.execute("SELECT id FROM users WHERE role IN ('owner','admin','project_manager') AND is_active = 1").fetchall()
            conn2.close()
            for u in users_to_notify:
                create_notification(
                    u['id'], 'system',
                    f'Lien Waiver Needed: {job_label}',
                    f'Payment received for {job_label}. Please create a lien waiver.',
                    f'/lien-waivers'
                )
        else:
            conn2.close()
    # Team Pay auto-prompt: if job has a team pay schedule, notify owners
    if job_id_val:
        conn3 = get_db()
        tp_sched = conn3.execute('SELECT id FROM team_pay_schedules WHERE job_id = ?', (job_id_val,)).fetchone()
        if tp_sched:
            job_row3 = conn3.execute('SELECT name FROM jobs WHERE id = ?', (job_id_val,)).fetchone()
            job_label3 = job_row3['name'] if job_row3 else f'Job #{job_id_val}'
            owners = conn3.execute("SELECT id FROM users WHERE role = 'owner' AND is_active = 1").fetchall()
            conn3.close()
            for u in owners:
                create_notification(
                    u['id'], 'system',
                    f'Team Pay: Payment for {job_label3}',
                    f'Payment received for {job_label3}. Distribute to team members.',
                    f'/team-pay/job/{tp_sched["id"]}'
                )
        else:
            conn3.close()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/accounting/payments/<int:pid>', methods=['DELETE'])
@api_role_required('owner', 'admin')
def delete_payment(pid):
    conn = get_db()
    conn.execute('DELETE FROM payments WHERE id = ?', (pid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/accounting/invoices', methods=['POST'])
@api_role_required('owner', 'admin')
def create_invoice():
    data = request.get_json()
    conn = get_db()
    conn.execute(
        '''INSERT INTO client_invoices (job_id, invoice_number, amount, status, description, issue_date, due_date, created_by)
           VALUES (?,?,?,?,?,?,?,?)''',
        (data['job_id'], data.get('invoice_number',''), float(data.get('amount', 0)),
         data.get('status', 'Draft'), data.get('description',''),
         data.get('issue_date', datetime.now().strftime('%Y-%m-%d')),
         data.get('due_date',''), session.get('user_id'))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/accounting/invoices/<int:iid>', methods=['PUT'])
@api_role_required('owner', 'admin')
def update_invoice(iid):
    data = request.get_json()
    conn = get_db()
    fields = []
    values = []
    for f in ('status', 'amount', 'invoice_number', 'description', 'due_date', 'paid_date'):
        if f in data:
            fields.append(f'{f} = ?')
            values.append(data[f])
    if fields:
        values.append(iid)
        conn.execute(f"UPDATE client_invoices SET {', '.join(fields)} WHERE id = ?", values)
        # Calculate days_to_pay if status changed to Paid
        if data.get('status') == 'Paid':
            inv = conn.execute('SELECT issue_date, paid_date FROM client_invoices WHERE id = ?', (iid,)).fetchone()
            if inv and inv['issue_date'] and inv['paid_date']:
                try:
                    issue = datetime.strptime(inv['issue_date'], '%Y-%m-%d')
                    paid = datetime.strptime(inv['paid_date'], '%Y-%m-%d')
                    days = (paid - issue).days
                    conn.execute('UPDATE client_invoices SET days_to_pay = ? WHERE id = ?', (days, iid))
                except ValueError:
                    pass
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/accounting/invoices/<int:iid>', methods=['DELETE'])
@api_role_required('owner', 'admin')
def delete_invoice(iid):
    conn = get_db()
    conn.execute('DELETE FROM client_invoices WHERE id = ?', (iid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Daily Log (Crew Tracking) ─────────────────────────────────

@app.route('/daily-log')
@role_required('owner', 'admin', 'project_manager')
def daily_log_page():
    return render_template('daily_log.html')

@app.route('/api/daily-log')
@api_role_required('owner', 'admin', 'project_manager')
def api_get_daily_log():
    log_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    db = get_db()
    # Get all active employees
    employees = db.execute("""
        SELECT u.id, u.display_name, u.hourly_rate, u.role,
               COALESCE(ep.employment_status, 'Active') as status
        FROM users u
        LEFT JOIN employee_profiles ep ON ep.user_id = u.id
        WHERE u.is_active = 1
        ORDER BY u.display_name
    """).fetchall()
    # Get logs for this date
    logs = db.execute("""
        SELECT dl.id, dl.user_id, dl.job_id, dl.hours, dl.notes, dl.time_entry_id,
               j.name as job_name
        FROM daily_logs dl
        JOIN jobs j ON j.id = dl.job_id
        WHERE dl.log_date = ?
        ORDER BY dl.user_id
    """, (log_date,)).fetchall()
    # Get active jobs
    jobs = db.execute("""
        SELECT id, name FROM jobs
        WHERE status IN ('In Progress', 'Needs Bid', 'Bid Complete')
        ORDER BY name
    """).fetchall()
    return jsonify({
        'date': log_date,
        'employees': [dict(e) for e in employees],
        'logs': [dict(l) for l in logs],
        'jobs': [dict(j) for j in jobs]
    })

@app.route('/api/daily-log', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_save_daily_log():
    """Save daily log entries and sync to time_entries for payroll."""
    data = request.get_json()
    log_date = data.get('date')
    entries = data.get('entries', [])
    if not log_date:
        return jsonify({'error': 'Date required'}), 400

    db = get_db()
    user_id = session['user_id']
    saved = 0
    removed = 0

    # Get existing logs for this date to track what to remove
    existing = db.execute("SELECT id, user_id, job_id, time_entry_id FROM daily_logs WHERE log_date = ?", (log_date,)).fetchall()
    existing_keys = {(r['user_id'], r['job_id']): {'id': int(r['id']), 'user_id': r['user_id'], 'job_id': r['job_id'], 'time_entry_id': int(r['time_entry_id']) if r['time_entry_id'] else None} for r in existing}
    incoming_keys = set()

    for entry in entries:
        emp_id = entry.get('user_id')
        job_id = entry.get('job_id')
        hours = float(entry.get('hours', 0))
        notes = entry.get('notes', '')
        if not emp_id or not job_id or hours <= 0:
            continue
        incoming_keys.add((emp_id, job_id))

        # Get employee rate
        emp = db.execute("SELECT hourly_rate FROM users WHERE id = ?", (emp_id,)).fetchone()
        rate = emp['hourly_rate'] if emp else 0

        key = (emp_id, job_id)
        if key in existing_keys:
            # Update existing log
            row = existing_keys[key]
            db.execute("""
                UPDATE daily_logs SET hours = ?, notes = ?, updated_at = datetime('now','localtime')
                WHERE id = ?
            """, (hours, notes, row['id']))
            # Update linked time entry
            if row['time_entry_id']:
                db.execute("""
                    UPDATE time_entries SET hours = ?, hourly_rate = ?, description = ?
                    WHERE id = ?
                """, (hours, rate, notes or 'Daily log', row['time_entry_id']))
            else:
                # Create time entry
                cur = db.execute("""
                    INSERT INTO time_entries (user_id, job_id, hours, hourly_rate, work_date, description, approved)
                    VALUES (?, ?, ?, ?, ?, ?, 0)
                """, (emp_id, job_id, hours, rate, log_date, notes or 'Daily log'))
                db.execute("UPDATE daily_logs SET time_entry_id = ? WHERE id = ?", (cur.lastrowid, row['id']))
        else:
            # Create time entry first
            cur = db.execute("""
                INSERT INTO time_entries (user_id, job_id, hours, hourly_rate, work_date, description, approved)
                VALUES (?, ?, ?, ?, ?, ?, 0)
            """, (emp_id, job_id, hours, rate, log_date, notes or 'Daily log'))
            te_id = cur.lastrowid
            # Create daily log
            db.execute("""
                INSERT INTO daily_logs (log_date, user_id, job_id, hours, notes, time_entry_id, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (log_date, emp_id, job_id, hours, notes, te_id, user_id))
        saved += 1

    # Remove logs that were deleted (not in incoming)
    for key, row in existing_keys.items():
        if key not in incoming_keys:
            te_id = row['time_entry_id']
            # Delete the daily_log row first (has FK to time_entries)
            db.execute("DELETE FROM daily_logs WHERE id = ?", (row['id'],))
            # Then delete the time entry if it hasn't been approved/added to payroll
            if te_id:
                te = db.execute("SELECT approved, payroll_run_id FROM time_entries WHERE id = ?", (te_id,)).fetchone()
                if te and not te['approved'] and not te['payroll_run_id']:
                    db.execute("DELETE FROM time_entries WHERE id = ?", (te_id,))
            removed += 1

    db.commit()
    return jsonify({'ok': True, 'saved': saved, 'removed': removed})

@app.route('/api/daily-log/summary')
@api_role_required('owner', 'admin', 'project_manager')
def api_daily_log_summary():
    """Get summary of crew assignments for a date range (for dashboard)."""
    start = request.args.get('start', datetime.now().strftime('%Y-%m-%d'))
    end = request.args.get('end', start)
    db = get_db()
    summary = db.execute("""
        SELECT dl.log_date, COUNT(DISTINCT dl.user_id) as crew_count,
               SUM(dl.hours) as total_hours,
               COUNT(DISTINCT dl.job_id) as job_count
        FROM daily_logs dl
        WHERE dl.log_date BETWEEN ? AND ?
        GROUP BY dl.log_date
        ORDER BY dl.log_date
    """, (start, end)).fetchall()
    return jsonify([dict(s) for s in summary])

# ─── Payroll / Time Entry ───────────────────────────────────────

@app.route('/payroll')
@role_required('owner', 'admin')
def payroll_overview():
    return render_template('payroll/overview.html')

@app.route('/payroll/employee/<int:user_id>')
@role_required('owner', 'admin')
def payroll_employee(user_id):
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    conn.close()
    if not user:
        return 'Employee not found', 404
    return render_template('payroll/employee.html', employee=user)

@app.route('/time-entry')
@login_required
def time_entry():
    conn = get_db()
    jobs = conn.execute('SELECT id, name FROM jobs ORDER BY name').fetchall()
    conn.close()
    return render_template('payroll/time_entry.html', jobs=jobs)

@app.route('/api/payroll/summary')
@api_role_required('owner', 'admin')
def api_payroll_summary():
    conn = get_db()
    users = conn.execute(
        '''SELECT u.*, ep.employee_number FROM users u
           LEFT JOIN employee_profiles ep ON u.id = ep.user_id
           WHERE u.is_active = 1 ORDER BY u.display_name'''
    ).fetchall()
    result = []
    for u in users:
        entries = conn.execute(
            'SELECT COALESCE(SUM(hours), 0) as total_hours, COALESCE(SUM(hours * hourly_rate), 0) as total_pay FROM time_entries WHERE user_id = ?',
            (u['id'],)
        ).fetchone()
        pending = conn.execute(
            'SELECT COALESCE(SUM(hours), 0) FROM time_entries WHERE user_id = ? AND approved = 0',
            (u['id'],)
        ).fetchone()[0]
        result.append({
            'id': u['id'],
            'display_name': u['display_name'],
            'first_name': u['first_name'] or '',
            'last_name': u['last_name'] or '',
            'home_base_city': u['home_base_city'] or '',
            'username': u['username'],
            'role': u['role'],
            'hourly_rate': u['hourly_rate'] or 0,
            'email': u['email'] or '',
            'phone': u['phone'] or '',
            'employee_number': u['employee_number'] or '',
            'total_hours': round(entries['total_hours'], 1),
            'total_pay': round(entries['total_pay'], 2),
            'pending_hours': round(pending, 1),
        })
    conn.close()
    return jsonify(result)

@app.route('/api/payroll/employee/<int:user_id>')
@api_role_required('owner', 'admin')
def api_payroll_employee(user_id):
    conn = get_db()
    entries = conn.execute(
        '''SELECT te.*, j.name as job_name FROM time_entries te
           LEFT JOIN jobs j ON te.job_id = j.id
           WHERE te.user_id = ? ORDER BY te.work_date DESC''',
        (user_id,)
    ).fetchall()
    conn.close()
    return jsonify([dict(e) for e in entries])

@app.route('/api/time-entries', methods=['GET'])
@api_login_required
def get_time_entries():
    conn = get_db()
    user_id = session['user_id']
    if session.get('role') in ('owner', 'admin'):
        user_id = request.args.get('user_id', user_id, type=int)
    entries = conn.execute(
        '''SELECT te.*, j.name as job_name FROM time_entries te
           LEFT JOIN jobs j ON te.job_id = j.id
           WHERE te.user_id = ? ORDER BY te.work_date DESC LIMIT 100''',
        (user_id,)
    ).fetchall()
    conn.close()
    return jsonify([dict(e) for e in entries])

@app.route('/api/time-entries', methods=['POST'])
@api_login_required
def create_time_entry():
    data = request.get_json()
    user_id = session['user_id']
    conn = get_db()
    user = conn.execute('SELECT hourly_rate FROM users WHERE id = ?', (user_id,)).fetchone()
    rate = float(data.get('hourly_rate', 0)) or (user['hourly_rate'] or 0) if user else 0
    conn.execute(
        'INSERT INTO time_entries (user_id, job_id, hours, hourly_rate, work_date, description) VALUES (?,?,?,?,?,?)',
        (user_id, data['job_id'], float(data.get('hours', 0)), rate,
         data.get('work_date', datetime.now().strftime('%Y-%m-%d')), data.get('description', ''))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/time-entries/<int:tid>/approve', methods=['POST'])
@api_role_required('owner', 'admin')
def approve_time_entry(tid):
    conn = get_db()
    conn.execute('UPDATE time_entries SET approved = 1, approved_by = ? WHERE id = ?',
                 (session['user_id'], tid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/time-entries/<int:tid>', methods=['DELETE'])
@api_login_required
def delete_time_entry(tid):
    conn = get_db()
    entry = conn.execute('SELECT user_id FROM time_entries WHERE id = ?', (tid,)).fetchone()
    if not entry:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if entry['user_id'] != session['user_id'] and session.get('role') not in ('owner', 'admin'):
        conn.close()
        return jsonify({'error': 'Access denied'}), 403
    conn.execute('DELETE FROM time_entries WHERE id = ?', (tid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Payroll Entry Management (owner/admin managing employee time) ────

@app.route('/api/payroll/entries', methods=['POST'])
@api_role_required('owner', 'admin')
def api_payroll_create_entry():
    """Create a time entry for any employee."""
    data = request.get_json(force=True)
    user_id = data.get('user_id')
    job_id = data.get('job_id')
    if not user_id or not job_id:
        return jsonify({'error': 'user_id and job_id required'}), 400
    conn = get_db()
    rate = float(data.get('hourly_rate', 0))
    if not rate:
        user = conn.execute('SELECT hourly_rate FROM users WHERE id = ?', (user_id,)).fetchone()
        rate = user['hourly_rate'] or 0 if user else 0
    conn.execute(
        '''INSERT INTO time_entries (user_id, job_id, hours, hourly_rate, work_date, description, entry_type)
           VALUES (?,?,?,?,?,?,?)''',
        (user_id, job_id, float(data.get('hours', 0)), rate,
         data.get('work_date', datetime.now().strftime('%Y-%m-%d')),
         data.get('description', ''), data.get('entry_type', 'regular'))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/payroll/entries/<int:tid>', methods=['PUT'])
@api_role_required('owner', 'admin')
def api_payroll_update_entry(tid):
    """Update a time entry."""
    data = request.get_json(force=True)
    conn = get_db()
    entry = conn.execute('SELECT id FROM time_entries WHERE id = ?', (tid,)).fetchone()
    if not entry:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    fields = []
    values = []
    for f in ('job_id', 'hours', 'hourly_rate', 'work_date', 'description', 'entry_type'):
        if f in data:
            fields.append(f'{f} = ?')
            values.append(data[f])
    if fields:
        values.append(tid)
        conn.execute(f"UPDATE time_entries SET {', '.join(fields)} WHERE id = ?", values)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/payroll/entries/<int:tid>', methods=['DELETE'])
@api_role_required('owner', 'admin')
def api_payroll_delete_entry(tid):
    """Delete a time entry."""
    conn = get_db()
    conn.execute('DELETE FROM time_entries WHERE id = ?', (tid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/time-entries/week')
@api_login_required
def api_time_entries_week():
    """Get weekly grid data for a user."""
    from datetime import datetime, timedelta
    date_str = request.args.get('date', '')
    user_id = request.args.get('user_id', session.get('user_id'))
    # Only owners can view others' time
    if str(user_id) != str(session.get('user_id')) and session.get('role') != 'owner':
        return jsonify({'error': 'Access denied'}), 403
    try:
        target = datetime.strptime(date_str, '%Y-%m-%d') if date_str else datetime.now()
    except ValueError:
        target = datetime.now()
    # Get Monday of the week
    monday = target - timedelta(days=target.weekday())
    sunday = monday + timedelta(days=6)
    conn = get_db()
    entries = conn.execute(
        '''SELECT te.*, j.name as job_name FROM time_entries te
           LEFT JOIN jobs j ON te.job_id = j.id
           WHERE te.user_id = ? AND te.work_date BETWEEN ? AND ?
           ORDER BY j.name, te.work_date''',
        (user_id, monday.strftime('%Y-%m-%d'), sunday.strftime('%Y-%m-%d'))
    ).fetchall()
    jobs = conn.execute("SELECT id, name FROM jobs WHERE status IN ('In Progress','Awarded','Pre-Construction') ORDER BY name").fetchall()
    conn.close()
    # Build grid: {job_id: {date: hours}}
    grid = {}
    for e in entries:
        jid = e['job_id']
        if jid not in grid:
            grid[jid] = {'job_id': jid, 'job_name': e['job_name'], 'entries': {}}
        grid[jid]['entries'][e['work_date']] = {
            'id': e['id'], 'hours': e['hours'], 'entry_type': e['entry_type'] or 'regular',
            'description': e['description'] or ''
        }
    dates = [(monday + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(7)]
    return jsonify({
        'week_start': monday.strftime('%Y-%m-%d'),
        'week_end': sunday.strftime('%Y-%m-%d'),
        'dates': dates,
        'grid': list(grid.values()),
        'available_jobs': [{'id': j['id'], 'name': j['name']} for j in jobs],
    })

@app.route('/api/time-entries/batch', methods=['POST'])
@api_login_required
def api_time_entries_batch():
    """Bulk save a week of time entries."""
    data = request.get_json()
    user_id = data.get('user_id', session.get('user_id'))
    if str(user_id) != str(session.get('user_id')) and session.get('role') != 'owner':
        return jsonify({'error': 'Access denied'}), 403
    entries = data.get('entries', [])
    conn = get_db()
    user = conn.execute('SELECT hourly_rate FROM users WHERE id = ?', (user_id,)).fetchone()
    rate = user['hourly_rate'] if user else 0
    for entry in entries:
        job_id = entry.get('job_id')
        work_date = entry.get('work_date')
        hours = float(entry.get('hours', 0) or 0)
        entry_type = entry.get('entry_type', 'regular')
        desc = entry.get('description', '')
        if hours <= 0:
            conn.execute('DELETE FROM time_entries WHERE user_id = ? AND job_id = ? AND work_date = ?',
                         (user_id, job_id, work_date))
        else:
            existing = conn.execute('SELECT id FROM time_entries WHERE user_id = ? AND job_id = ? AND work_date = ?',
                                    (user_id, job_id, work_date)).fetchone()
            if existing:
                conn.execute('''UPDATE time_entries SET hours = ?, hourly_rate = ?, entry_type = ?, description = ?
                                WHERE id = ?''', (hours, rate, entry_type, desc, existing['id']))
            else:
                conn.execute('''INSERT INTO time_entries (user_id, job_id, hours, hourly_rate, work_date, entry_type, description)
                                VALUES (?,?,?,?,?,?,?)''', (user_id, job_id, hours, rate, work_date, entry_type, desc))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/payroll/period-summary')
@api_role_required('owner', 'admin')
def api_payroll_period_summary():
    """Summary by employee for a date range."""
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    if not start or not end:
        return jsonify({'error': 'start and end required'}), 400
    conn = get_db()
    rows = conn.execute(
        '''SELECT u.id, u.display_name, u.hourly_rate,
                  SUM(te.hours) as total_hours,
                  SUM(CASE WHEN te.entry_type = 'overtime' THEN te.hours ELSE 0 END) as ot_hours,
                  COUNT(DISTINCT te.work_date) as days_worked,
                  COUNT(DISTINCT te.job_id) as jobs_worked
           FROM time_entries te JOIN users u ON te.user_id = u.id
           WHERE te.work_date BETWEEN ? AND ?
           GROUP BY u.id ORDER BY u.display_name''', (start, end)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/payroll/approve-period', methods=['POST'])
@api_role_required('owner', 'admin')
def api_payroll_approve_period():
    """Bulk approve all entries in a period."""
    data = request.get_json()
    start = data.get('start', '')
    end = data.get('end', '')
    if not start or not end:
        return jsonify({'error': 'start and end required'}), 400
    conn = get_db()
    conn.execute(
        '''UPDATE time_entries SET approved = 1, approved_by = ?
           WHERE work_date BETWEEN ? AND ? AND approved = 0''',
        (session.get('user_id'), start, end)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Payroll Runs ──────────────────────────────────────────────

@app.route('/payroll/runs/<int:run_id>')
@role_required('owner', 'admin')
def payroll_run_detail(run_id):
    conn = get_db()
    run = conn.execute('SELECT * FROM payroll_runs WHERE id = ?', (run_id,)).fetchone()
    conn.close()
    if not run:
        return 'Payroll run not found', 404
    return render_template('payroll/run_detail.html', run=run)

@app.route('/api/payroll/runs', methods=['GET'])
@api_role_required('owner', 'admin')
def api_payroll_runs_list():
    conn = get_db()
    runs = conn.execute(
        '''SELECT pr.*, u.display_name as created_by_name,
                  (SELECT COUNT(*) FROM payroll_run_employees WHERE payroll_run_id = pr.id) as employee_count
           FROM payroll_runs pr
           LEFT JOIN users u ON pr.created_by = u.id
           ORDER BY pr.created_at DESC'''
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in runs])

@app.route('/api/payroll/runs', methods=['POST'])
@api_role_required('owner', 'admin')
def api_payroll_runs_create():
    data = request.get_json(force=True)
    period_start = data.get('period_start', '')
    period_end = data.get('period_end', '')
    check_date = data.get('check_date', '')
    notes = data.get('notes', '')
    employee_ids = data.get('employee_ids', [])
    if not period_start or not period_end:
        return jsonify({'error': 'Period start and end dates are required'}), 400
    if not employee_ids:
        return jsonify({'error': 'At least one employee must be selected'}), 400
    conn = get_db()
    # Auto-increment run_number
    last = conn.execute('SELECT MAX(run_number) as mx FROM payroll_runs').fetchone()
    run_number = (last['mx'] or 0) + 1
    cur = conn.execute(
        '''INSERT INTO payroll_runs (run_number, period_start, period_end, check_date, notes, created_by)
           VALUES (?,?,?,?,?,?)''',
        (run_number, period_start, period_end, check_date, notes, session.get('user_id'))
    )
    run_id = cur.lastrowid
    # Add employees with their current hourly rate
    for uid in employee_ids:
        user = conn.execute('SELECT hourly_rate FROM users WHERE id = ?', (uid,)).fetchone()
        rate = user['hourly_rate'] if user else 0
        conn.execute(
            '''INSERT INTO payroll_run_employees (payroll_run_id, user_id, hourly_rate)
               VALUES (?,?,?)''',
            (run_id, uid, rate)
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': run_id}), 201

@app.route('/api/payroll/runs/<int:run_id>', methods=['GET'])
@api_role_required('owner', 'admin')
def api_payroll_run_detail(run_id):
    conn = get_db()
    run = conn.execute('SELECT * FROM payroll_runs WHERE id = ?', (run_id,)).fetchone()
    if not run:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    # Get employees in this run
    employees = conn.execute(
        '''SELECT pre.*, u.display_name, u.role FROM payroll_run_employees pre
           JOIN users u ON pre.user_id = u.id
           WHERE pre.payroll_run_id = ? ORDER BY u.display_name''',
        (run_id,)
    ).fetchall()
    # Get all time entries for these employees in the pay period
    emp_ids = [e['user_id'] for e in employees]
    entries = []
    if emp_ids:
        placeholders = ','.join('?' * len(emp_ids))
        entries = conn.execute(
            f'''SELECT te.*, j.name as job_name FROM time_entries te
                LEFT JOIN jobs j ON te.job_id = j.id
                WHERE te.user_id IN ({placeholders})
                AND te.work_date BETWEEN ? AND ?
                ORDER BY te.user_id, te.work_date''',
            emp_ids + [run['period_start'], run['period_end']]
        ).fetchall()
    # Build dates array for the period
    from datetime import datetime as dt, timedelta
    start = dt.strptime(run['period_start'], '%Y-%m-%d')
    end = dt.strptime(run['period_end'], '%Y-%m-%d')
    dates = []
    current = start
    while current <= end:
        dates.append(current.strftime('%Y-%m-%d'))
        current += timedelta(days=1)
    # Available jobs
    jobs = conn.execute("SELECT id, name FROM jobs ORDER BY name").fetchall()
    conn.close()
    return jsonify({
        'run': dict(run),
        'employees': [dict(e) for e in employees],
        'entries': [dict(e) for e in entries],
        'dates': dates,
        'available_jobs': [{'id': j['id'], 'name': j['name']} for j in jobs],
    })

@app.route('/api/payroll/runs/<int:run_id>', methods=['PUT'])
@api_role_required('owner', 'admin')
def api_payroll_run_update(run_id):
    conn = get_db()
    run = conn.execute('SELECT * FROM payroll_runs WHERE id = ?', (run_id,)).fetchone()
    if not run:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if run['status'] != 'Draft':
        conn.close()
        return jsonify({'error': 'Cannot edit a finalized run'}), 400
    data = request.get_json(force=True)
    # Update basic fields
    for field in ('period_start', 'period_end', 'check_date', 'notes'):
        if field in data:
            conn.execute(f"UPDATE payroll_runs SET {field} = ? WHERE id = ?", (data[field], run_id))
    # Add/remove employees
    if 'add_employee_ids' in data:
        for uid in data['add_employee_ids']:
            user = conn.execute('SELECT hourly_rate FROM users WHERE id = ?', (uid,)).fetchone()
            rate = user['hourly_rate'] if user else 0
            conn.execute(
                'INSERT OR IGNORE INTO payroll_run_employees (payroll_run_id, user_id, hourly_rate) VALUES (?,?,?)',
                (run_id, uid, rate)
            )
    if 'remove_employee_ids' in data:
        for uid in data['remove_employee_ids']:
            conn.execute(
                'DELETE FROM payroll_run_employees WHERE payroll_run_id = ? AND user_id = ?',
                (run_id, uid)
            )
    conn.execute("UPDATE payroll_runs SET updated_at = datetime('now','localtime') WHERE id = ?", (run_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/payroll/runs/<int:run_id>', methods=['DELETE'])
@api_role_required('owner', 'admin')
def api_payroll_run_delete(run_id):
    data = request.get_json(force=True) if request.data else {}
    password = data.get('password', '')
    if not password:
        return jsonify({'error': 'Password is required to delete a payroll run'}), 400
    # Verify password
    conn = get_db()
    user = conn.execute('SELECT password_hash FROM users WHERE id = ?', (session.get('user_id'),)).fetchone()
    if not user or not check_password_hash(user['password_hash'], password):
        conn.close()
        return jsonify({'error': 'Incorrect password'}), 403
    run = conn.execute('SELECT status FROM payroll_runs WHERE id = ?', (run_id,)).fetchone()
    if not run:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    # Unlink any time entries
    conn.execute('UPDATE time_entries SET payroll_run_id = NULL WHERE payroll_run_id = ?', (run_id,))
    conn.execute('DELETE FROM payroll_run_employees WHERE payroll_run_id = ?', (run_id,))
    conn.execute('DELETE FROM payroll_runs WHERE id = ?', (run_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/payroll/runs/<int:run_id>/timesheet', methods=['POST'])
@api_role_required('owner', 'admin')
def api_payroll_run_timesheet(run_id):
    """Mass save all employees' time entries for a payroll run."""
    conn = get_db()
    run = conn.execute('SELECT * FROM payroll_runs WHERE id = ?', (run_id,)).fetchone()
    if not run:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if run['status'] != 'Draft':
        conn.close()
        return jsonify({'error': 'Cannot edit a finalized run'}), 400
    data = request.get_json(force=True)
    entries = data.get('entries', [])
    # entries: [{user_id, job_id OR job_name, work_date, hours}, ...]
    # Get employees in this run for rate lookup
    emp_rows = conn.execute(
        'SELECT user_id, hourly_rate FROM payroll_run_employees WHERE payroll_run_id = ?', (run_id,)
    ).fetchall()
    emp_rates = {e['user_id']: e['hourly_rate'] for e in emp_rows}
    # Cache for resolving job names → IDs (and auto-creating new jobs)
    job_name_cache = {}
    for j in conn.execute('SELECT id, name FROM jobs').fetchall():
        job_name_cache[j['name'].strip().lower()] = j['id']
    new_jobs = []  # track newly created jobs to return to frontend
    for entry in entries:
        uid = int(entry.get('user_id', 0))
        work_date = entry.get('work_date', '')
        hours = float(entry.get('hours', 0) or 0)
        # Resolve job_id from either job_id or job_name
        job_id = int(entry.get('job_id', 0) or 0)
        job_name = entry.get('job_name', '').strip()
        if not job_id and job_name:
            key = job_name.lower()
            if key in job_name_cache:
                job_id = job_name_cache[key]
            else:
                # Auto-create job
                cur = conn.execute(
                    "INSERT INTO jobs (name, status) VALUES (?, 'In Progress')",
                    (job_name,)
                )
                job_id = cur.lastrowid
                job_name_cache[key] = job_id
                new_jobs.append({'id': job_id, 'name': job_name})
        if not uid or not job_id or not work_date:
            continue
        rate = emp_rates.get(uid, 0)
        if hours <= 0:
            conn.execute(
                '''DELETE FROM time_entries
                   WHERE user_id = ? AND job_id = ? AND work_date = ?
                   AND (payroll_run_id = ? OR payroll_run_id IS NULL)''',
                (uid, job_id, work_date, run_id)
            )
        else:
            existing = conn.execute(
                '''SELECT id FROM time_entries
                   WHERE user_id = ? AND job_id = ? AND work_date = ?''',
                (uid, job_id, work_date)
            ).fetchone()
            if existing:
                conn.execute(
                    '''UPDATE time_entries SET hours = ?, hourly_rate = ?, payroll_run_id = ?
                       WHERE id = ?''',
                    (hours, rate, run_id, existing['id'])
                )
            else:
                conn.execute(
                    '''INSERT INTO time_entries (user_id, job_id, hours, hourly_rate, work_date, payroll_run_id)
                       VALUES (?,?,?,?,?,?)''',
                    (uid, job_id, hours, rate, work_date, run_id)
                )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'new_jobs': new_jobs})

@app.route('/api/payroll/runs/<int:run_id>/copy-previous', methods=['GET'])
@api_role_required('owner', 'admin')
def api_payroll_run_copy_previous(run_id):
    from datetime import datetime as dt, timedelta
    conn = get_db()
    run = conn.execute('SELECT * FROM payroll_runs WHERE id = ?', (run_id,)).fetchone()
    if not run:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if run['status'] != 'Draft':
        conn.close()
        return jsonify({'error': 'Run is not in Draft status'}), 400
    # Find the most recent previous run
    prev_run = conn.execute(
        '''SELECT * FROM payroll_runs WHERE period_end < ? ORDER BY period_end DESC LIMIT 1''',
        (run['period_start'],)
    ).fetchone()
    if not prev_run:
        conn.close()
        return jsonify({'error': 'No previous payroll run found'}), 404
    # Get employees in the current run
    cur_emps = conn.execute(
        'SELECT user_id FROM payroll_run_employees WHERE payroll_run_id = ?', (run_id,)
    ).fetchall()
    cur_emp_ids = [e['user_id'] for e in cur_emps]
    if not cur_emp_ids:
        conn.close()
        return jsonify({'entries': []})
    # Get time entries from previous run for matching employees
    placeholders = ','.join('?' * len(cur_emp_ids))
    prev_entries = conn.execute(
        f'''SELECT te.user_id, te.work_date, te.hours, te.job_id, j.name as job_name
            FROM time_entries te
            LEFT JOIN jobs j ON te.job_id = j.id
            WHERE te.user_id IN ({placeholders})
            AND te.work_date BETWEEN ? AND ?
            ORDER BY te.user_id, te.work_date''',
        cur_emp_ids + [prev_run['period_start'], prev_run['period_end']]
    ).fetchall()
    # Map by day offset: prev day 0 → current day 0, etc.
    prev_start = dt.strptime(prev_run['period_start'], '%Y-%m-%d')
    cur_start = dt.strptime(run['period_start'], '%Y-%m-%d')
    cur_end = dt.strptime(run['period_end'], '%Y-%m-%d')
    num_days = (cur_end - cur_start).days + 1
    mapped = []
    for entry in prev_entries:
        entry_date = dt.strptime(entry['work_date'], '%Y-%m-%d')
        offset = (entry_date - prev_start).days
        if offset < 0 or offset >= num_days:
            continue
        new_date = (cur_start + timedelta(days=offset)).strftime('%Y-%m-%d')
        mapped.append({
            'user_id': entry['user_id'],
            'work_date': new_date,
            'hours': entry['hours'],
            'job_id': entry['job_id'],
            'job_name': entry['job_name'] or '',
        })
    conn.close()
    return jsonify({'entries': mapped, 'from_run': prev_run['run_number']})

@app.route('/api/payroll/runs/<int:run_id>/finalize', methods=['POST'])
@api_role_required('owner', 'admin')
def api_payroll_run_finalize(run_id):
    conn = get_db()
    run = conn.execute('SELECT * FROM payroll_runs WHERE id = ?', (run_id,)).fetchone()
    if not run:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if run['status'] != 'Draft':
        conn.close()
        return jsonify({'error': 'Already finalized'}), 400
    # Snapshot totals per employee
    employees = conn.execute(
        'SELECT * FROM payroll_run_employees WHERE payroll_run_id = ?', (run_id,)
    ).fetchall()
    grand_hours = 0
    grand_pay = 0
    for emp in employees:
        entries = conn.execute(
            '''SELECT COALESCE(SUM(hours), 0) as total_hours
               FROM time_entries
               WHERE user_id = ? AND work_date BETWEEN ? AND ?''',
            (emp['user_id'], run['period_start'], run['period_end'])
        ).fetchone()
        total_hours = entries['total_hours']
        # Calculate OT: anything over 40 per week would need weekly breakdown
        # For now, simple: all hours at regular rate
        reg_hours = total_hours
        ot_hours = 0
        gross = total_hours * emp['hourly_rate']
        conn.execute(
            '''UPDATE payroll_run_employees
               SET regular_hours = ?, overtime_hours = ?, total_hours = ?, gross_pay = ?
               WHERE id = ?''',
            (reg_hours, ot_hours, total_hours, gross, emp['id'])
        )
        grand_hours += total_hours
        grand_pay += gross
    # Stamp time entries with payroll_run_id and approve
    emp_ids = [e['user_id'] for e in employees]
    if emp_ids:
        placeholders = ','.join('?' * len(emp_ids))
        conn.execute(
            f'''UPDATE time_entries SET payroll_run_id = ?, approved = 1, approved_by = ?
                WHERE user_id IN ({placeholders})
                AND work_date BETWEEN ? AND ?
                AND (payroll_run_id IS NULL OR payroll_run_id = ?)''',
            [run_id, session.get('user_id')] + emp_ids + [run['period_start'], run['period_end'], run_id]
        )
    # Update run totals and status
    conn.execute(
        '''UPDATE payroll_runs
           SET status = 'Finalized', total_hours = ?, total_gross_pay = ?,
               finalized_by = ?, finalized_at = datetime('now','localtime'),
               updated_at = datetime('now','localtime')
           WHERE id = ?''',
        (grand_hours, grand_pay, session.get('user_id'), run_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/payroll/runs/<int:run_id>/reopen', methods=['POST'])
@api_role_required('owner')
def api_payroll_run_reopen(run_id):
    conn = get_db()
    run = conn.execute('SELECT status FROM payroll_runs WHERE id = ?', (run_id,)).fetchone()
    if not run:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if run['status'] != 'Finalized':
        conn.close()
        return jsonify({'error': 'Run is not finalized'}), 400
    conn.execute(
        '''UPDATE payroll_runs
           SET status = 'Draft', finalized_by = NULL, finalized_at = '',
               updated_at = datetime('now','localtime')
           WHERE id = ?''',
        (run_id,)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Employees (HR) ──────────────────────────────────────────────

def _get_fernet():
    """Get or create Fernet cipher for SSN encryption."""
    from cryptography.fernet import Fernet
    key_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'ssn.key')
    os.makedirs(os.path.dirname(key_path), exist_ok=True)
    if os.path.exists(key_path):
        with open(key_path, 'rb') as f:
            key = f.read()
    else:
        key = Fernet.generate_key()
        with open(key_path, 'wb') as f:
            f.write(key)
    return Fernet(key)

def _encrypt_ssn(ssn_plain):
    """Encrypt SSN and return (encrypted, last4)."""
    if not ssn_plain:
        return '', ''
    clean = ssn_plain.replace('-', '').replace(' ', '')
    last4 = clean[-4:] if len(clean) >= 4 else clean
    fernet = _get_fernet()
    encrypted = fernet.encrypt(clean.encode()).decode()
    return encrypted, last4

def _decrypt_ssn(ssn_encrypted):
    """Decrypt SSN and return formatted string."""
    if not ssn_encrypted:
        return ''
    fernet = _get_fernet()
    decrypted = fernet.decrypt(ssn_encrypted.encode()).decode()
    if len(decrypted) == 9:
        return f'{decrypted[:3]}-{decrypted[3:5]}-{decrypted[5:]}'
    return decrypted

@app.route('/employees')
@role_required('owner', 'admin')
def employees_list():
    return render_template('employees/list.html')

@app.route('/employees/<int:uid>')
@role_required('owner', 'admin')
def employees_detail(uid):
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (uid,)).fetchone()
    conn.close()
    if not user:
        return 'Employee not found', 404
    return render_template('employees/detail.html', employee=user)

@app.route('/api/employees')
@api_role_required('owner', 'admin')
def api_employees_list():
    status_filter = request.args.get('status', 'Active')
    conn = get_db()
    if status_filter == 'all':
        rows = conn.execute(
            '''SELECT u.*, ep.employee_number, ep.employment_status, ep.hire_date, ep.termination_date
               FROM users u LEFT JOIN employee_profiles ep ON u.id = ep.user_id
               ORDER BY u.display_name'''
        ).fetchall()
    else:
        rows = conn.execute(
            '''SELECT u.*, ep.employee_number, ep.employment_status, ep.hire_date, ep.termination_date
               FROM users u LEFT JOIN employee_profiles ep ON u.id = ep.user_id
               WHERE ep.employment_status = ? ORDER BY u.display_name''', (status_filter,)
        ).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d.pop('password_hash', None)
        result.append(d)
    return jsonify(result)

@app.route('/api/employees', methods=['POST'])
@api_role_required('owner', 'admin')
def api_employees_create():
    data = request.get_json()
    username = (data.get('username') or '').strip()
    password = data.get('password', '')
    first_name = (data.get('first_name') or '').strip()
    last_name = (data.get('last_name') or '').strip()
    display_name = (data.get('display_name') or '').strip() or f'{first_name} {last_name}'.strip()
    if not display_name:
        return jsonify({'error': 'First name is required'}), 400
    if username and not password:
        return jsonify({'error': 'Password is required when setting a username'}), 400

    conn = get_db()

    # If username provided, check for duplicates and create with login access
    if username:
        existing = conn.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
        if existing:
            conn.close()
            return jsonify({'error': 'Username already taken'}), 400
        pw_hash = generate_password_hash(password)
        must_change = 1
    else:
        # No login access — generate a unique placeholder username
        import uuid
        username = '_nologin_' + uuid.uuid4().hex[:8]
        pw_hash = generate_password_hash(uuid.uuid4().hex)
        must_change = 0

    cursor = conn.execute(
        '''INSERT INTO users (username, display_name, first_name, last_name, home_base_city,
           password_hash, role, email, phone, hourly_rate, must_change_password)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
        (username, display_name, first_name, last_name, data.get('home_base_city', ''),
         pw_hash,
         data.get('role', 'employee'), data.get('email', ''), data.get('phone', ''),
         float(data.get('hourly_rate', 0)), must_change)
    )
    uid = cursor.lastrowid

    # Create employee profile
    ssn_enc, ssn_last4 = _encrypt_ssn(data.get('ssn', ''))
    conn.execute(
        '''INSERT INTO employee_profiles (user_id, employee_number, ssn_encrypted, ssn_last4,
           date_of_birth, hire_date, employment_status, address_street, address_city, address_state,
           address_zip, shirt_size, emergency_contact_name, emergency_contact_phone,
           emergency_contact_relationship, notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (uid, data.get('employee_number', ''), ssn_enc, ssn_last4,
         data.get('date_of_birth', ''), data.get('hire_date', datetime.now().strftime('%Y-%m-%d')),
         'Active', data.get('address_street', ''), data.get('address_city', ''),
         data.get('address_state', ''), data.get('address_zip', ''), data.get('shirt_size', ''),
         data.get('emergency_contact_name', ''), data.get('emergency_contact_phone', ''),
         data.get('emergency_contact_relationship', ''), data.get('notes', ''))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': uid}), 201

@app.route('/api/employees/<int:uid>')
@api_role_required('owner', 'admin')
def api_employees_detail(uid):
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (uid,)).fetchone()
    if not user:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    profile = conn.execute('SELECT * FROM employee_profiles WHERE user_id = ?', (uid,)).fetchone()
    # Get time entry stats
    stats = conn.execute(
        '''SELECT COALESCE(SUM(hours), 0) as total_hours,
                  COALESCE(SUM(hours * hourly_rate), 0) as total_pay,
                  COUNT(*) as entry_count
           FROM time_entries WHERE user_id = ?''', (uid,)
    ).fetchone()
    recent_entries = conn.execute(
        '''SELECT te.*, j.name as job_name FROM time_entries te
           LEFT JOIN jobs j ON te.job_id = j.id
           WHERE te.user_id = ? ORDER BY te.work_date DESC LIMIT 20''', (uid,)
    ).fetchall()
    conn.close()

    d = dict(user)
    d.pop('password_hash', None)
    if profile:
        for k in profile.keys():
            if k not in ('id', 'user_id'):
                d[k] = profile[k]
    d['ssn_display'] = f'***-**-{profile["ssn_last4"]}' if profile and profile['ssn_last4'] else ''
    d['total_hours'] = round(stats['total_hours'], 1)
    d['total_pay'] = round(stats['total_pay'], 2)
    d['entry_count'] = stats['entry_count']
    d['recent_entries'] = [dict(e) for e in recent_entries]
    return jsonify(d)

@app.route('/api/employees/<int:uid>', methods=['PUT'])
@api_role_required('owner', 'admin')
def api_employees_update(uid):
    data = request.get_json()
    conn = get_db()

    # Update user fields
    user_fields = []
    user_values = []
    for f in ('display_name', 'first_name', 'last_name', 'home_base_city', 'role', 'email', 'phone', 'hourly_rate'):
        if f in data:
            user_fields.append(f'{f} = ?')
            user_values.append(data[f])
    # Auto-compose display_name from first/last if provided
    if 'first_name' in data or 'last_name' in data:
        cur = conn.execute('SELECT first_name, last_name FROM users WHERE id = ?', (uid,)).fetchone()
        fn = data.get('first_name', cur['first_name'] if cur else '')
        ln = data.get('last_name', cur['last_name'] if cur else '')
        composed = f'{fn} {ln}'.strip()
        if composed:
            user_fields.append('display_name = ?')
            user_values.append(composed)
    if 'password' in data and data['password']:
        user_fields.append('password_hash = ?')
        user_values.append(generate_password_hash(data['password']))
    if user_fields:
        user_fields.append("updated_at = datetime('now','localtime')")
        user_values.append(uid)
        conn.execute(f"UPDATE users SET {', '.join(user_fields)} WHERE id = ?", user_values)

    # Update profile fields
    profile = conn.execute('SELECT id FROM employee_profiles WHERE user_id = ?', (uid,)).fetchone()
    if profile:
        prof_fields = []
        prof_values = []
        for f in ('employee_number', 'date_of_birth', 'hire_date', 'address_street',
                   'address_city', 'address_state', 'address_zip', 'shirt_size',
                   'emergency_contact_name', 'emergency_contact_phone',
                   'emergency_contact_relationship', 'notes'):
            if f in data:
                prof_fields.append(f'{f} = ?')
                prof_values.append(data[f])
        if 'ssn' in data and data['ssn']:
            ssn_enc, ssn_last4 = _encrypt_ssn(data['ssn'])
            prof_fields.append('ssn_encrypted = ?')
            prof_values.append(ssn_enc)
            prof_fields.append('ssn_last4 = ?')
            prof_values.append(ssn_last4)
        if prof_fields:
            prof_fields.append("updated_at = datetime('now','localtime')")
            prof_values.append(uid)
            conn.execute(f"UPDATE employee_profiles SET {', '.join(prof_fields)} WHERE user_id = ?", prof_values)
    else:
        # Create profile if missing
        ssn_enc, ssn_last4 = _encrypt_ssn(data.get('ssn', ''))
        conn.execute(
            '''INSERT INTO employee_profiles (user_id, employee_number, ssn_encrypted, ssn_last4,
               date_of_birth, hire_date, employment_status, address_street, address_city,
               address_state, address_zip, shirt_size, emergency_contact_name,
               emergency_contact_phone, emergency_contact_relationship, notes)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (uid, data.get('employee_number', ''), ssn_enc, ssn_last4,
             data.get('date_of_birth', ''), data.get('hire_date', ''), 'Active',
             data.get('address_street', ''), data.get('address_city', ''),
             data.get('address_state', ''), data.get('address_zip', ''),
             data.get('shirt_size', ''), data.get('emergency_contact_name', ''),
             data.get('emergency_contact_phone', ''),
             data.get('emergency_contact_relationship', ''), data.get('notes', ''))
        )

    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/employees/<int:uid>/status', methods=['PUT'])
@api_role_required('owner', 'admin')
def api_employees_status(uid):
    data = request.get_json()
    new_status = data.get('status', '')
    if new_status not in ('Active', 'Inactive', 'Terminated'):
        return jsonify({'error': 'Invalid status'}), 400

    conn = get_db()
    if new_status == 'Terminated':
        conn.execute("UPDATE employee_profiles SET employment_status = 'Terminated', termination_date = date('now','localtime'), updated_at = datetime('now','localtime') WHERE user_id = ?", (uid,))
        conn.execute('UPDATE users SET is_active = 0 WHERE id = ?', (uid,))
    elif new_status == 'Inactive':
        conn.execute("UPDATE employee_profiles SET employment_status = 'Inactive', updated_at = datetime('now','localtime') WHERE user_id = ?", (uid,))
        conn.execute('UPDATE users SET is_active = 0 WHERE id = ?', (uid,))
    elif new_status == 'Active':
        conn.execute("UPDATE employee_profiles SET employment_status = 'Active', termination_date = '', updated_at = datetime('now','localtime') WHERE user_id = ?", (uid,))
        conn.execute('UPDATE users SET is_active = 1 WHERE id = ?', (uid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/employees/<int:uid>/ssn')
@api_role_required('owner')
def api_employees_ssn(uid):
    conn = get_db()
    profile = conn.execute('SELECT ssn_encrypted FROM employee_profiles WHERE user_id = ?', (uid,)).fetchone()
    conn.close()
    if not profile or not profile['ssn_encrypted']:
        return jsonify({'ssn': ''})
    try:
        ssn = _decrypt_ssn(profile['ssn_encrypted'])
    except Exception:
        return jsonify({'error': 'Failed to decrypt SSN'}), 500
    return jsonify({'ssn': ssn})

# ─── Tax Forms (W4 / W9) ────────────────────────────────────────

TAX_FORMS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'tax_forms')
os.makedirs(TAX_FORMS_DIR, exist_ok=True)

@app.route('/tax-forms')
@role_required('owner', 'admin')
def tax_forms_list():
    return render_template('tax_forms/list.html')

@app.route('/api/tax-forms')
@api_role_required('owner', 'admin')
def api_tax_forms_list():
    entity = request.args.get('entity_type', '')
    form_type = request.args.get('form_type', '')
    user_id = request.args.get('user_id', type=int)
    conn = get_db()
    sql = '''SELECT tf.*, u.display_name as employee_name
             FROM tax_forms tf LEFT JOIN users u ON tf.user_id = u.id WHERE 1=1'''
    params = []
    if entity:
        sql += ' AND tf.entity_type = ?'
        params.append(entity)
    if form_type:
        sql += ' AND tf.form_type = ?'
        params.append(form_type)
    if user_id:
        sql += ' AND tf.user_id = ?'
        params.append(user_id)
    sql += ' ORDER BY tf.updated_at DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d['has_file'] = bool(d.get('file_path'))
        result.append(d)
    return jsonify(result)

@app.route('/api/tax-forms', methods=['POST'])
@api_role_required('owner', 'admin')
def api_tax_forms_create():
    if request.content_type and 'multipart' in request.content_type:
        data = request.form.to_dict()
    else:
        data = request.get_json()
    conn = get_db()

    # Build insert from data
    fields = ['form_type','entity_type','status','user_id',
              'w9_name','w9_business_name','w9_tax_class','w9_exemptions',
              'w9_address','w9_city_state_zip','w9_account_numbers','w9_tin','w9_tin_type',
              'w9_signature_name','w9_signature_date',
              'w4_first_name','w4_last_name','w4_address','w4_city_state_zip','w4_ssn',
              'w4_filing_status','w4_multiple_jobs','w4_dependents_amount',
              'w4_other_income','w4_deductions','w4_extra_withholding','w4_exempt',
              'w4_signature_name','w4_signature_date',
              'w4_employer_name','w4_employer_ein','w4_first_date_employment',
              'f1099_payer_name','f1099_payer_tin','f1099_recipient_name','f1099_recipient_tin',
              'f1099_recipient_address','f1099_recipient_city_state_zip','f1099_amount',
              'f1099_tax_year','f1099_type',
              'notes']
    cols = ['created_by']
    vals = [session.get('user_id')]
    for f in fields:
        if f in data and data[f] != '':
            cols.append(f)
            vals.append(data[f])
    placeholders = ','.join(['?'] * len(vals))
    col_names = ','.join(cols)
    cursor = conn.execute(f'INSERT INTO tax_forms ({col_names}) VALUES ({placeholders})', vals)
    tf_id = cursor.lastrowid

    file = request.files.get('file')
    if file and file.filename:
        from werkzeug.utils import secure_filename
        original_name = file.filename
        fname = secure_filename(file.filename)
        fname = f"{int(datetime.now().timestamp())}_{fname}"
        file.save(os.path.join(TAX_FORMS_DIR, fname))
        conn.execute('UPDATE tax_forms SET file_path = ?, original_filename = ? WHERE id = ?',
                     (fname, original_name, tf_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': tf_id}), 201

@app.route('/api/tax-forms/<int:tf_id>')
@api_role_required('owner', 'admin')
def api_tax_form_detail(tf_id):
    conn = get_db()
    row = conn.execute(
        '''SELECT tf.*, u.display_name as employee_name
           FROM tax_forms tf LEFT JOIN users u ON tf.user_id = u.id
           WHERE tf.id = ?''', (tf_id,)
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    d = dict(row)
    d['has_file'] = bool(d.get('file_path'))
    return jsonify(d)

@app.route('/api/tax-forms/<int:tf_id>', methods=['PUT'])
@api_role_required('owner', 'admin')
def api_tax_forms_update(tf_id):
    if request.content_type and 'multipart' in request.content_type:
        data = request.form.to_dict()
    else:
        data = request.get_json()
    conn = get_db()
    allowed = ['form_type','entity_type','status','user_id',
               'w9_name','w9_business_name','w9_tax_class','w9_exemptions',
               'w9_address','w9_city_state_zip','w9_account_numbers','w9_tin','w9_tin_type',
               'w9_signature_name','w9_signature_date',
               'w4_first_name','w4_last_name','w4_address','w4_city_state_zip','w4_ssn',
               'w4_filing_status','w4_multiple_jobs','w4_dependents_amount',
               'w4_other_income','w4_deductions','w4_extra_withholding','w4_exempt',
               'w4_signature_name','w4_signature_date',
               'w4_employer_name','w4_employer_ein','w4_first_date_employment',
               'notes']
    fields = []
    values = []
    for f in allowed:
        if f in data:
            fields.append(f'{f} = ?')
            values.append(data[f] if data[f] != '' else None if f == 'user_id' else data[f])
    fields.append("updated_at = datetime('now','localtime')")
    if fields:
        values.append(tf_id)
        conn.execute(f"UPDATE tax_forms SET {', '.join(fields)} WHERE id = ?", values)

    file = request.files.get('file')
    if file and file.filename:
        from werkzeug.utils import secure_filename
        original_name = file.filename
        fname = secure_filename(file.filename)
        fname = f"{int(datetime.now().timestamp())}_{fname}"
        file.save(os.path.join(TAX_FORMS_DIR, fname))
        conn.execute('UPDATE tax_forms SET file_path = ?, original_filename = ? WHERE id = ?',
                     (fname, original_name, tf_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/tax-forms/<int:tf_id>', methods=['DELETE'])
@api_role_required('owner', 'admin')
def api_tax_forms_delete(tf_id):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM tax_forms WHERE id = ?', (tf_id,)).fetchone()
    if row and row['file_path']:
        fpath = os.path.join(TAX_FORMS_DIR, row['file_path'])
        if os.path.exists(fpath):
            os.remove(fpath)
    conn.execute('DELETE FROM tax_forms WHERE id = ?', (tf_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/tax-forms/<int:tf_id>/file')
@api_role_required('owner', 'admin')
def api_tax_form_file(tf_id):
    conn = get_db()
    row = conn.execute('SELECT file_path, original_filename FROM tax_forms WHERE id = ?', (tf_id,)).fetchone()
    conn.close()
    if not row or not row['file_path']:
        return 'File not found', 404
    fpath = os.path.join(TAX_FORMS_DIR, row['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    from flask import send_file
    return send_file(fpath, download_name=row['original_filename'] or row['file_path'])

# ─── Warranty ───────────────────────────────────────────────────

WARRANTY_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'warranty')
os.makedirs(WARRANTY_DIR, exist_ok=True)

@app.route('/warranty')
@role_required('owner', 'admin', 'project_manager')
def warranty_list():
    return render_template('warranty/list.html')

@app.route('/warranty/job/<int:job_id>')
@role_required('owner', 'admin', 'project_manager')
def warranty_job(job_id):
    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    conn.close()
    if not job:
        return 'Job not found', 404
    return render_template('warranty/job.html', job=job)

@app.route('/warranty/claims/<int:claim_id>')
@role_required('owner', 'admin', 'project_manager')
def warranty_claim_detail(claim_id):
    return render_template('warranty/claim_detail.html', claim_id=claim_id)

@app.route('/api/warranty')
@api_role_required('owner', 'admin', 'project_manager')
def api_warranty_list():
    conn = get_db()
    items = conn.execute(
        '''SELECT wi.*, j.name as job_name,
           (SELECT COUNT(*) FROM warranty_claims wc WHERE wc.warranty_id = wi.id) as claim_count,
           (SELECT COUNT(*) FROM warranty_claims wc WHERE wc.warranty_id = wi.id AND wc.status NOT IN ('Resolved','Denied')) as open_claims
           FROM warranty_items wi
           LEFT JOIN jobs j ON wi.job_id = j.id ORDER BY wi.warranty_end ASC'''
    ).fetchall()
    conn.close()
    result = []
    for i in items:
        d = dict(i)
        d['has_file'] = bool(d.get('file_path'))
        result.append(d)
    return jsonify(result)

@app.route('/api/warranty/job/<int:job_id>')
@api_role_required('owner', 'admin', 'project_manager')
def api_warranty_job(job_id):
    conn = get_db()
    items = conn.execute('SELECT * FROM warranty_items WHERE job_id = ? ORDER BY building, unit_number, warranty_end', (job_id,)).fetchall()
    result = []
    for item in items:
        claims = conn.execute(
            '''SELECT wc.*, u.display_name as assigned_name
               FROM warranty_claims wc
               LEFT JOIN users u ON wc.assigned_to = u.id
               WHERE wc.warranty_id = ? ORDER BY
               CASE wc.status WHEN 'Open' THEN 0 WHEN 'In Progress' THEN 1 WHEN 'Resolved' THEN 2 ELSE 3 END,
               wc.claim_date DESC''',
            (item['id'],)
        ).fetchall()
        d = dict(item)
        d['has_file'] = bool(d.get('file_path'))
        d['claims'] = [dict(c) for c in claims]
        result.append(d)
    conn.close()
    return jsonify(result)

@app.route('/api/warranty/items', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def create_warranty_item():
    # Support both JSON and multipart (file upload)
    if request.content_type and 'multipart' in request.content_type:
        data = request.form.to_dict()
    else:
        data = request.get_json()
    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO warranty_items (job_id, item_description, manufacturer, warranty_start, warranty_end,
           coverage_details, status, building, unit_number, model_number, serial_number, equipment_type)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
        (data.get('job_id'), data.get('item_description',''), data.get('manufacturer',''),
         data.get('warranty_start',''), data.get('warranty_end',''),
         data.get('coverage_details',''), data.get('status','Active'),
         data.get('building',''), data.get('unit_number',''),
         data.get('model_number',''), data.get('serial_number',''),
         data.get('equipment_type',''))
    )
    wid = cursor.lastrowid

    # Handle file upload
    file = request.files.get('file')
    if file and file.filename:
        from werkzeug.utils import secure_filename
        original_name = file.filename
        fname = secure_filename(file.filename)
        fname = f"{int(datetime.now().timestamp())}_{fname}"
        file.save(os.path.join(WARRANTY_DIR, fname))
        conn.execute('UPDATE warranty_items SET file_path = ?, original_filename = ? WHERE id = ?',
                     (fname, original_name, wid))

        # Auto-add to closeout checklist if not already there
        job_id = data.get('job_id')
        if job_id:
            existing = conn.execute(
                "SELECT id FROM closeout_checklists WHERE job_id = ? AND item_type = 'Warranty Letter' AND item_name = ?",
                (job_id, data.get('item_description', original_name))
            ).fetchone()
            if not existing:
                max_sort = conn.execute('SELECT COALESCE(MAX(sort_order), -1) FROM closeout_checklists WHERE job_id = ?', (job_id,)).fetchone()[0]
                conn.execute(
                    '''INSERT INTO closeout_checklists (job_id, item_name, item_type, status, file_path, sort_order, created_by)
                       VALUES (?,?,?,?,?,?,?)''',
                    (job_id, data.get('item_description', original_name), 'Warranty Letter', 'Complete',
                     fname, max_sort + 1, session.get('user_id'))
                )

    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': wid}), 201

@app.route('/api/warranty/items/<int:wid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def update_warranty_item(wid):
    if request.content_type and 'multipart' in request.content_type:
        data = request.form.to_dict()
    else:
        data = request.get_json()
    conn = get_db()
    fields = []
    values = []
    for f in ('item_description','manufacturer','warranty_start','warranty_end','coverage_details','status',
              'building','unit_number','model_number','serial_number','equipment_type'):
        if f in data:
            fields.append(f'{f} = ?')
            values.append(data[f])
    if fields:
        values.append(wid)
        conn.execute(f"UPDATE warranty_items SET {', '.join(fields)} WHERE id = ?", values)

    # Handle file upload on edit
    file = request.files.get('file')
    if file and file.filename:
        from werkzeug.utils import secure_filename
        original_name = file.filename
        fname = secure_filename(file.filename)
        fname = f"{int(datetime.now().timestamp())}_{fname}"
        file.save(os.path.join(WARRANTY_DIR, fname))
        conn.execute('UPDATE warranty_items SET file_path = ?, original_filename = ? WHERE id = ?',
                     (fname, original_name, wid))

        # Auto-add to closeout
        item = conn.execute('SELECT job_id, item_description FROM warranty_items WHERE id = ?', (wid,)).fetchone()
        if item and item['job_id']:
            existing = conn.execute(
                "SELECT id FROM closeout_checklists WHERE job_id = ? AND item_type = 'Warranty Letter' AND item_name = ?",
                (item['job_id'], item['item_description'])
            ).fetchone()
            if not existing:
                max_sort = conn.execute('SELECT COALESCE(MAX(sort_order), -1) FROM closeout_checklists WHERE job_id = ?', (item['job_id'],)).fetchone()[0]
                conn.execute(
                    '''INSERT INTO closeout_checklists (job_id, item_name, item_type, status, file_path, sort_order, created_by)
                       VALUES (?,?,?,?,?,?,?)''',
                    (item['job_id'], item['item_description'], 'Warranty Letter', 'Complete',
                     fname, max_sort + 1, session.get('user_id'))
                )

    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/warranty/items/<int:wid>/upload', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def upload_warranty_file(wid):
    """Upload a warranty letter file to an existing warranty item."""
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'No file provided'}), 400
    from werkzeug.utils import secure_filename
    original_name = file.filename
    fname = secure_filename(file.filename)
    fname = f"{int(datetime.now().timestamp())}_{fname}"
    file.save(os.path.join(WARRANTY_DIR, fname))
    conn = get_db()
    conn.execute('UPDATE warranty_items SET file_path = ?, original_filename = ? WHERE id = ?',
                 (fname, original_name, wid))
    # Auto-add to closeout
    item = conn.execute('SELECT job_id, item_description FROM warranty_items WHERE id = ?', (wid,)).fetchone()
    if item and item['job_id']:
        existing = conn.execute(
            "SELECT id FROM closeout_checklists WHERE job_id = ? AND item_type = 'Warranty Letter' AND item_name = ?",
            (item['job_id'], item['item_description'] or original_name)
        ).fetchone()
        if not existing:
            max_sort = conn.execute('SELECT COALESCE(MAX(sort_order), -1) FROM closeout_checklists WHERE job_id = ?', (item['job_id'],)).fetchone()[0]
            conn.execute(
                '''INSERT INTO closeout_checklists (job_id, item_name, item_type, status, file_path, sort_order, created_by)
                   VALUES (?,?,?,?,?,?,?)''',
                (item['job_id'], item['item_description'] or original_name, 'Warranty Letter', 'Complete',
                 fname, max_sort + 1, session.get('user_id'))
            )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'filename': fname})

@app.route('/api/warranty/items/<int:wid>/file')
@api_role_required('owner', 'admin', 'project_manager')
def serve_warranty_file(wid):
    """Serve the warranty letter file."""
    conn = get_db()
    item = conn.execute('SELECT file_path, original_filename FROM warranty_items WHERE id = ?', (wid,)).fetchone()
    conn.close()
    if not item or not item['file_path']:
        return 'File not found', 404
    fpath = os.path.join(WARRANTY_DIR, item['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    from flask import send_file
    return send_file(fpath, download_name=item['original_filename'] or item['file_path'])

@app.route('/api/warranty/items/<int:wid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def delete_warranty_item(wid):
    conn = get_db()
    item = conn.execute('SELECT file_path FROM warranty_items WHERE id = ?', (wid,)).fetchone()
    if item and item['file_path']:
        fpath = os.path.join(WARRANTY_DIR, item['file_path'])
        if os.path.exists(fpath):
            os.remove(fpath)
    conn.execute('DELETE FROM warranty_claims WHERE warranty_id = ?', (wid,))
    conn.execute('DELETE FROM warranty_items WHERE id = ?', (wid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/warranty/claims', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def create_warranty_claim():
    data = request.get_json()
    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO warranty_claims (warranty_id, claim_date, description, resolution, status,
           priority, assigned_to, caller_name, caller_phone, caller_email,
           building, unit_number, scheduled_date, created_by)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (data['warranty_id'], data.get('claim_date', datetime.now().strftime('%Y-%m-%d')),
         data.get('description',''), data.get('resolution',''), data.get('status','Open'),
         data.get('priority','Normal'), data.get('assigned_to') or None,
         data.get('caller_name',''), data.get('caller_phone',''), data.get('caller_email',''),
         data.get('building',''), data.get('unit_number',''),
         data.get('scheduled_date',''), session.get('user_id'))
    )
    # Notify assigned user
    assigned_to = data.get('assigned_to')
    if assigned_to:
        try:
            wi = conn.execute('SELECT item_description, job_id FROM warranty_items WHERE id = ?', (data['warranty_id'],)).fetchone()
            job = conn.execute('SELECT name FROM jobs WHERE id = ?', (wi['job_id'],)).fetchone() if wi else None
            conn.execute(
                'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
                (int(assigned_to), 'warranty_claim',
                 'New Warranty Claim Assigned',
                 f"Warranty claim for {wi['item_description'] if wi else 'unknown'} on {job['name'] if job else 'unknown'}: {data.get('description', '')[:100]}",
                 f"/warranty/claims/{cursor.lastrowid}")
            )
            conn.commit()
        except Exception:
            pass
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cursor.lastrowid}), 201

@app.route('/api/warranty/claims/<int:cid>')
@api_role_required('owner', 'admin', 'project_manager')
def api_warranty_claim_detail(cid):
    conn = get_db()
    claim = conn.execute(
        '''SELECT wc.*, wi.item_description, wi.manufacturer, wi.model_number, wi.serial_number,
           wi.equipment_type, wi.building as equip_building, wi.unit_number as equip_unit,
           wi.job_id, wi.file_path as warranty_file, wi.original_filename as warranty_filename,
           j.name as job_name, u.display_name as assigned_name
           FROM warranty_claims wc
           JOIN warranty_items wi ON wc.warranty_id = wi.id
           LEFT JOIN jobs j ON wi.job_id = j.id
           LEFT JOIN users u ON wc.assigned_to = u.id
           WHERE wc.id = ?''', (cid,)
    ).fetchone()
    conn.close()
    if not claim:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(dict(claim))

@app.route('/api/warranty/claims/<int:cid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def update_warranty_claim(cid):
    data = request.get_json()
    conn = get_db()
    fields = []
    values = []
    for f in ('description','resolution','status','priority','assigned_to',
              'caller_name','caller_phone','caller_email','building','unit_number',
              'scheduled_date','resolved_date'):
        if f in data:
            fields.append(f'{f} = ?')
            val = data[f]
            if f == 'assigned_to' and (val == '' or val is None):
                val = None
            values.append(val)
    # Auto-set resolved_date when status changes to Resolved
    if 'status' in data and data['status'] == 'Resolved' and 'resolved_date' not in data:
        fields.append('resolved_date = ?')
        values.append(datetime.now().strftime('%Y-%m-%d'))
    if fields:
        values.append(cid)
        conn.execute(f"UPDATE warranty_claims SET {', '.join(fields)} WHERE id = ?", values)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/warranty/claims/<int:cid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def delete_warranty_claim(cid):
    conn = get_db()
    conn.execute('DELETE FROM warranty_claims WHERE id = ?', (cid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Service Calls ──────────────────────────────────────────────

@app.route('/service-calls')
@login_required
def service_calls_list():
    return render_template('service_calls/list.html')

@app.route('/service-calls/<int:call_id>')
@login_required
def service_call_detail(call_id):
    return render_template('service_calls/detail.html', call_id=call_id)

@app.route('/api/service-calls')
@api_login_required
def api_service_calls():
    conn = get_db()
    calls = conn.execute(
        '''SELECT sc.*, j.name as job_name, u.display_name as assigned_name
           FROM service_calls sc
           LEFT JOIN jobs j ON sc.job_id = j.id
           LEFT JOIN users u ON sc.assigned_to = u.id
           ORDER BY
             CASE sc.priority WHEN 'Urgent' THEN 0 WHEN 'High' THEN 1 WHEN 'Normal' THEN 2 ELSE 3 END,
             sc.created_at DESC'''
    ).fetchall()
    conn.close()
    return jsonify([dict(c) for c in calls])

@app.route('/api/service-calls', methods=['POST'])
@api_login_required
def create_service_call():
    data = request.get_json()
    conn = get_db()
    conn.execute(
        '''INSERT INTO service_calls (job_id, caller_name, caller_phone, caller_email, description, priority, status, assigned_to, scheduled_date, created_by)
           VALUES (?,?,?,?,?,?,?,?,?,?)''',
        (data.get('job_id') or None, data.get('caller_name',''), data.get('caller_phone',''),
         data.get('caller_email',''), data.get('description',''),
         data.get('priority','Normal'), data.get('status','Open'),
         data.get('assigned_to') or None, data.get('scheduled_date',''), session.get('user_id'))
    )
    conn.commit()

    # Create notification for assigned user
    assigned_to = data.get('assigned_to')
    if assigned_to:
        try:
            conn.execute(
                'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
                (int(assigned_to), 'service_call',
                 'New Service Call Assigned',
                 f"A new service call has been assigned to you: {data.get('description', '')[:100]}",
                 '/service-calls')
            )
            conn.commit()
        except Exception:
            pass

    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/service-calls/<int:cid>')
@api_login_required
def api_service_call_detail(cid):
    conn = get_db()
    call = conn.execute(
        '''SELECT sc.*, j.name as job_name, u.display_name as assigned_name
           FROM service_calls sc
           LEFT JOIN jobs j ON sc.job_id = j.id
           LEFT JOIN users u ON sc.assigned_to = u.id
           WHERE sc.id = ?''', (cid,)
    ).fetchone()
    conn.close()
    if not call:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(dict(call))

@app.route('/api/service-calls/<int:cid>', methods=['PUT'])
@api_login_required
def update_service_call(cid):
    data = request.get_json()
    conn = get_db()
    fields = []
    values = []
    for f in ('caller_name','caller_phone','caller_email','description','priority','status','assigned_to','resolution','scheduled_date','resolved_date','job_id'):
        if f in data:
            fields.append(f'{f} = ?')
            values.append(data[f] if data[f] != '' else None if f in ('assigned_to','job_id') else data[f])
    if fields:
        values.append(cid)
        conn.execute(f"UPDATE service_calls SET {', '.join(fields)} WHERE id = ?", values)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/service-calls/<int:cid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def delete_service_call(cid):
    conn = get_db()
    conn.execute('DELETE FROM service_calls WHERE id = ?', (cid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── How To's ───────────────────────────────────────────────────

@app.route('/howtos')
@login_required
def howtos_list():
    return render_template('howtos/list.html')

@app.route('/howtos/new')
@role_required('owner', 'admin', 'project_manager')
def howtos_new():
    return render_template('howtos/edit.html', article=None)

@app.route('/howtos/<int:article_id>')
@login_required
def howtos_article(article_id):
    conn = get_db()
    article = conn.execute('SELECT * FROM howto_articles WHERE id = ?', (article_id,)).fetchone()
    conn.close()
    if not article:
        return 'Article not found', 404
    return render_template('howtos/article.html', article=article)

@app.route('/howtos/<int:article_id>/edit')
@role_required('owner', 'admin', 'project_manager')
def howtos_edit(article_id):
    conn = get_db()
    article = conn.execute('SELECT * FROM howto_articles WHERE id = ?', (article_id,)).fetchone()
    conn.close()
    if not article:
        return 'Article not found', 404
    return render_template('howtos/edit.html', article=article)

@app.route('/api/howtos')
@api_login_required
def api_howtos():
    conn = get_db()
    articles = conn.execute(
        'SELECT id, title, category, tags, created_at, updated_at FROM howto_articles ORDER BY updated_at DESC'
    ).fetchall()
    conn.close()
    return jsonify([dict(a) for a in articles])

@app.route('/api/howtos/<int:aid>')
@api_login_required
def api_howto_detail(aid):
    conn = get_db()
    article = conn.execute('SELECT * FROM howto_articles WHERE id = ?', (aid,)).fetchone()
    conn.close()
    if not article:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(dict(article))

@app.route('/api/howtos', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def create_howto():
    data = request.get_json()
    conn = get_db()
    cursor = conn.execute(
        'INSERT INTO howto_articles (title, category, content, tags, created_by) VALUES (?,?,?,?,?)',
        (data.get('title',''), data.get('category',''), data.get('content',''),
         data.get('tags',''), session.get('user_id'))
    )
    conn.commit()
    aid = cursor.lastrowid
    conn.close()
    return jsonify({'ok': True, 'id': aid}), 201

@app.route('/api/howtos/<int:aid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def update_howto(aid):
    data = request.get_json()
    conn = get_db()
    conn.execute(
        '''UPDATE howto_articles SET title=?, category=?, content=?, tags=?,
           updated_at=datetime('now','localtime') WHERE id=?''',
        (data.get('title',''), data.get('category',''), data.get('content',''),
         data.get('tags',''), aid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/howtos/<int:aid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def delete_howto(aid):
    conn = get_db()
    conn.execute('DELETE FROM howto_articles WHERE id = ?', (aid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Code Books ─────────────────────────────────────────────────

@app.route('/codebooks')
@login_required
def codebooks_list():
    return render_template('codebooks/list.html')

@app.route('/codebooks/<int:book_id>')
@login_required
def codebooks_browse(book_id):
    conn = get_db()
    book = conn.execute('SELECT * FROM code_books WHERE id = ?', (book_id,)).fetchone()
    conn.close()
    if not book:
        return 'Code book not found', 404
    return render_template('codebooks/browse.html', book=book)

@app.route('/api/codebooks')
@api_login_required
def api_codebooks():
    conn = get_db()
    books = conn.execute('SELECT * FROM code_books ORDER BY code').fetchall()
    conn.close()
    return jsonify([dict(b) for b in books])

@app.route('/api/codebooks/<int:book_id>/sections')
@api_login_required
def api_codebook_sections(book_id):
    conn = get_db()
    sections = conn.execute(
        'SELECT * FROM code_sections WHERE book_id = ? ORDER BY sort_order', (book_id,)
    ).fetchall()
    conn.close()
    return jsonify([dict(s) for s in sections])

@app.route('/api/codebooks/search')
@api_login_required
def api_codebook_search():
    q = request.args.get('q', '').strip()
    book_id = request.args.get('book_id', type=int)
    if not q:
        return jsonify([])
    conn = get_db()
    if book_id:
        sections = conn.execute(
            '''SELECT cs.*, cb.code as book_code, cb.name as book_name
               FROM code_sections cs JOIN code_books cb ON cs.book_id = cb.id
               WHERE cs.book_id = ? AND (cs.section_number LIKE ? OR cs.title LIKE ? OR cs.content LIKE ?)
               ORDER BY cs.sort_order LIMIT 50''',
            (book_id, f'%{q}%', f'%{q}%', f'%{q}%')
        ).fetchall()
    else:
        sections = conn.execute(
            '''SELECT cs.*, cb.code as book_code, cb.name as book_name
               FROM code_sections cs JOIN code_books cb ON cs.book_id = cb.id
               WHERE cs.section_number LIKE ? OR cs.title LIKE ? OR cs.content LIKE ?
               ORDER BY cb.code, cs.sort_order LIMIT 50''',
            (f'%{q}%', f'%{q}%', f'%{q}%')
        ).fetchall()
    conn.close()
    return jsonify([dict(s) for s in sections])

@app.route('/api/codebooks/bookmarks')
@api_login_required
def api_codebook_bookmarks():
    conn = get_db()
    bookmarks = conn.execute(
        '''SELECT cb.*, cs.section_number, cs.title, bk.code as book_code
           FROM code_bookmarks cb
           JOIN code_sections cs ON cb.section_id = cs.id
           JOIN code_books bk ON cs.book_id = bk.id
           WHERE cb.user_id = ? ORDER BY cb.created_at DESC''',
        (session['user_id'],)
    ).fetchall()
    conn.close()
    return jsonify([dict(b) for b in bookmarks])

@app.route('/api/codebooks/bookmarks', methods=['POST'])
@api_login_required
def toggle_bookmark():
    data = request.get_json()
    section_id = data.get('section_id')
    note = data.get('note', '')
    conn = get_db()
    existing = conn.execute(
        'SELECT id FROM code_bookmarks WHERE user_id = ? AND section_id = ?',
        (session['user_id'], section_id)
    ).fetchone()
    if existing:
        conn.execute('DELETE FROM code_bookmarks WHERE id = ?', (existing['id'],))
        action = 'removed'
    else:
        conn.execute(
            'INSERT INTO code_bookmarks (user_id, section_id, note) VALUES (?,?,?)',
            (session['user_id'], section_id, note)
        )
        action = 'added'
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'action': action})

# ─── Customers (Phase 1) ─────────────────────────────────────────

@app.route('/customers')
@role_required('owner', 'admin', 'project_manager')
def customers_list():
    return render_template('customers/list.html')

@app.route('/customers/<int:cid>')
@role_required('owner', 'admin', 'project_manager')
def customers_detail(cid):
    return render_template('customers/detail.html', customer_id=cid)

@app.route('/api/customers')
@api_role_required('owner', 'admin', 'project_manager')
def api_customers():
    conn = get_db()
    customers = conn.execute('SELECT * FROM customers ORDER BY company_name').fetchall()
    result = []
    for c in customers:
        total_bids = conn.execute('SELECT COUNT(*) FROM bids WHERE customer_id = ?', (c['id'],)).fetchone()[0]
        accepted_bids = conn.execute("SELECT COUNT(*) FROM bids WHERE customer_id = ? AND status = 'Accepted'", (c['id'],)).fetchone()[0]
        d = dict(c)
        d['total_bids'] = total_bids
        d['accepted_bids'] = accepted_bids
        result.append(d)
    conn.close()
    return jsonify(result)

@app.route('/api/customers', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_create_customer():
    data = request.get_json()
    name = (data.get('company_name') or '').strip()
    if not name:
        return jsonify({'error': 'Company name is required'}), 400
    if data.get('company_type') == 'Supplier':
        return jsonify({'error': 'Use the Vendors page to manage suppliers'}), 400
    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO customers (company_name, company_type, primary_contact, contact_email,
           contact_phone, address, city, state, zip_code, website, notes, created_by)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
        (name, data.get('company_type', 'General Contractor'),
         data.get('primary_contact', ''), data.get('contact_email', ''),
         data.get('contact_phone', ''), data.get('address', ''),
         data.get('city', ''), data.get('state', ''),
         data.get('zip_code', ''), data.get('website', ''),
         data.get('notes', ''), session.get('user_id'))
    )
    cid = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cid}), 201

@app.route('/api/customers/<int:cid>')
@api_role_required('owner', 'admin', 'project_manager')
def api_customer_detail(cid):
    conn = get_db()
    c = conn.execute('SELECT * FROM customers WHERE id = ?', (cid,)).fetchone()
    if not c:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    contacts = conn.execute('SELECT * FROM customer_contacts WHERE customer_id = ? ORDER BY is_primary DESC, name', (cid,)).fetchall()
    bids = conn.execute(
        '''SELECT b.id, b.bid_name, b.status, b.total_bid, b.bid_date, j.name as job_name
           FROM bids b LEFT JOIN jobs j ON b.job_id = j.id
           WHERE b.customer_id = ? ORDER BY b.bid_date DESC''', (cid,)
    ).fetchall()
    jobs = conn.execute(
        'SELECT * FROM jobs WHERE customer_id = ? ORDER BY updated_at DESC', (cid,)
    ).fetchall()
    total_bids = len(bids)
    accepted_bids = sum(1 for b in bids if b['status'] == 'Accepted')
    active_jobs = sum(1 for j in jobs if j['status'] in ('In Progress', 'Awarded', 'Pre-Construction'))
    total_revenue = sum(b['total_bid'] or 0 for b in bids if b['status'] == 'Accepted')
    conn.close()
    result = dict(c)
    result['contacts'] = [dict(ct) for ct in contacts]
    result['bids'] = [dict(b) for b in bids]
    result['jobs'] = [dict(j) for j in jobs]
    result['total_bids'] = total_bids
    result['accepted_bids'] = accepted_bids
    result['active_jobs'] = active_jobs
    result['total_revenue'] = total_revenue
    return jsonify(result)

@app.route('/api/customers/<int:cid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_customer(cid):
    data = request.get_json()
    if data.get('company_type') == 'Supplier':
        return jsonify({'error': 'Use the Vendors page to manage suppliers'}), 400
    conn = get_db()
    fields = []
    params = []
    for col in ('company_name', 'company_type', 'primary_contact', 'contact_email',
                'contact_phone', 'address', 'city', 'state', 'zip_code', 'website', 'notes'):
        if col in data:
            fields.append(f'{col} = ?')
            params.append(data[col])
    if 'is_active' in data:
        fields.append('is_active = ?')
        params.append(int(data['is_active']))
    if fields:
        fields.append("updated_at = datetime('now','localtime')")
        params.append(cid)
        conn.execute(f"UPDATE customers SET {', '.join(fields)} WHERE id = ?", params)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/customers/<int:cid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_deactivate_customer(cid):
    conn = get_db()
    conn.execute("UPDATE customers SET is_active = 0, updated_at = datetime('now','localtime') WHERE id = ?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/customers/<int:cid>/contacts', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_add_customer_contact(cid):
    data = request.get_json()
    conn = get_db()
    if data.get('is_primary'):
        conn.execute('UPDATE customer_contacts SET is_primary = 0 WHERE customer_id = ?', (cid,))
    cursor = conn.execute(
        'INSERT INTO customer_contacts (customer_id, name, title, email, phone, is_primary) VALUES (?,?,?,?,?,?)',
        (cid, data.get('name', ''), data.get('title', ''), data.get('email', ''),
         data.get('phone', ''), 1 if data.get('is_primary') else 0)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cursor.lastrowid}), 201

@app.route('/api/customers/<int:cid>/contacts/<int:ctid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_customer_contact(cid, ctid):
    data = request.get_json()
    conn = get_db()
    if data.get('is_primary'):
        conn.execute('UPDATE customer_contacts SET is_primary = 0 WHERE customer_id = ?', (cid,))
    conn.execute(
        'UPDATE customer_contacts SET name=?, title=?, email=?, phone=?, is_primary=? WHERE id=? AND customer_id=?',
        (data.get('name', ''), data.get('title', ''), data.get('email', ''),
         data.get('phone', ''), 1 if data.get('is_primary') else 0, ctid, cid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/customers/<int:cid>/contacts/<int:ctid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_customer_contact(cid, ctid):
    conn = get_db()
    conn.execute('DELETE FROM customer_contacts WHERE id = ? AND customer_id = ?', (ctid, cid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/customers/search')
@api_role_required('owner', 'admin', 'project_manager')
def api_customers_search():
    q = request.args.get('q', '').strip()
    if len(q) < 1:
        return jsonify([])
    conn = get_db()
    results = conn.execute(
        "SELECT id, company_name, company_type, primary_contact FROM customers WHERE is_active = 1 AND company_name LIKE ? ORDER BY company_name LIMIT 15",
        (f'%{q}%',)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in results])

# ─── Vendors ─────────────────────────────────────────────────

@app.route('/vendors')
@role_required('owner', 'admin', 'project_manager')
def vendors_list():
    return render_template('vendors/list.html')

@app.route('/vendors/<int:vid>')
@role_required('owner', 'admin', 'project_manager')
def vendors_detail(vid):
    return render_template('vendors/detail.html', vendor_id=vid)

@app.route('/api/vendors')
@api_role_required('owner', 'admin', 'project_manager')
def api_vendors():
    conn = get_db()
    vendors = conn.execute('SELECT * FROM vendors ORDER BY company_name').fetchall()
    result = []
    for v in vendors:
        quote_count = conn.execute('SELECT COUNT(*) FROM supplier_quotes WHERE vendor_id = ?', (v['id'],)).fetchone()[0]
        invoice_count = conn.execute(
            '''SELECT COUNT(*) FROM supplier_invoices si
               JOIN billtrust_config bc ON si.supplier_config_id = bc.id
               WHERE bc.vendor_id = ?''', (v['id'],)
        ).fetchone()[0]
        d = dict(v)
        d['quote_count'] = quote_count
        d['invoice_count'] = invoice_count
        result.append(d)
    conn.close()
    return jsonify(result)

@app.route('/api/vendors', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_create_vendor():
    data = request.get_json()
    name = (data.get('company_name') or '').strip()
    if not name:
        return jsonify({'error': 'Company name is required'}), 400
    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO vendors (company_name, vendor_type, account_number, payment_terms,
           primary_contact, contact_email, contact_phone, address, city, state, zip_code,
           website, notes, created_by)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (name, data.get('vendor_type', 'Supplier'),
         data.get('account_number', ''), data.get('payment_terms', ''),
         data.get('primary_contact', ''), data.get('contact_email', ''),
         data.get('contact_phone', ''), data.get('address', ''),
         data.get('city', ''), data.get('state', ''),
         data.get('zip_code', ''), data.get('website', ''),
         data.get('notes', ''), session.get('user_id'))
    )
    vid = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': vid}), 201

@app.route('/api/vendors/<int:vid>')
@api_role_required('owner', 'admin', 'project_manager')
def api_vendor_detail(vid):
    conn = get_db()
    v = conn.execute('SELECT * FROM vendors WHERE id = ?', (vid,)).fetchone()
    if not v:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    contacts = conn.execute('SELECT * FROM vendor_contacts WHERE vendor_id = ? ORDER BY is_primary DESC, name', (vid,)).fetchall()
    quotes = conn.execute(
        '''SELECT sq.id, sq.quote_number, sq.supplier_name, sq.status, sq.total, sq.quote_date, j.name as job_name
           FROM supplier_quotes sq LEFT JOIN jobs j ON sq.job_id = j.id
           WHERE sq.vendor_id = ? ORDER BY sq.quote_date DESC''', (vid,)
    ).fetchall()
    invoices = conn.execute(
        '''SELECT si.id, si.invoice_number, si.invoice_date, si.status, si.total, si.balance_due, j.name as job_name
           FROM supplier_invoices si
           JOIN billtrust_config bc ON si.supplier_config_id = bc.id
           LEFT JOIN jobs j ON si.job_id = j.id
           WHERE bc.vendor_id = ? ORDER BY si.invoice_date DESC''', (vid,)
    ).fetchall()
    total_quoted = sum(q['total'] or 0 for q in quotes)
    total_invoiced = sum(i['total'] or 0 for i in invoices)
    open_balance = sum(i['balance_due'] or 0 for i in invoices)
    conn.close()
    result = dict(v)
    result['contacts'] = [dict(ct) for ct in contacts]
    result['quotes'] = [dict(q) for q in quotes]
    result['invoices'] = [dict(i) for i in invoices]
    result['total_quoted'] = total_quoted
    result['total_invoiced'] = total_invoiced
    result['open_balance'] = open_balance
    return jsonify(result)

@app.route('/api/vendors/<int:vid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_vendor(vid):
    data = request.get_json()
    conn = get_db()
    fields = []
    params = []
    for col in ('company_name', 'vendor_type', 'account_number', 'payment_terms',
                'primary_contact', 'contact_email', 'contact_phone',
                'address', 'city', 'state', 'zip_code', 'website', 'notes'):
        if col in data:
            fields.append(f'{col} = ?')
            params.append(data[col])
    if 'is_active' in data:
        fields.append('is_active = ?')
        params.append(int(data['is_active']))
    if fields:
        fields.append("updated_at = datetime('now','localtime')")
        params.append(vid)
        conn.execute(f"UPDATE vendors SET {', '.join(fields)} WHERE id = ?", params)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/vendors/<int:vid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_deactivate_vendor(vid):
    conn = get_db()
    conn.execute("UPDATE vendors SET is_active = 0, updated_at = datetime('now','localtime') WHERE id = ?", (vid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/vendors/<int:vid>/contacts', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_add_vendor_contact(vid):
    data = request.get_json()
    conn = get_db()
    if data.get('is_primary'):
        conn.execute('UPDATE vendor_contacts SET is_primary = 0 WHERE vendor_id = ?', (vid,))
    cursor = conn.execute(
        'INSERT INTO vendor_contacts (vendor_id, name, title, email, phone, is_primary) VALUES (?,?,?,?,?,?)',
        (vid, data.get('name', ''), data.get('title', ''), data.get('email', ''),
         data.get('phone', ''), 1 if data.get('is_primary') else 0)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cursor.lastrowid}), 201

@app.route('/api/vendors/<int:vid>/contacts/<int:ctid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_vendor_contact(vid, ctid):
    data = request.get_json()
    conn = get_db()
    if data.get('is_primary'):
        conn.execute('UPDATE vendor_contacts SET is_primary = 0 WHERE vendor_id = ?', (vid,))
    conn.execute(
        'UPDATE vendor_contacts SET name=?, title=?, email=?, phone=?, is_primary=? WHERE id=? AND vendor_id=?',
        (data.get('name', ''), data.get('title', ''), data.get('email', ''),
         data.get('phone', ''), 1 if data.get('is_primary') else 0, ctid, vid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/vendors/<int:vid>/contacts/<int:ctid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_vendor_contact(vid, ctid):
    conn = get_db()
    conn.execute('DELETE FROM vendor_contacts WHERE id = ? AND vendor_id = ?', (ctid, vid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/vendors/search')
@api_role_required('owner', 'admin', 'project_manager')
def api_vendors_search():
    q = request.args.get('q', '').strip()
    if len(q) < 1:
        return jsonify([])
    conn = get_db()
    results = conn.execute(
        "SELECT id, company_name, vendor_type, primary_contact FROM vendors WHERE is_active = 1 AND company_name LIKE ? ORDER BY company_name LIMIT 15",
        (f'%{q}%',)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in results])

# ─── Projects ───────────────────────────────────────────────────

@app.route('/projects')
@role_required('owner', 'admin', 'project_manager')
def projects_overview():
    return render_template('projects/overview.html')

@app.route('/projects/<int:job_id>')
@role_required('owner', 'admin', 'project_manager')
def projects_detail(job_id):
    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    conn.close()
    if not job:
        return 'Project not found', 404
    return render_template('projects/detail.html', job=job)

@app.route('/api/projects')
@api_role_required('owner', 'admin', 'project_manager')
def api_projects():
    conn = get_db()
    jobs = conn.execute('SELECT * FROM jobs ORDER BY updated_at DESC').fetchall()
    result = []
    for job in jobs:
        items = conn.execute(
            'SELECT total_net_price, qty_ordered, price_per FROM line_items WHERE job_id = ?',
            (job['id'],)
        ).fetchall()
        material_cost = sum(
            (row['total_net_price'] or 0) if (row['total_net_price'] or 0)
            else (row['qty_ordered'] or 0) * (row['price_per'] or 0)
            for row in items
        )
        expenses = conn.execute('SELECT COALESCE(SUM(amount),0) FROM expenses WHERE job_id = ?', (job['id'],)).fetchone()[0]
        service_count = conn.execute("SELECT COUNT(*) FROM service_calls WHERE job_id = ? AND status NOT IN ('Resolved','Closed')", (job['id'],)).fetchone()[0]
        warranty_count = conn.execute('SELECT COUNT(*) FROM warranty_items WHERE job_id = ?', (job['id'],)).fetchone()[0]

        customer_name = None
        customer_id = job['customer_id']
        if customer_id:
            cust = conn.execute('SELECT company_name FROM customers WHERE id = ?', (customer_id,)).fetchone()
            if cust:
                customer_name = cust['company_name']

        jid = job['id']
        co_count = conn.execute('SELECT COUNT(*) FROM change_orders WHERE job_id = ?', (jid,)).fetchone()[0]
        rfi_count = conn.execute('SELECT COUNT(*) FROM rfis WHERE job_id = ?', (jid,)).fetchone()[0]
        submittal_count = conn.execute('SELECT COUNT(*) FROM submittals WHERE job_id = ?', (jid,)).fetchone()[0]
        lw_count = conn.execute('SELECT COUNT(*) FROM lien_waivers WHERE job_id = ?', (jid,)).fetchone()[0]
        payapp_count = conn.execute('SELECT COUNT(*) FROM pay_applications pa JOIN pay_app_contracts pac ON pa.contract_id = pac.id WHERE pac.job_id = ?', (jid,)).fetchone()[0]
        plan_count = conn.execute('SELECT COUNT(*) FROM plans WHERE job_id = ?', (jid,)).fetchone()[0]
        contract_count = conn.execute('SELECT COUNT(*) FROM contracts WHERE job_id = ?', (jid,)).fetchone()[0]
        permit_count = conn.execute('SELECT COUNT(*) FROM permits WHERE job_id = ?', (jid,)).fetchone()[0]

        result.append({
            'id': jid,
            'name': job['name'],
            'status': job['status'],
            'location': f"{job['city'] or ''} {job['state'] or ''}".strip() or '-',
            'tax_rate': job['tax_rate'] or 0,
            'material_cost': round(material_cost, 2),
            'expenses': round(expenses, 2),
            'open_service_calls': service_count,
            'warranty_items': warranty_count,
            'updated_at': job['updated_at'],
            'customer_id': customer_id,
            'customer_name': customer_name,
            'co_count': co_count,
            'rfi_count': rfi_count,
            'submittal_count': submittal_count,
            'lw_count': lw_count,
            'payapp_count': payapp_count,
            'plan_count': plan_count,
            'contract_count': contract_count,
            'permit_count': permit_count,
        })
    conn.close()
    return jsonify(result)

@app.route('/api/projects/<int:job_id>')
@api_role_required('owner', 'admin', 'project_manager')
def api_project_detail(job_id):
    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    if not job:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    items = conn.execute(
        'SELECT total_net_price, qty_ordered, price_per FROM line_items WHERE job_id = ?', (job_id,)
    ).fetchall()
    material_cost = sum(
        (row['total_net_price'] or 0) if (row['total_net_price'] or 0)
        else (row['qty_ordered'] or 0) * (row['price_per'] or 0)
        for row in items
    )

    expenses = conn.execute('SELECT * FROM expenses WHERE job_id = ? ORDER BY expense_date DESC', (job_id,)).fetchall()
    payments = conn.execute('SELECT * FROM payments WHERE job_id = ? ORDER BY payment_date DESC', (job_id,)).fetchall()
    invoices = conn.execute('SELECT * FROM client_invoices WHERE job_id = ? ORDER BY issue_date DESC', (job_id,)).fetchall()
    service_calls = conn.execute(
        '''SELECT sc.*, u.display_name as assigned_name FROM service_calls sc
           LEFT JOIN users u ON sc.assigned_to = u.id
           WHERE sc.job_id = ? ORDER BY sc.created_at DESC''', (job_id,)
    ).fetchall()
    warranties = conn.execute('SELECT * FROM warranty_items WHERE job_id = ? ORDER BY warranty_end', (job_id,)).fetchall()
    time_entries = conn.execute(
        '''SELECT te.*, u.display_name as employee_name FROM time_entries te
           LEFT JOIN users u ON te.user_id = u.id
           WHERE te.job_id = ? ORDER BY te.work_date DESC''', (job_id,)
    ).fetchall()

    # Resolve customer name
    job_dict = dict(job)
    job_dict['customer_name'] = None
    if job['customer_id']:
        cust = conn.execute('SELECT company_name FROM customers WHERE id = ?', (job['customer_id'],)).fetchone()
        if cust:
            job_dict['customer_name'] = cust['company_name']

    conn.close()
    return jsonify({
        'job': job_dict,
        'material_cost': round(material_cost, 2),
        'expenses': [dict(e) for e in expenses],
        'payments': [dict(p) for p in payments],
        'invoices': [dict(i) for i in invoices],
        'service_calls': [dict(c) for c in service_calls],
        'warranties': [dict(w) for w in warranties],
        'time_entries': [dict(t) for t in time_entries],
    })

# ─── User Management ────────────────────────────────────────────

@app.route('/admin/users')
@role_required('owner')
def admin_users():
    return render_template('admin/users.html')

@app.route('/api/admin/users')
@api_role_required('owner', 'admin')
def api_admin_users():
    conn = get_db()
    users = conn.execute('SELECT id, username, display_name, first_name, last_name, home_base_city, role, email, phone, hourly_rate, is_active, created_at FROM users ORDER BY display_name').fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])

@app.route('/api/admin/users', methods=['POST'])
@api_role_required('owner', 'admin')
def api_create_user():
    data = request.get_json()
    username = (data.get('username') or '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400

    conn = get_db()
    existing = conn.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
    if existing:
        conn.close()
        return jsonify({'error': 'Username already taken'}), 400

    first_name = (data.get('first_name') or '').strip()
    last_name = (data.get('last_name') or '').strip()
    display_name = data.get('display_name', '').strip() or f'{first_name} {last_name}'.strip() or username
    cursor = conn.execute(
        '''INSERT INTO users (username, display_name, first_name, last_name, home_base_city,
           password_hash, role, email, phone, hourly_rate, must_change_password)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
        (username, display_name, first_name, last_name, data.get('home_base_city', ''),
         generate_password_hash(password),
         data.get('role', 'employee'), data.get('email', ''), data.get('phone', ''),
         float(data.get('hourly_rate', 0)), 1 if data.get('must_change_password', False) else 0)
    )
    new_uid = cursor.lastrowid
    # Auto-create employee profile
    conn.execute(
        "INSERT OR IGNORE INTO employee_profiles (user_id, employment_status, hire_date) VALUES (?, 'Active', date('now','localtime'))",
        (new_uid,)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/admin/users/<int:uid>', methods=['PUT'])
@api_role_required('owner', 'admin')
def api_update_user(uid):
    data = request.get_json()
    conn = get_db()
    fields = []
    values = []
    for f in ('display_name', 'first_name', 'last_name', 'home_base_city', 'role', 'email', 'phone', 'hourly_rate', 'is_active'):
        if f in data:
            fields.append(f'{f} = ?')
            values.append(data[f])
    # Auto-compose display_name from first/last if provided
    if 'first_name' in data or 'last_name' in data:
        cur = conn.execute('SELECT first_name, last_name FROM users WHERE id = ?', (uid,)).fetchone()
        fn = data.get('first_name', cur['first_name'] if cur else '')
        ln = data.get('last_name', cur['last_name'] if cur else '')
        composed = f'{fn} {ln}'.strip()
        if composed:
            fields.append('display_name = ?')
            values.append(composed)
    if 'password' in data and data['password']:
        fields.append('password_hash = ?')
        values.append(generate_password_hash(data['password']))
    if fields:
        fields.append("updated_at = datetime('now','localtime')")
        values.append(uid)
        conn.execute(f"UPDATE users SET {', '.join(fields)} WHERE id = ?", values)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/admin/users/<int:uid>', methods=['DELETE'])
@api_role_required('owner', 'admin')
def api_delete_user(uid):
    if uid == session.get('user_id'):
        return jsonify({'error': 'Cannot delete your own account'}), 400
    conn = get_db()
    # Deactivate instead of hard-delete to preserve FK references
    conn.execute("UPDATE users SET is_active = 0, updated_at = datetime('now','localtime') WHERE id = ?", (uid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/admin/users/<int:uid>/reset-password', methods=['POST'])
@api_role_required('owner', 'admin')
def api_reset_user_password(uid):
    """Reset a user's password to 'password' and force change on next login."""
    conn = get_db()
    user = conn.execute('SELECT username, display_name FROM users WHERE id = ?', (uid,)).fetchone()
    if not user:
        conn.close()
        return jsonify({'error': 'User not found'}), 404
    conn.execute(
        'UPDATE users SET password_hash = ?, must_change_password = 1 WHERE id = ?',
        (generate_password_hash('password'), uid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'message': f'Password reset for {user["display_name"] or user["username"]}. Temporary password: password'})

# ─── Session Heartbeat ───────────────────────────────────────────

@app.route('/api/heartbeat', methods=['POST'])
@api_login_required
def api_heartbeat():
    try:
        conn = get_db()
        today = datetime.now().strftime('%Y-%m-%d')
        conn.execute(
            '''INSERT INTO user_sessions (user_id, session_date, first_seen, last_seen, page_views)
               VALUES (?, ?, datetime('now','localtime'), datetime('now','localtime'), 1)
               ON CONFLICT(user_id, session_date) DO UPDATE SET
                   last_seen = datetime('now','localtime'),
                   page_views = page_views + 1''',
            (session['user_id'], today)
        )
        conn.commit()
        # Check schedule notifications once per user per day
        sched_key = f'sched_notif_{session["user_id"]}_{today}'
        if not session.get(sched_key):
            try:
                events = [dict(r) for r in conn.execute(
                    "SELECT * FROM job_schedule_events WHERE status NOT IN ('Complete','Cancelled')"
                ).fetchall()]
                conn.close()
                if events:
                    _check_schedule_notifications(events)
                session[sched_key] = True
            except Exception:
                pass
        else:
            conn.close()
    except Exception:
        pass
    return jsonify({'ok': True, '_noToast': True})

# ─── Activity Log (Owner Only) ──────────────────────────────────

@app.route('/admin/activity-log')
@role_required('owner')
def admin_activity_log():
    return render_template('admin/activity_log.html')

@app.route('/api/admin/activity-log')
@api_role_required('owner')
def api_admin_activity_log():
    user_id = request.args.get('user_id')
    action = request.args.get('action')
    entity_type = request.args.get('entity_type')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    limit = min(int(request.args.get('limit', 500)), 2000)

    where, params = ['1=1'], []
    if user_id:
        where.append('a.user_id = ?'); params.append(int(user_id))
    if action:
        where.append('a.action = ?'); params.append(action)
    if entity_type:
        where.append('a.entity_type = ?'); params.append(entity_type)
    if date_from:
        where.append('a.created_at >= ?'); params.append(date_from)
    if date_to:
        where.append('a.created_at <= ?'); params.append(date_to + ' 23:59:59')
    params.append(limit)

    conn = get_db()
    rows = conn.execute(
        f'''SELECT a.*, u.display_name, u.username
            FROM activity_logs a
            JOIN users u ON u.id = a.user_id
            WHERE {' AND '.join(where)}
            ORDER BY a.created_at DESC
            LIMIT ?''',
        params
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/admin/user-stats')
@api_role_required('owner')
def api_admin_user_stats():
    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')

    users = conn.execute(
        'SELECT id, username, display_name, role FROM users WHERE is_active = 1 ORDER BY display_name'
    ).fetchall()

    stats = []
    for u in users:
        uid = u['id']
        total_actions = conn.execute('SELECT COUNT(*) FROM activity_logs WHERE user_id = ?', (uid,)).fetchone()[0]
        today_actions = conn.execute(
            "SELECT COUNT(*) FROM activity_logs WHERE user_id = ? AND created_at >= ?", (uid, today)
        ).fetchone()[0]
        week_actions = conn.execute(
            "SELECT COUNT(*) FROM activity_logs WHERE user_id = ? AND created_at >= ?", (uid, week_ago)
        ).fetchone()[0]

        # Time today from user_sessions
        s_today = conn.execute(
            'SELECT first_seen, last_seen, page_views FROM user_sessions WHERE user_id = ? AND session_date = ?',
            (uid, today)
        ).fetchone()
        time_today_min = 0
        if s_today:
            try:
                fs = datetime.strptime(s_today['first_seen'], '%Y-%m-%d %H:%M:%S')
                ls = datetime.strptime(s_today['last_seen'], '%Y-%m-%d %H:%M:%S')
                time_today_min = max(int((ls - fs).total_seconds() / 60), 0)
            except Exception:
                pass

        # Time this week
        week_sessions = conn.execute(
            'SELECT first_seen, last_seen FROM user_sessions WHERE user_id = ? AND session_date >= ?',
            (uid, week_ago)
        ).fetchall()
        time_week_min = 0
        for ws in week_sessions:
            try:
                fs = datetime.strptime(ws['first_seen'], '%Y-%m-%d %H:%M:%S')
                ls = datetime.strptime(ws['last_seen'], '%Y-%m-%d %H:%M:%S')
                time_week_min += max(int((ls - fs).total_seconds() / 60), 0)
            except Exception:
                pass

        last_active = conn.execute(
            'SELECT created_at FROM activity_logs WHERE user_id = ? ORDER BY created_at DESC LIMIT 1', (uid,)
        ).fetchone()

        stats.append({
            'user_id': uid,
            'username': u['username'],
            'display_name': u['display_name'],
            'role': u['role'],
            'total_actions': total_actions,
            'today_actions': today_actions,
            'week_actions': week_actions,
            'time_today_min': time_today_min,
            'time_week_min': time_week_min,
            'last_active': last_active['created_at'] if last_active else None,
        })

    conn.close()
    return jsonify(stats)

# Helper: list of users for dropdowns
@app.route('/api/users/list')
@api_login_required
def api_users_list():
    conn = get_db()
    users = conn.execute('SELECT id, username, display_name, role FROM users WHERE is_active = 1 ORDER BY display_name').fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])

# Helper: list of jobs for dropdowns
@app.route('/api/jobs/list')
@app.route('/api/jobs')
@api_login_required
def api_jobs_list():
    conn = get_db()
    jobs = conn.execute('SELECT id, name, status FROM jobs ORDER BY name').fetchall()
    conn.close()
    return jsonify([dict(j) for j in jobs])

# ─── Activity Logging ────────────────────────────────────────

def log_activity(user_id, action, entity_type='', entity_id=None, description=''):
    """Log a user action to the activity_logs table."""
    try:
        conn = get_db()
        conn.execute(
            '''INSERT INTO activity_logs (user_id, action, entity_type, entity_id, description, ip_address)
               VALUES (?,?,?,?,?,?)''',
            (user_id, action, entity_type, entity_id, description, request.remote_addr or '')
        )
        conn.commit()
        conn.close()
    except Exception:
        pass

# ─── Email Autocomplete ─────────────────────────────────────

def save_email_autocomplete(recipients):
    """Save/update email addresses for autocomplete suggestions."""
    conn = get_db()
    for email in recipients:
        conn.execute('''INSERT INTO email_autocomplete (email, used_count, last_used_at)
            VALUES (?, 1, datetime('now','localtime'))
            ON CONFLICT(email) DO UPDATE SET
                used_count = used_count + 1,
                last_used_at = datetime('now','localtime')''',
            (email.lower().strip(),))
    conn.commit()
    conn.close()

@app.route('/api/email-autocomplete')
@api_login_required
def api_email_autocomplete():
    q = request.args.get('q', '').strip()
    conn = get_db()
    if q:
        rows = conn.execute(
            'SELECT email FROM email_autocomplete WHERE email LIKE ? ORDER BY used_count DESC, last_used_at DESC LIMIT 10',
            (f'{q}%',)).fetchall()
    else:
        rows = conn.execute(
            'SELECT email FROM email_autocomplete ORDER BY used_count DESC, last_used_at DESC LIMIT 10').fetchall()
    conn.close()
    return jsonify({'emails': [r['email'] for r in rows]})

# ─── Notifications ──────────────────────────────────────────

def create_notification(user_id, ntype, title, message='', link=''):
    """Helper to create a notification for a user."""
    conn = get_db()
    conn.execute(
        'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
        (user_id, ntype, title, message, link)
    )
    conn.commit()
    conn.close()

@app.route('/api/notifications')
@api_login_required
def api_notifications():
    conn = get_db()
    notifs = conn.execute(
        '''SELECT * FROM notifications WHERE user_id = ?
           ORDER BY created_at DESC LIMIT 50''',
        (session['user_id'],)
    ).fetchall()
    conn.close()
    return jsonify([dict(n) for n in notifs])

@app.route('/api/notifications/unread-count')
@api_login_required
def api_notifications_unread_count():
    conn = get_db()
    count = conn.execute(
        'SELECT COUNT(*) FROM notifications WHERE user_id = ? AND is_read = 0',
        (session['user_id'],)
    ).fetchone()[0]
    conn.close()
    return jsonify({'count': count})

@app.route('/api/notifications/<int:nid>/read', methods=['PUT'])
@api_login_required
def api_notification_read(nid):
    conn = get_db()
    conn.execute(
        'UPDATE notifications SET is_read = 1 WHERE id = ? AND user_id = ?',
        (nid, session['user_id'])
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/notifications/mark-all-read', methods=['POST'])
@api_login_required
def api_notifications_mark_all_read():
    conn = get_db()
    conn.execute(
        'UPDATE notifications SET is_read = 1 WHERE user_id = ? AND is_read = 0',
        (session['user_id'],)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Bids ───────────────────────────────────────────────────

def calculate_bid(data):
    """Calculate all derived bid fields from system counts."""
    # System counts
    num_apartments = int(data.get('num_apartments', 0) or 0)
    num_non_apt = int(data.get('num_non_apartment_systems', 0) or 0)
    num_mini_splits = int(data.get('num_mini_splits', 0) or 0)
    has_clubhouse = int(data.get('has_clubhouse', 0) or 0)
    clubhouse_systems = int(data.get('clubhouse_systems', 0) or 0) if has_clubhouse else 0
    # Mini splits count as 0.75 of a standard system for labor calculation
    total_systems = num_apartments + num_non_apt + (num_mini_splits * 0.75) + clubhouse_systems

    # Man hours breakdown per system
    rough_in = float(data.get('rough_in_hours', 15) or 15)
    ahu = float(data.get('ahu_install_hours', 1) or 1)
    condenser = float(data.get('condenser_install_hours', 1) or 1)
    trim = float(data.get('trim_out_hours', 1) or 1)
    startup = float(data.get('startup_hours', 2) or 2)
    man_hours_per_system = rough_in + ahu + condenser + trim + startup
    total_man_hours = total_systems * man_hours_per_system

    # Labor
    crew_size = int(data.get('crew_size', 4) or 4)
    hours_per_day = float(data.get('hours_per_day', 8) or 8)
    labor_rate = float(data.get('labor_rate_per_hour', 0) or 0)
    labor_cost_per_unit = float(data.get('labor_cost_per_unit', 0) or 0)

    # Labor cost: override > per-unit > hourly rate
    labor_cost_override = float(data.get('labor_cost_override', 0) or 0)
    if labor_cost_override > 0:
        labor_cost = round(labor_cost_override, 2)
    elif labor_cost_per_unit > 0:
        labor_cost = round(total_systems * labor_cost_per_unit, 2)
    else:
        labor_cost = round(total_man_hours * labor_rate, 2)

    # Duration
    duration_days = round(total_man_hours / (crew_size * hours_per_day), 2) if crew_size > 0 and hours_per_day > 0 else 0
    num_weeks = round(duration_days / 5, 2)

    # Per diem (auto from mileage)
    job_mileage = float(data.get('job_mileage', 0) or 0)
    per_diem_rate = float(data.get('per_diem_rate', 0) or 0)
    # Auto-set per diem rate from mileage if rate is 0
    if per_diem_rate == 0 and job_mileage > 0:
        if job_mileage >= 250:
            per_diem_rate = 75
        elif job_mileage >= 101:
            per_diem_rate = 60
    per_diem_total = round(per_diem_rate * duration_days * crew_size, 2)

    # Material cost breakdown
    material_subtotal = float(data.get('material_subtotal', 0) or 0)
    material_shipping = float(data.get('material_shipping', 0) or 0)
    material_tax_rate = float(data.get('material_tax_rate', 0) or 0)
    material_tax_amount = round(material_subtotal * material_tax_rate / 100, 2)
    material_cost = round(material_subtotal + material_shipping + material_tax_amount, 2)
    # Backward compat: if no breakdown fields but material_cost was sent directly
    if material_subtotal == 0 and material_shipping == 0 and float(data.get('material_cost', 0) or 0) > 0:
        material_cost = float(data.get('material_cost', 0) or 0)
    insurance_cost = float(data.get('insurance_cost', 0) or 0)
    permit_cost = float(data.get('permit_cost', 0) or 0)
    management_fee = float(data.get('management_fee', 0) or 0)

    # Admin costs
    admin_costs = float(data.get('admin_costs', 0) or 0)

    # Housing (monthly rate × duration in months from bid timeline)
    housing_rate = float(data.get('housing_rate', 0) or 0)
    housing_months = round(num_weeks / 4.33, 2) if num_weeks > 0 else 0
    housing_total = round(housing_rate * housing_months, 2)

    # Totals
    total_cost_to_build = round(material_cost + labor_cost + insurance_cost + permit_cost + management_fee + per_diem_total + admin_costs + housing_total, 2)
    subtotal = total_cost_to_build
    profit_pct = float(data.get('company_profit_pct', 0) or 0)
    profit_mode = data.get('profit_mode', 'percentage') or 'percentage'
    profit_per_system = float(data.get('profit_per_system', 0) or 0)

    if profit_mode == 'per_system' and profit_per_system > 0:
        company_profit = round(total_systems * profit_per_system, 2)
    else:
        company_profit = round(subtotal * (profit_pct / 100), 2)
    total_bid = round(subtotal + company_profit, 2)
    net_profit = round(total_bid - total_cost_to_build, 2)

    # Per-unit calcs
    cost_per_apartment = round(total_cost_to_build / num_apartments, 2) if num_apartments > 0 else 0
    cost_per_system = round(total_cost_to_build / total_systems, 2) if total_systems > 0 else 0
    labor_cost_per_apt = round(labor_cost / num_apartments, 2) if num_apartments > 0 else 0
    labor_cost_per_sys = round(labor_cost / total_systems, 2) if total_systems > 0 else 0

    # Suggested bids (proportional split if clubhouse)
    if has_clubhouse and clubhouse_systems > 0 and total_systems > 0:
        apt_share = (total_systems - clubhouse_systems) / total_systems
        suggested_apartment_bid = round(total_bid * apt_share, 2)
        suggested_clubhouse_bid = round(total_bid * (1 - apt_share), 2)
    else:
        suggested_apartment_bid = total_bid
        suggested_clubhouse_bid = 0

    return {
        'total_systems': total_systems,
        'man_hours_per_system': man_hours_per_system,
        'total_man_hours': total_man_hours,
        'duration_days': duration_days,
        'num_weeks': num_weeks,
        'labor_cost': labor_cost,
        'material_cost': material_cost,
        'per_diem_rate': per_diem_rate,
        'per_diem_total': per_diem_total,
        'total_cost_to_build': total_cost_to_build,
        'subtotal': subtotal,
        'company_profit': company_profit,
        'total_bid': total_bid,
        'net_profit': net_profit,
        'cost_per_apartment': cost_per_apartment,
        'cost_per_system': cost_per_system,
        'labor_cost_per_apartment': labor_cost_per_apt,
        'labor_cost_per_system': labor_cost_per_sys,
        'suggested_apartment_bid': suggested_apartment_bid,
        'suggested_clubhouse_bid': suggested_clubhouse_bid,
        'housing_months': housing_months,
        'housing_total': housing_total,
    }

# ─── Bid Cost Stripping (non-owner) ──────────────────────────
_BID_COST_KEYS = {
    'total_bid', 'subtotal', 'labor_rate_per_hour', 'labor_cost_per_unit',
    'labor_cost', 'labor_cost_per_apartment', 'labor_cost_per_system',
    'company_profit_pct', 'company_profit', 'net_profit',
    'material_cost', 'total_cost_to_build', 'cost_per_apartment', 'cost_per_system',
    'insurance_cost', 'permit_cost', 'management_fee', 'pay_schedule_pct',
    'per_diem_rate', 'per_diem_days', 'per_diem_total', 'job_mileage',
    'rough_in_hours', 'ahu_install_hours', 'condenser_install_hours',
    'trim_out_hours', 'startup_hours', 'man_hours_per_system', 'total_man_hours',
    'crew_size', 'hours_per_day', 'duration_days', 'num_weeks',
    'suggested_apartment_bid', 'suggested_clubhouse_bid', 'price_per_ton',
    'profit_mode', 'profit_per_system', 'actual_bid_override', 'bid_type',
    'labor_cost_override', 'admin_costs', 'admin_costs_notes',
    'housing_rate', 'housing_months', 'housing_total',
}

def _strip_bid_costs(bid_dict):
    """Remove all financial/cost fields from a bid dict for non-owner roles."""
    for key in _BID_COST_KEYS:
        bid_dict.pop(key, None)
    bid_dict.pop('partners', None)
    bid_dict.pop('personnel', None)
    return bid_dict

@app.route('/bids')
@role_required('owner', 'admin', 'project_manager')
def bids_list():
    return render_template('bids/list.html')

@app.route('/bids/new')
@role_required('owner', 'admin', 'project_manager')
def bids_new():
    return render_template('bids/detail.html', bid_id=0)

@app.route('/bids/<int:bid_id>')
@role_required('owner', 'admin', 'project_manager')
def bids_detail(bid_id):
    return render_template('bids/detail.html', bid_id=bid_id)

@app.route('/api/bids')
@api_role_required('owner', 'admin', 'project_manager')
def api_bids():
    conn = get_db()
    bids = conn.execute(
        '''SELECT b.*, j.name as job_name FROM bids b
           LEFT JOIN jobs j ON b.job_id = j.id
           ORDER BY b.updated_at DESC'''
    ).fetchall()
    conn.close()

    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    is_owner = session.get('role') == 'owner'
    results = []
    for b in bids:
        d = dict(b)
        if os.path.isdir(proposals_dir):
            pdf_files = sorted([f for f in os.listdir(proposals_dir) if f.startswith('Proposal_') and f.endswith(f'_{b["id"]}.pdf')])
            if pdf_files:
                d['proposal_pdf'] = f'/api/bids/{b["id"]}/proposal/{pdf_files[-1]}'
        if not is_owner:
            _strip_bid_costs(d)
        results.append(d)

    return jsonify(results)

def _bid_fields(data, calcs):
    """Extract all bid fields from request data + calculated values."""
    f = lambda key, default=0: float(data.get(key, default) or default)
    i = lambda key, default=0: int(data.get(key, default) or default)
    s = lambda key, default='': data.get(key, default) or default
    return (
        data.get('job_id') or None, s('bid_name'), s('status', 'Draft'), s('project_type', 'Multi-Family'),
        i('num_apartments'), i('num_non_apartment_systems'), i('num_mini_splits'),
        i('has_clubhouse'), i('clubhouse_systems'), f('clubhouse_tons'), calcs['total_systems'],
        f('total_tons'), f('price_per_ton'), calcs['material_cost'],
        calcs['man_hours_per_system'], f('rough_in_hours', 15), f('ahu_install_hours', 1),
        f('condenser_install_hours', 1), f('trim_out_hours', 1), f('startup_hours', 2),
        calcs['total_man_hours'], i('crew_size', 4), f('hours_per_day', 8),
        calcs['duration_days'], calcs['num_weeks'],
        f('labor_rate_per_hour', 37), f('labor_cost_per_unit'), calcs['labor_cost'],
        f('job_mileage'), calcs['per_diem_rate'], calcs['duration_days'], calcs['per_diem_total'],
        f('insurance_cost'), f('permit_cost'), f('management_fee'), f('pay_schedule_pct', 0.33),
        f('company_profit_pct'), calcs['company_profit'], calcs['subtotal'], calcs['total_bid'],
        calcs['total_cost_to_build'], calcs['net_profit'],
        calcs['cost_per_apartment'], calcs['cost_per_system'],
        calcs['labor_cost_per_apartment'], calcs['labor_cost_per_system'],
        calcs['suggested_apartment_bid'], calcs['suggested_clubhouse_bid'],
        s('contracting_gc'), s('gc_attention'), s('bid_number'), s('bid_date'),
        s('bid_workup_date'), s('bid_due_date'), s('bid_submitted_date'), s('lead_name'),
        s('inclusions'), s('exclusions'), s('bid_description'), s('notes'),
        s('profit_mode', 'percentage'), f('profit_per_system'),
        f('actual_bid_override'), s('bid_type'),
        f('labor_cost_override'), f('admin_costs'), s('admin_costs_notes'),
        f('housing_rate'), calcs['housing_months'], calcs['housing_total'],
        f('material_subtotal'), f('material_shipping'), f('material_tax_rate'),
    )

_BID_INSERT_COLS = '''job_id, bid_name, status, project_type,
    num_apartments, num_non_apartment_systems, num_mini_splits,
    has_clubhouse, clubhouse_systems, clubhouse_tons, total_systems,
    total_tons, price_per_ton, material_cost,
    man_hours_per_system, rough_in_hours, ahu_install_hours,
    condenser_install_hours, trim_out_hours, startup_hours,
    total_man_hours, crew_size, hours_per_day,
    duration_days, num_weeks,
    labor_rate_per_hour, labor_cost_per_unit, labor_cost,
    job_mileage, per_diem_rate, per_diem_days, per_diem_total,
    insurance_cost, permit_cost, management_fee, pay_schedule_pct,
    company_profit_pct, company_profit, subtotal, total_bid,
    total_cost_to_build, net_profit,
    cost_per_apartment, cost_per_system,
    labor_cost_per_apartment, labor_cost_per_system,
    suggested_apartment_bid, suggested_clubhouse_bid,
    contracting_gc, gc_attention, bid_number, bid_date,
    bid_workup_date, bid_due_date, bid_submitted_date, lead_name,
    inclusions, exclusions, bid_description, notes,
    profit_mode, profit_per_system, actual_bid_override, bid_type,
    labor_cost_override, admin_costs, admin_costs_notes,
    housing_rate, housing_months, housing_total,
    material_subtotal, material_shipping, material_tax_rate'''

def _next_bid_number(conn):
    """Get next bid number starting from 2600."""
    row = conn.execute("SELECT MAX(CAST(bid_number AS INTEGER)) FROM bids WHERE bid_number GLOB '[0-9]*'").fetchone()
    max_num = row[0] or 0
    return str(max(max_num + 1, 2600))

@app.route('/api/bids/next-number')
@api_role_required('owner', 'admin', 'project_manager')
def api_next_bid_number():
    conn = get_db()
    num = _next_bid_number(conn)
    conn.close()
    return jsonify({'bid_number': num})

@app.route('/api/bids', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_create_bid():
    data = request.get_json()
    conn = get_db()
    # Auto-assign bid number if not provided
    if not data.get('bid_number', '').strip():
        data['bid_number'] = _next_bid_number(conn)
    calcs = calculate_bid(data)
    fields = _bid_fields(data, calcs)
    placeholders = ','.join(['?'] * (len(fields) + 2))  # +2 for created_by, customer_id
    customer_id = data.get('customer_id') or None
    cursor = conn.execute(
        f'INSERT INTO bids ({_BID_INSERT_COLS}, created_by, customer_id) VALUES ({placeholders})',
        fields + (session.get('user_id'), customer_id)
    )
    bid_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': bid_id}), 201

@app.route('/api/bids/<int:bid_id>')
@api_role_required('owner', 'admin', 'project_manager')
def api_bid_detail(bid_id):
    conn = get_db()
    bid = conn.execute(
        '''SELECT b.*, j.name as job_name FROM bids b
           LEFT JOIN jobs j ON b.job_id = j.id WHERE b.id = ?''', (bid_id,)
    ).fetchone()
    if not bid:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    partners = conn.execute('SELECT * FROM bid_partners WHERE bid_id = ? ORDER BY id', (bid_id,)).fetchall()
    personnel = conn.execute(
        '''SELECT bp.*, u.display_name as user_display_name FROM bid_personnel bp
           LEFT JOIN users u ON bp.user_id = u.id WHERE bp.bid_id = ? ORDER BY bp.id''', (bid_id,)
    ).fetchall()
    proposal_lines = conn.execute('SELECT * FROM bid_proposal_lines WHERE bid_id = ? ORDER BY sort_order, id', (bid_id,)).fetchall()
    conn.close()
    result = dict(bid)
    result['partners'] = [dict(p) for p in partners]
    result['personnel'] = [dict(p) for p in personnel]
    result['proposal_lines'] = [dict(l) for l in proposal_lines]

    # Check if a proposal PDF exists for this bid
    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    if os.path.isdir(proposals_dir):
        pdf_files = sorted([f for f in os.listdir(proposals_dir) if f.startswith('Proposal_') and f.endswith(f'_{bid_id}.pdf')])
        if pdf_files:
            result['proposal_pdf'] = f'/api/bids/{bid_id}/proposal/{pdf_files[-1]}'

    if session.get('role') != 'owner':
        _strip_bid_costs(result)

    return jsonify(result)

@app.route('/api/bids/<int:bid_id>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_bid(bid_id):
    data = request.get_json()
    calcs = calculate_bid(data)
    fields = _bid_fields(data, calcs)
    set_clause = ', '.join([f'{col.strip()}=?' for col in _BID_INSERT_COLS.split(',')])
    customer_id = data.get('customer_id') or None
    conn = get_db()
    conn.execute(
        f"UPDATE bids SET {set_clause}, customer_id=?, updated_at=datetime('now','localtime') WHERE id=?",
        fields + (customer_id, bid_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_bid(bid_id):
    conn = get_db()
    conn.execute('DELETE FROM bids WHERE id = ?', (bid_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/calculate', methods=['POST'])
@api_role_required('owner')
def api_bid_calculate(bid_id):
    data = request.get_json()
    calcs = calculate_bid(data)
    return jsonify(calcs)

@app.route('/api/bids/<int:bid_id>/partners', methods=['POST'])
@api_role_required('owner')
def api_add_bid_partner(bid_id):
    data = request.get_json()
    conn = get_db()
    cursor = conn.execute(
        'INSERT INTO bid_partners (bid_id, partner_name, profit_pct, profit_amount) VALUES (?,?,?,?)',
        (bid_id, data.get('partner_name', ''), float(data.get('profit_pct', 0) or 0),
         float(data.get('profit_amount', 0) or 0))
    )
    pid = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': pid}), 201

@app.route('/api/bids/<int:bid_id>/partners/<int:pid>', methods=['DELETE'])
@api_role_required('owner')
def api_delete_bid_partner(bid_id, pid):
    conn = get_db()
    conn.execute('DELETE FROM bid_partners WHERE id = ? AND bid_id = ?', (pid, bid_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/personnel', methods=['POST'])
@api_role_required('owner')
def api_add_bid_personnel(bid_id):
    data = request.get_json()
    conn = get_db()
    cursor = conn.execute(
        'INSERT INTO bid_personnel (bid_id, user_id, name, role, hourly_rate) VALUES (?,?,?,?,?)',
        (bid_id, data.get('user_id') or None, data.get('name', ''),
         data.get('role', ''), float(data.get('hourly_rate', 0) or 0))
    )
    pid = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': pid}), 201

@app.route('/api/bids/<int:bid_id>/personnel/<int:pid>', methods=['DELETE'])
@api_role_required('owner')
def api_delete_bid_personnel(bid_id, pid):
    conn = get_db()
    conn.execute('DELETE FROM bid_personnel WHERE id = ? AND bid_id = ?', (pid, bid_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Proposals (PDF + Email) ─────────────────────────────────

@app.route('/api/bids/<int:bid_id>/preview-proposal')
@api_role_required('owner', 'admin', 'project_manager')
def api_preview_proposal(bid_id):
    """Preview the proposal HTML before generating PDF."""
    conn = get_db()
    bid = conn.execute(
        'SELECT b.*, j.name as job_name FROM bids b LEFT JOIN jobs j ON b.job_id = j.id WHERE b.id = ?',
        (bid_id,)
    ).fetchone()
    if not bid:
        conn.close()
        return jsonify({'error': 'Bid not found'}), 404
    bid = dict(bid)
    proposal_lines = conn.execute('SELECT * FROM bid_proposal_lines WHERE bid_id = ? ORDER BY sort_order, id', (bid_id,)).fetchall()
    conn.close()
    today = datetime.now().strftime('%B %d, %Y')
    logo_path = os.path.abspath(os.path.join(app.static_folder, 'logo.jpg'))
    html = render_template('bids/proposal_pdf.html', bid=bid, today=today, logo_path='file://' + logo_path, proposal_lines=[dict(l) for l in proposal_lines])
    return html

@app.route('/api/bids/<int:bid_id>/generate-proposal', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_generate_proposal(bid_id):
    """Generate a PDF proposal from bid data."""
    conn = get_db()
    bid = conn.execute(
        'SELECT b.*, j.name as job_name FROM bids b LEFT JOIN jobs j ON b.job_id = j.id WHERE b.id = ?',
        (bid_id,)
    ).fetchone()
    if not bid:
        conn.close()
        return jsonify({'error': 'Bid not found'}), 404

    bid = dict(bid)
    proposal_lines = conn.execute('SELECT * FROM bid_proposal_lines WHERE bid_id = ? ORDER BY sort_order, id', (bid_id,)).fetchall()
    conn.close()
    today = datetime.now().strftime('%B %d, %Y')
    logo_path = os.path.abspath(os.path.join(app.static_folder, 'logo.jpg'))

    html = render_template('bids/proposal_pdf.html', bid=bid, today=today, logo_path='file://' + logo_path, proposal_lines=[dict(l) for l in proposal_lines])

    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    os.makedirs(proposals_dir, exist_ok=True)

    safe_name = ''.join(c if c.isalnum() or c in ' -_' else '' for c in (bid.get('bid_name') or 'proposal')).strip()
    filename = f"Proposal_{safe_name}_{bid_id}.pdf"
    filepath = os.path.join(proposals_dir, filename)

    try:
        wp = weasyprint.HTML(string=html, base_url=os.path.dirname(__file__))
        wp.write_pdf(filepath)
    except Exception as e:
        return jsonify({'error': f'PDF generation failed: {str(e)[:200]}'}), 500

    return jsonify({'ok': True, 'filename': filename, 'path': f'/api/bids/{bid_id}/proposal/{filename}'})


@app.route('/api/bids/<int:bid_id>/proposal/<filename>')
@api_role_required('owner', 'admin', 'project_manager')
def api_download_proposal(bid_id, filename):
    """View or download a generated proposal PDF."""
    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    filepath = os.path.join(proposals_dir, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    # Serve inline for preview (browser renders PDF), download param forces download
    if request.args.get('download'):
        return send_file(filepath, as_attachment=True, download_name=filename)
    return send_file(filepath, mimetype='application/pdf')


@app.route('/api/bids/<int:bid_id>/proposal-lines', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_save_proposal_lines(bid_id):
    """Save proposal line items (replaces all existing lines)."""
    data = request.get_json(force=True)
    lines = data.get('lines', [])
    conn = get_db()
    conn.execute('DELETE FROM bid_proposal_lines WHERE bid_id = ?', (bid_id,))
    for i, line in enumerate(lines):
        desc = (line.get('description') or '').strip()
        amount = float(line.get('amount', 0))
        if desc:
            conn.execute(
                'INSERT INTO bid_proposal_lines (bid_id, description, amount, sort_order) VALUES (?,?,?,?)',
                (bid_id, desc, amount, i)
            )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/email-proposal', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_email_proposal(bid_id):
    """Email a proposal PDF to specified recipients."""
    data = request.get_json()
    recipients = [e.strip() for e in data.get('recipients', []) if e.strip()]
    subject = data.get('subject', 'HVAC Installation Proposal')
    body_text = data.get('body', '')

    # Load saved SMTP settings as fallback (password is never sent to browser)
    saved_settings = {}
    settings_path = os.path.join(os.path.dirname(__file__), 'data', 'email_settings.json')
    if os.path.exists(settings_path):
        with open(settings_path) as f:
            saved_settings = json.load(f)

    smtp_host = data.get('smtp_host') or saved_settings.get('smtp_host', '')
    smtp_port = int(data.get('smtp_port') or saved_settings.get('smtp_port', 587) or 587)
    smtp_user = data.get('smtp_user') or saved_settings.get('smtp_user', '')
    smtp_pass = data.get('smtp_pass') or saved_settings.get('smtp_pass', '')
    from_email = data.get('from_email') or saved_settings.get('from_email', '') or smtp_user

    if not recipients:
        return jsonify({'error': 'No recipients specified'}), 400
    if not smtp_host or not smtp_user:
        return jsonify({'error': 'SMTP settings required. Configure in Settings or provide in request.'}), 400

    # Find the most recent proposal PDF for this bid
    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    pdf_files = [f for f in os.listdir(proposals_dir) if f.startswith('Proposal_') and f.endswith(f'_{bid_id}.pdf')]
    if not pdf_files:
        return jsonify({'error': 'No proposal PDF found. Generate the proposal first.'}), 404
    pdf_path = os.path.join(proposals_dir, sorted(pdf_files)[-1])

    try:
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = subject

        msg.attach(MIMEText(body_text, 'plain'))

        with open(pdf_path, 'rb') as f:
            part = MIMEBase('application', 'pdf')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(pdf_path)}"')
            msg.attach(part)

        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)

        save_email_autocomplete(recipients)
        return jsonify({'ok': True, 'sent_to': recipients})

    except Exception as e:
        return jsonify({'error': f'Email failed: {str(e)}'}), 500

# ─── Bid Follow-ups (Phase 4) ───────────────────────────────────

@app.route('/api/bids/<int:bid_id>/followups')
@api_role_required('owner', 'admin', 'project_manager')
def api_bid_followups(bid_id):
    conn = get_db()
    rows = conn.execute(
        '''SELECT bf.*, u.display_name as assigned_name FROM bid_followups bf
           LEFT JOIN users u ON bf.assigned_to = u.id
           WHERE bf.bid_id = ? ORDER BY bf.followup_date''', (bid_id,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/bids/<int:bid_id>/followups', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_create_bid_followup(bid_id):
    data = request.get_json()
    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO bid_followups (bid_id, followup_date, followup_type, notes, assigned_to, created_by)
           VALUES (?,?,?,?,?,?)''',
        (bid_id, data.get('followup_date',''), data.get('followup_type','Call'),
         data.get('notes',''), data.get('assigned_to') or None, session.get('user_id'))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cursor.lastrowid}), 201

@app.route('/api/bids/<int:bid_id>/followups/<int:fid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_bid_followup(bid_id, fid):
    data = request.get_json()
    conn = get_db()
    conn.execute(
        '''UPDATE bid_followups SET followup_date=?, followup_type=?, notes=?, result=?,
           status=?, assigned_to=? WHERE id=? AND bid_id=?''',
        (data.get('followup_date',''), data.get('followup_type','Call'), data.get('notes',''),
         data.get('result',''), data.get('status','Scheduled'),
         data.get('assigned_to') or None, fid, bid_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/followups/<int:fid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_bid_followup(bid_id, fid):
    conn = get_db()
    conn.execute('DELETE FROM bid_followups WHERE id = ? AND bid_id = ?', (fid, bid_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/followups/upcoming')
@api_role_required('owner', 'admin', 'project_manager')
def api_upcoming_followups():
    from datetime import datetime, timedelta
    today = datetime.now().strftime('%Y-%m-%d')
    week_out = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
    conn = get_db()
    rows = conn.execute(
        '''SELECT bf.*, b.bid_name, j.name as job_name, u.display_name as assigned_name
           FROM bid_followups bf
           JOIN bids b ON bf.bid_id = b.id
           LEFT JOIN jobs j ON b.job_id = j.id
           LEFT JOIN users u ON bf.assigned_to = u.id
           WHERE bf.status = 'Scheduled' AND bf.followup_date BETWEEN ? AND ?
           ORDER BY bf.followup_date''', (today, week_out)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/bids/<int:bid_id>/award', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_award_bid(bid_id):
    from datetime import datetime
    conn = get_db()
    bid = conn.execute('SELECT * FROM bids WHERE id = ?', (bid_id,)).fetchone()
    if not bid:
        conn.close()
        return jsonify({'error': 'Bid not found'}), 404
    job_id = bid['job_id']

    today = datetime.now().strftime('%Y-%m-%d')

    # Auto-create job from bid data if no job linked
    if not job_id:
        job_name = bid['bid_name'] or f'Bid #{bid_id}'
        # Pull address info from customer if available
        address = city = state = zip_code = ''
        tax_rate = 0
        customer_id = bid.get('customer_id')
        if customer_id:
            cust = conn.execute('SELECT * FROM customers WHERE id = ?', (customer_id,)).fetchone()
            if cust:
                address = cust['address'] or ''
                city = cust['city'] or ''
                state = cust['state'] or ''
                zip_code = cust['zip_code'] or ''
                if zip_code:
                    tax_info = lookup_tax(zip_code)
                    tax_rate = tax_info['tax_rate']
                    if not city:
                        city = tax_info['city']
                    if not state:
                        state = tax_info['state']
        cursor = conn.execute(
            'INSERT INTO jobs (name, status, address, city, state, zip_code, tax_rate, customer_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
            (job_name, 'Needs Bid', address, city, state, zip_code, tax_rate, customer_id)
        )
        job_id = cursor.lastrowid
        conn.execute('UPDATE bids SET job_id = ? WHERE id = ?', (job_id, bid_id))

    # 1. Set bid status to Accepted
    conn.execute("UPDATE bids SET status = 'Accepted', updated_at = datetime('now','localtime') WHERE id = ?", (bid_id,))
    # 2. Set job status to Awarded
    conn.execute("UPDATE jobs SET status = 'Awarded', awarded_date = ?, updated_at = datetime('now','localtime') WHERE id = ?", (today, job_id))
    # 3. Create precon meeting record
    existing_precon = conn.execute('SELECT id FROM precon_meetings WHERE job_id = ?', (job_id,)).fetchone()
    if not existing_precon:
        conn.execute(
            "INSERT INTO precon_meetings (job_id, status, created_by) VALUES (?, 'Not Scheduled', ?)",
            (job_id, session.get('user_id'))
        )
    # 4. Auto-create schedule phases
    phases = [
        ('Pre-Construction', 0), ('Rough-In', 1), ('Trim Out', 2), ('Startup', 3), ('Closeout', 4)
    ]
    for phase_name, sort in phases:
        existing = conn.execute('SELECT id FROM job_schedule_events WHERE job_id = ? AND phase_name = ?', (job_id, phase_name)).fetchone()
        if not existing:
            conn.execute(
                '''INSERT INTO job_schedule_events (job_id, phase_name, status, sort_order, created_by)
                   VALUES (?,?,?,?,?)''',
                (job_id, phase_name, 'Pending', sort, session.get('user_id'))
            )
    # 5. Seed default PM benchmarks
    default_benchmarks = [
        ('Pre-Construction', 'Precon meeting completed', 0),
        ('Pre-Construction', 'Contract signed', 1),
        ('Pre-Construction', 'Insurance submitted', 2),
        ('Pre-Construction', 'Permits applied for', 3),
        ('Pre-Construction', 'Material ordered', 4),
        ('Pre-Construction', 'Schedule set with GC', 5),
        ('Rough-In', 'Ductwork installed', 6),
        ('Rough-In', 'Line sets run', 7),
        ('Rough-In', 'Condensate drains installed', 8),
        ('Rough-In', 'Low voltage wire pulled', 9),
        ('Rough-In', 'Disconnects installed', 10),
        ('Rough-In', 'Gas piping complete', 11),
        ('Rough-In', 'Rough inspection passed', 12),
        ('Trim Out', 'AHUs installed', 13),
        ('Trim Out', 'Condensers set', 14),
        ('Trim Out', 'Registers and grilles installed', 15),
        ('Trim Out', 'Thermostats installed', 16),
        ('Trim Out', 'Refrigerant lines connected', 17),
        ('Trim Out', 'Condensate tested', 18),
        ('Startup', 'System startup complete', 19),
        ('Startup', 'Charge verified', 20),
        ('Startup', 'Airflow balanced', 21),
        ('Startup', 'Thermostat programmed', 22),
        ('Startup', 'Startup report completed', 23),
        ('Closeout', 'O&M manuals delivered', 24),
        ('Closeout', 'Warranty letters submitted', 25),
        ('Closeout', 'As-builts delivered', 26),
        ('Closeout', 'Lien waiver submitted', 27),
        ('Closeout', 'Final inspection passed', 28),
        ('Closeout', 'Punch list complete', 29),
    ]
    existing_benchmarks = conn.execute('SELECT COUNT(*) FROM pm_benchmarks WHERE job_id = ?', (job_id,)).fetchone()[0]
    if existing_benchmarks == 0:
        for phase, task, sort in default_benchmarks:
            conn.execute(
                'INSERT INTO pm_benchmarks (job_id, phase, task_name, sort_order) VALUES (?,?,?,?)',
                (job_id, phase, task, sort)
            )
    # 6. Seed 32-step pipeline
    seed_pipeline_steps(conn, job_id)
    # Mark bidding steps complete on award
    for s in range(1, 9):
        conn.execute("UPDATE job_pipeline_steps SET status='complete', completed_date=date('now','localtime') WHERE job_id=? AND step_number=? AND status='pending'", (job_id, s))
    # 7. Set project manager to user awarding the bid (if not already set)
    conn.execute("UPDATE jobs SET project_manager_id = COALESCE(project_manager_id, ?), customer_id = COALESCE(customer_id, ?) WHERE id = ?",
                 (session.get('user_id'), bid.get('customer_id'), job_id))

    # 8. Auto-create pay app contract (if none exists for this job)
    existing_contract = conn.execute('SELECT id FROM pay_app_contracts WHERE job_id = ?', (job_id,)).fetchone()
    if not existing_contract:
        gc_name = bid.get('contracting_gc') or ''
        gc_contact = bid.get('gc_attention') or ''
        gc_address = gc_email = gc_phone = ''
        cust_id = bid.get('customer_id')
        if cust_id:
            cust = conn.execute('SELECT * FROM customers WHERE id = ?', (cust_id,)).fetchone()
            if cust:
                gc_address = f"{cust['address'] or ''}, {cust['city'] or ''} {cust['state'] or ''} {cust['zip_code'] or ''}".strip(', ')
                gc_email = cust['contact_email'] or ''
                gc_phone = cust['contact_phone'] or ''
                if not gc_name:
                    gc_name = cust['company_name'] or ''
                if not gc_contact:
                    gc_contact = cust['primary_contact'] or ''
        job_row = conn.execute('SELECT name, address, city, state FROM jobs WHERE id = ?', (job_id,)).fetchone()
        project_name = job_row['name'] if job_row else bid['bid_name'] or ''
        project_address = f"{job_row['address'] or ''}, {job_row['city'] or ''} {job_row['state'] or ''}".strip(', ') if job_row else ''
        conn.execute(
            '''INSERT INTO pay_app_contracts (job_id, gc_name, gc_address, gc_contact, gc_email, gc_phone,
               project_name, project_address, contract_for, contract_date,
               original_contract_sum, retainage_work_pct, retainage_stored_pct, created_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (job_id, gc_name, gc_address, gc_contact, gc_email, gc_phone,
             project_name, project_address, 'HVAC', today,
             bid.get('total_bid') or 0, 10, 0, session.get('user_id'))
        )

    # 9. Auto-seed closeout checklist defaults
    existing_checklist = conn.execute('SELECT id FROM closeout_checklists WHERE job_id = ?', (job_id,)).fetchone()
    if not existing_checklist:
        closeout_defaults = [
            ('O&M Manual', 'O&M Manual'), ('Warranty Letter', 'Warranty Letter'),
            ('As-Built Drawings', 'As-Built'), ('Start-Up Report', 'Start-Up Report'),
            ('Balancing Report', 'Balancing Report'), ('Test Report', 'Test Report'),
            ('Lien Waiver', 'Lien Waiver'), ('Certificate of Completion', 'Certificate of Completion'),
            ('Permit Closeout', 'Permit'),
        ]
        for idx, (name, item_type) in enumerate(closeout_defaults):
            conn.execute(
                'INSERT INTO closeout_checklists (job_id, item_name, item_type, sort_order, created_by) VALUES (?,?,?,?,?)',
                (job_id, name, item_type, idx, session.get('user_id'))
            )

    # 10. Notify owners + PMs
    users = conn.execute("SELECT id FROM users WHERE role IN ('owner','project_manager') AND is_active = 1").fetchall()
    job = conn.execute('SELECT name FROM jobs WHERE id = ?', (job_id,)).fetchone()
    conn.commit()
    conn.close()
    for u in users:
        create_notification(u['id'], 'bid', f'Bid Awarded: {bid["bid_name"]}',
                          f'Job "{job["name"]}" has been awarded.', f'/projects/{job_id}')
    return jsonify({'ok': True, 'job_id': job_id})

# ─── Bid Takeoff ──────────────────────────────────────────────────

@app.route('/takeoffs')
@role_required('owner')
def takeoffs_list():
    return render_template('bids/takeoffs.html')

@app.route('/api/takeoffs')
@api_role_required('owner')
def api_takeoffs_list():
    conn = get_db()
    # Pull from projects, join to bids to find takeoff data
    jobs = conn.execute(
        '''SELECT j.id as job_id, j.name as job_name, j.status as job_status, j.city, j.state,
           j.updated_at,
           b.id as bid_id, b.bid_name, b.material_subtotal,
           (SELECT COUNT(*) FROM bid_takeoff_items WHERE bid_id = b.id) as item_count,
           (SELECT COUNT(*) FROM bid_takeoff_items WHERE bid_id = b.id AND enabled = 1) as enabled_count,
           (SELECT COUNT(*) FROM bid_takeoff_unit_types WHERE bid_id = b.id) as unit_type_count,
           (SELECT COALESCE(SUM(unit_count), 0) FROM bid_takeoff_unit_types WHERE bid_id = b.id) as total_units
           FROM jobs j
           LEFT JOIN bids b ON b.job_id = j.id AND b.id = (
               SELECT b2.id FROM bids b2 WHERE b2.job_id = j.id ORDER BY b2.updated_at DESC LIMIT 1
           )
           ORDER BY j.updated_at DESC'''
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in jobs])

@app.route('/api/projects/<int:job_id>/takeoff/start', methods=['POST'])
@api_role_required('owner')
def api_start_takeoff_for_project(job_id):
    """Find or create a bid for this project, return the bid_id for the takeoff page."""
    conn = get_db()
    # Find existing bid for this job
    bid = conn.execute('SELECT id FROM bids WHERE job_id = ? ORDER BY updated_at DESC LIMIT 1', (job_id,)).fetchone()
    if bid:
        bid_id = bid['id']
    else:
        # Auto-create a bid from the project
        job = conn.execute('SELECT name FROM jobs WHERE id = ?', (job_id,)).fetchone()
        job_name = job['name'] if job else f'Project #{job_id}'
        cursor = conn.execute(
            "INSERT INTO bids (job_id, bid_name, status, created_by) VALUES (?, ?, 'Draft', ?)",
            (job_id, job_name, session.get('user_id'))
        )
        bid_id = cursor.lastrowid
        conn.commit()
    conn.close()
    return jsonify({'ok': True, 'bid_id': bid_id})

@app.route('/api/bids/<int:bid_id>/takeoff', methods=['DELETE'])
@api_role_required('owner')
def api_delete_takeoff(bid_id):
    conn = get_db()
    conn.execute('DELETE FROM bid_takeoff_items WHERE bid_id = ?', (bid_id,))
    conn.execute('DELETE FROM bid_takeoff_unit_types WHERE bid_id = ?', (bid_id,))
    conn.execute("UPDATE bids SET takeoff_config = '{}' WHERE id = ?", (bid_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

DEFAULT_TAKEOFF_ITEMS = [
    # ═══ Rough-In — from Takeoff_Template.xlsm rows 12-56 ═══
    {'phase':'Rough-In','category':'CRD','part_name':'12x14 CRD W/165° Link','sku':'I-CRD50 12X14','unit_price':43,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':1},
    {'phase':'Rough-In','category':'CRD','part_name':'Fire/Smoke Radiation Damper 12"x14"','sku':'FSD-111-1','unit_price':608,'calc_basis':'fixed','qty_multiplier':0,'waste_pct':7.5,'enabled':0,'sort_order':2},
    {'phase':'Rough-In','category':'CRD','part_name':'6x12x8 90° CRD Boot W/165° Link','sku':'50CRD-95-BT','unit_price':35,'calc_basis':'per_total_drop','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':3},
    {'phase':'Rough-In','category':'Boots','part_name':'8" Foam Boot','sku':'L7001','unit_price':16.06,'calc_basis':'per_8in_drop','qty_multiplier':1,'waste_pct':7.5,'enabled':0,'sort_order':4},
    {'phase':'Rough-In','category':'Boots','part_name':'6" Foam Boot','sku':'','unit_price':12,'calc_basis':'per_6in_drop','qty_multiplier':1,'waste_pct':7.5,'enabled':0,'sort_order':5},
    {'phase':'Rough-In','category':'Exhaust Fan','part_name':'Broan 688','sku':'688','unit_price':18.45,'calc_basis':'per_bathroom','qty_multiplier':1,'waste_pct':7.5,'enabled':0,'sort_order':6},
    {'phase':'Rough-In','category':'CRD','part_name':'80CFM Exhaust Fan','sku':'QTXEG080','unit_price':100.50,'calc_basis':'per_bathroom','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':7},
    {'phase':'Rough-In','category':'CRD','part_name':'4" Round CRD 165FB','sku':'55CRD 4"','unit_price':32,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':0,'sort_order':8},
    {'phase':'Rough-In','category':'Duct Adapter','part_name':'6" Finger Saver ST Collar','sku':'L0090','unit_price':1.30,'calc_basis':'per_6in_drop','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':9},
    {'phase':'Rough-In','category':'Duct Adapter','part_name':'8" Finger Saver ST Collar','sku':'L0092','unit_price':1.65,'calc_basis':'per_8in_drop','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':10},
    {'phase':'Rough-In','category':'Duct Adapter','part_name':'8x6 Reducer 28GA','sku':'L0292','unit_price':5.80,'calc_basis':'per_6in_drop','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':11},
    {'phase':'Rough-In','category':'Flex','part_name':'6" x 25\' Foil Flex R6 Bag','sku':'L1972','unit_price':31.46,'calc_basis':'flex_6r6','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':12},
    {'phase':'Rough-In','category':'Flex','part_name':'6" x 25\' Foil Flex R8 Bag','sku':'L1939','unit_price':40.54,'calc_basis':'flex_6r8','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':13},
    {'phase':'Rough-In','category':'Flex','part_name':'8" x 25\' Foil Flex R6 Bag','sku':'L1974','unit_price':37.56,'calc_basis':'flex_8r6','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':14},
    {'phase':'Rough-In','category':'Flex','part_name':'8" x 25\' Foil Flex R8 Bag','sku':'L1941','unit_price':49.82,'calc_basis':'flex_8r8','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':15},
    {'phase':'Rough-In','category':'Flex','part_name':'3M Fire Wrap','sku':'SA','unit_price':170,'calc_basis':'per_system','qty_multiplier':1.5,'waste_pct':7.5,'enabled':0,'sort_order':16},
    {'phase':'Rough-In','category':'Flex','part_name':'FSK Duct Wrap R-6.0','sku':'L0475','unit_price':118,'calc_basis':'per_system','qty_multiplier':0.148,'waste_pct':7.5,'enabled':0,'sort_order':17},
    {'phase':'Rough-In','category':'Line Set','part_name':'3/4 x 3/8 Tube Insulation Proflex','sku':'L0484','unit_price':0.284,'calc_basis':'per_system','qty_multiplier':50,'waste_pct':7.5,'enabled':1,'sort_order':18},
    {'phase':'Rough-In','category':'Line Set','part_name':'3/8 OD x 50\' Refrig Tube','sku':'H0716','unit_price':57,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':19},
    {'phase':'Rough-In','category':'Line Set','part_name':'3/4 OD x 50\' Refrig Tube','sku':'H0719','unit_price':130,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':20},
    {'phase':'Rough-In','category':'Line Set','part_name':'Mini Split Line Set','sku':'','unit_price':217,'calc_basis':'fixed','qty_multiplier':3,'waste_pct':7.5,'enabled':0,'sort_order':21},
    {'phase':'Rough-In','category':'Line Set','part_name':'3" Hanging Duct Strap Silver 300\'','sku':'M0069','unit_price':8.50,'calc_basis':'per_system','qty_multiplier':0.1,'waste_pct':7.5,'enabled':0,'sort_order':22},
    {'phase':'Rough-In','category':'Round Pipe','part_name':'4" Adjustable 90 Ell','sku':'L0121','unit_price':2.99,'calc_basis':'adj_90','qty_multiplier':3,'waste_pct':7.5,'enabled':1,'sort_order':23},
    {'phase':'Rough-In','category':'Round Pipe','part_name':'3" Adjustable 90','sku':'L0120','unit_price':2.99,'calc_basis':'per_bathroom','qty_multiplier':1,'waste_pct':7.5,'enabled':0,'sort_order':24},
    {'phase':'Rough-In','category':'Round Pipe','part_name':'4x3 Reducer','sku':'','unit_price':5.65,'calc_basis':'per_bathroom','qty_multiplier':1,'waste_pct':7.5,'enabled':0,'sort_order':25},
    {'phase':'Rough-In','category':'Round Pipe','part_name':'3" Conductor Pipe','sku':'L0463','unit_price':1.25,'calc_basis':'per_bathroom','qty_multiplier':2,'waste_pct':7.5,'enabled':0,'sort_order':26},
    {'phase':'Rough-In','category':'Round Pipe','part_name':'4" Conductor Pipe','sku':'L0464','unit_price':1.60,'calc_basis':'conductor_pipe','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':27},
    {'phase':'Rough-In','category':'Shorts & Smalls','part_name':'14/4 SOOWA Cord 250\'','sku':'P3833','unit_price':1.495,'calc_basis':'fixed','qty_multiplier':150,'waste_pct':7.5,'enabled':0,'sort_order':28},
    {'phase':'Rough-In','category':'Shorts & Smalls','part_name':'18/8 Thermostat Wire Plenum 500\'','sku':'P2240','unit_price':0.565,'calc_basis':'per_system','qty_multiplier':100,'waste_pct':7.5,'enabled':1,'sort_order':29},
    {'phase':'Rough-In','category':'Shorts & Smalls','part_name':'4.25" Dryer Box 2x6','sku':'L3526','unit_price':32.50,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':30},
    {'phase':'Rough-In','category':'Shorts & Smalls','part_name':'1x48x120 Duct Board','sku':'L0470','unit_price':58,'calc_basis':'per_ductboard_config','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':31},
    {'phase':'Rough-In','category':'Shorts & Smalls','part_name':'#8 x 3/4" Hex Washer Screw','sku':'Q4659','unit_price':32.65,'calc_basis':'per_system','qty_multiplier':0.1,'waste_pct':7.5,'enabled':1,'sort_order':32},
    {'phase':'Rough-In','category':'Shorts & Smalls','part_name':'5"x18" 16GA Boca Plate Strap','sku':'M1322','unit_price':4.65,'calc_basis':'per_system','qty_multiplier':6,'waste_pct':7.5,'enabled':1,'sort_order':33},
    {'phase':'Rough-In','category':'Shorts & Smalls','part_name':'Black Duct Tape 2" 60YDS','sku':'L0444','unit_price':5.60,'calc_basis':'per_system','qty_multiplier':0.25,'waste_pct':7.5,'enabled':1,'sort_order':34},
    {'phase':'Rough-In','category':'Shorts & Smalls','part_name':'Foil Tape 2.5" x 60YDS','sku':'L0494','unit_price':15.15,'calc_basis':'per_system','qty_multiplier':1.5,'waste_pct':7.5,'enabled':1,'sort_order':35},
    {'phase':'Rough-In','category':'Shorts & Smalls','part_name':'Flex Fix 3" x 120YDS','sku':'L0445','unit_price':13.17,'calc_basis':'per_system','qty_multiplier':0.5,'waste_pct':7.5,'enabled':1,'sort_order':36},
    {'phase':'Rough-In','category':'Shorts & Smalls','part_name':'3/4 x 100\' Galv Hanger Iron Strap','sku':'M0091','unit_price':9,'calc_basis':'per_system','qty_multiplier':0.25,'waste_pct':7.5,'enabled':1,'sort_order':37},
    {'phase':'Rough-In','category':'Shorts & Smalls','part_name':'26" Snap-On Rails','sku':'L0158','unit_price':0.70,'calc_basis':'rails','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':38},
    {'phase':'Rough-In','category':'Shorts & Smalls','part_name':'36" Nylon Duct Strap','sku':'L0526','unit_price':0.30,'calc_basis':'zip_ties','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':39},
    {'phase':'Rough-In','category':'Tools','part_name':'12" 6T Rough-In Blade 5PK','sku':'T3710','unit_price':26,'calc_basis':'per_system','qty_multiplier':0.05,'waste_pct':7.5,'enabled':1,'sort_order':40},
    {'phase':'Rough-In','category':'Tools','part_name':'4-3/8" Hole Saw','sku':'T0996','unit_price':44,'calc_basis':'per_system','qty_multiplier':0.1,'waste_pct':7.5,'enabled':1,'sort_order':41},
    {'phase':'Rough-In','category':'Other','part_name':'Metacaulk MC150+ 5G Bucket','sku':'M1346','unit_price':258,'calc_basis':'per_system','qty_multiplier':0.071,'waste_pct':7.5,'enabled':1,'sort_order':42},
    {'phase':'Rough-In','category':'Other','part_name':'Gray 1GAL Duct Sealant','sku':'M6003','unit_price':13,'calc_basis':'per_system','qty_multiplier':0.25,'waste_pct':7.5,'enabled':1,'sort_order':43},
    {'phase':'Rough-In','category':'Other','part_name':'10.1oz White Silicone','sku':'M1842','unit_price':5.90,'calc_basis':'per_venthood','qty_multiplier':0.167,'waste_pct':7.5,'enabled':1,'sort_order':44},
    {'phase':'Rough-In','category':'Other','part_name':'2-1/2" Paint Chip Brush','sku':'M0725','unit_price':0.70,'calc_basis':'per_system','qty_multiplier':0.25,'waste_pct':7.5,'enabled':1,'sort_order':45},

    # ═══ Trim Out — from Takeoff_Template.xlsm rows 57-96 ═══
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'3" 2-Piece Pump-Up','sku':'L0681','unit_price':3.80,'calc_basis':'per_system','qty_multiplier':4,'waste_pct':7.5,'enabled':1,'sort_order':1},
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'6" Pump Ups','sku':'L3311','unit_price':3.98,'calc_basis':'per_system','qty_multiplier':6,'waste_pct':7.5,'enabled':0,'sort_order':2},
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'1QT Plumb-Tite Blue Cement','sku':'R0042','unit_price':14.95,'calc_basis':'per_system','qty_multiplier':0.1,'waste_pct':7.5,'enabled':1,'sort_order':3},
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'3/4 x 20\' PVC SCH 40 Pipe','sku':'R0071','unit_price':0.34,'calc_basis':'per_system','qty_multiplier':10,'waste_pct':7.5,'enabled':1,'sort_order':4},
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'3/4 PVC 90 Ell','sku':'R0311','unit_price':0.40,'calc_basis':'per_system','qty_multiplier':6,'waste_pct':7.5,'enabled':1,'sort_order':5},
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'3/4 PVC Coupling','sku':'R0341','unit_price':0.35,'calc_basis':'per_system','qty_multiplier':6,'waste_pct':7.5,'enabled':1,'sort_order':6},
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'3/4 PVC Male Adapter','sku':'R0351','unit_price':0.40,'calc_basis':'per_system','qty_multiplier':2,'waste_pct':7.5,'enabled':1,'sort_order':7},
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'3/4" P-Trap Cond Drain PVC','sku':'L0680','unit_price':1.62,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':8},
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'3/4 PVC Tee','sku':'R0331','unit_price':0.50,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':9},
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'Condensate Switch Elbow Pipe Mount','sku':'K2501','unit_price':17.70,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':10},
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'Orange Wire Nuts Jar 100','sku':'P1491','unit_price':5.48,'calc_basis':'per_system','qty_multiplier':0.08,'waste_pct':7.5,'enabled':1,'sort_order':11},
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'1# 15% Silver Solder','sku':'M0434','unit_price':97,'calc_basis':'per_system','qty_multiplier':0.05,'waste_pct':7.5,'enabled':1,'sort_order':12},
    {'phase':'Trim Out','category':'Drain','part_name':'30x30 Drain Pan','sku':'','unit_price':38,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':0,'sort_order':13},
    {'phase':'Trim Out','category':'Drain','part_name':'30x60 Drain Pan','sku':'','unit_price':44,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':0,'sort_order':14},
    {'phase':'Trim Out','category':'Brazing','part_name':'Acetylene 10 Refill','sku':'M3511','unit_price':20,'calc_basis':'per_system','qty_multiplier':0.1,'waste_pct':7.5,'enabled':1,'sort_order':15},
    {'phase':'Trim Out','category':'Brazing','part_name':'Oxygen 20 Refill','sku':'M3513','unit_price':10,'calc_basis':'per_system','qty_multiplier':0.1,'waste_pct':7.5,'enabled':1,'sort_order':16},
    {'phase':'Trim Out','category':'Refrigerant','part_name':'R454B Refrigerant 20# Drum','sku':'K0286','unit_price':371.61,'calc_basis':'per_system','qty_multiplier':0.067,'waste_pct':7.5,'enabled':1,'sort_order':17},
    {'phase':'Trim Out','category':'Fittings','part_name':'Silver Locking Caps 50 Pak','sku':'','unit_price':289,'calc_basis':'per_system','qty_multiplier':0.04,'waste_pct':7.5,'enabled':1,'sort_order':18},
    {'phase':'Trim Out','category':'Shorts & Smalls','part_name':'Tie Wire','sku':'','unit_price':15.65,'calc_basis':'per_system','qty_multiplier':0.1,'waste_pct':7.5,'enabled':0,'sort_order':19},
    {'phase':'Trim Out','category':'Hardware','part_name':'#10 x 1" Anchor Kit 24ct','sku':'M0051','unit_price':7.22,'calc_basis':'per_venthood','qty_multiplier':0.04,'waste_pct':7.5,'enabled':1,'sort_order':20},
    {'phase':'Trim Out','category':'Covers','part_name':'4" Black Venthood Widemouth','sku':'L3516','unit_price':17,'calc_basis':'venthood_covers','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':21},
    {'phase':'Trim Out','category':'Filters','part_name':'20x20x1 Fiberglass Filter','sku':'L0453','unit_price':2.13,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':22},
    {'phase':'Trim Out','category':'Register/Grill','part_name':'12"x6" 3-Way Stamped Supply Grill','sku':'L1736','unit_price':5.40,'calc_basis':'per_total_drop','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':23},
    {'phase':'Trim Out','category':'Register/Grill','part_name':'8" Supply Register','sku':'','unit_price':6.11,'calc_basis':'per_8in_drop','qty_multiplier':1,'waste_pct':7.5,'enabled':0,'sort_order':24},
    {'phase':'Trim Out','category':'Register/Grill','part_name':'6" Supply Register','sku':'','unit_price':4.79,'calc_basis':'per_6in_drop','qty_multiplier':1,'waste_pct':7.5,'enabled':0,'sort_order':25},
    {'phase':'Trim Out','category':'Register/Grill','part_name':'16x8 Return Grill (Pass Through)','sku':'L2280','unit_price':8.90,'calc_basis':'per_bedroom','qty_multiplier':2,'waste_pct':7.5,'enabled':0,'sort_order':26},
    {'phase':'Trim Out','category':'Register/Grill','part_name':'24"x12" Stamped Return Grille','sku':'L1778','unit_price':7.40,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':27},
    {'phase':'Trim Out','category':'Register/Grill','part_name':'30x14 Stamped Return','sku':'','unit_price':20,'calc_basis':'per_system','qty_multiplier':0,'waste_pct':7.5,'enabled':0,'sort_order':28},
    {'phase':'Trim Out','category':'Register/Grill','part_name':'30x20 Stamped Return','sku':'L9989','unit_price':18.10,'calc_basis':'per_system','qty_multiplier':0,'waste_pct':7.5,'enabled':0,'sort_order':29},
    {'phase':'Trim Out','category':'Zoning','part_name':'Two Story Zone','sku':'','unit_price':550,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':30},

    # ═══ Equipment — from Takeoff_Template.xlsm rows 97-142 ═══
    {'phase':'Equipment','category':'Air Handlers','part_name':'1.5T Front Return AHU 5KW R454B','sku':'K3730','unit_price':500,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':1.5,'waste_pct':0,'enabled':1,'sort_order':1},
    {'phase':'Equipment','category':'Air Handlers','part_name':'2.0 Ton Air Handler','sku':'','unit_price':506,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':2.0,'waste_pct':0,'enabled':1,'sort_order':2},
    {'phase':'Equipment','category':'Air Handlers','part_name':'2.5 Ton Air Handler','sku':'','unit_price':665.40,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':2.5,'waste_pct':0,'enabled':1,'sort_order':3},
    {'phase':'Equipment','category':'Air Handlers','part_name':'3.0 Ton Air Handler','sku':'','unit_price':770,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':3.0,'waste_pct':0,'enabled':1,'sort_order':4},
    {'phase':'Equipment','category':'Air Handlers','part_name':'3.5 Ton Air Handler','sku':'','unit_price':864,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':3.5,'waste_pct':0,'enabled':1,'sort_order':5},
    {'phase':'Equipment','category':'Air Handlers','part_name':'4.0 Ton Air Handler','sku':'','unit_price':958,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':4.0,'waste_pct':0,'enabled':1,'sort_order':6},
    {'phase':'Equipment','category':'Air Handlers','part_name':'5.0 Ton Air Handler','sku':'','unit_price':1195.84,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':5.0,'waste_pct':0,'enabled':1,'sort_order':7},
    {'phase':'Equipment','category':'Condensers','part_name':'1.5 Ton Condenser','sku':'','unit_price':820,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':1.5,'waste_pct':0,'enabled':1,'sort_order':8},
    {'phase':'Equipment','category':'Condensers','part_name':'2.0 Ton Condenser','sku':'','unit_price':883,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':2.0,'waste_pct':0,'enabled':1,'sort_order':9},
    {'phase':'Equipment','category':'Condensers','part_name':'2.5 Ton Condenser','sku':'','unit_price':1082.34,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':2.5,'waste_pct':0,'enabled':1,'sort_order':10},
    {'phase':'Equipment','category':'Condensers','part_name':'3.0 Ton Condenser','sku':'','unit_price':1053,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':3.0,'waste_pct':0,'enabled':1,'sort_order':11},
    {'phase':'Equipment','category':'Condensers','part_name':'3.5 Ton Condenser','sku':'','unit_price':1170,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':3.5,'waste_pct':0,'enabled':1,'sort_order':12},
    {'phase':'Equipment','category':'Condensers','part_name':'4.0 Ton Condenser','sku':'','unit_price':1287,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':4.0,'waste_pct':0,'enabled':1,'sort_order':13},
    {'phase':'Equipment','category':'Condensers','part_name':'5.0 Ton Condenser','sku':'','unit_price':1920.60,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':5.0,'waste_pct':0,'enabled':1,'sort_order':14},
    {'phase':'Equipment','category':'Heat Pumps','part_name':'1.5T Heat Pump 208-230V','sku':'K0529','unit_price':1090,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':1.5,'waste_pct':0,'enabled':0,'sort_order':15},
    {'phase':'Equipment','category':'Heat Pumps','part_name':'2.0 Ton Heat Pump','sku':'','unit_price':1095,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':2.0,'waste_pct':0,'enabled':0,'sort_order':16},
    {'phase':'Equipment','category':'Heat Pumps','part_name':'2.5 Ton Heat Pump','sku':'','unit_price':1190,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':2.5,'waste_pct':0,'enabled':0,'sort_order':17},
    {'phase':'Equipment','category':'Heat Pumps','part_name':'3.0 Ton Heat Pump','sku':'','unit_price':1465,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':3.0,'waste_pct':0,'enabled':0,'sort_order':18},
    {'phase':'Equipment','category':'Heat Pumps','part_name':'3.5 Ton Heat Pump','sku':'','unit_price':1610,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':3.5,'waste_pct':0,'enabled':0,'sort_order':19},
    {'phase':'Equipment','category':'Heat Pumps','part_name':'4.0 Ton Heat Pump','sku':'','unit_price':1755,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':4.0,'waste_pct':0,'enabled':0,'sort_order':20},
    {'phase':'Equipment','category':'Heat Pumps','part_name':'5.0 Ton Heat Pump','sku':'','unit_price':2148,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':5.0,'waste_pct':0,'enabled':0,'sort_order':21},
    {'phase':'Equipment','category':'Thermostat','part_name':'Heat Pump Thermostat','sku':'L0741','unit_price':39,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':0,'sort_order':22},
    {'phase':'Equipment','category':'Thermostat','part_name':'Lyric T6 Programmable Thermostat','sku':'L5286','unit_price':52.50,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':7.5,'enabled':1,'sort_order':23},

    # ═══ Startup/Other — from Takeoff_Template.xlsm rows 10-11, 144 ═══
    {'phase':'Startup/Other','category':'License/Permit','part_name':'License','sku':'','unit_price':175,'calc_basis':'fixed','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':1},
    {'phase':'Startup/Other','category':'License/Permit','part_name':'Permits','sku':'','unit_price':74.50,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':2},
    {'phase':'Startup/Other','category':'Shipping','part_name':'Freight to Jobsite','sku':'','unit_price':3000,'calc_basis':'fixed','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':3},
    {'phase':'Startup/Other','category':'Mini Split','part_name':'9K Mini Split Indoor','sku':'L9078','unit_price':255,'calc_basis':'fixed','qty_multiplier':0,'waste_pct':0,'enabled':0,'sort_order':4},
    {'phase':'Startup/Other','category':'Mini Split','part_name':'9K Mini Split Outdoor','sku':'L9071','unit_price':615,'calc_basis':'fixed','qty_multiplier':0,'waste_pct':0,'enabled':0,'sort_order':5},
    {'phase':'Startup/Other','category':'Mini Split','part_name':'12K Mini Split Indoor','sku':'L9080','unit_price':237,'calc_basis':'fixed','qty_multiplier':0,'waste_pct':0,'enabled':0,'sort_order':6},
    {'phase':'Startup/Other','category':'Mini Split','part_name':'12K Mini Split Outdoor','sku':'L9073','unit_price':565,'calc_basis':'fixed','qty_multiplier':0,'waste_pct':0,'enabled':0,'sort_order':7},
    {'phase':'Startup/Other','category':'Mini Split','part_name':'18K Mini Split Indoor','sku':'L6448','unit_price':573,'calc_basis':'fixed','qty_multiplier':0,'waste_pct':0,'enabled':0,'sort_order':8},
    {'phase':'Startup/Other','category':'Mini Split','part_name':'18K Mini Split Outdoor HP','sku':'L6432','unit_price':1140,'calc_basis':'fixed','qty_multiplier':0,'waste_pct':0,'enabled':0,'sort_order':9},
    {'phase':'Startup/Other','category':'Mini Split','part_name':'24K Mini Split Indoor','sku':'L9082','unit_price':305,'calc_basis':'fixed','qty_multiplier':0,'waste_pct':0,'enabled':0,'sort_order':10},
    {'phase':'Startup/Other','category':'Mini Split','part_name':'24K Mini Split Outdoor','sku':'L9075','unit_price':720,'calc_basis':'fixed','qty_multiplier':0,'waste_pct':0,'enabled':0,'sort_order':11},

    # ═══ Suggested Parts — common HVAC apartment items not in standard takeoff ═══
    {'phase':'Suggested Parts','category':'Electrical','part_name':'30A Non-Fuse Disconnect','sku':'','unit_price':18,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':1},
    {'phase':'Suggested Parts','category':'Electrical','part_name':'6\' Whip Kit 3/4" x 6\'','sku':'','unit_price':12,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':2},
    {'phase':'Suggested Parts','category':'Electrical','part_name':'60A Non-Fuse Disconnect','sku':'','unit_price':24,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':3},
    {'phase':'Suggested Parts','category':'Pads & Supports','part_name':'Condenser Pad 24x36x3','sku':'','unit_price':28,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':4},
    {'phase':'Suggested Parts','category':'Pads & Supports','part_name':'Condenser Pad 36x36x3','sku':'','unit_price':38,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':5},
    {'phase':'Suggested Parts','category':'Pads & Supports','part_name':'Anti-Vibration Pads 4pk','sku':'','unit_price':8,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':6},
    {'phase':'Suggested Parts','category':'Pads & Supports','part_name':'Roof Condenser Stand/Curb','sku':'','unit_price':185,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':7},
    {'phase':'Suggested Parts','category':'Pads & Supports','part_name':'Wall Hanger Bracket Kit','sku':'','unit_price':75,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':8},
    {'phase':'Suggested Parts','category':'Penetrations','part_name':'4" Roof Flashing Boot','sku':'','unit_price':12,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':9},
    {'phase':'Suggested Parts','category':'Penetrations','part_name':'3" Roof Flashing Boot','sku':'','unit_price':10,'calc_basis':'per_bathroom','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':10},
    {'phase':'Suggested Parts','category':'Penetrations','part_name':'Galv Vent Screen 4"','sku':'','unit_price':2.50,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':11},
    {'phase':'Suggested Parts','category':'Gas','part_name':'1/2" Gas Flex Connector 24"','sku':'','unit_price':16,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':12},
    {'phase':'Suggested Parts','category':'Gas','part_name':'3/4" Gas Flex Connector 36"','sku':'','unit_price':22,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':13},
    {'phase':'Suggested Parts','category':'Gas','part_name':'Gas Shut-Off Valve 1/2"','sku':'','unit_price':9,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':14},
    {'phase':'Suggested Parts','category':'Condensate','part_name':'Condensate Pump','sku':'','unit_price':65,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':15},
    {'phase':'Suggested Parts','category':'Condensate','part_name':'1/4" Vinyl Tubing 50\'','sku':'','unit_price':8,'calc_basis':'per_system','qty_multiplier':0.2,'waste_pct':0,'enabled':0,'sort_order':16},
    {'phase':'Suggested Parts','category':'Line Set Fittings','part_name':'1/4" Flare Nut','sku':'','unit_price':1.50,'calc_basis':'per_system','qty_multiplier':4,'waste_pct':0,'enabled':0,'sort_order':17},
    {'phase':'Suggested Parts','category':'Line Set Fittings','part_name':'3/8" Flare Nut','sku':'','unit_price':1.75,'calc_basis':'per_system','qty_multiplier':4,'waste_pct':0,'enabled':0,'sort_order':18},
    {'phase':'Suggested Parts','category':'Line Set Fittings','part_name':'1/2" Flare Nut','sku':'','unit_price':2.25,'calc_basis':'per_system','qty_multiplier':4,'waste_pct':0,'enabled':0,'sort_order':19},
    {'phase':'Suggested Parts','category':'Line Set Fittings','part_name':'3/4" Flare Nut','sku':'','unit_price':2.75,'calc_basis':'per_system','qty_multiplier':4,'waste_pct':0,'enabled':0,'sort_order':20},
    {'phase':'Suggested Parts','category':'Venting','part_name':'Range Hood Duct Kit 3.25x10','sku':'','unit_price':22,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':21},
    {'phase':'Suggested Parts','category':'Venting','part_name':'Dryer Vent Hose 4" x 8\'','sku':'','unit_price':14,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':22},
    {'phase':'Suggested Parts','category':'Safety','part_name':'Earthquake/Seismic Strap Kit','sku':'','unit_price':24,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':23},
    {'phase':'Suggested Parts','category':'Safety','part_name':'CO Detector Combo','sku':'','unit_price':35,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':24},
    {'phase':'Suggested Parts','category':'Dampers','part_name':'6" Round Balancing Damper','sku':'','unit_price':5.50,'calc_basis':'per_6in_drop','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':25},
    {'phase':'Suggested Parts','category':'Dampers','part_name':'8" Round Balancing Damper','sku':'','unit_price':6.50,'calc_basis':'per_8in_drop','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':26},
    {'phase':'Suggested Parts','category':'Dampers','part_name':'Motorized Zone Damper 10"','sku':'','unit_price':85,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':27},
    {'phase':'Suggested Parts','category':'Insulation','part_name':'Insulation Hanger 12" 100pk','sku':'','unit_price':32,'calc_basis':'per_system','qty_multiplier':0.05,'waste_pct':0,'enabled':0,'sort_order':28},
    {'phase':'Suggested Parts','category':'Insulation','part_name':'3/8" Armaflex Pipe Insulation 6\'','sku':'','unit_price':4.50,'calc_basis':'per_system','qty_multiplier':2,'waste_pct':0,'enabled':0,'sort_order':29},
    {'phase':'Suggested Parts','category':'Consumables','part_name':'Nitrogen Tank Refill','sku':'','unit_price':35,'calc_basis':'per_system','qty_multiplier':0.05,'waste_pct':0,'enabled':0,'sort_order':30},
    {'phase':'Suggested Parts','category':'Consumables','part_name':'Nylog Gasket/Thread Sealant','sku':'','unit_price':12,'calc_basis':'per_system','qty_multiplier':0.04,'waste_pct':0,'enabled':0,'sort_order':31},
    {'phase':'Suggested Parts','category':'Consumables','part_name':'Vacuum Pump Oil Quart','sku':'','unit_price':16,'calc_basis':'per_system','qty_multiplier':0.04,'waste_pct':0,'enabled':0,'sort_order':32},
    {'phase':'Suggested Parts','category':'Sheet Metal','part_name':'24x24x8 Drop-In Ceiling Box','sku':'','unit_price':18,'calc_basis':'per_total_drop','qty_multiplier':1,'waste_pct':7.5,'enabled':0,'sort_order':33},
    {'phase':'Suggested Parts','category':'Sheet Metal','part_name':'Return Air Box','sku':'','unit_price':45,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':34},
    {'phase':'Suggested Parts','category':'Sheet Metal','part_name':'Supply Plenum','sku':'','unit_price':55,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':35},
]

@app.route('/bids/<int:bid_id>/takeoff')
@role_required('owner')
def bids_takeoff(bid_id):
    return render_template('bids/takeoff.html', bid_id=bid_id)

@app.route('/api/bids/<int:bid_id>/takeoff/unit-types')
@api_role_required('owner')
def api_takeoff_unit_types(bid_id):
    conn = get_db()
    rows = conn.execute('SELECT * FROM bid_takeoff_unit_types WHERE bid_id = ? ORDER BY sort_order', (bid_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/bids/<int:bid_id>/takeoff/unit-types', methods=['POST'])
@api_role_required('owner')
def api_create_takeoff_unit_type(bid_id):
    d = request.get_json()
    conn = get_db()
    max_sort = conn.execute('SELECT COALESCE(MAX(sort_order),0) FROM bid_takeoff_unit_types WHERE bid_id = ?', (bid_id,)).fetchone()[0]
    cursor = conn.execute(
        '''INSERT INTO bid_takeoff_unit_types (bid_id, name, unit_count, bedrooms, bathrooms,
           drops_8in, drops_6in, stories, tons, cfm, sort_order, heat_kit)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
        (bid_id, d.get('name',''), d.get('unit_count',0), d.get('bedrooms',1), d.get('bathrooms',1),
         d.get('drops_8in',0), d.get('drops_6in',0), d.get('stories',1),
         d.get('tons',0), d.get('cfm',0), max_sort + 1, d.get('heat_kit',''))
    )
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return jsonify({'ok': True, 'id': new_id})

@app.route('/api/bids/<int:bid_id>/takeoff/unit-types/<int:ut_id>', methods=['PUT'])
@api_role_required('owner')
def api_update_takeoff_unit_type(bid_id, ut_id):
    d = request.get_json()
    conn = get_db()
    conn.execute(
        '''UPDATE bid_takeoff_unit_types SET name=?, unit_count=?, bedrooms=?, bathrooms=?,
           drops_8in=?, drops_6in=?, stories=?, tons=?, cfm=?, sort_order=?, heat_kit=?
           WHERE id=? AND bid_id=?''',
        (d.get('name',''), d.get('unit_count',0), d.get('bedrooms',1), d.get('bathrooms',1),
         d.get('drops_8in',0), d.get('drops_6in',0), d.get('stories',1),
         d.get('tons',0), d.get('cfm',0), d.get('sort_order',0), d.get('heat_kit',''), ut_id, bid_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/takeoff/unit-types/<int:ut_id>', methods=['DELETE'])
@api_role_required('owner')
def api_delete_takeoff_unit_type(bid_id, ut_id):
    conn = get_db()
    conn.execute('DELETE FROM bid_takeoff_unit_types WHERE id = ? AND bid_id = ?', (ut_id, bid_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/takeoff/items')
@api_role_required('owner')
def api_takeoff_items(bid_id):
    conn = get_db()
    rows = conn.execute('SELECT * FROM bid_takeoff_items WHERE bid_id = ? ORDER BY phase, sort_order', (bid_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/bids/<int:bid_id>/takeoff/items', methods=['POST'])
@api_role_required('owner')
def api_create_takeoff_item(bid_id):
    d = request.get_json()
    conn = get_db()
    max_sort = conn.execute('SELECT COALESCE(MAX(sort_order),0) FROM bid_takeoff_items WHERE bid_id = ? AND phase = ?',
                            (bid_id, d.get('phase','Rough-In'))).fetchone()[0]
    cursor = conn.execute(
        '''INSERT INTO bid_takeoff_items (bid_id, phase, category, part_name, sku, unit_price,
           calc_basis, qty_multiplier, tons_match, waste_pct, enabled, qty_override, sort_order, notes)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (bid_id, d.get('phase','Rough-In'), d.get('category',''), d.get('part_name','New Item'),
         d.get('sku',''), d.get('unit_price',0), d.get('calc_basis','per_system'),
         d.get('qty_multiplier',1), d.get('tons_match'), d.get('waste_pct',0),
         d.get('enabled',1), d.get('qty_override'), max_sort + 1, d.get('notes',''))
    )
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return jsonify({'ok': True, 'id': new_id})

@app.route('/api/bids/<int:bid_id>/takeoff/items/bulk', methods=['PUT'])
@api_role_required('owner')
def api_bulk_save_takeoff(bid_id):
    data = request.get_json()
    conn = get_db()
    # Save unit types
    for ut in data.get('unit_types', []):
        if ut.get('id'):
            conn.execute(
                '''UPDATE bid_takeoff_unit_types SET name=?, unit_count=?, bedrooms=?, bathrooms=?,
                   drops_8in=?, drops_6in=?, stories=?, tons=?, cfm=?, sort_order=?, heat_kit=?
                   WHERE id=? AND bid_id=?''',
                (ut.get('name',''), ut.get('unit_count',0), ut.get('bedrooms',1), ut.get('bathrooms',1),
                 ut.get('drops_8in',0), ut.get('drops_6in',0), ut.get('stories',1),
                 ut.get('tons',0), ut.get('cfm',0), ut.get('sort_order',0), ut.get('heat_kit',''), ut['id'], bid_id)
            )
    # Save items
    for item in data.get('items', []):
        if item.get('id'):
            conn.execute(
                '''UPDATE bid_takeoff_items SET phase=?, category=?, part_name=?, sku=?, unit_price=?,
                   calc_basis=?, qty_multiplier=?, tons_match=?, waste_pct=?, enabled=?,
                   qty_override=?, sort_order=?, notes=?
                   WHERE id=? AND bid_id=?''',
                (item.get('phase','Rough-In'), item.get('category',''), item.get('part_name',''),
                 item.get('sku',''), item.get('unit_price',0), item.get('calc_basis','per_system'),
                 item.get('qty_multiplier',1), item.get('tons_match'), item.get('waste_pct',0),
                 item.get('enabled',1), item.get('qty_override'), item.get('sort_order',0),
                 item.get('notes',''), item['id'], bid_id)
            )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/takeoff/items/<int:item_id>', methods=['DELETE'])
@api_role_required('owner')
def api_delete_takeoff_item(bid_id, item_id):
    conn = get_db()
    conn.execute('DELETE FROM bid_takeoff_items WHERE id = ? AND bid_id = ?', (item_id, bid_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/takeoff/seed-defaults', methods=['POST'])
@api_role_required('owner')
def api_seed_takeoff_defaults(bid_id):
    conn = get_db()
    existing = conn.execute('SELECT COUNT(*) FROM bid_takeoff_items WHERE bid_id = ?', (bid_id,)).fetchone()[0]
    if existing > 0:
        conn.close()
        return jsonify({'ok': False, 'error': 'Items already exist'})
    for item in DEFAULT_TAKEOFF_ITEMS:
        conn.execute(
            '''INSERT INTO bid_takeoff_items (bid_id, phase, category, part_name, sku, unit_price,
               calc_basis, qty_multiplier, tons_match, waste_pct, enabled, sort_order)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (bid_id, item['phase'], item['category'], item['part_name'], item.get('sku',''),
             item['unit_price'], item['calc_basis'], item['qty_multiplier'],
             item.get('tons_match'), item['waste_pct'], item['enabled'], item['sort_order'])
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/takeoff/reset-defaults', methods=['POST'])
@api_role_required('owner','admin')
def api_reset_takeoff_defaults(bid_id):
    """Delete all existing takeoff items and re-seed with current defaults."""
    conn = get_db()
    conn.execute('DELETE FROM bid_takeoff_items WHERE bid_id = ?', (bid_id,))
    for item in DEFAULT_TAKEOFF_ITEMS:
        conn.execute(
            '''INSERT INTO bid_takeoff_items (bid_id, phase, category, part_name, sku, unit_price,
               calc_basis, qty_multiplier, tons_match, waste_pct, enabled, sort_order)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (bid_id, item['phase'], item['category'], item['part_name'], item.get('sku',''),
             item['unit_price'], item['calc_basis'], item['qty_multiplier'],
             item.get('tons_match'), item['waste_pct'], item['enabled'], item['sort_order'])
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/takeoff/config')
@api_role_required('owner')
def api_takeoff_config(bid_id):
    conn = get_db()
    row = conn.execute('SELECT takeoff_config FROM bids WHERE id = ?', (bid_id,)).fetchone()
    conn.close()
    import json as _json
    cfg = _json.loads(row['takeoff_config'] or '{}') if row else {}
    return jsonify(cfg)

@app.route('/api/bids/<int:bid_id>/takeoff/config', methods=['PUT'])
@api_role_required('owner')
def api_update_takeoff_config(bid_id):
    import json as _json
    data = request.get_json()
    conn = get_db()
    conn.execute("UPDATE bids SET takeoff_config = ?, updated_at = datetime('now','localtime') WHERE id = ?",
                 (_json.dumps(data), bid_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/takeoff/push-to-bid', methods=['POST'])
@api_role_required('owner')
def api_push_takeoff_to_bid(bid_id):
    data = request.get_json()
    total = data.get('total', 0)
    conn = get_db()
    conn.execute("UPDATE bids SET material_subtotal = ?, updated_at = datetime('now','localtime') WHERE id = ?",
                 (total, bid_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Takeoff PDF / Email ───────────────────────────────────────

@app.route('/api/bids/<int:bid_id>/takeoff/generate-pdf', methods=['POST'])
@api_role_required('owner', 'admin')
def api_takeoff_generate_pdf(bid_id):
    """Generate a material takeoff PDF from client-computed data."""
    import json as _json
    data = request.get_json()
    conn = get_db()
    bid = conn.execute('SELECT * FROM bids WHERE id = ?', (bid_id,)).fetchone()
    unit_types_rows = conn.execute('SELECT * FROM bid_takeoff_unit_types WHERE bid_id = ? ORDER BY sort_order', (bid_id,)).fetchall()
    cfg_row = conn.execute('SELECT takeoff_config FROM bids WHERE id = ?', (bid_id,)).fetchone()
    conn.close()
    if not bid:
        return jsonify({'error': 'Bid not found'}), 404
    bid = dict(bid)
    unit_types = [dict(r) for r in unit_types_rows]
    config = _json.loads(cfg_row['takeoff_config'] or '{}') if cfg_row else {}

    # Items with calculated quantities come from the client
    phase_items_data = data.get('phases', {})
    phase_totals = data.get('phase_totals', {})
    grand_total = data.get('grand_total', 0)

    # Build phases list for template
    phase_order = ['Rough-In', 'Trim Out', 'Equipment', 'Startup/Other', 'Suggested Parts']
    phases = []
    for p in phase_order:
        items_list = phase_items_data.get(p, [])
        if items_list:
            phases.append((p, items_list))

    # Unit type totals
    ut_totals = {'systems': 0, 'bedrooms': 0, 'bathrooms': 0, 'drops_8': 0, 'drops_6': 0}
    for ut in unit_types:
        c = ut.get('unit_count', 0)
        ut_totals['systems'] += c
        ut_totals['bedrooms'] += c * (ut.get('bedrooms', 0))
        ut_totals['bathrooms'] += c * (ut.get('bathrooms', 0))
        ut_totals['drops_8'] += c * (ut.get('drops_8in', 0))
        ut_totals['drops_6'] += c * (ut.get('drops_6in', 0))

    # Config display
    config_labels = {
        'cfgBuildType': 'Build Type', 'cfgCRD': 'CRD', 'cfgOrientation': 'AHU Orientation',
        'cfgAHUType': 'AHU Type', 'cfgDrainPan': 'Drain Pan', 'cfgMiniSplits': 'Mini Splits',
        'cfgDuctboard': 'Ductboard', 'cfgExhaustType': 'Exhaust Fan', 'cfgWrapping': 'Wrapping',
        'cfgDryerFireWrap': 'Dryer Fire Wrap', 'cfgPassThroughs': 'Pass-Throughs',
        'cfgCondenserLoc': 'Condenser Loc', 'cfgOutsideAir': 'Outside Air',
        'cfgRangeHoods': 'Range Hoods', 'cfgZoned': 'Zoned System',
        'cfgBagsPerDrop': 'Bags/Drop', 'cfgDuctboardPerUnit': 'Ductboard/Unit'
    }
    config_display = [(config_labels.get(k, k), v) for k, v in config.items() if k in config_labels]

    # Tax calculation
    tax_rate = float(bid.get('material_tax_rate', 0) or 0)
    tax_amount = round(grand_total * tax_rate / 100, 2)
    total_with_tax = round(grand_total + tax_amount, 2)

    # Job location for tax label
    tax_location = ''
    if bid.get('job_id'):
        conn2 = get_db()
        job_row = conn2.execute('SELECT city, state FROM jobs WHERE id = ?', (bid['job_id'],)).fetchone()
        conn2.close()
        if job_row:
            parts = [job_row['city'] or '', job_row['state'] or '']
            tax_location = ', '.join(p for p in parts if p)

    today = datetime.now().strftime('%B %d, %Y')
    logo_path = os.path.abspath(os.path.join(app.static_folder, 'logo.jpg'))

    html = render_template('bids/takeoff_pdf.html',
        bid=bid, today=today, logo_path='file://' + logo_path,
        unit_types=unit_types, ut_totals=ut_totals,
        config_display=config_display, phases=phases,
        phase_totals=phase_totals, grand_total=grand_total,
        tax_rate=tax_rate, tax_amount=tax_amount,
        total_with_tax=total_with_tax, tax_location=tax_location
    )

    takeoffs_dir = os.path.join(os.path.dirname(__file__), 'data', 'takeoffs')
    os.makedirs(takeoffs_dir, exist_ok=True)
    safe_name = ''.join(c if c.isalnum() or c in ' -_' else '' for c in (bid.get('bid_name') or 'takeoff')).strip()
    filename = f"Takeoff_{safe_name}_{bid_id}.pdf"
    filepath = os.path.join(takeoffs_dir, filename)

    try:
        wp = weasyprint.HTML(string=html, base_url=os.path.dirname(__file__))
        wp.write_pdf(filepath)
    except Exception as e:
        return jsonify({'error': f'PDF generation failed: {str(e)[:200]}'}), 500

    return jsonify({'ok': True, 'filename': filename, 'path': f'/api/bids/{bid_id}/takeoff/pdf/{filename}'})


@app.route('/api/bids/<int:bid_id>/takeoff/pdf/<filename>')
@api_role_required('owner', 'admin')
def api_download_takeoff_pdf(bid_id, filename):
    """View or download a generated takeoff PDF."""
    takeoffs_dir = os.path.join(os.path.dirname(__file__), 'data', 'takeoffs')
    filepath = os.path.join(takeoffs_dir, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'PDF not found'}), 404
    return send_file(filepath, mimetype='application/pdf')


@app.route('/api/bids/<int:bid_id>/takeoff/email', methods=['POST'])
@api_role_required('owner', 'admin')
def api_email_takeoff(bid_id):
    """Email the takeoff PDF to specified recipients."""
    data = request.get_json()
    recipients = [e.strip() for e in data.get('recipients', []) if e.strip()]
    subject = data.get('subject', 'HVAC Material Takeoff')
    body_text = data.get('body', '')

    saved_settings = {}
    settings_path = os.path.join(os.path.dirname(__file__), 'data', 'email_settings.json')
    if os.path.exists(settings_path):
        with open(settings_path) as f:
            saved_settings = json.load(f)

    smtp_host = data.get('smtp_host') or saved_settings.get('smtp_host', '')
    smtp_port = int(data.get('smtp_port') or saved_settings.get('smtp_port', 587) or 587)
    smtp_user = data.get('smtp_user') or saved_settings.get('smtp_user', '')
    smtp_pass = data.get('smtp_pass') or saved_settings.get('smtp_pass', '')
    from_email = data.get('from_email') or saved_settings.get('from_email', '') or smtp_user

    if not recipients:
        return jsonify({'error': 'No recipients specified'}), 400
    if not smtp_host or not smtp_user:
        return jsonify({'error': 'SMTP settings required. Configure in Settings or provide in request.'}), 400

    takeoffs_dir = os.path.join(os.path.dirname(__file__), 'data', 'takeoffs')
    pdf_files = [f for f in os.listdir(takeoffs_dir) if f.startswith('Takeoff_') and f.endswith(f'_{bid_id}.pdf')]
    if not pdf_files:
        return jsonify({'error': 'No takeoff PDF found. Generate the PDF first.'}), 404
    pdf_path = os.path.join(takeoffs_dir, sorted(pdf_files)[-1])

    try:
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = subject
        msg.attach(MIMEText(body_text, 'plain'))

        with open(pdf_path, 'rb') as f:
            part = MIMEBase('application', 'pdf')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(pdf_path)}"')
            msg.attach(part)

        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)

        save_email_autocomplete(recipients)
        return jsonify({'ok': True, 'sent_to': recipients})
    except Exception as e:
        return jsonify({'error': f'Email failed: {str(e)}'}), 500


# ─── Commercial Takeoff ──────────────────────────────────────────

DEFAULT_COMMERCIAL_TAKEOFF_ITEMS = [
    # ═══ Rough-In (~20 items) ═══
    {'phase':'Rough-In','category':'Curbs','part_name':'RTU Curb/Adapter','sku':'','unit_price':450,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':1},
    {'phase':'Rough-In','category':'Ductwork','part_name':'Sheet Metal Supply Duct (4\' section)','sku':'','unit_price':65,'calc_basis':'per_supply_run','qty_multiplier':3,'waste_pct':10,'enabled':1,'sort_order':2},
    {'phase':'Rough-In','category':'Ductwork','part_name':'Sheet Metal Return Duct (4\' section)','sku':'','unit_price':75,'calc_basis':'per_return_run','qty_multiplier':2,'waste_pct':10,'enabled':1,'sort_order':3},
    {'phase':'Rough-In','category':'Ductwork','part_name':'Spiral Pipe 6" x 5\'','sku':'','unit_price':18,'calc_basis':'per_supply_run','qty_multiplier':2,'waste_pct':10,'enabled':1,'sort_order':4},
    {'phase':'Rough-In','category':'Ductwork','part_name':'Spiral Pipe 8" x 5\'','sku':'','unit_price':24,'calc_basis':'per_supply_run','qty_multiplier':1,'waste_pct':10,'enabled':1,'sort_order':5},
    {'phase':'Rough-In','category':'Flex','part_name':'6" x 25\' Flex Duct R6','sku':'L1972','unit_price':31.46,'calc_basis':'per_supply_run','qty_multiplier':1,'waste_pct':10,'enabled':1,'sort_order':6},
    {'phase':'Rough-In','category':'Flex','part_name':'8" x 25\' Flex Duct R6','sku':'L1974','unit_price':37.56,'calc_basis':'per_supply_run','qty_multiplier':1,'waste_pct':10,'enabled':1,'sort_order':7},
    {'phase':'Rough-In','category':'Flex','part_name':'10" x 25\' Flex Duct R6','sku':'','unit_price':48,'calc_basis':'per_supply_run','qty_multiplier':0.5,'waste_pct':10,'enabled':0,'sort_order':8},
    {'phase':'Rough-In','category':'Fittings','part_name':'Duct Transitions','sku':'','unit_price':28,'calc_basis':'per_system','qty_multiplier':2,'waste_pct':10,'enabled':1,'sort_order':9},
    {'phase':'Rough-In','category':'Dampers','part_name':'Fire Dampers','sku':'','unit_price':85,'calc_basis':'per_total_run','qty_multiplier':0.25,'waste_pct':0,'enabled':1,'sort_order':10},
    {'phase':'Rough-In','category':'Dampers','part_name':'Volume Dampers','sku':'','unit_price':22,'calc_basis':'per_supply_run','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':11},
    {'phase':'Rough-In','category':'Tape & Sealant','part_name':'Duct Sealant 1 GAL','sku':'M6003','unit_price':13,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':12},
    {'phase':'Rough-In','category':'Tape & Sealant','part_name':'Foil Tape 2.5" x 60YDS','sku':'L0494','unit_price':15.15,'calc_basis':'per_system','qty_multiplier':2,'waste_pct':0,'enabled':1,'sort_order':13},
    {'phase':'Rough-In','category':'Tape & Sealant','part_name':'Black Duct Tape 2" 60YDS','sku':'L0444','unit_price':5.60,'calc_basis':'per_system','qty_multiplier':0.5,'waste_pct':0,'enabled':1,'sort_order':14},
    {'phase':'Rough-In','category':'Hangers','part_name':'Duct Hangers / Trapeze','sku':'','unit_price':8,'calc_basis':'per_total_run','qty_multiplier':3,'waste_pct':10,'enabled':1,'sort_order':15},
    {'phase':'Rough-In','category':'Hangers','part_name':'Hanger Strap 3/4" x 100\'','sku':'M0091','unit_price':9,'calc_basis':'per_system','qty_multiplier':0.5,'waste_pct':0,'enabled':1,'sort_order':16},
    {'phase':'Rough-In','category':'Wire','part_name':'18/8 Thermostat Wire 500\'','sku':'P2240','unit_price':0.565,'calc_basis':'per_system','qty_multiplier':150,'waste_pct':0,'enabled':1,'sort_order':17},
    {'phase':'Rough-In','category':'Insulation','part_name':'Duct Wrap R-6.0','sku':'L0475','unit_price':118,'calc_basis':'per_system','qty_multiplier':0.5,'waste_pct':10,'enabled':1,'sort_order':18},
    {'phase':'Rough-In','category':'Penetrations','part_name':'Roof Jacks / Penetrations','sku':'','unit_price':35,'calc_basis':'per_system','qty_multiplier':2,'waste_pct':0,'enabled':1,'sort_order':19},
    {'phase':'Rough-In','category':'Screws','part_name':'#8 x 3/4" Hex Washer Screws','sku':'Q4659','unit_price':32.65,'calc_basis':'per_system','qty_multiplier':0.25,'waste_pct':0,'enabled':1,'sort_order':20},

    # ═══ Trim Out (~15 items) ═══
    {'phase':'Trim Out','category':'Diffusers','part_name':'Supply Diffuser 24x24 4-Way','sku':'','unit_price':32,'calc_basis':'per_supply_run','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':1},
    {'phase':'Trim Out','category':'Grilles','part_name':'Return Grille 24x24','sku':'','unit_price':28,'calc_basis':'per_return_run','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':2},
    {'phase':'Trim Out','category':'Filters','part_name':'Filters (per system)','sku':'','unit_price':8,'calc_basis':'per_system','qty_multiplier':2,'waste_pct':0,'enabled':1,'sort_order':3},
    {'phase':'Trim Out','category':'Condensate','part_name':'3/4" PVC Condensate Pipe 20\'','sku':'R0071','unit_price':0.34,'calc_basis':'per_system','qty_multiplier':20,'waste_pct':0,'enabled':1,'sort_order':4},
    {'phase':'Trim Out','category':'Condensate','part_name':'3/4 PVC 90 Ell','sku':'R0311','unit_price':0.40,'calc_basis':'per_system','qty_multiplier':6,'waste_pct':0,'enabled':1,'sort_order':5},
    {'phase':'Trim Out','category':'Condensate','part_name':'PVC Fittings Assorted','sku':'','unit_price':12,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':6},
    {'phase':'Trim Out','category':'Condensate','part_name':'Condensate Pump','sku':'','unit_price':65,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':7},
    {'phase':'Trim Out','category':'Condensate','part_name':'PVC Cement 1 QT','sku':'R0042','unit_price':14.95,'calc_basis':'per_system','qty_multiplier':0.1,'waste_pct':0,'enabled':1,'sort_order':8},
    {'phase':'Trim Out','category':'Line Sets','part_name':'Refrigerant Line Set 3/8 x 3/4 x 25\'','sku':'','unit_price':85,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':9},
    {'phase':'Trim Out','category':'Line Sets','part_name':'Refrigerant Line Set 3/8 x 3/4 x 50\'','sku':'','unit_price':155,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':10},
    {'phase':'Trim Out','category':'Insulation','part_name':'Line Set Insulation 3/4"','sku':'L0484','unit_price':0.284,'calc_basis':'per_system','qty_multiplier':50,'waste_pct':0,'enabled':0,'sort_order':11},
    {'phase':'Trim Out','category':'Sealant','part_name':'10.1oz White Silicone','sku':'M1842','unit_price':5.90,'calc_basis':'per_system','qty_multiplier':0.5,'waste_pct':0,'enabled':1,'sort_order':12},
    {'phase':'Trim Out','category':'Electrical','part_name':'Wire Nuts Jar 100ct','sku':'P1491','unit_price':5.48,'calc_basis':'per_system','qty_multiplier':0.1,'waste_pct':0,'enabled':1,'sort_order':13},
    {'phase':'Trim Out','category':'Brazing','part_name':'15% Silver Solder 1#','sku':'M0434','unit_price':97,'calc_basis':'per_system','qty_multiplier':0.1,'waste_pct':0,'enabled':1,'sort_order':14},
    {'phase':'Trim Out','category':'Refrigerant','part_name':'R-410A Refrigerant 25# Drum','sku':'','unit_price':185,'calc_basis':'per_system','qty_multiplier':0.2,'waste_pct':0,'enabled':1,'sort_order':15},

    # ═══ Equipment (~15 items) ═══
    {'phase':'Equipment','category':'RTU','part_name':'3 Ton RTU','sku':'','unit_price':3200,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':3,'waste_pct':0,'enabled':1,'sort_order':1},
    {'phase':'Equipment','category':'RTU','part_name':'5 Ton RTU','sku':'','unit_price':4800,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':5,'waste_pct':0,'enabled':1,'sort_order':2},
    {'phase':'Equipment','category':'RTU','part_name':'7.5 Ton RTU','sku':'','unit_price':7200,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':7.5,'waste_pct':0,'enabled':1,'sort_order':3},
    {'phase':'Equipment','category':'RTU','part_name':'10 Ton RTU','sku':'','unit_price':9500,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':10,'waste_pct':0,'enabled':1,'sort_order':4},
    {'phase':'Equipment','category':'RTU','part_name':'12.5 Ton RTU','sku':'','unit_price':12000,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':12.5,'waste_pct':0,'enabled':1,'sort_order':5},
    {'phase':'Equipment','category':'RTU','part_name':'15 Ton RTU','sku':'','unit_price':14500,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':15,'waste_pct':0,'enabled':1,'sort_order':6},
    {'phase':'Equipment','category':'RTU','part_name':'20 Ton RTU','sku':'','unit_price':19000,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':20,'waste_pct':0,'enabled':1,'sort_order':7},
    {'phase':'Equipment','category':'RTU','part_name':'25 Ton RTU','sku':'','unit_price':24000,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':25,'waste_pct':0,'enabled':1,'sort_order':8},
    {'phase':'Equipment','category':'Split','part_name':'3 Ton Split Condenser','sku':'','unit_price':1800,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':3,'waste_pct':0,'enabled':0,'sort_order':9},
    {'phase':'Equipment','category':'Split','part_name':'5 Ton Split Condenser','sku':'','unit_price':2800,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':5,'waste_pct':0,'enabled':0,'sort_order':10},
    {'phase':'Equipment','category':'Split','part_name':'7.5 Ton Split Condenser','sku':'','unit_price':4200,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':7.5,'waste_pct':0,'enabled':0,'sort_order':11},
    {'phase':'Equipment','category':'Split','part_name':'10 Ton Split Condenser','sku':'','unit_price':5500,'calc_basis':'by_tonnage','qty_multiplier':1,'tons_match':10,'waste_pct':0,'enabled':0,'sort_order':12},
    {'phase':'Equipment','category':'Thermostat','part_name':'Programmable Thermostat','sku':'','unit_price':52.50,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':13},
    {'phase':'Equipment','category':'Thermostat','part_name':'Smart Thermostat / BMS Interface','sku':'','unit_price':185,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':14},

    # ═══ Startup/Other (~8 items) ═══
    {'phase':'Startup/Other','category':'Permits','part_name':'Mechanical Permit','sku':'','unit_price':500,'calc_basis':'fixed','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':1},
    {'phase':'Startup/Other','category':'Crane','part_name':'Crane Rental (RTU Set)','sku':'','unit_price':2500,'calc_basis':'fixed','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':2},
    {'phase':'Startup/Other','category':'Shipping','part_name':'Freight / Delivery','sku':'','unit_price':3000,'calc_basis':'fixed','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':3},
    {'phase':'Startup/Other','category':'Gas','part_name':'Gas Piping Allowance','sku':'','unit_price':1500,'calc_basis':'fixed','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':4},
    {'phase':'Startup/Other','category':'Startup','part_name':'Startup / Commissioning','sku':'','unit_price':350,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':1,'sort_order':5},
    {'phase':'Startup/Other','category':'Consumables','part_name':'Nitrogen','sku':'','unit_price':35,'calc_basis':'per_system','qty_multiplier':0.5,'waste_pct':0,'enabled':1,'sort_order':6},
    {'phase':'Startup/Other','category':'Consumables','part_name':'Refrigerant (startup)','sku':'','unit_price':185,'calc_basis':'per_system','qty_multiplier':0.1,'waste_pct':0,'enabled':1,'sort_order':7},
    {'phase':'Startup/Other','category':'Testing','part_name':'Test & Balance Allowance','sku':'','unit_price':500,'calc_basis':'fixed','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':8},

    # ═══ Suggested Parts (~8 items) ═══
    {'phase':'Suggested Parts','category':'Electrical','part_name':'30A Non-Fuse Disconnect','sku':'','unit_price':18,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':1},
    {'phase':'Suggested Parts','category':'Electrical','part_name':'60A Non-Fuse Disconnect','sku':'','unit_price':24,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':2},
    {'phase':'Suggested Parts','category':'Electrical','part_name':'Whip Kit 3/4" x 6\'','sku':'','unit_price':12,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':3},
    {'phase':'Suggested Parts','category':'Pads','part_name':'Condenser Pad 36x36x3','sku':'','unit_price':38,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':4},
    {'phase':'Suggested Parts','category':'Controls','part_name':'Economizer','sku':'','unit_price':450,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':5},
    {'phase':'Suggested Parts','category':'Safety','part_name':'Smoke Detector','sku':'','unit_price':45,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':6},
    {'phase':'Suggested Parts','category':'Safety','part_name':'CO Detector','sku':'','unit_price':35,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':7},
    {'phase':'Suggested Parts','category':'Seismic','part_name':'Seismic Straps','sku':'','unit_price':24,'calc_basis':'per_system','qty_multiplier':1,'waste_pct':0,'enabled':0,'sort_order':8},
]

@app.route('/bids/<int:bid_id>/commercial-takeoff')
@role_required('owner')
def bids_commercial_takeoff(bid_id):
    return render_template('bids/commercial_takeoff.html', bid_id=bid_id)

@app.route('/api/bids/<int:bid_id>/commercial-takeoff/config')
@api_role_required('owner')
def api_commercial_takeoff_config(bid_id):
    conn = get_db()
    row = conn.execute('SELECT commercial_takeoff_config FROM bids WHERE id = ?', (bid_id,)).fetchone()
    conn.close()
    import json as _json
    cfg = _json.loads(row['commercial_takeoff_config'] or '{}') if row else {}
    return jsonify(cfg)

@app.route('/api/bids/<int:bid_id>/commercial-takeoff/config', methods=['PUT'])
@api_role_required('owner')
def api_update_commercial_takeoff_config(bid_id):
    import json as _json
    data = request.get_json()
    conn = get_db()
    conn.execute("UPDATE bids SET commercial_takeoff_config = ?, updated_at = datetime('now','localtime') WHERE id = ?",
                 (_json.dumps(data), bid_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/commercial-takeoff/systems')
@api_role_required('owner')
def api_commercial_takeoff_systems(bid_id):
    conn = get_db()
    rows = conn.execute('SELECT * FROM bid_commercial_takeoff_systems WHERE bid_id = ? ORDER BY sort_order', (bid_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/bids/<int:bid_id>/commercial-takeoff/systems', methods=['POST'])
@api_role_required('owner')
def api_create_commercial_takeoff_system(bid_id):
    d = request.get_json()
    conn = get_db()
    max_sort = conn.execute('SELECT COALESCE(MAX(sort_order),0) FROM bid_commercial_takeoff_systems WHERE bid_id = ?', (bid_id,)).fetchone()[0]
    cursor = conn.execute(
        '''INSERT INTO bid_commercial_takeoff_systems (bid_id, name, system_count, tons, cfm, supply_runs, return_runs, sort_order)
           VALUES (?,?,?,?,?,?,?,?)''',
        (bid_id, d.get('name',''), d.get('system_count',1), d.get('tons',0),
         d.get('cfm',0), d.get('supply_runs',0), d.get('return_runs',0), max_sort + 1)
    )
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return jsonify({'ok': True, 'id': new_id})

@app.route('/api/bids/<int:bid_id>/commercial-takeoff/systems/<int:sys_id>', methods=['DELETE'])
@api_role_required('owner')
def api_delete_commercial_takeoff_system(bid_id, sys_id):
    conn = get_db()
    conn.execute('DELETE FROM bid_commercial_takeoff_systems WHERE id = ? AND bid_id = ?', (sys_id, bid_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/commercial-takeoff/items')
@api_role_required('owner')
def api_commercial_takeoff_items(bid_id):
    conn = get_db()
    rows = conn.execute('SELECT * FROM bid_commercial_takeoff_items WHERE bid_id = ? ORDER BY phase, sort_order', (bid_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/bids/<int:bid_id>/commercial-takeoff/items', methods=['POST'])
@api_role_required('owner')
def api_create_commercial_takeoff_item(bid_id):
    d = request.get_json()
    conn = get_db()
    max_sort = conn.execute('SELECT COALESCE(MAX(sort_order),0) FROM bid_commercial_takeoff_items WHERE bid_id = ? AND phase = ?',
                            (bid_id, d.get('phase','Rough-In'))).fetchone()[0]
    cursor = conn.execute(
        '''INSERT INTO bid_commercial_takeoff_items (bid_id, phase, category, part_name, sku, unit_price,
           calc_basis, qty_multiplier, tons_match, waste_pct, enabled, qty_override, sort_order, notes)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (bid_id, d.get('phase','Rough-In'), d.get('category',''), d.get('part_name','New Item'),
         d.get('sku',''), d.get('unit_price',0), d.get('calc_basis','per_system'),
         d.get('qty_multiplier',1), d.get('tons_match'), d.get('waste_pct',0),
         d.get('enabled',1), d.get('qty_override'), max_sort + 1, d.get('notes',''))
    )
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return jsonify({'ok': True, 'id': new_id})

@app.route('/api/bids/<int:bid_id>/commercial-takeoff/items/<int:item_id>', methods=['DELETE'])
@api_role_required('owner')
def api_delete_commercial_takeoff_item(bid_id, item_id):
    conn = get_db()
    conn.execute('DELETE FROM bid_commercial_takeoff_items WHERE id = ? AND bid_id = ?', (item_id, bid_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/commercial-takeoff/items/bulk', methods=['PUT'])
@api_role_required('owner')
def api_bulk_save_commercial_takeoff(bid_id):
    data = request.get_json()
    conn = get_db()
    for sys in data.get('systems', []):
        if sys.get('id'):
            conn.execute(
                '''UPDATE bid_commercial_takeoff_systems SET name=?, system_count=?, tons=?, cfm=?,
                   supply_runs=?, return_runs=?, sort_order=?
                   WHERE id=? AND bid_id=?''',
                (sys.get('name',''), sys.get('system_count',1), sys.get('tons',0),
                 sys.get('cfm',0), sys.get('supply_runs',0), sys.get('return_runs',0),
                 sys.get('sort_order',0), sys['id'], bid_id)
            )
    for item in data.get('items', []):
        if item.get('id'):
            conn.execute(
                '''UPDATE bid_commercial_takeoff_items SET phase=?, category=?, part_name=?, sku=?, unit_price=?,
                   calc_basis=?, qty_multiplier=?, tons_match=?, waste_pct=?, enabled=?,
                   qty_override=?, sort_order=?, notes=?
                   WHERE id=? AND bid_id=?''',
                (item.get('phase','Rough-In'), item.get('category',''), item.get('part_name',''),
                 item.get('sku',''), item.get('unit_price',0), item.get('calc_basis','per_system'),
                 item.get('qty_multiplier',1), item.get('tons_match'), item.get('waste_pct',0),
                 item.get('enabled',1), item.get('qty_override'), item.get('sort_order',0),
                 item.get('notes',''), item['id'], bid_id)
            )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/commercial-takeoff/seed-defaults', methods=['POST'])
@api_role_required('owner')
def api_seed_commercial_takeoff_defaults(bid_id):
    conn = get_db()
    existing = conn.execute('SELECT COUNT(*) FROM bid_commercial_takeoff_items WHERE bid_id = ?', (bid_id,)).fetchone()[0]
    if existing > 0:
        conn.close()
        return jsonify({'ok': False, 'error': 'Items already exist'})
    for item in DEFAULT_COMMERCIAL_TAKEOFF_ITEMS:
        conn.execute(
            '''INSERT INTO bid_commercial_takeoff_items (bid_id, phase, category, part_name, sku, unit_price,
               calc_basis, qty_multiplier, tons_match, waste_pct, enabled, sort_order)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (bid_id, item['phase'], item['category'], item['part_name'], item.get('sku',''),
             item['unit_price'], item['calc_basis'], item['qty_multiplier'],
             item.get('tons_match'), item['waste_pct'], item['enabled'], item['sort_order'])
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/commercial-takeoff/reset-defaults', methods=['POST'])
@api_role_required('owner', 'admin')
def api_reset_commercial_takeoff_defaults(bid_id):
    conn = get_db()
    conn.execute('DELETE FROM bid_commercial_takeoff_items WHERE bid_id = ?', (bid_id,))
    for item in DEFAULT_COMMERCIAL_TAKEOFF_ITEMS:
        conn.execute(
            '''INSERT INTO bid_commercial_takeoff_items (bid_id, phase, category, part_name, sku, unit_price,
               calc_basis, qty_multiplier, tons_match, waste_pct, enabled, sort_order)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (bid_id, item['phase'], item['category'], item['part_name'], item.get('sku',''),
             item['unit_price'], item['calc_basis'], item['qty_multiplier'],
             item.get('tons_match'), item['waste_pct'], item['enabled'], item['sort_order'])
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bids/<int:bid_id>/commercial-takeoff/push-to-bid', methods=['POST'])
@api_role_required('owner')
def api_push_commercial_takeoff_to_bid(bid_id):
    data = request.get_json()
    total = data.get('total', 0)
    conn = get_db()
    conn.execute("UPDATE bids SET material_subtotal = ?, updated_at = datetime('now','localtime') WHERE id = ?",
                 (total, bid_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ─── Precon Meetings (Phase 4) ──────────────────────────────────

@app.route('/api/jobs/<int:job_id>/precon')
@api_role_required('owner', 'admin', 'project_manager')
def api_precon_meeting(job_id):
    conn = get_db()
    meeting = conn.execute('SELECT * FROM precon_meetings WHERE job_id = ?', (job_id,)).fetchone()
    conn.close()
    return jsonify(dict(meeting) if meeting else {'status': 'No Meeting'})

@app.route('/api/jobs/<int:job_id>/precon', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_precon_meeting(job_id):
    data = request.get_json()
    conn = get_db()
    existing = conn.execute('SELECT id FROM precon_meetings WHERE job_id = ?', (job_id,)).fetchone()
    if existing:
        conn.execute(
            '''UPDATE precon_meetings SET meeting_date=?, attendees=?, location=?, agenda=?,
               minutes=?, gc_contact=?, status=?, updated_at=datetime('now','localtime') WHERE job_id=?''',
            (data.get('meeting_date',''), data.get('attendees',''), data.get('location',''),
             data.get('agenda',''), data.get('minutes',''), data.get('gc_contact',''),
             data.get('status','Not Scheduled'), job_id)
        )
    else:
        conn.execute(
            '''INSERT INTO precon_meetings (job_id, meeting_date, attendees, location, agenda, minutes, gc_contact, status, created_by)
               VALUES (?,?,?,?,?,?,?,?,?)''',
            (job_id, data.get('meeting_date',''), data.get('attendees',''), data.get('location',''),
             data.get('agenda',''), data.get('minutes',''), data.get('gc_contact',''),
             data.get('status','Not Scheduled'), session.get('user_id'))
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/jobs/<int:job_id>/precon/complete', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_complete_precon_meeting(job_id):
    conn = get_db()
    conn.execute("UPDATE precon_meetings SET status = 'Completed', updated_at=datetime('now','localtime') WHERE job_id = ?", (job_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── PM Benchmarks (Phase 4/6) ──────────────────────────────────

@app.route('/api/jobs/<int:job_id>/benchmarks')
@api_role_required('owner', 'admin', 'project_manager')
def api_job_benchmarks(job_id):
    conn = get_db()
    rows = conn.execute(
        '''SELECT pb.*, u.display_name as completed_by_name FROM pm_benchmarks pb
           LEFT JOIN users u ON pb.completed_by = u.id
           WHERE pb.job_id = ? ORDER BY pb.sort_order''', (job_id,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/jobs/<int:job_id>/benchmarks/<int:bid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_benchmark(job_id, bid):
    data = request.get_json()
    from datetime import datetime
    conn = get_db()
    new_status = data.get('status', 'Not Started')
    completed_by = session.get('user_id') if new_status == 'Complete' else None
    completed_date = datetime.now().strftime('%Y-%m-%d') if new_status == 'Complete' else ''
    conn.execute(
        '''UPDATE pm_benchmarks SET status=?, completed_date=?, completed_by=?, notes=?,
           updated_at=datetime('now','localtime') WHERE id=? AND job_id=?''',
        (new_status, completed_date, completed_by, data.get('notes',''), bid, job_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/email-settings', methods=['GET', 'POST'])
@api_role_required('owner', 'admin')
def api_email_settings():
    """Get or save email/SMTP settings."""
    settings_path = os.path.join(os.path.dirname(__file__), 'data', 'email_settings.json')
    if request.method == 'GET':
        if os.path.exists(settings_path):
            with open(settings_path) as f:
                settings = json.load(f)
            settings.pop('smtp_pass', None)  # never return password
            return jsonify(settings)
        return jsonify({})
    else:
        data = request.get_json()
        # Load existing to preserve password if not provided
        existing = {}
        if os.path.exists(settings_path):
            with open(settings_path) as f:
                existing = json.load(f)
        settings = {
            'smtp_host': data.get('smtp_host', existing.get('smtp_host', '')),
            'smtp_port': data.get('smtp_port', existing.get('smtp_port', 587)),
            'smtp_user': data.get('smtp_user', existing.get('smtp_user', '')),
            'from_email': data.get('from_email', existing.get('from_email', '')),
            'team_emails': data.get('team_emails', existing.get('team_emails', '')),
        }
        if data.get('smtp_pass'):
            settings['smtp_pass'] = data['smtp_pass']
        elif existing.get('smtp_pass'):
            settings['smtp_pass'] = existing['smtp_pass']
        with open(settings_path, 'w') as f:
            json.dump(settings, f)
        return jsonify({'ok': True})


# ─── Chatbot ────────────────────────────────────────────────

@app.route('/chatbot')
@login_required
def chatbot_page():
    return render_template('chatbot.html')

@app.route('/api/chatbot/sessions')
@api_login_required
def api_chat_sessions():
    conn = get_db()
    sessions = conn.execute(
        'SELECT * FROM chat_sessions WHERE user_id = ? ORDER BY created_at DESC',
        (session['user_id'],)
    ).fetchall()
    conn.close()
    return jsonify([dict(s) for s in sessions])

@app.route('/api/chatbot/sessions', methods=['POST'])
@api_login_required
def api_create_chat_session():
    conn = get_db()
    cursor = conn.execute(
        'INSERT INTO chat_sessions (user_id, title) VALUES (?, ?)',
        (session['user_id'], 'New Chat')
    )
    sid = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': sid}), 201

@app.route('/api/chatbot/sessions/<int:sid>/messages')
@api_login_required
def api_chat_messages(sid):
    conn = get_db()
    msgs = conn.execute(
        'SELECT * FROM chat_messages WHERE session_id = ? ORDER BY created_at ASC',
        (sid,)
    ).fetchall()
    conn.close()
    return jsonify([dict(m) for m in msgs])

@app.route('/api/chatbot/sessions/<int:sid>/messages', methods=['POST'])
@api_login_required
def api_send_chat_message(sid):
    data = request.get_json()
    user_msg = (data.get('content') or '').strip()
    if not user_msg:
        return jsonify({'error': 'Empty message'}), 400

    conn = get_db()
    # Verify session belongs to user
    sess = conn.execute('SELECT * FROM chat_sessions WHERE id = ? AND user_id = ?',
                        (sid, session['user_id'])).fetchone()
    if not sess:
        conn.close()
        return jsonify({'error': 'Session not found'}), 404

    # Save user message
    conn.execute('INSERT INTO chat_messages (session_id, role, content) VALUES (?,?,?)',
                 (sid, 'user', user_msg))

    # Update session title from first message
    msg_count = conn.execute('SELECT COUNT(*) FROM chat_messages WHERE session_id = ?', (sid,)).fetchone()[0]
    if msg_count <= 1:
        title = user_msg[:50] + ('...' if len(user_msg) > 50 else '')
        conn.execute('UPDATE chat_sessions SET title = ? WHERE id = ?', (title, sid))

    # Generate bot response -- try Claude first, fall back to rule-based
    bot_response = generate_claude_response(
        conn, user_msg,
        role=session.get('role', 'employee'),
        user_id=session.get('user_id'),
        session_id=sid
    )
    if bot_response is None:
        # Fallback to rule-based engine
        bot_response = generate_bot_response(conn, user_msg, role=session.get('role', 'employee'), user_id=session.get('user_id'))

    # Save bot message
    conn.execute('INSERT INTO chat_messages (session_id, role, content) VALUES (?,?,?)',
                 (sid, 'assistant', bot_response))

    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'response': bot_response})


# ─── Permits ─────────────────────────────────────────────────────

@app.route('/permits')
@role_required('owner', 'admin', 'project_manager')
def permits_page():
    return render_template('permits/list.html')

@app.route('/api/permits')
@api_role_required('owner', 'admin', 'project_manager')
def api_permits():
    job_id = request.args.get('job_id', '')
    status = request.args.get('status', '')
    conn = get_db()
    query = '''SELECT p.*, j.name as job_name FROM permits p
               LEFT JOIN jobs j ON p.job_id = j.id WHERE 1=1'''
    params = []
    if job_id:
        query += ' AND p.job_id = ?'
        params.append(job_id)
    if status:
        query += ' AND p.status = ?'
        params.append(status)
    query += ' ORDER BY p.updated_at DESC'
    permits = conn.execute(query, params).fetchall()
    result = []
    for p in permits:
        d = dict(p)
        inspections = conn.execute(
            '''SELECT pi.*, u.display_name as requested_by_name
               FROM permit_inspections pi
               LEFT JOIN users u ON pi.requested_by = u.id
               WHERE pi.permit_id = ? ORDER BY pi.scheduled_date''',
            (p['id'],)
        ).fetchall()
        d['inspections'] = [dict(i) for i in inspections]
        d['inspection_count'] = len(inspections)
        d['passed_count'] = sum(1 for i in inspections if i['status'] == 'Passed')
        result.append(d)
    # Check for permit notifications
    _check_permit_notifications(result, conn)
    conn.close()
    return jsonify(result)

@app.route('/api/permits', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_create_permit():
    data = request.get_json()
    job_id = data.get('job_id')
    if not job_id:
        return jsonify({'error': 'Job is required'}), 400
    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO permits (job_id, permit_type, permit_number, issuing_authority,
           status, applied_date, approved_date, expiration_date, cost,
           inspector_name, inspector_phone, notes, created_by)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (job_id, data.get('permit_type', 'Mechanical'),
         data.get('permit_number', ''), data.get('issuing_authority', ''),
         data.get('status', 'Not Applied'), data.get('applied_date', ''),
         data.get('approved_date', ''), data.get('expiration_date', ''),
         float(data.get('cost', 0)), data.get('inspector_name', ''),
         data.get('inspector_phone', ''), data.get('notes', ''),
         session.get('user_id'))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cursor.lastrowid}), 201

@app.route('/api/permits/<int:pid>')
@api_role_required('owner', 'admin', 'project_manager')
def api_permit_detail(pid):
    conn = get_db()
    p = conn.execute(
        '''SELECT p.*, j.name as job_name FROM permits p
           LEFT JOIN jobs j ON p.job_id = j.id WHERE p.id = ?''', (pid,)
    ).fetchone()
    if not p:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    inspections = conn.execute('SELECT * FROM permit_inspections WHERE permit_id = ? ORDER BY scheduled_date', (pid,)).fetchall()
    conn.close()
    result = dict(p)
    result['inspections'] = [dict(i) for i in inspections]
    return jsonify(result)

@app.route('/api/permits/<int:pid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_permit(pid):
    data = request.get_json()
    conn = get_db()
    fields = []
    params = []
    for col in ('permit_type', 'permit_number', 'issuing_authority', 'status',
                'applied_date', 'approved_date', 'expiration_date', 'cost',
                'inspector_name', 'inspector_phone', 'notes'):
        if col in data:
            fields.append(f'{col} = ?')
            params.append(data[col] if col != 'cost' else float(data[col] or 0))
    if fields:
        fields.append("updated_at = datetime('now','localtime')")
        params.append(pid)
        conn.execute(f"UPDATE permits SET {', '.join(fields)} WHERE id = ?", params)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/permits/<int:pid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_permit(pid):
    conn = get_db()
    conn.execute('DELETE FROM permits WHERE id = ?', (pid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/permits/<int:pid>/inspections', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_add_inspection(pid):
    data = request.get_json()
    status = data.get('status', 'Scheduled')
    valid_statuses = ('Requested', 'Scheduled', 'Passed', 'Failed', 'Cancelled', 'Re-Inspect')
    if status not in valid_statuses:
        return jsonify({'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'}), 400

    conn = get_db()
    requested_by = None
    requested_date = ''
    if status == 'Requested':
        requested_by = session.get('user_id')
        requested_date = datetime.now().strftime('%Y-%m-%d %H:%M')

    cursor = conn.execute(
        '''INSERT INTO permit_inspections (permit_id, inspection_type, status,
           scheduled_date, completed_date, inspector, result_notes, created_by,
           requested_by, requested_date)
           VALUES (?,?,?,?,?,?,?,?,?,?)''',
        (pid, data.get('inspection_type', 'Rough-In'),
         status, data.get('scheduled_date', ''),
         data.get('completed_date', ''), data.get('inspector', ''),
         data.get('result_notes', ''), session.get('user_id'),
         requested_by, requested_date)
    )
    conn.commit()

    # If Requested, notify all owner/admin users
    insp_id = cursor.lastrowid
    if status == 'Requested':
        permit = conn.execute(
            'SELECT p.permit_type, j.name as job_name FROM permits p LEFT JOIN jobs j ON p.job_id = j.id WHERE p.id = ?',
            (pid,)
        ).fetchone()
        insp_type = data.get('inspection_type', 'Rough-In')
        job_name = permit['job_name'] if permit else 'Unknown Job'
        permit_type = permit['permit_type'] if permit else ''
        requester = session.get('display_name', session.get('username', ''))
        admins = conn.execute(
            "SELECT id FROM users WHERE role IN ('owner','admin') AND is_active = 1 AND id != ?",
            (session.get('user_id'),)
        ).fetchall()
        for admin in admins:
            conn.execute(
                'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
                (admin['id'], 'inspection',
                 f'Inspection Request: {insp_type}',
                 f'{requester} requested a {insp_type} inspection for {job_name} — {permit_type} permit. Please call in to schedule.',
                 '/permits')
            )
        conn.commit()

    conn.close()
    return jsonify({'ok': True, 'id': insp_id}), 201

@app.route('/api/permits/<int:pid>/inspections/<int:iid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_inspection(pid, iid):
    data = request.get_json()
    conn = get_db()

    # Get current inspection state before update
    old_insp = conn.execute('SELECT * FROM permit_inspections WHERE id = ? AND permit_id = ?', (iid, pid)).fetchone()

    fields = []
    params = []
    for col in ('inspection_type', 'status', 'scheduled_date', 'completed_date',
                'inspector', 'result_notes'):
        if col in data:
            fields.append(f'{col} = ?')
            params.append(data[col])
    if fields:
        params.extend([iid, pid])
        conn.execute(f"UPDATE permit_inspections SET {', '.join(fields)} WHERE id = ? AND permit_id = ?", params)
        conn.commit()

    # Notify requesting PM when Requested → Scheduled
    new_status = data.get('status')
    if old_insp and old_insp['status'] == 'Requested' and new_status == 'Scheduled' and old_insp['requested_by']:
        permit = conn.execute(
            'SELECT p.permit_type, j.name as job_name FROM permits p LEFT JOIN jobs j ON p.job_id = j.id WHERE p.id = ?',
            (pid,)
        ).fetchone()
        insp_type = data.get('inspection_type', old_insp['inspection_type'])
        job_name = permit['job_name'] if permit else 'Unknown Job'
        sched_date = data.get('scheduled_date', '')
        conn.execute(
            'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
            (old_insp['requested_by'], 'inspection',
             f'Inspection Scheduled: {insp_type}',
             f'Your {insp_type} inspection for {job_name} has been scheduled{" for " + sched_date if sched_date else ""}.',
             '/permits')
        )
        conn.commit()

    conn.close()
    return jsonify({'ok': True})

@app.route('/api/permits/<int:pid>/inspections/<int:iid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_inspection(pid, iid):
    conn = get_db()
    conn.execute('DELETE FROM permit_inspections WHERE id = ? AND permit_id = ?', (iid, pid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

PERMITS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'permits')

@app.route('/api/permits/upload', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_permit_upload():
    """Upload a permit document (drag-and-drop support)."""
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'No file provided'}), 400
    job_id = request.form.get('job_id')
    if not job_id:
        return jsonify({'error': 'Job is required'}), 400
    conn = get_db()
    # Save file
    os.makedirs(PERMITS_DIR, exist_ok=True)
    from werkzeug.utils import secure_filename
    fname = secure_filename(file.filename)
    fname = f"{int(datetime.now().timestamp())}_{fname}"
    file.save(os.path.join(PERMITS_DIR, fname))
    # Create permit record
    permit_type = request.form.get('permit_type', 'Mechanical')
    if permit_type not in ('Mechanical','Building','Plumbing','Electrical','Fire','Roofing','Demolition','Other'):
        permit_type = 'Mechanical'
    cursor = conn.execute(
        '''INSERT INTO permits (job_id, permit_type, status, file_path, notes, created_by)
           VALUES (?, ?, 'Not Applied', ?, ?, ?)''',
        (job_id, permit_type, fname, f'Uploaded: {file.filename}', session.get('user_id'))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cursor.lastrowid})

@app.route('/api/permits/file/<path:filename>')
@login_required
def api_permit_file(filename):
    return send_from_directory(PERMITS_DIR, filename)

def _check_permit_notifications(permits, conn):
    """Send weekly notifications for permits not yet approved."""
    from datetime import timedelta
    now = datetime.now()
    week_ago = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    owners = conn.execute("SELECT id FROM users WHERE role IN ('owner','admin','project_manager') AND is_active = 1").fetchall()
    for p in permits:
        if p.get('status') in ('Approved', 'N/A'):
            continue
        job_name = p.get('job_name', 'Unknown')
        for owner in owners:
            existing = conn.execute(
                "SELECT id FROM notifications WHERE user_id = ? AND type = 'permit' AND message LIKE ? AND created_at >= ?",
                (owner['id'], f'%permit #{p["id"]}%', week_ago)
            ).fetchone()
            if not existing:
                status_msg = p['status']
                if p['status'] == 'Not Applied':
                    status_msg = 'needs to be applied for'
                elif p['status'] == 'Applied':
                    status_msg = 'applied — awaiting approval'
                elif p['status'] == 'Under Review':
                    status_msg = 'under review'
                elif p['status'] == 'Denied':
                    status_msg = 'DENIED — action needed'
                elif p['status'] == 'Expired':
                    status_msg = 'EXPIRED — renewal needed'
                conn.execute(
                    'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
                    (owner['id'], 'permit',
                     f'Permit: {p["permit_type"]} — {job_name}',
                     f'{p["permit_type"]} permit for {job_name} {status_msg} (permit #{p["id"]})',
                     '/permits')
                )
    conn.commit()

# ─── Pay Applications (AIA G702/G703) ────────────────────────────

@app.route('/payapps')
@role_required('owner', 'admin', 'project_manager')
def payapps_page():
    return render_template('payapps/list.html')

@app.route('/payapps/contract/<int:contract_id>')
@role_required('owner', 'admin', 'project_manager')
def payapps_contract_page(contract_id):
    return render_template('payapps/contract.html', contract_id=contract_id)

@app.route('/payapps/application/<int:app_id>')
@role_required('owner', 'admin', 'project_manager')
def payapps_application_page(app_id):
    return render_template('payapps/application.html', app_id=app_id)

@app.route('/api/payapps/contracts')
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_contracts():
    conn = get_db()
    rows = conn.execute('''
        SELECT c.*, j.name as job_name,
            (SELECT COUNT(*) FROM pay_applications WHERE contract_id = c.id) as app_count,
            (SELECT status FROM pay_applications WHERE contract_id = c.id ORDER BY application_number DESC LIMIT 1) as latest_status
        FROM pay_app_contracts c
        LEFT JOIN jobs j ON c.job_id = j.id
        ORDER BY j.name ASC
    ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/jobs/<int:job_id>/customer-info')
@api_role_required('owner', 'admin', 'project_manager')
def api_job_customer_info(job_id):
    """Get customer + bid + contract info linked to a job for auto-filling pay app contracts."""
    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    if not job:
        conn.close()
        return jsonify({})
    result = {}
    # Pull bid data (most recent accepted or any bid linked to this job)
    bid = conn.execute(
        "SELECT * FROM bids WHERE job_id = ? ORDER BY CASE WHEN status='Accepted' THEN 0 ELSE 1 END, id DESC LIMIT 1",
        (job_id,)
    ).fetchone()
    if bid:
        result['contracting_gc'] = bid['contracting_gc'] or ''
        result['gc_attention'] = bid['gc_attention'] or ''
        result['total_bid'] = bid['total_bid'] or 0
        result['bid_name'] = bid['bid_name'] or ''
        if bid['bid_number']:
            result['project_no'] = bid['bid_number']
    # Pull customer data — try job's customer_id, then bid's customer_id, then match by GC name
    customer_id = job['customer_id']
    if not customer_id and bid and 'customer_id' in bid.keys():
        customer_id = bid['customer_id']
    cust = None
    if customer_id:
        cust = conn.execute('SELECT * FROM customers WHERE id = ?', (customer_id,)).fetchone()
    # Fallback: look up customer by contracting GC name
    if not cust and result.get('contracting_gc'):
        cust = conn.execute(
            'SELECT * FROM customers WHERE company_name = ? AND is_active = 1',
            (result['contracting_gc'],)
        ).fetchone()
    if cust:
        addr_parts = [cust['address'] or '', (cust['city'] or '') + ' ' + (cust['state'] or '') + ' ' + (cust['zip_code'] or '')]
        full_addr = ', '.join(p.strip() for p in addr_parts if p.strip())
        result['company_name'] = cust['company_name'] or ''
        result['address'] = full_addr
        result['primary_contact'] = cust['primary_contact'] or ''
        result['contact_email'] = cust['contact_email'] or ''
        result['contact_phone'] = cust['contact_phone'] or ''
    # Pull contract data (active contract for this job)
    contract = conn.execute(
        'SELECT * FROM contracts WHERE job_id = ? AND status = ? ORDER BY value DESC LIMIT 1',
        (job_id, 'Active')
    ).fetchone()
    if not contract:
        contract = conn.execute(
            'SELECT * FROM contracts WHERE job_id = ? ORDER BY value DESC LIMIT 1',
            (job_id,)
        ).fetchone()
    if contract:
        # Use contract value as fallback for contract sum
        if not result.get('total_bid') and contract['value']:
            result['total_bid'] = contract['value']
        # Contract date — prefer the explicit contract_date field, fall back to upload_date
        contract_cols = contract.keys()
        if 'contract_date' in contract_cols and contract['contract_date']:
            result['contract_date'] = contract['contract_date']
        elif contract['upload_date']:
            result['contract_date'] = contract['upload_date']
        # Contractor name as another fallback for GC name
        if not result.get('contracting_gc') and contract['contractor']:
            result['contracting_gc'] = contract['contractor']
        # If still no customer address, try looking up contractor as a customer
        if not result.get('address') and contract['contractor']:
            cust2 = conn.execute(
                'SELECT * FROM customers WHERE company_name = ? AND is_active = 1',
                (contract['contractor'],)
            ).fetchone()
            if cust2:
                addr_parts = [cust2['address'] or '', (cust2['city'] or '') + ' ' + (cust2['state'] or '') + ' ' + (cust2['zip_code'] or '')]
                full_addr = ', '.join(p.strip() for p in addr_parts if p.strip())
                result['address'] = full_addr
                if not result.get('company_name'):
                    result['company_name'] = cust2['company_name'] or ''
                if not result.get('primary_contact'):
                    result['primary_contact'] = cust2['primary_contact'] or ''
                if not result.get('contact_email'):
                    result['contact_email'] = cust2['contact_email'] or ''
                if not result.get('contact_phone'):
                    result['contact_phone'] = cust2['contact_phone'] or ''
    # Job awarded date as fallback for contract date
    if not result.get('contract_date') and job.get('awarded_date'):
        result['contract_date'] = job['awarded_date']
    # Project address from job
    job_addr_parts = [job['address'] or '', (job['city'] or '') + ' ' + (job['state'] or '')]
    result['project_address'] = ', '.join(p.strip() for p in job_addr_parts if p.strip())
    result['project_name'] = job['name'] or ''
    conn.close()
    return jsonify(result)

@app.route('/api/payapps/contracts', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_contracts_create():
    data = request.get_json(force=True)
    conn = get_db()
    job = conn.execute('SELECT name, address, city, state FROM jobs WHERE id = ?', (data.get('job_id'),)).fetchone()
    project_name = job['name'] if job else ''
    project_address = f"{job['address'] or ''}, {job['city'] or ''} {job['state'] or ''}".strip(', ') if job else ''
    cursor = conn.execute(
        '''INSERT INTO pay_app_contracts (job_id, gc_name, gc_address, project_name, project_address,
           project_no, contract_for, contract_date, original_contract_sum,
           retainage_work_pct, retainage_stored_pct, gc_contact, gc_email, gc_phone, created_by)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (data.get('job_id'), data.get('gc_name', ''), data.get('gc_address', ''),
         project_name, project_address,
         data.get('project_no', ''), data.get('contract_for', ''), data.get('contract_date', ''),
         data.get('original_contract_sum', 0),
         data.get('retainage_work_pct', 10), data.get('retainage_stored_pct', 0),
         data.get('gc_contact', ''), data.get('gc_email', ''), data.get('gc_phone', ''),
         session['user_id'])
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cursor.lastrowid})

@app.route('/api/payapps/contracts/<int:cid>/detail')
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_contract_detail(cid):
    conn = get_db()
    c = conn.execute('SELECT * FROM pay_app_contracts WHERE id = ?', (cid,)).fetchone()
    if not c:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    job = conn.execute('SELECT name FROM jobs WHERE id = ?', (c['job_id'],)).fetchone()
    result = dict(c)
    result['job_name'] = job['name'] if job else ''
    conn.close()
    return jsonify(result)

@app.route('/api/payapps/contracts/<int:cid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_contracts_update(cid):
    data = request.get_json(force=True)
    conn = get_db()
    fields = []
    params = []
    for key in ('gc_name', 'gc_address', 'project_no', 'contract_for', 'contract_date',
                'original_contract_sum', 'retainage_work_pct', 'retainage_stored_pct',
                'gc_contact', 'gc_email', 'gc_phone'):
        if key in data:
            fields.append(f'{key} = ?')
            params.append(data[key])
    if fields:
        params.append(cid)
        conn.execute(f'UPDATE pay_app_contracts SET {", ".join(fields)} WHERE id = ?', params)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/payapps/contracts/<int:cid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_contracts_delete(cid):
    conn = get_db()
    conn.execute('DELETE FROM pay_app_contracts WHERE id = ?', (cid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# --- SOV Items ---

@app.route('/api/payapps/contracts/<int:cid>/sov')
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_sov_list(cid):
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM pay_app_sov_items WHERE contract_id = ? ORDER BY sort_order, id', (cid,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/payapps/contracts/<int:cid>/sov', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_sov_create(cid):
    data = request.get_json(force=True)
    conn = get_db()
    insert_at = data.get('insert_at')  # sort_order position to insert before
    if insert_at is not None:
        insert_at = int(insert_at)
        # Shift all items at or after this position down by 1
        conn.execute('UPDATE pay_app_sov_items SET sort_order = sort_order + 1 WHERE contract_id = ? AND sort_order >= ?', (cid, insert_at))
        sort_order = insert_at
    else:
        max_sort = conn.execute('SELECT COALESCE(MAX(sort_order), -1) FROM pay_app_sov_items WHERE contract_id = ?', (cid,)).fetchone()[0]
        sort_order = max_sort + 1
    conn.execute(
        '''INSERT INTO pay_app_sov_items (contract_id, item_number, description, scheduled_value,
           is_header, retainage_exempt, sort_order)
           VALUES (?,?,?,?,?,?,?)''',
        (cid, 0, data.get('description', ''), data.get('scheduled_value', 0),
         data.get('is_header', 0), data.get('retainage_exempt', 0), sort_order)
    )
    # Renumber all item_numbers sequentially by sort_order
    rows = conn.execute('SELECT id FROM pay_app_sov_items WHERE contract_id = ? ORDER BY sort_order, id', (cid,)).fetchall()
    for idx, row in enumerate(rows):
        conn.execute('UPDATE pay_app_sov_items SET item_number = ? WHERE id = ?', (idx + 1, row['id']))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/payapps/sov/<int:sid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_sov_update(sid):
    data = request.get_json(force=True)
    conn = get_db()
    fields = []
    params = []
    for key in ('description', 'scheduled_value', 'is_header', 'retainage_exempt', 'sort_order', 'item_number'):
        if key in data:
            fields.append(f'{key} = ?')
            params.append(data[key])
    if fields:
        params.append(sid)
        conn.execute(f'UPDATE pay_app_sov_items SET {", ".join(fields)} WHERE id = ?', params)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/payapps/sov/<int:sid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_sov_delete(sid):
    conn = get_db()
    item = conn.execute('SELECT contract_id FROM pay_app_sov_items WHERE id = ?', (sid,)).fetchone()
    conn.execute('DELETE FROM pay_app_sov_items WHERE id = ?', (sid,))
    # Renumber remaining items
    if item:
        rows = conn.execute('SELECT id FROM pay_app_sov_items WHERE contract_id = ? ORDER BY sort_order, id', (item['contract_id'],)).fetchall()
        for idx, row in enumerate(rows):
            conn.execute('UPDATE pay_app_sov_items SET item_number = ?, sort_order = ? WHERE id = ?', (idx + 1, idx, row['id']))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/payapps/contracts/<int:cid>/sov/reorder', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_sov_reorder(cid):
    """Reorder SOV items. Expects {order: [{id, sort_order, item_number}, ...]}"""
    data = request.get_json(force=True)
    conn = get_db()
    for item in data.get('order', []):
        conn.execute('UPDATE pay_app_sov_items SET sort_order = ?, item_number = ? WHERE id = ? AND contract_id = ?',
                     (item['sort_order'], item['item_number'], item['id'], cid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# --- Applications ---

@app.route('/api/payapps/contracts/<int:cid>/applications')
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_apps_list(cid):
    conn = get_db()
    apps = conn.execute(
        'SELECT * FROM pay_applications WHERE contract_id = ? ORDER BY application_number', (cid,)
    ).fetchall()
    contract = conn.execute('SELECT * FROM pay_app_contracts WHERE id = ?', (cid,)).fetchone()
    result = []
    for a in apps:
        row = dict(a)
        row['current_payment_due'] = _calc_current_payment(conn, contract, a)
        result.append(row)
    conn.close()
    return jsonify(result)

@app.route('/api/payapps/contracts/<int:cid>/applications', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_apps_create(cid):
    conn = get_db()
    max_num = conn.execute(
        'SELECT MAX(application_number) FROM pay_applications WHERE contract_id = ?', (cid,)
    ).fetchone()[0] or 0
    today = datetime.now().strftime('%Y-%m-%d')
    cursor = conn.execute(
        '''INSERT INTO pay_applications (contract_id, application_number, period_to, application_date, created_by)
           VALUES (?,?,?,?,?)''',
        (cid, max_num + 1, today, today, session['user_id'])
    )
    app_id = cursor.lastrowid
    # Create empty entries for all SOV items
    sov_items = conn.execute('SELECT id FROM pay_app_sov_items WHERE contract_id = ? AND is_header = 0', (cid,)).fetchall()
    for item in sov_items:
        conn.execute(
            'INSERT INTO pay_app_line_entries (pay_app_id, sov_item_id, work_this_period, materials_stored) VALUES (?,?,0,0)',
            (app_id, item['id'])
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': app_id})

@app.route('/api/payapps/applications/<int:aid>')
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_app_detail(aid):
    conn = get_db()
    app_row = conn.execute('SELECT * FROM pay_applications WHERE id = ?', (aid,)).fetchone()
    if not app_row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    contract = conn.execute('SELECT * FROM pay_app_contracts WHERE id = ?', (app_row['contract_id'],)).fetchone()
    job = conn.execute('SELECT name FROM jobs WHERE id = ?', (contract['job_id'],)).fetchone()

    sov_items = conn.execute(
        'SELECT * FROM pay_app_sov_items WHERE contract_id = ? ORDER BY sort_order, id',
        (contract['id'],)
    ).fetchall()

    # Current entries
    current_entries = {}
    for entry in conn.execute('SELECT * FROM pay_app_line_entries WHERE pay_app_id = ?', (aid,)).fetchall():
        current_entries[entry['sov_item_id']] = dict(entry)

    # Prior apps
    prior_apps = conn.execute(
        'SELECT id FROM pay_applications WHERE contract_id = ? AND application_number < ? ORDER BY application_number',
        (contract['id'], app_row['application_number'])
    ).fetchall()
    prior_ids = [p['id'] for p in prior_apps]

    # "From previous" for each SOV item
    from_previous = {}
    if prior_ids:
        placeholders = ','.join('?' * len(prior_ids))
        for row in conn.execute(
            f'SELECT sov_item_id, SUM(work_this_period) as total_prev FROM pay_app_line_entries WHERE pay_app_id IN ({placeholders}) GROUP BY sov_item_id',
            prior_ids
        ).fetchall():
            from_previous[row['sov_item_id']] = row['total_prev'] or 0

    # Auto-calculate CO additions/deductions from change_orders table
    # Determine cutoff: period_to of previous application (or None if first app)
    co_cutoff = None
    if prior_ids:
        prev_app = conn.execute(
            'SELECT period_to FROM pay_applications WHERE contract_id = ? AND application_number = ?',
            (contract['id'], app_row['application_number'] - 1)
        ).fetchone()
        if prev_app and prev_app['period_to']:
            co_cutoff = prev_app['period_to']

    # Query all non-Void COs for this contract
    all_cos = conn.execute(
        "SELECT amount, created_at FROM change_orders WHERE pay_app_contract_id = ? AND status != 'Void'",
        (contract['id'],)
    ).fetchall()

    co_this_add = 0
    co_this_ded = 0
    co_prev_add = 0
    co_prev_ded = 0
    for co_row in all_cos:
        amt = co_row['amount'] or 0
        created = (co_row['created_at'] or '')[:10]
        is_previous = co_cutoff and created <= co_cutoff
        if is_previous:
            if amt >= 0:
                co_prev_add += amt
            else:
                co_prev_ded += abs(amt)
        else:
            if amt >= 0:
                co_this_add += amt
            else:
                co_this_ded += abs(amt)

    # Previous Line 6
    prev_line_6 = 0
    if prior_ids:
        prev_line_6 = _calc_earned_less_retainage(conn, contract, app_row['application_number'] - 1)

    # Build lines
    lines = []
    t = dict(scheduled_value=0, from_previous=0, this_period=0, materials_stored=0, total_completed=0, balance=0, retainage=0)

    for item in sov_items:
        entry = current_entries.get(item['id'], {})
        prev = from_previous.get(item['id'], 0)
        this_p = entry.get('work_this_period', 0) or 0
        mats = entry.get('materials_stored', 0) or 0
        sv = item['scheduled_value'] or 0

        completed = prev + this_p + mats
        pct = (completed / sv * 100) if sv else 0
        bal = sv - completed

        ret = 0
        if not item['is_header'] and not item['retainage_exempt'] and sv > 0:
            ret = (contract['retainage_work_pct'] / 100) * (prev + this_p)
            ret += (contract['retainage_stored_pct'] / 100) * mats
            ret = round(ret, 2)

        lines.append({
            'sov_item_id': item['id'], 'item_number': item['item_number'],
            'description': item['description'], 'scheduled_value': sv,
            'is_header': item['is_header'], 'retainage_exempt': item['retainage_exempt'],
            'from_previous': prev, 'this_period': this_p, 'materials_stored': mats,
            'total_completed': completed, 'pct_complete': round(pct, 0),
            'balance': bal, 'retainage': ret,
        })

        if not item['is_header']:
            t['scheduled_value'] += sv
            t['from_previous'] += prev
            t['this_period'] += this_p
            t['materials_stored'] += mats
            t['total_completed'] += completed
            t['balance'] += bal
            t['retainage'] += ret

    t['retainage'] = round(t['retainage'], 2)

    conn.close()
    return jsonify({
        'application': dict(app_row),
        'contract': dict(contract),
        'job_name': job['name'] if job else '',
        'lines': lines,
        'totals': t,
        'g702': {
            'co_this_additions': co_this_add,
            'co_this_deductions': co_this_ded,
            'co_prev_additions': co_prev_add,
            'co_prev_deductions': co_prev_ded,
            'previous_certificates': prev_line_6,
        }
    })

@app.route('/api/payapps/applications/<int:aid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_app_update(aid):
    data = request.get_json(force=True)
    conn = get_db()

    # Update meta fields
    for field in ('period_to', 'application_date', 'co_additions', 'co_deductions', 'status'):
        if field in data:
            conn.execute(f'UPDATE pay_applications SET {field} = ? WHERE id = ?', (data[field], aid))

    # Update line entries
    if 'entries' in data:
        for entry in data['entries']:
            conn.execute(
                '''INSERT INTO pay_app_line_entries (pay_app_id, sov_item_id, work_this_period, materials_stored)
                   VALUES (?,?,?,?)
                   ON CONFLICT(pay_app_id, sov_item_id) DO UPDATE SET
                   work_this_period = excluded.work_this_period,
                   materials_stored = excluded.materials_stored''',
                (aid, entry['sov_item_id'], entry.get('work_this_period', 0), entry.get('materials_stored', 0))
            )

    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/payapps/applications/<int:aid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_app_delete(aid):
    conn = get_db()
    conn.execute('DELETE FROM pay_applications WHERE id = ?', (aid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/payapps/applications/<int:aid>/generate-pdf', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_generate_pdf(aid):
    """Generate G702/G703 PDF for a pay application."""

    # Re-use the detail logic to get all data
    conn = get_db()
    app_row = conn.execute('SELECT * FROM pay_applications WHERE id = ?', (aid,)).fetchone()
    if not app_row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    contract = conn.execute('SELECT * FROM pay_app_contracts WHERE id = ?', (app_row['contract_id'],)).fetchone()
    job = conn.execute('SELECT name FROM jobs WHERE id = ?', (contract['job_id'],)).fetchone()
    job_name = job['name'] if job else ''

    sov_items = conn.execute(
        'SELECT * FROM pay_app_sov_items WHERE contract_id = ? ORDER BY sort_order, id',
        (contract['id'],)
    ).fetchall()

    current_entries = {}
    for entry in conn.execute('SELECT * FROM pay_app_line_entries WHERE pay_app_id = ?', (aid,)).fetchall():
        current_entries[entry['sov_item_id']] = dict(entry)

    prior_apps = conn.execute(
        'SELECT id FROM pay_applications WHERE contract_id = ? AND application_number < ? ORDER BY application_number',
        (contract['id'], app_row['application_number'])
    ).fetchall()
    prior_ids = [p['id'] for p in prior_apps]

    from_previous = {}
    if prior_ids:
        placeholders = ','.join('?' * len(prior_ids))
        for row in conn.execute(
            f'SELECT sov_item_id, SUM(work_this_period) as total_prev FROM pay_app_line_entries WHERE pay_app_id IN ({placeholders}) GROUP BY sov_item_id',
            prior_ids
        ).fetchall():
            from_previous[row['sov_item_id']] = row['total_prev'] or 0

    # Auto-calculate CO additions/deductions from change_orders table
    co_cutoff = None
    if prior_ids:
        prev_app = conn.execute(
            'SELECT period_to FROM pay_applications WHERE contract_id = ? AND application_number = ?',
            (contract['id'], app_row['application_number'] - 1)
        ).fetchone()
        if prev_app and prev_app['period_to']:
            co_cutoff = prev_app['period_to']

    all_cos = conn.execute(
        "SELECT amount, created_at FROM change_orders WHERE pay_app_contract_id = ? AND status != 'Void'",
        (contract['id'],)
    ).fetchall()

    co_this_add = 0
    co_this_ded = 0
    prior_co_add = 0
    prior_co_ded = 0
    for co_row in all_cos:
        amt = co_row['amount'] or 0
        created = (co_row['created_at'] or '')[:10]
        is_previous = co_cutoff and created <= co_cutoff
        if is_previous:
            if amt >= 0:
                prior_co_add += amt
            else:
                prior_co_ded += abs(amt)
        else:
            if amt >= 0:
                co_this_add += amt
            else:
                co_this_ded += abs(amt)

    prev_line_6 = 0
    if prior_ids:
        prev_line_6 = _calc_earned_less_retainage(conn, contract, app_row['application_number'] - 1)

    lines = []
    t = dict(scheduled_value=0, from_previous=0, this_period=0, materials_stored=0, total_completed=0, balance=0, retainage=0)

    for item in sov_items:
        entry = current_entries.get(item['id'], {})
        prev = from_previous.get(item['id'], 0)
        this_p = entry.get('work_this_period', 0) or 0
        mats = entry.get('materials_stored', 0) or 0
        sv = item['scheduled_value'] or 0

        completed = prev + this_p + mats
        pct = (completed / sv * 100) if sv else 0
        bal = sv - completed

        ret = 0
        if not item['is_header'] and not item['retainage_exempt'] and sv > 0:
            ret = (contract['retainage_work_pct'] / 100) * (prev + this_p)
            ret += (contract['retainage_stored_pct'] / 100) * mats
            ret = round(ret, 2)

        lines.append({
            'sov_item_id': item['id'], 'item_number': item['item_number'],
            'description': item['description'], 'scheduled_value': sv,
            'is_header': item['is_header'], 'retainage_exempt': item['retainage_exempt'],
            'from_previous': prev, 'this_period': this_p, 'materials_stored': mats,
            'total_completed': completed, 'pct_complete': round(pct, 0),
            'balance': bal, 'retainage': ret,
        })

        if not item['is_header']:
            t['scheduled_value'] += sv
            t['from_previous'] += prev
            t['this_period'] += this_p
            t['materials_stored'] += mats
            t['total_completed'] += completed
            t['balance'] += bal
            t['retainage'] += ret

    t['retainage'] = round(t['retainage'], 2)

    # Calculate G702 values using auto-calculated CO data
    co_add = co_this_add
    co_ded = co_this_ded
    net_co = (prior_co_add + co_add) - (prior_co_ded + co_ded)

    l1 = contract['original_contract_sum'] or 0
    l2 = net_co
    l3 = l1 + l2
    l4 = t['total_completed']
    l5 = t['retainage']
    l6 = l4 - l5
    l7 = prev_line_6
    l8 = l6 - l7
    l9 = l3 - l6

    ret_work = 0
    ret_stored = 0
    for line in lines:
        if not line['is_header'] and not line['retainage_exempt']:
            ret_work += (contract['retainage_work_pct'] / 100) * (line['from_previous'] + line['this_period'])
        if not line['is_header']:
            ret_stored += (contract['retainage_stored_pct'] / 100) * (line['materials_stored'] or 0)

    g702 = {
        'l1': l1, 'l2': l2, 'l3': l3, 'l4': l4, 'l5': l5, 'l5a': ret_work, 'l5b': ret_stored,
        'l6': l6, 'l7': l7, 'l8': l8, 'l9': l9,
        'co_prev_add': prior_co_add, 'co_prev_ded': prior_co_ded,
        'co_this_add': co_this_add, 'co_this_ded': co_this_ded,
    }

    def fmt_money(n):
        return '${:,.2f}'.format(n or 0)

    def fmt_date(d):
        """Convert YYYY-MM-DD to M-D-YYYY (e.g. 2026-03-03 → 3-3-2026)."""
        if not d or len(d) < 10:
            return d or ''
        try:
            parts = d[:10].split('-')
            return f"{int(parts[1])}-{int(parts[2])}-{parts[0]}"
        except (ValueError, IndexError):
            return d

    logo_path = os.path.abspath(os.path.join(app.static_folder, 'logo.jpg'))

    # Calculate total pages: page 1 = G702, then G703 pages
    non_header_count = sum(1 for l in lines if not l['is_header'])
    header_count = sum(1 for l in lines if l['is_header'])
    total_lines = non_header_count + header_count + 1  # +1 for grand totals row
    total_pages = 1 + max(1, -(-total_lines // 35))  # ceil division, ~35 lines per G703 page

    # Check for contractor signature image
    sig_dir = os.path.join(os.path.dirname(__file__), 'data', 'signatures')
    sig_path = os.path.join(sig_dir, 'contractor_signature.png')
    signature_file_url = 'file://' + os.path.abspath(sig_path) if os.path.exists(sig_path) else None

    html = render_template('payapps/payapp_pdf.html',
        application=dict(app_row), contract=dict(contract), job_name=job_name,
        lines=lines, totals=t, g702=g702,
        fmt=fmt_money, fmt_date=fmt_date, logo_path='file://' + logo_path,
        total_pages=total_pages, signature_path=signature_file_url
    )

    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    os.makedirs(proposals_dir, exist_ok=True)

    safe_name = ''.join(c if c.isalnum() or c in ' -_' else '' for c in job_name).strip()
    filename = f"PayApp_{safe_name}_App{app_row['application_number']}.pdf"
    filepath = os.path.join(proposals_dir, filename)

    if weasyprint is None:
        conn.close()
        return jsonify({'error': 'PDF generation is not available. WeasyPrint is not installed on the server. Please contact your administrator.'}), 500

    try:
        wp = weasyprint.HTML(string=html, base_url=os.path.dirname(__file__))
        wp.write_pdf(filepath)
    except Exception as e:
        conn.close()
        return jsonify({'error': f'PDF generation failed: {str(e)[:200]}'}), 500

    # Save filename to DB
    conn.execute('UPDATE pay_applications SET pdf_file = ? WHERE id = ?', (filename, aid))
    conn.commit()
    conn.close()

    return jsonify({'ok': True, 'filename': filename, 'path': f'/api/payapps/applications/{aid}/pdf/{filename}'})


@app.route('/api/payapps/applications/<int:aid>/pdf/<filename>')
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_download_pdf(aid, filename):
    """View or download a generated pay app PDF."""
    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    filepath = os.path.join(proposals_dir, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    if request.args.get('download'):
        return send_file(filepath, as_attachment=True, download_name=filename)
    return send_file(filepath, mimetype='application/pdf')


@app.route('/api/settings/signature', methods=['POST'])
@api_role_required('owner', 'admin')
def api_upload_signature():
    """Upload a contractor signature image (PNG/JPG)."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400
    sig_dir = os.path.join(os.path.dirname(__file__), 'data', 'signatures')
    os.makedirs(sig_dir, exist_ok=True)
    sig_path = os.path.join(sig_dir, 'contractor_signature.png')
    f.save(sig_path)
    return jsonify({'ok': True})

@app.route('/api/settings/signature')
@api_role_required('owner', 'admin', 'project_manager')
def api_get_signature():
    """Serve the contractor signature image."""
    sig_dir = os.path.join(os.path.dirname(__file__), 'data', 'signatures')
    sig_path = os.path.join(sig_dir, 'contractor_signature.png')
    if not os.path.exists(sig_path):
        return jsonify({'error': 'No signature uploaded'}), 404
    return send_file(sig_path, mimetype='image/png')

@app.route('/api/settings/signature', methods=['DELETE'])
@api_role_required('owner', 'admin')
def api_delete_signature():
    """Delete the contractor signature image."""
    sig_dir = os.path.join(os.path.dirname(__file__), 'data', 'signatures')
    sig_path = os.path.join(sig_dir, 'contractor_signature.png')
    if os.path.exists(sig_path):
        os.remove(sig_path)
    return jsonify({'ok': True})

@app.route('/api/payapps/applications/<int:aid>/email', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_email(aid):
    """Email pay app PDF to specified recipients."""
    data = request.get_json()
    recipients = [e.strip() for e in data.get('recipients', []) if e.strip()]
    subject = data.get('subject', 'Pay Application')
    body_text = data.get('body', '')

    if not recipients:
        return jsonify({'error': 'No recipients specified'}), 400

    # Load saved SMTP settings
    saved_settings = {}
    settings_path = os.path.join(os.path.dirname(__file__), 'data', 'email_settings.json')
    if os.path.exists(settings_path):
        with open(settings_path) as f:
            saved_settings = json.load(f)

    smtp_host = saved_settings.get('smtp_host', '')
    smtp_port = int(saved_settings.get('smtp_port', 587) or 587)
    smtp_user = saved_settings.get('smtp_user', '')
    smtp_pass = saved_settings.get('smtp_pass', '')
    from_email = saved_settings.get('from_email', '') or smtp_user

    if not smtp_host or not smtp_user:
        return jsonify({'error': 'SMTP not configured. Go to Settings to set up email.'}), 400

    # Find the PDF
    conn = get_db()
    app_row = conn.execute('SELECT pdf_file FROM pay_applications WHERE id = ?', (aid,)).fetchone()
    conn.close()
    if not app_row or not app_row['pdf_file']:
        return jsonify({'error': 'No PDF generated yet. Generate the PDF first.'}), 404

    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    pdf_path = os.path.join(proposals_dir, app_row['pdf_file'])
    if not os.path.exists(pdf_path):
        return jsonify({'error': 'PDF file not found on disk.'}), 404

    try:
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = subject
        msg.attach(MIMEText(body_text, 'plain'))

        with open(pdf_path, 'rb') as f:
            part = MIMEBase('application', 'pdf')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(pdf_path)}"')
            msg.attach(part)

        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)

        save_email_autocomplete(recipients)
        return jsonify({'ok': True, 'sent_to': recipients})
    except Exception as e:
        return jsonify({'error': f'Email failed: {str(e)}'}), 500

@app.route('/api/payapps/applications/<int:aid>/signed', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_upload_signed(aid):
    """Upload a signed/notarized pay app document."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400

    signed_dir = os.path.join(os.path.dirname(__file__), 'data', 'payapps_signed')
    os.makedirs(signed_dir, exist_ok=True)

    ext = os.path.splitext(f.filename)[1].lower() or '.pdf'
    filename = f"signed_app_{aid}{ext}"
    filepath = os.path.join(signed_dir, filename)
    f.save(filepath)

    conn = get_db()
    conn.execute('UPDATE pay_applications SET signed_file = ? WHERE id = ?', (filename, aid))
    conn.commit()
    conn.close()

    return jsonify({'ok': True, 'filename': filename})

@app.route('/api/payapps/applications/<int:aid>/signed/<filename>')
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_view_signed(aid, filename):
    """View a signed/notarized pay app document."""
    signed_dir = os.path.join(os.path.dirname(__file__), 'data', 'payapps_signed')
    filepath = os.path.join(signed_dir, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    ext = os.path.splitext(filename)[1].lower()
    mime = 'application/pdf' if ext == '.pdf' else 'image/png' if ext == '.png' else 'image/jpeg'
    return send_file(filepath, mimetype=mime)

@app.route('/api/payapps/contracts/<int:cid>/upload-existing', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_upload_existing(cid):
    """Upload an already-submitted pay app PDF and create it as PA#N."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400
    period_to = request.form.get('period_to', '')
    status = request.form.get('status', 'Submitted')
    if status not in ('Draft', 'Submitted', 'Approved', 'Paid'):
        status = 'Submitted'
    conn = get_db()
    max_num = conn.execute(
        'SELECT MAX(application_number) FROM pay_applications WHERE contract_id = ?', (cid,)
    ).fetchone()[0] or 0
    app_number = max_num + 1
    today = datetime.now().strftime('%Y-%m-%d')
    cursor = conn.execute(
        '''INSERT INTO pay_applications (contract_id, application_number, period_to, application_date, status, created_by)
           VALUES (?,?,?,?,?,?)''',
        (cid, app_number, period_to or today, today, status, session['user_id'])
    )
    aid = cursor.lastrowid
    # Save the uploaded file as the signed copy
    signed_dir = os.path.join(os.path.dirname(__file__), 'data', 'payapps_signed')
    os.makedirs(signed_dir, exist_ok=True)
    ext = os.path.splitext(f.filename)[1].lower() or '.pdf'
    filename = f"signed_app_{aid}{ext}"
    filepath = os.path.join(signed_dir, filename)
    f.save(filepath)
    conn.execute('UPDATE pay_applications SET signed_file = ? WHERE id = ?', (filename, aid))
    # Create empty line entries for all SOV items so user can enter draw amounts
    sov_items = conn.execute('SELECT id FROM pay_app_sov_items WHERE contract_id = ? AND is_header = 0', (cid,)).fetchall()
    for item in sov_items:
        conn.execute(
            'INSERT OR IGNORE INTO pay_app_line_entries (pay_app_id, sov_item_id, work_this_period, materials_stored) VALUES (?,?,0,0)',
            (aid, item['id']))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': aid, 'application_number': app_number})

@app.route('/payapps/analytics')
@role_required('owner', 'admin', 'project_manager')
def payapps_analytics_page():
    return render_template('payapps/analytics.html')


@app.route('/api/payapps/analytics')
@api_role_required('owner', 'admin', 'project_manager')
def api_payapps_analytics():
    """Return analytics data for all pay app contracts."""
    conn = get_db()
    contracts = conn.execute('''
        SELECT c.*, j.name as job_name
        FROM pay_app_contracts c
        LEFT JOIN jobs j ON c.job_id = j.id
        ORDER BY c.created_at DESC
    ''').fetchall()

    projects = []
    agg = {'total_contract_value': 0, 'total_billed': 0, 'total_retainage': 0, 'total_balance': 0}
    status_counts = {'Draft': 0, 'Submitted': 0, 'Approved': 0, 'Paid': 0}

    for c in contracts:
        cid = c['id']
        original = c['original_contract_sum'] or 0

        # Latest application
        latest_app = conn.execute(
            'SELECT * FROM pay_applications WHERE contract_id = ? ORDER BY application_number DESC LIMIT 1',
            (cid,)
        ).fetchone()

        # Sum COs from all applications
        co_row = conn.execute(
            'SELECT SUM(co_additions) as adds, SUM(co_deductions) as deds FROM pay_applications WHERE contract_id = ?',
            (cid,)
        ).fetchone()
        co_additions = (co_row['adds'] or 0) if co_row else 0
        co_deductions = (co_row['deds'] or 0) if co_row else 0
        net_co = co_additions - co_deductions
        contract_sum_to_date = original + net_co

        # Totals from all SOV items + all entries
        sov_items = conn.execute(
            'SELECT * FROM pay_app_sov_items WHERE contract_id = ? AND is_header = 0',
            (cid,)
        ).fetchall()

        all_apps = conn.execute(
            'SELECT id, application_number FROM pay_applications WHERE contract_id = ? ORDER BY application_number',
            (cid,)
        ).fetchall()
        all_app_ids = [a['id'] for a in all_apps]

        total_work = 0
        total_stored = 0
        total_retainage = 0

        if all_app_ids:
            ph = ','.join('?' * len(all_app_ids))
            # Sum work_this_period across all apps per SOV item
            work_sums = {}
            for row in conn.execute(
                f'SELECT sov_item_id, SUM(work_this_period) as tw FROM pay_app_line_entries WHERE pay_app_id IN ({ph}) GROUP BY sov_item_id',
                all_app_ids
            ).fetchall():
                work_sums[row['sov_item_id']] = row['tw'] or 0

            # Get latest app's materials_stored
            latest_id = all_app_ids[-1] if all_app_ids else None
            mat_sums = {}
            if latest_id:
                for row in conn.execute(
                    'SELECT sov_item_id, materials_stored FROM pay_app_line_entries WHERE pay_app_id = ?',
                    (latest_id,)
                ).fetchall():
                    mat_sums[row['sov_item_id']] = row['materials_stored'] or 0

            for item in sov_items:
                sid = item['id']
                work = work_sums.get(sid, 0)
                mats = mat_sums.get(sid, 0)
                total_work += work
                total_stored += mats
                if not item['retainage_exempt'] and (item['scheduled_value'] or 0) > 0:
                    total_retainage += (c['retainage_work_pct'] / 100) * work
                    total_retainage += (c['retainage_stored_pct'] / 100) * mats

        total_retainage = round(total_retainage, 2)
        total_completed = total_work + total_stored
        total_billed = total_completed - total_retainage
        balance = contract_sum_to_date - total_billed
        pct = (total_completed / contract_sum_to_date * 100) if contract_sum_to_date else 0

        latest_status = latest_app['status'] if latest_app else None
        if latest_status:
            status_counts[latest_status] = status_counts.get(latest_status, 0) + 1

        projects.append({
            'contract_id': cid,
            'job_name': c['job_name'] or '',
            'project_name': c['project_name'] or '',
            'gc_name': c['gc_name'] or '',
            'original_contract_sum': original,
            'co_additions': co_additions,
            'co_deductions': co_deductions,
            'net_co': net_co,
            'contract_sum_to_date': contract_sum_to_date,
            'total_completed': round(total_completed, 2),
            'total_retainage': total_retainage,
            'total_billed': round(total_billed, 2),
            'balance_to_finish': round(balance, 2),
            'pct_complete': round(pct, 1),
            'latest_app_number': latest_app['application_number'] if latest_app else None,
            'latest_status': latest_status,
            'latest_period_to': latest_app['period_to'] if latest_app else None,
        })

        agg['total_contract_value'] += contract_sum_to_date
        agg['total_billed'] += total_billed
        agg['total_retainage'] += total_retainage
        agg['total_balance'] += balance

    conn.close()
    # Round aggregates
    for k in agg:
        agg[k] = round(agg[k], 2)

    return jsonify({'projects': projects, 'kpis': agg, 'status_counts': status_counts})


def _calc_earned_less_retainage(conn, contract, up_to_app_number):
    """Calculate Total Earned Less Retainage (G702 Line 6) through a given app number."""
    sov_items = conn.execute(
        'SELECT * FROM pay_app_sov_items WHERE contract_id = ? ORDER BY sort_order',
        (contract['id'],)
    ).fetchall()
    apps = conn.execute(
        'SELECT id, application_number FROM pay_applications WHERE contract_id = ? AND application_number <= ? ORDER BY application_number',
        (contract['id'], up_to_app_number)
    ).fetchall()
    if not apps:
        return 0

    app_ids = [a['id'] for a in apps]
    latest_app_id = app_ids[-1]

    # Sum work_this_period per SOV item across all apps, and get latest materials_stored
    all_entries = {}
    placeholders = ','.join('?' * len(app_ids))
    for row in conn.execute(
        f'SELECT sov_item_id, pay_app_id, work_this_period, materials_stored FROM pay_app_line_entries WHERE pay_app_id IN ({placeholders})',
        app_ids
    ).fetchall():
        sid = row['sov_item_id']
        if sid not in all_entries:
            all_entries[sid] = {'work_total': 0, 'latest_materials': 0}
        all_entries[sid]['work_total'] += (row['work_this_period'] or 0)
        if row['pay_app_id'] == latest_app_id:
            all_entries[sid]['latest_materials'] = row['materials_stored'] or 0

    total_completed = 0
    total_retainage = 0
    for item in sov_items:
        if item['is_header']:
            continue
        ed = all_entries.get(item['id'], {'work_total': 0, 'latest_materials': 0})
        work = ed['work_total']
        mats = ed['latest_materials']
        total_completed += work + mats
        if not item['retainage_exempt'] and (item['scheduled_value'] or 0) > 0:
            total_retainage += (contract['retainage_work_pct'] / 100) * work
            total_retainage += (contract['retainage_stored_pct'] / 100) * mats

    return round(total_completed - total_retainage, 2)


def _calc_current_payment(conn, contract, app_row):
    """Quick calculation of current payment due for list display."""
    sov_items = conn.execute(
        'SELECT * FROM pay_app_sov_items WHERE contract_id = ? AND is_header = 0',
        (contract['id'],)
    ).fetchall()
    all_apps = conn.execute(
        'SELECT id, application_number FROM pay_applications WHERE contract_id = ? AND application_number <= ? ORDER BY application_number',
        (contract['id'], app_row['application_number'])
    ).fetchall()
    if not all_apps:
        return 0

    app_ids = [a['id'] for a in all_apps]
    current_id = app_row['id']
    prior_ids = [a['id'] for a in all_apps if a['id'] != current_id]

    # Current entries
    entries = {}
    for e in conn.execute('SELECT * FROM pay_app_line_entries WHERE pay_app_id = ?', (current_id,)).fetchall():
        entries[e['sov_item_id']] = e

    # From previous
    from_prev = {}
    if prior_ids:
        ph = ','.join('?' * len(prior_ids))
        for r in conn.execute(f'SELECT sov_item_id, SUM(work_this_period) as t FROM pay_app_line_entries WHERE pay_app_id IN ({ph}) GROUP BY sov_item_id', prior_ids).fetchall():
            from_prev[r['sov_item_id']] = r['t'] or 0

    total_completed = 0
    total_retainage = 0
    for item in sov_items:
        prev = from_prev.get(item['id'], 0)
        e = entries.get(item['id'])
        this_p = (e['work_this_period'] or 0) if e else 0
        mats = (e['materials_stored'] or 0) if e else 0
        total_completed += prev + this_p + mats
        if not item['retainage_exempt'] and (item['scheduled_value'] or 0) > 0:
            total_retainage += (contract['retainage_work_pct'] / 100) * (prev + this_p)
            total_retainage += (contract['retainage_stored_pct'] / 100) * mats

    line_6 = total_completed - total_retainage

    # Previous line 6
    prev_line_6 = 0
    if prior_ids:
        prev_line_6 = _calc_earned_less_retainage(conn, contract, app_row['application_number'] - 1)

    return round(line_6 - prev_line_6, 2)


# ─── Equipment Manuals ───────────────────────────────────────────

MANUALS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'manuals')
os.makedirs(MANUALS_DIR, exist_ok=True)

@app.route('/manuals')
@login_required
def manuals_page():
    return render_template('manuals/list.html')

@app.route('/api/manuals')
@api_login_required
def api_manuals_list():
    q = request.args.get('q', '').strip()
    mfg = request.args.get('manufacturer', '').strip()
    conn = get_db()
    sql = 'SELECT * FROM equipment_manuals WHERE 1=1'
    params = []
    if q:
        sql += ' AND (model_number LIKE ? OR title LIKE ? OR manufacturer LIKE ?)'
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if mfg:
        sql += ' AND manufacturer = ?'
        params.append(mfg)
    sql += ' ORDER BY manufacturer, model_number'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/manuals', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_manuals_create():
    conn = get_db()
    if request.content_type and 'multipart' in request.content_type:
        mfg = request.form.get('manufacturer', '')
        model = request.form.get('model_number', '')
        title = request.form.get('title', '')
        mtype = request.form.get('manual_type', 'Installation')
        file = request.files.get('file')
        file_path = ''
        if file and file.filename:
            from werkzeug.utils import secure_filename
            fname = secure_filename(file.filename)
            # Avoid collisions
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.save(os.path.join(MANUALS_DIR, fname))
            file_path = fname
        conn.execute(
            '''INSERT INTO equipment_manuals (manufacturer, model_number, manual_type, title, file_path, uploaded_by)
               VALUES (?,?,?,?,?,?)''',
            (mfg, model, mtype, title, file_path, session['user_id'])
        )
    else:
        data = request.get_json(force=True)
        conn.execute(
            '''INSERT INTO equipment_manuals (manufacturer, model_number, manual_type, title, external_url, brand, equipment_type, tonnage, fuel_type, tags, uploaded_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
            (data.get('manufacturer', ''), data.get('model_number', ''), data.get('manual_type', 'Installation'),
             data.get('title', ''), data.get('external_url', ''),
             data.get('brand', data.get('manufacturer', '')), data.get('equipment_type', ''),
             data.get('tonnage', ''), data.get('fuel_type', ''), data.get('tags', ''),
             session['user_id'])
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/manuals/<int:mid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_manuals_delete(mid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM equipment_manuals WHERE id = ?', (mid,)).fetchone()
    if row and row['file_path']:
        fpath = os.path.join(MANUALS_DIR, row['file_path'])
        if os.path.exists(fpath):
            os.remove(fpath)
    conn.execute('DELETE FROM equipment_manuals WHERE id = ?', (mid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/manuals/<int:mid>/file')
@api_login_required
def api_manuals_file(mid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM equipment_manuals WHERE id = ?', (mid,)).fetchone()
    conn.close()
    if not row or not row['file_path']:
        return 'Not found', 404
    fpath = os.path.join(MANUALS_DIR, row['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    return send_file(fpath, mimetype='application/pdf')

@app.route('/api/manuals/bulk-import', methods=['POST'])
@api_role_required('owner', 'admin')
def api_manuals_bulk_import():
    """Import manuals from JSON data."""
    data = request.get_json()
    items = data.get('items', [])
    conn = get_db()
    count = 0
    for item in items:
        conn.execute(
            '''INSERT INTO equipment_manuals (manufacturer, model_number, manual_type, title, external_url, brand, equipment_type, tonnage, fuel_type, tags, uploaded_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
            (item.get('manufacturer',''), item.get('model_number',''), item.get('manual_type','Installation'),
             item.get('title',''), item.get('external_url',''), item.get('brand',''),
             item.get('equipment_type',''), item.get('tonnage',''), item.get('fuel_type',''),
             item.get('tags',''), session.get('user_id'))
        )
        count += 1
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'imported': count})

@app.route('/api/manuals/brands')
@api_login_required
def api_manuals_brands():
    """Get distinct brand list."""
    conn = get_db()
    brands = conn.execute("SELECT DISTINCT brand FROM equipment_manuals WHERE brand != '' ORDER BY brand").fetchall()
    types = conn.execute("SELECT DISTINCT equipment_type FROM equipment_manuals WHERE equipment_type != '' ORDER BY equipment_type").fetchall()
    conn.close()
    return jsonify({
        'brands': [b['brand'] for b in brands],
        'equipment_types': [t['equipment_type'] for t in types]
    })

@app.route('/api/manuals/search')
@api_login_required
def api_manuals_search():
    """Enhanced search with brand/type/tonnage filters."""
    q = request.args.get('q', '').strip()
    brand = request.args.get('brand', '')
    etype = request.args.get('type', '')
    clauses = []
    params = []
    if q:
        clauses.append("(manufacturer LIKE ? OR model_number LIKE ? OR title LIKE ? OR tags LIKE ?)")
        params.extend([f'%{q}%'] * 4)
    if brand:
        clauses.append("brand = ?")
        params.append(brand)
    if etype:
        clauses.append("equipment_type = ?")
        params.append(etype)
    where = ' AND '.join(clauses) if clauses else '1=1'
    conn = get_db()
    results = conn.execute(f'SELECT * FROM equipment_manuals WHERE {where} ORDER BY manufacturer, model_number', params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in results])

@app.route('/api/codebooks/<int:book_id>/sections/<int:sid>/content', methods=['POST'])
@api_role_required('owner', 'admin')
def api_codebook_update_section_content(book_id, sid):
    """Update section content."""
    data = request.get_json()
    conn = get_db()
    conn.execute('UPDATE code_sections SET content = ?, is_complete = 1 WHERE id = ? AND book_id = ?',
                 (data.get('content', ''), sid, book_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/codebooks/import-sections', methods=['POST'])
@api_role_required('owner', 'admin')
def api_codebook_import_sections():
    """Bulk import code book sections."""
    data = request.get_json()
    sections = data.get('sections', [])
    conn = get_db()
    count = 0
    for s in sections:
        conn.execute(
            '''INSERT INTO code_sections (book_id, section_number, title, content, parent_section_id, depth, sort_order, is_complete)
               VALUES (?,?,?,?,?,?,?,?)''',
            (s.get('book_id'), s.get('section_number',''), s.get('title',''), s.get('content',''),
             s.get('parent_section_id'), s.get('depth', 0), s.get('sort_order', 0), 1 if s.get('content') else 0)
        )
        count += 1
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'imported': count})


# ─── Job Schedule ────────────────────────────────────────────────

@app.route('/schedule')
@login_required
def schedule_page():
    return render_template('schedule/list.html')

@app.route('/schedule/job/<int:job_id>')
@login_required
def schedule_job_page(job_id):
    return render_template('schedule/job.html', job_id=job_id)

@app.route('/api/schedule/events')
@api_login_required
def api_schedule_events():
    job_id = request.args.get('job_id', '')
    conn = get_db()
    if job_id:
        rows = conn.execute(
            'SELECT * FROM job_schedule_events WHERE job_id = ? ORDER BY sort_order, id',
            (job_id,)
        ).fetchall()
    else:
        rows = conn.execute(
            'SELECT * FROM job_schedule_events ORDER BY job_id, sort_order, id'
        ).fetchall()
    result = [dict(r) for r in rows]
    conn.close()
    return jsonify(result)

def _check_schedule_notifications(events):
    """Create notifications for schedule benchmarks: due soon, starting soon, overdue, missed start."""
    from datetime import timedelta
    now = datetime.now()
    today = now.date()
    today_str = now.strftime('%Y-%m-%d')
    two_days = now + timedelta(hours=48)
    conn = get_db()
    try:
        _do_schedule_notifications(conn, events, now, today, today_str, two_days)
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()

def _do_schedule_notifications(conn, events, now, today, today_str, two_days):
    from datetime import timedelta
    # Get all owner user IDs for escalation
    owner_ids = [r['id'] for r in conn.execute("SELECT id FROM users WHERE role = 'owner'").fetchall()]

    def _already_notified(user_id, event_id, ntype_suffix):
        return conn.execute(
            '''SELECT id FROM notifications WHERE user_id = ? AND type = 'schedule'
               AND message LIKE ? AND created_at >= ?''',
            (user_id, f'%event #{event_id}%{ntype_suffix}%', today_str)
        ).fetchone()

    def _get_job_name(job_id):
        job = conn.execute('SELECT name FROM jobs WHERE id = ?', (job_id,)).fetchone()
        return job['name'] if job else f'Job #{job_id}'

    for e in events:
        if e['status'] in ('Complete', 'Cancelled'):
            continue
        link = f'/schedule/job/{e["job_id"]}'
        job_name = _get_job_name(e['job_id'])
        assigned = e.get('assigned_to')

        # 1. End date within 24h (existing behavior)
        if e.get('end_date') and assigned:
            try:
                end_dt = datetime.strptime(e['end_date'], '%Y-%m-%d')
            except (ValueError, TypeError):
                end_dt = None
            if end_dt and now <= end_dt <= now + timedelta(hours=24):
                if not _already_notified(assigned, e['id'], 'due_soon'):
                    conn.execute(
                        'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
                        (assigned, 'schedule', 'Schedule: Due Soon',
                         f'{e["phase_name"]} on {job_name} due {e["end_date"]} (event #{e["id"]}) due_soon', link)
                    )

        # 2. Start date within 48 hours - notify assigned user
        if e.get('start_date') and assigned and e['status'] == 'Pending':
            try:
                start_dt = datetime.strptime(e['start_date'], '%Y-%m-%d')
            except (ValueError, TypeError):
                start_dt = None
            if start_dt and now <= start_dt <= two_days:
                if not _already_notified(assigned, e['id'], 'start_soon'):
                    conn.execute(
                        'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
                        (assigned, 'schedule', 'Schedule: Starting Soon',
                         f'{e["phase_name"]} on {job_name} starts in 2 days (event #{e["id"]}) start_soon', link)
                    )

        # 3. Overdue: In Progress phase past end_date
        if e.get('end_date') and e['status'] == 'In Progress':
            try:
                end_dt = datetime.strptime(e['end_date'], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                end_dt = None
            if end_dt and today > end_dt:
                notify_users = set(owner_ids)
                if assigned:
                    notify_users.add(assigned)
                for uid in notify_users:
                    if not _already_notified(uid, e['id'], 'overdue'):
                        conn.execute(
                            'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
                            (uid, 'schedule', 'Schedule: Phase Overdue',
                             f'{e["phase_name"]} on {job_name} is overdue (was due {e["end_date"]}) (event #{e["id"]}) overdue', link)
                        )

        # 4. Missed start: Pending phase past start_date
        if e.get('start_date') and e['status'] == 'Pending':
            try:
                start_dt = datetime.strptime(e['start_date'], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                start_dt = None
            if start_dt and today > start_dt:
                notify_users = set(owner_ids)
                if assigned:
                    notify_users.add(assigned)
                for uid in notify_users:
                    if not _already_notified(uid, e['id'], 'missed_start'):
                        conn.execute(
                            'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
                            (uid, 'schedule', 'Schedule: Missed Start',
                             f'{e["phase_name"]} on {job_name} was scheduled to start {e["start_date"]} but hasn\'t begun (event #{e["id"]}) missed_start', link)
                        )

@app.route('/api/schedule/events', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_schedule_create():
    data = request.get_json(force=True)
    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO job_schedule_events (job_id, phase_name, description, start_date, end_date,
           assigned_to, sort_order, created_by, depends_on, estimated_hours, crew_size)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
        (data.get('job_id'), data.get('phase_name', ''), data.get('description', ''),
         data.get('start_date', ''), data.get('end_date', ''),
         data.get('assigned_to') or None, data.get('sort_order', 0), session['user_id'],
         data.get('depends_on') or None, data.get('estimated_hours', 0), data.get('crew_size', 1))
    )
    event_id = cursor.lastrowid
    conn.commit()
    # Notify assigned user
    if data.get('assigned_to'):
        job = conn.execute('SELECT name FROM jobs WHERE id = ?', (data['job_id'],)).fetchone()
        job_name = job['name'] if job else f'Job #{data["job_id"]}'
        start = data.get('start_date', '')
        end = data.get('end_date', '')
        date_range = f'{start} to {end}' if start and end else start or end or 'TBD'
        create_notification(
            int(data['assigned_to']), 'schedule', 'Schedule: Phase Assigned',
            f'{data["phase_name"]} on {job_name} — {date_range}',
            f'/schedule/job/{data["job_id"]}'
        )
    conn.close()
    return jsonify({'ok': True, 'id': event_id})

@app.route('/api/schedule/events/bulk', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_schedule_create_bulk():
    """Create multiple schedule phases in one request."""
    import time as _time
    data = request.get_json(force=True)
    job_id = data.get('job_id')
    phases = data.get('phases', [])
    if not job_id or not phases:
        return jsonify({'error': 'job_id and phases required'}), 400
    # Retry loop to handle SQLite lock contention
    for attempt in range(5):
        try:
            conn = get_db()
            created = 0
            for p in phases:
                conn.execute(
                    '''INSERT INTO job_schedule_events (job_id, phase_name, description, start_date, end_date,
                       assigned_to, sort_order, created_by, estimated_hours, crew_size)
                       VALUES (?,?,?,?,?,?,?,?,?,?)''',
                    (job_id, p.get('phase_name', ''), p.get('description', ''),
                     p.get('start_date', ''), p.get('end_date', ''),
                     p.get('assigned_to') or None, p.get('sort_order', 0), session['user_id'],
                     p.get('estimated_hours', 0), p.get('crew_size', 1))
                )
                created += 1
            conn.commit()
            conn.close()
            return jsonify({'ok': True, 'created': created})
        except Exception as exc:
            try:
                conn.close()
            except Exception:
                pass
            if 'locked' in str(exc) and attempt < 4:
                _time.sleep(0.5)
                continue
            return jsonify({'error': str(exc)}), 500
    return jsonify({'error': 'Database busy, please try again'}), 500

@app.route('/api/schedule/events/<int:eid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_schedule_update(eid):
    data = request.get_json(force=True)
    conn = get_db()
    old = conn.execute('SELECT * FROM job_schedule_events WHERE id = ?', (eid,)).fetchone()
    if not old:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    phase_name = data.get('phase_name', old['phase_name'])
    description = data.get('description', old['description'])
    start_date = data.get('start_date', old['start_date'])
    end_date = data.get('end_date', old['end_date'])
    assigned_to = data.get('assigned_to', old['assigned_to']) or None
    status = data.get('status', old['status'])
    sort_order = data.get('sort_order', old['sort_order'])
    depends_on = data.get('depends_on', old['depends_on']) or None
    estimated_hours = data.get('estimated_hours', old['estimated_hours'] or 0)
    crew_size = data.get('crew_size', old['crew_size'] or 1)
    pct_complete = data.get('pct_complete', old['pct_complete'] if old['pct_complete'] is not None else 0)

    conn.execute(
        '''UPDATE job_schedule_events SET phase_name=?, description=?, start_date=?, end_date=?,
           assigned_to=?, status=?, sort_order=?, depends_on=?, estimated_hours=?, crew_size=?,
           pct_complete=?, updated_at=datetime('now','localtime')
           WHERE id=?''',
        (phase_name, description, start_date, end_date, assigned_to, status, sort_order,
         depends_on, estimated_hours, crew_size, pct_complete, eid)
    )
    conn.commit()

    # Notify if assigned_to changed
    new_assigned = int(assigned_to) if assigned_to else None
    old_assigned = old['assigned_to']
    if new_assigned and new_assigned != old_assigned:
        job = conn.execute('SELECT name FROM jobs WHERE id = ?', (old['job_id'],)).fetchone()
        job_name = job['name'] if job else f'Job #{old["job_id"]}'
        date_range = f'{start_date} to {end_date}' if start_date and end_date else start_date or end_date or 'TBD'
        create_notification(
            new_assigned, 'schedule', 'Schedule: Phase Assigned',
            f'{phase_name} on {job_name} — {date_range}',
            f'/schedule/job/{old["job_id"]}'
        )
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/schedule/events/reorder', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_schedule_reorder():
    data = request.get_json(force=True)
    order = data.get('order', [])  # list of {id, sort_order}
    if not order:
        return jsonify({'error': 'No order provided'}), 400
    conn = get_db()
    for item in order:
        conn.execute('UPDATE job_schedule_events SET sort_order=?, updated_at=datetime(\'now\',\'localtime\') WHERE id=?',
                     (item['sort_order'], item['id']))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/schedule/events/<int:eid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_schedule_delete(eid):
    conn = get_db()
    conn.execute('DELETE FROM job_schedule_events WHERE id = ?', (eid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/jobs/<int:job_id>/bid-labor')
@api_role_required('owner', 'admin', 'project_manager')
def api_bid_labor(job_id):
    """Get bid labor data for a job — systems, man-hours, per-phase breakdown."""
    conn = get_db()
    bid = conn.execute(
        "SELECT * FROM bids WHERE job_id = ? AND status IN ('Draft','Sent','Accepted') ORDER BY id DESC LIMIT 1",
        (job_id,)
    ).fetchone()
    if not bid:
        conn.close()
        return jsonify({'found': False, 'total_systems': 0, 'man_hours_per_system': 20, 'phases': {}})

    bid = dict(bid)
    PHASE_HOUR_MAP = {
        'Rough-In': 'rough_in_hours',
        'AHU Install': 'ahu_install_hours',
        'Condenser Install': 'condenser_install_hours',
        'Trim-Out': 'trim_out_hours',
        'Startup': 'startup_hours',
    }
    PHASE_DEFAULTS = {
        'rough_in_hours': 15, 'ahu_install_hours': 1,
        'condenser_install_hours': 1, 'trim_out_hours': 1, 'startup_hours': 2,
    }
    total_systems = bid.get('total_systems') or 0
    man_hours_per_system = bid.get('man_hours_per_system') or 20
    phases = {}
    for phase_name, col in PHASE_HOUR_MAP.items():
        hrs_per_sys = bid.get(col) or PHASE_DEFAULTS.get(col, 0)
        phases[phase_name] = {
            'hours_per_system': hrs_per_sys,
            'total_hours': round(total_systems * hrs_per_sys, 1),
        }
    conn.close()
    return jsonify({
        'found': True,
        'bid_id': bid['id'],
        'bid_name': bid.get('bid_name', ''),
        'total_systems': total_systems,
        'man_hours_per_system': man_hours_per_system,
        'total_man_hours': bid.get('total_man_hours') or round(total_systems * man_hours_per_system, 1),
        'phases': phases,
    })


@app.route('/api/weather/forecast')
@api_login_required
def api_weather_forecast():
    """Proxy to Open-Meteo for weather forecast / historical data with delay risk flags."""
    import urllib.request, urllib.error
    lat = request.args.get('lat', type=float)
    lng = request.args.get('lng', type=float)
    start = request.args.get('start')
    end = request.args.get('end')
    if not lat or not lng or not start or not end:
        return jsonify({'error': 'lat, lng, start, end required'}), 400

    start_dt = datetime.strptime(start, '%Y-%m-%d').date()
    end_dt = datetime.strptime(end, '%Y-%m-%d').date()
    today = datetime.now().date()
    forecast_limit = today + timedelta(days=16)

    days_result = []

    # Split into forecast range and historical-estimate range
    if start_dt <= forecast_limit:
        fc_end = min(end_dt, forecast_limit)
        fc_url = (
            f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lng}"
            f"&daily=precipitation_probability_max,precipitation_sum,temperature_2m_min,"
            f"temperature_2m_max,wind_gusts_10m_max,weather_code"
            f"&temperature_unit=fahrenheit&wind_speed_unit=mph&precipitation_unit=inch"
            f"&timezone=America/Chicago&start_date={start_dt.strftime('%Y-%m-%d')}&end_date={fc_end.strftime('%Y-%m-%d')}"
        )
        try:
            req = urllib.request.Request(fc_url, headers={'User-Agent': 'JobTracker/1.0'})
            with urllib.request.urlopen(req, timeout=10) as resp:
                fc_data = json.loads(resp.read().decode())
            daily = fc_data.get('daily', {})
            dates = daily.get('time', [])
            for i, d in enumerate(dates):
                precip_prob = (daily.get('precipitation_probability_max') or [])[i] if i < len(daily.get('precipitation_probability_max') or []) else 0
                precip_sum = (daily.get('precipitation_sum') or [])[i] if i < len(daily.get('precipitation_sum') or []) else 0
                temp_min = (daily.get('temperature_2m_min') or [])[i] if i < len(daily.get('temperature_2m_min') or []) else 50
                temp_max = (daily.get('temperature_2m_max') or [])[i] if i < len(daily.get('temperature_2m_max') or []) else 80
                wind_gust = (daily.get('wind_gusts_10m_max') or [])[i] if i < len(daily.get('wind_gusts_10m_max') or []) else 0
                wcode = (daily.get('weather_code') or [])[i] if i < len(daily.get('weather_code') or []) else 0

                risk_reasons = []
                if precip_prob and precip_prob > 60:
                    risk_reasons.append('rain likely')
                if precip_sum and precip_sum > 0.25:
                    risk_reasons.append('rain')
                if temp_min is not None and temp_min < 25:
                    risk_reasons.append('freeze')
                if temp_max is not None and temp_max > 105:
                    risk_reasons.append('extreme heat')
                if wind_gust and wind_gust > 35:
                    risk_reasons.append('high wind')

                days_result.append({
                    'date': d, 'high': temp_max, 'low': temp_min,
                    'precip_prob': precip_prob or 0, 'precip': precip_sum or 0,
                    'wind_gust': wind_gust or 0, 'weather_code': wcode or 0,
                    'delay_risk': len(risk_reasons) > 0,
                    'risk_reasons': risk_reasons, 'source': 'forecast',
                })
        except Exception:
            pass  # Forecast failed — we'll still try historical

    # For dates beyond 16-day forecast, pull prior year historical as estimate
    if end_dt > forecast_limit:
        hist_start = max(start_dt, forecast_limit + timedelta(days=1))
        # Map to prior year
        try:
            hist_start_ly = hist_start.replace(year=hist_start.year - 1)
            hist_end_ly = end_dt.replace(year=end_dt.year - 1)
        except ValueError:
            hist_start_ly = hist_start.replace(year=hist_start.year - 1, day=28)
            hist_end_ly = end_dt.replace(year=end_dt.year - 1, day=28)

        hist_url = (
            f"https://archive-api.open-meteo.com/v1/archive?latitude={lat}&longitude={lng}"
            f"&daily=precipitation_sum,temperature_2m_min,temperature_2m_max,wind_gusts_10m_max,weather_code"
            f"&temperature_unit=fahrenheit&wind_speed_unit=mph&precipitation_unit=inch"
            f"&timezone=America/Chicago&start_date={hist_start_ly.strftime('%Y-%m-%d')}&end_date={hist_end_ly.strftime('%Y-%m-%d')}"
        )
        try:
            req = urllib.request.Request(hist_url, headers={'User-Agent': 'JobTracker/1.0'})
            with urllib.request.urlopen(req, timeout=10) as resp:
                hist_data = json.loads(resp.read().decode())
            daily = hist_data.get('daily', {})
            dates = daily.get('time', [])
            # Map historical dates back to current year
            day_offset = 0
            current_date = hist_start
            for i, d in enumerate(dates):
                if current_date > end_dt:
                    break
                precip_sum = (daily.get('precipitation_sum') or [])[i] if i < len(daily.get('precipitation_sum') or []) else 0
                temp_min = (daily.get('temperature_2m_min') or [])[i] if i < len(daily.get('temperature_2m_min') or []) else 50
                temp_max = (daily.get('temperature_2m_max') or [])[i] if i < len(daily.get('temperature_2m_max') or []) else 80
                wind_gust = (daily.get('wind_gusts_10m_max') or [])[i] if i < len(daily.get('wind_gusts_10m_max') or []) else 0
                wcode = (daily.get('weather_code') or [])[i] if i < len(daily.get('weather_code') or []) else 0

                risk_reasons = []
                if precip_sum and precip_sum > 0.25:
                    risk_reasons.append('rain (historical)')
                if temp_min is not None and temp_min < 25:
                    risk_reasons.append('freeze (historical)')
                if temp_max is not None and temp_max > 105:
                    risk_reasons.append('extreme heat (historical)')
                if wind_gust and wind_gust > 35:
                    risk_reasons.append('high wind (historical)')

                days_result.append({
                    'date': current_date.strftime('%Y-%m-%d'), 'high': temp_max, 'low': temp_min,
                    'precip_prob': None, 'precip': precip_sum or 0,
                    'wind_gust': wind_gust or 0, 'weather_code': wcode or 0,
                    'delay_risk': len(risk_reasons) > 0,
                    'risk_reasons': risk_reasons, 'source': 'historical_estimate',
                })
                current_date += timedelta(days=1)
        except Exception:
            pass

    return jsonify({'days': days_result})


@app.route('/api/geocode')
@api_login_required
def api_geocode():
    """Geocode a location string using Open-Meteo geocoding API."""
    import urllib.request, urllib.error, urllib.parse
    q = request.args.get('q', '')
    if not q:
        return jsonify({'lat': 35.6528, 'lng': -97.4781})  # Default: Edmond, OK
    try:
        encoded_q = urllib.parse.quote(q)
        url = f"https://geocoding-api.open-meteo.com/v1/search?name={encoded_q}&count=1&language=en&format=json"
        req = urllib.request.Request(url, headers={'User-Agent': 'JobTracker/1.0'})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode())
        results = data.get('results', [])
        if results:
            return jsonify({'lat': results[0]['latitude'], 'lng': results[0]['longitude']})
    except Exception:
        pass
    return jsonify({'lat': 35.6528, 'lng': -97.4781})


@app.route('/api/schedule/backwards-plan', methods=['POST'])
@api_login_required
def api_schedule_backwards_plan():
    """Smart backwards planning: bid-driven, weather-aware, crew calculator."""
    import math, urllib.request, urllib.error, urllib.parse
    data = request.get_json(force=True)
    job_id = data.get('job_id')
    deadline_date = data.get('deadline_date')
    hours_per_day = data.get('hours_per_day', 10)
    crew_override = data.get('crew_override')  # None = auto, int = fixed crew
    project_type = data.get('project_type', 'apartment')  # apartment | commercial
    commercial = data.get('commercial', {})
    apartment_details = data.get('apartment_details', {})  # { units, buildings }

    if not job_id or not deadline_date:
        return jsonify({'error': 'job_id and deadline_date required'}), 400

    conn = get_db()

    # 1. Fetch phases
    rows = conn.execute(
        'SELECT * FROM job_schedule_events WHERE job_id = ? ORDER BY sort_order, id',
        (job_id,)
    ).fetchall()
    phases = [dict(r) for r in rows]
    if not phases:
        conn.close()
        return jsonify({'error': 'No phases found'}), 404

    # 2. Fetch bid labor data
    bid = conn.execute(
        "SELECT * FROM bids WHERE job_id = ? AND status IN ('Draft','Sent','Accepted') ORDER BY id DESC LIMIT 1",
        (job_id,)
    ).fetchone()
    bid = dict(bid) if bid else None

    # ── Commercial project hour calculations ──
    # Industry-standard labor rates for commercial HVAC
    if project_type == 'commercial' and commercial:
        ptacs = commercial.get('ptac_count', 0)
        vtacs = commercial.get('vtac_count', 0)
        splits = commercial.get('split_count', 0)
        rtus = commercial.get('rtu_count', 0)
        metal_lf = commercial.get('metal_duct_lf', 0)
        comm_stories = commercial.get('stories', 4)
        exhaust_fans = commercial.get('exhaust_fan_count', 0)
        mau_count = commercial.get('makeup_air_count', 0)
        sleeve_install = commercial.get('sleeve_install', True)

        # PTAC labor hours:
        #   Sleeve install (if we do it): ~1.5-2 hrs per unit (cut, frame, sleeve, seal, flash)
        #   PTAC set & connect: ~0.75 hr per unit (slide in, electrical, condensate, test)
        #   Pigtail/electrical: ~0.25 hr per unit
        ptac_sleeve_hrs = ptacs * 1.75 if sleeve_install else 0
        ptac_set_hrs = ptacs * 1.0  # set, connect, test

        # VTAC labor hours:
        #   Similar to PTAC but vertical, closet-mounted, slightly more complex
        #   Install: ~2.5 hrs per unit (frame closet penetrations, set, connect, duct)
        vtac_hrs = vtacs * 2.5

        # Split system labor hours (common areas - lobby, restaurant, etc.):
        #   Full duct system per split: rough-in ~12-16 hrs, AHU set ~2 hrs,
        #   condenser ~3 hrs, trim ~2 hrs, startup ~1.5 hrs
        split_rough_hrs = splits * 14
        split_ahu_hrs = splits * 2
        split_cond_hrs = splits * 3
        split_trim_hrs = splits * 2
        split_startup_hrs = splits * 1.5

        # RTU labor hours:
        #   Crane/set: ~4 hrs per unit
        #   Curb adapter & duct connect: ~6-8 hrs per unit
        #   Electrical (disconnect, whip, controls): ~4 hrs per unit
        #   Startup & commissioning: ~3 hrs per unit
        rtu_set_hrs = rtus * 4
        rtu_duct_hrs = rtus * 7
        rtu_elec_hrs = rtus * 4
        rtu_startup_hrs = rtus * 3

        # Metal ductwork:
        #   Fabrication & install rate: ~15-20 LF per man-day (8-10 hrs)
        #   Includes hangers, sealing, insulation
        #   ~0.5 man-hours per linear foot installed
        metal_duct_hrs = metal_lf * 0.5

        # Exhaust fans: ~3 hrs per fan (set, duct, wire, test)
        exhaust_hrs = exhaust_fans * 3

        # Make-up air units: ~8 hrs per MAU (set, duct, wire, controls, balance)
        mau_hrs = mau_count * 8

        # Map to phases using commercial logic:
        # Rough-In: sleeves + metal duct + split rough-in + RTU curb/duct prep
        # Equipment Set: PTAC set + VTAC set + AHU set + RTU crane/set + exhaust + MAU
        # Condenser/Outdoor: split condensers + RTU startup prep
        # Trim-Out: all trim, controls, registers + split trim
        # Startup: all startup + commissioning + test & balance
        COMMERCIAL_PHASE_HOURS = {
            'Rough-In': ptac_sleeve_hrs + metal_duct_hrs + split_rough_hrs + rtu_duct_hrs + vtac_hrs * 0.4,
            'AHU Install': ptac_set_hrs + vtac_hrs * 0.6 + split_ahu_hrs + rtu_set_hrs + exhaust_hrs + mau_hrs,
            'Condenser Install': split_cond_hrs + rtu_elec_hrs,
            'Trim-Out': split_trim_hrs + (ptacs + vtacs) * 0.25,  # minimal trim for PTACs
            'Startup': split_startup_hrs + rtu_startup_hrs + (ptacs + vtacs) * 0.15,  # quick PTAC test
        }

        total_systems = ptacs + vtacs + splits + rtus

    PHASE_HOUR_MAP = {
        'Rough-In': 'rough_in_hours',
        'AHU Install': 'ahu_install_hours',
        'Condenser Install': 'condenser_install_hours',
        'Trim-Out': 'trim_out_hours',
        'Startup': 'startup_hours',
    }
    PHASE_DEFAULTS = {
        'rough_in_hours': 15, 'ahu_install_hours': 1,
        'condenser_install_hours': 1, 'trim_out_hours': 1, 'startup_hours': 2,
    }
    if project_type != 'commercial':
        total_systems = bid.get('total_systems', 0) if bid else 0
        # Allow apartment_details.units to override when provided
        if apartment_details.get('units'):
            total_systems = apartment_details['units']

    # 3. Calculate remaining hours per phase using bid data + pct_complete
    for p in phases:
        pct = p.get('pct_complete') or 0

        if project_type == 'commercial' and commercial:
            # Use commercial calculated hours
            total_hrs = COMMERCIAL_PHASE_HOURS.get(p['phase_name'], 0)
            if total_hrs == 0:
                total_hrs = p.get('estimated_hours') or 0
        else:
            bid_col = PHASE_HOUR_MAP.get(p['phase_name'])
            if bid and total_systems and bid_col:
                hrs_per_sys = bid.get(bid_col) or PHASE_DEFAULTS.get(bid_col, 0)
                total_hrs = total_systems * hrs_per_sys
            else:
                total_hrs = p.get('estimated_hours') or 0
        p['_total_hours'] = round(total_hrs, 1)
        p['_remaining_hours'] = round(total_hrs * (1 - pct / 100), 1)

    # 4. Fetch job location for weather
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    lat, lng = 35.6528, -97.4781  # Default Edmond, OK
    if job:
        loc_str = ' '.join(filter(None, [job['city'] if job['city'] else None, job['state'] if job['state'] else None, job['zip_code'] if job['zip_code'] else None]))
        if loc_str.strip():
            try:
                encoded_q = urllib.parse.quote(loc_str.strip())
                geo_url = f"https://geocoding-api.open-meteo.com/v1/search?name={encoded_q}&count=1&language=en&format=json"
                req = urllib.request.Request(geo_url, headers={'User-Agent': 'JobTracker/1.0'})
                with urllib.request.urlopen(req, timeout=5) as resp:
                    geo_data = json.loads(resp.read().decode())
                results = geo_data.get('results', [])
                if results:
                    lat, lng = results[0]['latitude'], results[0]['longitude']
            except Exception:
                pass

    # 5. Fetch weather data for the planning window
    deadline = datetime.strptime(deadline_date, '%Y-%m-%d').date()
    today = datetime.now().date()
    plan_start = today
    forecast_limit = today + timedelta(days=16)
    weather_days = {}  # date_str -> {delay_risk, risk_reasons, ...}

    # Forecast portion
    if plan_start <= forecast_limit:
        fc_end = min(deadline, forecast_limit)
        fc_url = (
            f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lng}"
            f"&daily=precipitation_probability_max,precipitation_sum,temperature_2m_min,"
            f"temperature_2m_max,wind_gusts_10m_max,weather_code"
            f"&temperature_unit=fahrenheit&wind_speed_unit=mph&precipitation_unit=inch"
            f"&timezone=America/Chicago&start_date={plan_start.strftime('%Y-%m-%d')}&end_date={fc_end.strftime('%Y-%m-%d')}"
        )
        try:
            req = urllib.request.Request(fc_url, headers={'User-Agent': 'JobTracker/1.0'})
            with urllib.request.urlopen(req, timeout=10) as resp:
                fc_data = json.loads(resp.read().decode())
            daily = fc_data.get('daily', {})
            dates = daily.get('time', [])
            for i, d in enumerate(dates):
                pp = (daily.get('precipitation_probability_max') or [])[i] if i < len(daily.get('precipitation_probability_max') or []) else 0
                ps = (daily.get('precipitation_sum') or [])[i] if i < len(daily.get('precipitation_sum') or []) else 0
                tmin = (daily.get('temperature_2m_min') or [])[i] if i < len(daily.get('temperature_2m_min') or []) else 50
                tmax = (daily.get('temperature_2m_max') or [])[i] if i < len(daily.get('temperature_2m_max') or []) else 80
                wg = (daily.get('wind_gusts_10m_max') or [])[i] if i < len(daily.get('wind_gusts_10m_max') or []) else 0
                wc = (daily.get('weather_code') or [])[i] if i < len(daily.get('weather_code') or []) else 0
                reasons = []
                if pp and pp > 60: reasons.append('rain likely')
                if ps and ps > 0.25: reasons.append('rain')
                if tmin is not None and tmin < 25: reasons.append('freeze')
                if tmax is not None and tmax > 105: reasons.append('extreme heat')
                if wg and wg > 35: reasons.append('high wind')
                weather_days[d] = {
                    'high': tmax, 'low': tmin, 'precip_prob': pp or 0, 'precip': ps or 0,
                    'wind_gust': wg or 0, 'weather_code': wc or 0,
                    'delay_risk': len(reasons) > 0, 'risk_reasons': reasons, 'source': 'forecast',
                }
        except Exception:
            pass

    # Historical portion for dates beyond forecast
    if deadline > forecast_limit:
        hist_start = max(plan_start, forecast_limit + timedelta(days=1))
        try:
            hist_start_ly = hist_start.replace(year=hist_start.year - 1)
            hist_end_ly = deadline.replace(year=deadline.year - 1)
        except ValueError:
            hist_start_ly = hist_start.replace(year=hist_start.year - 1, day=28)
            hist_end_ly = deadline.replace(year=deadline.year - 1, day=28)
        hist_url = (
            f"https://archive-api.open-meteo.com/v1/archive?latitude={lat}&longitude={lng}"
            f"&daily=precipitation_sum,temperature_2m_min,temperature_2m_max,wind_gusts_10m_max,weather_code"
            f"&temperature_unit=fahrenheit&wind_speed_unit=mph&precipitation_unit=inch"
            f"&timezone=America/Chicago&start_date={hist_start_ly.strftime('%Y-%m-%d')}&end_date={hist_end_ly.strftime('%Y-%m-%d')}"
        )
        try:
            req = urllib.request.Request(hist_url, headers={'User-Agent': 'JobTracker/1.0'})
            with urllib.request.urlopen(req, timeout=10) as resp:
                hist_data = json.loads(resp.read().decode())
            daily = hist_data.get('daily', {})
            dates = daily.get('time', [])
            current_date = hist_start
            for i, d in enumerate(dates):
                if current_date > deadline: break
                ps = (daily.get('precipitation_sum') or [])[i] if i < len(daily.get('precipitation_sum') or []) else 0
                tmin = (daily.get('temperature_2m_min') or [])[i] if i < len(daily.get('temperature_2m_min') or []) else 50
                tmax = (daily.get('temperature_2m_max') or [])[i] if i < len(daily.get('temperature_2m_max') or []) else 80
                wg = (daily.get('wind_gusts_10m_max') or [])[i] if i < len(daily.get('wind_gusts_10m_max') or []) else 0
                wc = (daily.get('weather_code') or [])[i] if i < len(daily.get('weather_code') or []) else 0
                reasons = []
                if ps and ps > 0.25: reasons.append('rain (historical)')
                if tmin is not None and tmin < 25: reasons.append('freeze (historical)')
                if tmax is not None and tmax > 105: reasons.append('extreme heat (historical)')
                if wg and wg > 35: reasons.append('high wind (historical)')
                ds = current_date.strftime('%Y-%m-%d')
                weather_days[ds] = {
                    'high': tmax, 'low': tmin, 'precip_prob': None, 'precip': ps or 0,
                    'wind_gust': wg or 0, 'weather_code': wc or 0,
                    'delay_risk': len(reasons) > 0, 'risk_reasons': reasons, 'source': 'historical_estimate',
                }
                current_date += timedelta(days=1)
        except Exception:
            pass

    # 6. Build weather risk date set (weekdays only)
    weather_risk_dates = set()
    for ds, w in weather_days.items():
        if w['delay_risk']:
            d = datetime.strptime(ds, '%Y-%m-%d').date()
            if d.weekday() < 5:
                weather_risk_dates.add(d)

    # 7. Build dependency graph + topological sort (leaf-first for backwards planning)
    by_id = {p['id']: p for p in phases}
    successors = {p['id']: [] for p in phases}
    has_any_deps = any(p.get('depends_on') and p['depends_on'] in by_id for p in phases)

    if has_any_deps:
        # Use explicit dependencies
        for p in phases:
            if p.get('depends_on') and p['depends_on'] in by_id:
                successors[p['depends_on']].append(p['id'])
    else:
        # No explicit dependencies — auto-chain phases sequentially (sort order).
        # Each phase must finish before the next one starts.
        for i in range(len(phases) - 1):
            successors[phases[i]['id']].append(phases[i + 1]['id'])

    from collections import deque
    # For backwards planning: process phases with no successors first (they end at deadline),
    # then work backwards to phases that must finish before those start.
    # "Leaves" = phases with empty successor lists = end of the chain.
    successor_count = {p['id']: len(successors[p['id']]) for p in phases}
    queue = deque([pid for pid, cnt in successor_count.items() if cnt == 0])
    topo_order = []
    visited = set()
    while queue:
        pid = queue.popleft()
        if pid in visited:
            continue
        visited.add(pid)
        topo_order.append(pid)
        # Find predecessors: phases whose successor list includes this pid
        for p in phases:
            if pid in successors.get(p['id'], []):
                successor_count[p['id']] -= 1
                if successor_count[p['id']] == 0:
                    queue.append(p['id'])
    # If any phases weren't reached (orphans), add them
    for p in phases:
        if p['id'] not in visited:
            topo_order.append(p['id'])

    # 8. Compute dates backwards from deadline
    def count_work_days(start_d, end_d, skip_weather=True):
        """Count available work days between two dates (inclusive), skipping weekends and weather risk days."""
        count = 0
        d = start_d
        while d <= end_d:
            if d.weekday() < 5 and (not skip_weather or d not in weather_risk_dates):
                count += 1
            d += timedelta(days=1)
        return count

    def walk_back(end_d, work_days_needed):
        """Walk backwards from end_d to find start_d, skipping weekends and weather risk days."""
        current = end_d
        remaining = work_days_needed
        # First ensure end_d itself is a valid work day
        while current.weekday() >= 5 or current in weather_risk_dates:
            current -= timedelta(days=1)
        end_d = current
        remaining -= 1  # end_d counts as 1 work day
        while remaining > 0:
            current -= timedelta(days=1)
            if current.weekday() < 5 and current not in weather_risk_dates:
                remaining -= 1
        return current, end_d

    # ── Auto mode: pre-calculate a consistent crew size that fits all phases sequentially ──
    if not crew_override:
        total_active_hrs = sum(p['_remaining_hours'] for p in phases if p['_remaining_hours'] > 0)
        total_avail = count_work_days(today, deadline)
        if total_avail > 0 and total_active_hrs > 0:
            auto_crew = max(1, math.ceil(total_active_hrs / (total_avail * hours_per_day)))
            # Verify: total days at this crew must fit in available window
            total_days_at_crew = sum(
                max(1, math.ceil(p['_remaining_hours'] / (auto_crew * hours_per_day)))
                for p in phases if p['_remaining_hours'] > 0
            )
            while total_days_at_crew > total_avail and auto_crew < 50:
                auto_crew += 1
                total_days_at_crew = sum(
                    max(1, math.ceil(p['_remaining_hours'] / (auto_crew * hours_per_day)))
                    for p in phases if p['_remaining_hours'] > 0
                )
        else:
            auto_crew = max(1, math.ceil(total_active_hrs / hours_per_day)) if total_active_hrs > 0 else 1
        effective_crew = auto_crew
    else:
        effective_crew = crew_override

    computed_end = {}
    computed_start = {}
    phase_results = {}

    for pid in topo_order:
        p = by_id[pid]
        remaining_hrs = p['_remaining_hours']

        # Determine end date boundary
        succ_ids = successors[pid]
        if not succ_ids:
            boundary_end = deadline
        else:
            valid_succ_starts = [computed_start[s] for s in succ_ids if s in computed_start]
            if valid_succ_starts:
                boundary_end = min(valid_succ_starts) - timedelta(days=1)
                while boundary_end.weekday() >= 5:
                    boundary_end -= timedelta(days=1)
            else:
                boundary_end = deadline

        if remaining_hrs <= 0:
            # Phase already done — don't consume any calendar time.
            # Set start=end=boundary so the next phase upstream can use boundary_end directly.
            computed_start[pid] = boundary_end
            computed_end[pid] = boundary_end
            phase_results[pid] = {
                'work_days': 0, 'crew_needed': 0, 'remaining_hours': 0,
                'total_hours': p['_total_hours'],
                'weather_risk_days': 0, 'hours_per_day_needed': 0,
                'warning': 'Complete',
            }
            continue

        # Count available work days from today to boundary
        avail_work_days = count_work_days(today, boundary_end)

        # Calculate days needed with the effective crew
        work_days_needed = max(1, math.ceil(remaining_hrs / (effective_crew * hours_per_day)))
        crew_needed = effective_crew
        hrs_per_day_needed = hours_per_day
        warning = None
        if work_days_needed > avail_work_days:
            # Can't fit — calculate extended hours per day needed
            if avail_work_days > 0:
                hrs_per_day_needed = round(remaining_hrs / (effective_crew * avail_work_days), 1)
                work_days_needed = avail_work_days
                if hrs_per_day_needed > 14:
                    warning = f'Requires {hrs_per_day_needed}hr days — consider more crew or extending deadline'
                else:
                    warning = f'Extended to {hrs_per_day_needed}hr days to meet deadline'
            else:
                warning = 'No available work days before deadline'
                work_days_needed = 1

        start_d, end_d = walk_back(boundary_end, work_days_needed)
        computed_start[pid] = start_d
        computed_end[pid] = end_d

        # Count weather risk days in this phase's window
        phase_weather_risks = 0
        d = start_d
        while d <= end_d:
            if d.weekday() < 5 and d in weather_risk_dates:
                phase_weather_risks += 1
            d += timedelta(days=1)

        phase_results[pid] = {
            'work_days': work_days_needed,
            'crew_needed': crew_needed,
            'remaining_hours': remaining_hrs,
            'total_hours': p['_total_hours'],
            'weather_risk_days': phase_weather_risks,
            'hours_per_day_needed': hrs_per_day_needed,
            'warning': warning,
        }

    # 9. Update DB — only update estimated_hours and crew_size, NOT dates
    # Dates are calculated for display only. User manually sets/applies dates.
    apply_dates = data.get('apply_dates', False)  # only overwrite dates if explicitly requested
    for p in phases:
        pid = p['id']
        if pid in computed_start:
            pr = phase_results.get(pid, {})
            est_hrs = p['_total_hours'] if p['_total_hours'] > 0 else (p.get('estimated_hours') or 0)
            crew = pr.get('crew_needed', p.get('crew_size') or 1)
            if apply_dates:
                conn.execute(
                    '''UPDATE job_schedule_events SET start_date=?, end_date=?, estimated_hours=?, crew_size=?,
                       updated_at=datetime('now','localtime') WHERE id=?''',
                    (computed_start[pid].strftime('%Y-%m-%d'), computed_end[pid].strftime('%Y-%m-%d'),
                     est_hrs, crew, pid)
                )
            else:
                conn.execute(
                    '''UPDATE job_schedule_events SET estimated_hours=?, crew_size=?,
                       updated_at=datetime('now','localtime') WHERE id=?''',
                    (est_hrs, crew, pid)
                )
    conn.commit()

    # 10. Build response — use computed dates for display, actual DB dates for phases
    rows = conn.execute(
        'SELECT * FROM job_schedule_events WHERE job_id = ? ORDER BY sort_order, id',
        (job_id,)
    ).fetchall()
    updated_phases = [dict(r) for r in rows]
    conn.close()

    # Summary stats
    total_remaining = sum(pr.get('remaining_hours', 0) for pr in phase_results.values())
    calendar_days = (deadline - today).days
    total_weather_risk = len([d for d in weather_risk_dates if today <= d <= deadline])
    total_biz_days = count_work_days(today, deadline, skip_weather=False)
    total_avail_days = count_work_days(today, deadline, skip_weather=True)

    # Phase detail for results table — use computed dates for plan display
    phase_detail = []
    for p in updated_phases:
        pr = phase_results.get(p['id'], {})
        # Show computed dates for the plan, keep actual DB dates separate
        plan_start = computed_start[p['id']].strftime('%Y-%m-%d') if p['id'] in computed_start else p['start_date']
        plan_end = computed_end[p['id']].strftime('%Y-%m-%d') if p['id'] in computed_end else p['end_date']
        phase_detail.append({
            'id': p['id'],
            'phase_name': p['phase_name'],
            'start_date': plan_start,
            'end_date': plan_end,
            'actual_start_date': p['start_date'],  # what's in DB
            'actual_end_date': p['end_date'],       # what's in DB
            'work_days': pr.get('work_days', 0),
            'remaining_hours': pr.get('remaining_hours', 0),
            'total_hours': pr.get('total_hours', p.get('estimated_hours') or 0),
            'crew_needed': pr.get('crew_needed', p.get('crew_size') or 1),
            'weather_risk_days': pr.get('weather_risk_days', 0),
            'hours_per_day_needed': pr.get('hours_per_day_needed', hours_per_day),
            'pct_complete': p.get('pct_complete') or 0,
            'warning': pr.get('warning'),
            'status': p['status'],
        })

    # Weather days list for frontend
    weather_list = []
    for ds in sorted(weather_days.keys()):
        w = weather_days[ds]
        weather_list.append({'date': ds, **w})

    # Crew recommendation: max crew needed across active phases
    active_phases = [pr for pr in phase_results.values() if pr.get('remaining_hours', 0) > 0]
    max_crew = max((pr.get('crew_needed', 1) for pr in active_phases), default=1)

    # Find earliest start date across all active phases
    active_start_dates = [computed_start[pid] for pid in phase_results if phase_results[pid].get('remaining_hours', 0) > 0 and pid in computed_start]
    earliest_start = min(active_start_dates).strftime('%B %d, %Y') if active_start_dates else today.strftime('%B %d, %Y')
    # Find latest end date
    active_end_dates = [computed_end[pid] for pid in phase_results if phase_results[pid].get('remaining_hours', 0) > 0 and pid in computed_end]
    latest_end = max(active_end_dates).strftime('%B %d, %Y') if active_end_dates else deadline_date
    total_work_days = sum(pr.get('work_days', 0) for pr in active_phases)

    if project_type == 'commercial' and commercial:
        ptacs = commercial.get('ptac_count', 0)
        vtacs = commercial.get('vtac_count', 0)
        splits = commercial.get('split_count', 0)
        rtus = commercial.get('rtu_count', 0)
        room_units = ptacs + vtacs
        common_units = splits + rtus
        crew_recommendation = f"You need {max_crew} men working {hours_per_day}-hour days"
        if room_units > 0:
            crew_recommendation += f" — {room_units} PTAC/VTAC rooms + {common_units} common area systems"
    else:
        if crew_override:
            crew_recommendation = f"With {crew_override} men working {hours_per_day}-hour days — start {earliest_start}, finish by {latest_end} ({total_work_days} work days)"
        else:
            crew_recommendation = f"You need {max_crew} men working {hours_per_day}-hour days — start {earliest_start}, finish by {latest_end} ({total_work_days} work days)"

    # Override impact: compare auto vs override
    override_impact = None
    if crew_override and total_remaining > 0:
        # Calculate what auto mode would have given
        auto_crew = max(1, math.ceil(total_remaining / (total_avail_days * hours_per_day))) if total_avail_days > 0 else max(1, math.ceil(total_remaining / hours_per_day))
        auto_days = max(1, math.ceil(total_remaining / (auto_crew * hours_per_day)))
        override_days = sum(pr.get('work_days', 0) for pr in active_phases)
        # Check if any phase has extended hours
        max_hrs_per_day = max((pr.get('hours_per_day_needed', hours_per_day) for pr in active_phases), default=hours_per_day)
        override_impact = {
            'auto_crew': auto_crew,
            'auto_days': auto_days,
            'auto_hours_per_day': hours_per_day,
            'override_crew': crew_override,
            'override_days': override_days,
            'override_hours_per_day': max_hrs_per_day,
            'days_delta': override_days - auto_days,
            'hours_delta': max_hrs_per_day - hours_per_day,
        }

    # ── 11. Generate Detailed Weekly/Daily Benchmarks ──
    # Phase-specific task breakdowns for HVAC construction
    PHASE_TASKS = {
        'Rough-In': [
            'Layout & mark unit locations, chase openings, penetrations',
            'Frame & cut chases for supply/return plenums',
            'Hang main trunk lines and supply drops (flex duct)',
            'Install return air boots & return ductwork',
            'Run condensate drain piping (PVC/CPVC) to nearest drain',
            'Install fire dampers at fire-rated assemblies',
            'Install duct hangers, strapping, threaded rod supports',
            'Pull thermostat wire & low-voltage wire to each unit location',
            'Install exhaust fan rough-in (vent ducting to exterior)',
            'Seal all duct connections with mastic & tape per code',
            'Install CRDs / bath exhaust boots per unit',
            'Pass rough-in inspection before cover',
        ],
        'AHU Install': [
            'Stage air handlers at unit locations (attic/closet)',
            'Set AHU on mounting platform (hanger/shelf/stand)',
            'Connect supply & return plenums to AHU',
            'Connect condensate drain to AHU trap',
            'Install drain pan under unit (if above living space)',
            'Install Safe-T-Switch / float switch on condensate',
            'Wire thermostat wire to AHU terminal block',
            'Verify airflow direction and filter rack orientation',
        ],
        'Condenser Install': [
            'Set condensers/heat pumps on pads (ground) or curbs (roof)',
            'Braze line sets — suction & liquid lines per tonnage',
            'Insulate suction lines with Armaflex/fiberglass',
            'Install electrical disconnects at each condenser',
            'Pull whips from disconnect to condenser',
            'Pressure test with nitrogen (600psi min, hold 24hrs)',
            'Break vacuum and charge refrigerant per spec',
        ],
        'Trim-Out': [
            'Install supply registers (sized per drop diameter)',
            'Install return air grilles / filter grilles',
            'Mount and wire thermostats',
            'Install condensate pump-ups where required',
            'Install filter in each return air grille',
            'Install cork pads / vibration isolation',
            'Complete final condensate connections',
            'Touch-up duct sealing, register alignment',
            'Label all equipment, circuits, and thermostats',
            'Clean up work areas, remove debris',
        ],
        'Startup': [
            'Verify all electrical connections, tighten terminals',
            'Verify refrigerant charge (superheat/subcooling method)',
            'Measure airflow at each supply register (CFM)',
            'Test thermostat operation — heat & cool modes',
            'Verify condensate drain flow & safety switch operation',
            'Check duct leakage (blower door test if required)',
            'Document startup readings per unit (pressures, temps, CFM)',
            'Complete startup checklist for builder/owner',
            'Schedule & pass final mechanical inspection',
        ],
    }
    COMMERCIAL_PHASE_TASKS = {
        'Rough-In': [
            'Layout & mark PTAC/VTAC sleeve locations per plans (verify dimensions with architect)',
            'Cut, frame, and install PTAC sleeves — verify level, slope for condensate',
            'Flash and seal PTAC sleeves (interior & exterior weather barrier)',
            'Install metal ductwork trunk lines for common area systems',
            'Fabricate & install metal duct branches, transitions, and takeoffs',
            'Hang duct with trapeze hangers, threaded rod per SMACNA standards',
            'Install fire/smoke dampers at all fire-rated penetrations',
            'Run condensate drain piping for all systems to approved drains',
            'Pull thermostat/control wire for split systems and RTU controls',
            'Install exhaust duct risers for kitchen hoods and bath exhaust',
            'Seal all metal duct joints — mastic, tape, or gasket per spec',
            'Install VTAC closet rough-in — supply/return openings, condensate, electrical',
            'Pass rough-in inspection before walls close',
        ],
        'AHU Install': [
            'Set PTACs in sleeves — verify electrical, drain slope, secure to sleeve',
            'Connect PTAC pigtails and verify dedicated circuit',
            'Set VTACs in closets — connect supply/return, condensate, electrical',
            'Set split system air handlers in common areas (lobby, restaurant, office)',
            'Crane & set RTUs on rooftop curbs — verify curb alignment and gasket',
            'Connect RTU supply/return duct transitions to curb openings',
            'Install make-up air units (MAUs) — duct, electrical, controls',
            'Set and connect kitchen hood exhaust fans',
            'Install bath exhaust fans and connect to risers',
            'Wire all equipment disconnects and verify circuits',
        ],
        'Condenser Install': [
            'Set split system condensers on pads or brackets',
            'Braze line sets for split systems — nitrogen purge during brazing',
            'Insulate all suction lines and exposed line sets',
            'Install electrical disconnects and whips for condensers',
            'Wire RTU high-voltage connections and verify breaker sizing',
            'Pressure test all refrigerant systems (600psi nitrogen, 24hr hold)',
            'Pull vacuum on all refrigerant systems (500 microns min)',
            'Charge refrigerant per manufacturer specs — document charge amount',
        ],
        'Trim-Out': [
            'Install supply diffusers, registers, and grilles in common areas',
            'Install return air grilles with filters',
            'Mount and wire all thermostats and DDC controls',
            'Install condensate pump-ups where required',
            'Complete PTAC/VTAC final connections — thermostat, filter, cover',
            'Install access panels for all concealed valves and dampers',
            'Label all equipment, breakers, and control points',
            'Clean all supply and return openings — remove construction debris',
            'Touch-up sealant and insulation at all penetrations',
            'Install room temperature sensors for BMS (if applicable)',
        ],
        'Startup': [
            'Verify all electrical connections — torque check terminals',
            'Startup each PTAC/VTAC — verify heat, cool, fan modes',
            'Verify PTAC condensate drain flow (tilt test)',
            'Startup split systems — verify superheat/subcooling, airflow',
            'Startup RTUs — verify all stages, economizer operation',
            'Test and balance (TAB) all common area systems per spec',
            'Verify exhaust fan CFM at each kitchen hood and bath',
            'Document all startup readings — pressures, temps, CFM, amps',
            'Commission BMS/DDC controls (if applicable)',
            'Walk each room with GC — verify operation, labels, cleanliness',
            'Schedule & pass final mechanical inspection',
            'Provide O&M manuals, warranty cards, and as-built drawings',
        ],
    }

    DEFAULT_TASKS = [
        'Review scope and materials for this phase',
        'Execute daily work plan with crew',
        'Quality check completed work',
        'Document progress and update completion %',
        'Coordinate with GC on next-phase readiness',
    ]

    # Select task set based on project type
    active_phase_tasks = COMMERCIAL_PHASE_TASKS if project_type == 'commercial' else PHASE_TASKS

    benchmarks = []
    for pd in phase_detail:
        if pd['remaining_hours'] <= 0:
            continue  # phase complete, skip

        phase_name = pd['phase_name']
        start_str = pd['start_date']
        end_str = pd['end_date']
        if not start_str or not end_str:
            continue
        p_start = datetime.strptime(start_str, '%Y-%m-%d').date()
        p_end = datetime.strptime(end_str, '%Y-%m-%d').date()
        work_days = pd['work_days']
        crew = pd['crew_needed']
        remaining_hrs = pd['remaining_hours']
        total_hrs = pd['total_hours']
        hpd = pd['hours_per_day_needed']
        existing_pct = pd.get('pct_complete', 0) or 0  # current overall completion

        # Daily production rate
        hrs_per_day_total = crew * hpd  # total man-hours per day
        if total_systems > 0 and work_days > 0:
            units_per_day = round(total_systems / work_days, 1)
            units_per_day = max(0.5, units_per_day)
        else:
            units_per_day = 0

        # Phase-specific tasks
        tasks = active_phase_tasks.get(phase_name, DEFAULT_TASKS)

        # Build weekly breakdown
        weeks = []
        current_date = p_start
        week_num = 1
        while current_date <= p_end:
            # Find the week's Monday (or start date if mid-week)
            week_start = current_date
            # Week goes until Friday or end of phase
            week_end = week_start
            work_days_this_week = 0
            daily_plan = []
            while week_end <= p_end and (week_end.weekday() < 5 or week_end == week_start):
                if week_end.weekday() < 5:
                    is_weather_risk = week_end in weather_risk_dates
                    is_work_day = not is_weather_risk
                    if is_work_day:
                        work_days_this_week += 1
                    daily_plan.append({
                        'date': week_end.strftime('%Y-%m-%d'),
                        'day_name': week_end.strftime('%A'),
                        'is_work_day': is_work_day,
                        'weather_risk': is_weather_risk,
                        'weather_info': weather_days.get(week_end.strftime('%Y-%m-%d'), {}),
                    })
                # Move to next day but stop at Saturday
                if week_end.weekday() == 4:  # Friday
                    break
                week_end += timedelta(days=1)

            if not daily_plan:
                current_date = week_end + timedelta(days=1)
                # Skip weekends
                while current_date.weekday() >= 5 and current_date <= p_end:
                    current_date += timedelta(days=1)
                continue

            # Calculate cumulative progress at end of this week
            elapsed_work_days = 0
            d = p_start
            while d <= week_end:
                if d.weekday() < 5 and d not in weather_risk_dates:
                    elapsed_work_days += 1
                d += timedelta(days=1)
            pct_elapsed = round((elapsed_work_days / max(work_days, 1)) * 100, 0) if work_days > 0 else 100
            # Overall completion = existing progress + proportion of remaining work done
            cumulative_pct = min(100, round(existing_pct + pct_elapsed * (100 - existing_pct) / 100, 0))
            cumulative_units = round(total_systems * cumulative_pct / 100, 0) if total_systems > 0 else 0

            # Assign tasks to this week based on phase progression
            # Distribute phase tasks across weeks proportionally
            total_weeks_est = max(1, math.ceil(work_days / 5))
            task_start_idx = int((week_num - 1) / total_weeks_est * len(tasks))
            task_end_idx = min(len(tasks), int(week_num / total_weeks_est * len(tasks)) + 1)
            week_tasks = tasks[task_start_idx:task_end_idx] if task_start_idx < len(tasks) else [tasks[-1]]
            if not week_tasks:
                week_tasks = ['Continue phase work per daily targets']

            # Benchmark checkpoint
            if cumulative_pct <= 25:
                checkpoint_status = 'early'
            elif cumulative_pct <= 50:
                checkpoint_status = 'mid_early'
            elif cumulative_pct <= 75:
                checkpoint_status = 'mid_late'
            else:
                checkpoint_status = 'final'

            # Warning thresholds — never below existing progress
            behind_trigger_pct = max(existing_pct, cumulative_pct - 10)

            weeks.append({
                'week_num': week_num,
                'start_date': daily_plan[0]['date'],
                'end_date': daily_plan[-1]['date'],
                'work_days': work_days_this_week,
                'daily_plan': daily_plan,
                'target_pct_complete': cumulative_pct,
                'target_units_complete': cumulative_units,
                'target_hours_burned': round(remaining_hrs * pct_elapsed / 100, 1),
                'daily_target_units': units_per_day,
                'daily_target_hours': round(hrs_per_day_total, 1),
                'tasks': week_tasks,
                'checkpoint_status': checkpoint_status,
                'behind_trigger_pct': behind_trigger_pct,
                'behind_warning': f"If below {behind_trigger_pct:.0f}% by {daily_plan[-1]['date']}, you are falling behind. Add crew or extend hours immediately.",
            })

            week_num += 1
            current_date = week_end + timedelta(days=1)
            # Skip weekends
            while current_date.weekday() >= 5 and current_date <= p_end:
                current_date += timedelta(days=1)

        # Phase-level early warning signs (accounting for existing progress)
        early_warnings = []
        if existing_pct > 0:
            early_warnings.append(f"Phase is currently {existing_pct:.0f}% complete. Benchmarks below show targets for the remaining {100 - existing_pct:.0f}% of work.")
        if total_systems > 0:
            # Day 1 benchmark: by end of first day
            day1_target = units_per_day
            early_warnings.append(f"Day 1: Should complete {day1_target} units. If <{max(1, round(day1_target * 0.7))}, crew size or pace insufficient.")
        if work_days > 2:
            # End of day 2 benchmark — show overall %
            day2_remaining_pct = round(2 / max(work_days, 1) * 100, 0)
            day2_overall = round(existing_pct + day2_remaining_pct * (100 - existing_pct) / 100, 0)
            early_warnings.append(f"Day 2: Should be at ~{day2_overall}% overall. If crew hasn't found rhythm, address bottlenecks now.")
        if work_days >= 5:
            w1_remaining_pct = round(min(5, work_days) / max(work_days, 1) * 100, 0)
            w1_overall = round(existing_pct + w1_remaining_pct * (100 - existing_pct) / 100, 0)
            early_warnings.append(f"End of Week 1: Must be at {w1_overall}% overall to stay on track. This is your first major checkpoint.")
        early_warnings.append("Material shortage = immediate stop. Verify all materials on-site before phase start.")
        early_warnings.append("Inspection failures reset the clock. Pre-inspect your own work before calling inspector.")

        benchmarks.append({
            'phase_name': phase_name,
            'phase_id': pd['id'],
            'existing_pct_complete': existing_pct,
            'start_date': start_str,
            'end_date': end_str,
            'work_days': work_days,
            'crew_needed': crew,
            'remaining_hours': remaining_hrs,
            'total_hours': total_hrs,
            'hours_per_day': hpd,
            'units_per_day': units_per_day,
            'total_systems': total_systems,
            'weeks': weeks,
            'early_warnings': early_warnings,
            'tasks_overview': tasks,
            'key_milestone': f"Phase complete by {end_str}. All {phase_name.lower()} work finished, inspected, and ready for next phase.",
        })

    return jsonify({
        'phases': updated_phases,
        'plan': phase_detail,
        'summary': {
            'total_remaining_hours': round(total_remaining, 1),
            'calendar_days': calendar_days,
            'weather_risk_days': total_weather_risk,
            'business_days': total_biz_days,
            'available_work_days': total_avail_days,
            'deadline': deadline_date,
            'hours_per_day': hours_per_day,
            'crew_override': crew_override,
            'bid_found': bid is not None,
            'total_systems': total_systems,
            'project_type': project_type,
            'commercial': commercial if project_type == 'commercial' else None,
            'apartment_details': apartment_details if project_type == 'apartment' and apartment_details else None,
        },
        'crew_recommendation': crew_recommendation,
        'override_impact': override_impact,
        'weather': weather_list,
        'benchmarks': benchmarks,
    })


@app.route('/api/schedule/backwards-plan/save', methods=['POST'])
@api_login_required
def api_schedule_save_plan():
    """Save a backwards plan snapshot."""
    data = request.get_json(force=True)
    job_id = data.get('job_id')
    plan_name = data.get('plan_name', '').strip()
    if not job_id:
        return jsonify({'error': 'job_id required'}), 400
    if not plan_name:
        plan_name = f"Plan {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"

    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO schedule_plans (job_id, plan_name, deadline_date, hours_per_day, crew_override,
           plan_data, summary_data, weather_data, benchmarks_data, created_by)
           VALUES (?,?,?,?,?,?,?,?,?,?)''',
        (job_id, plan_name, data.get('deadline_date', ''), data.get('hours_per_day', 10),
         data.get('crew_override'), json.dumps(data.get('plan', [])),
         json.dumps(data.get('summary', {})), json.dumps(data.get('weather', [])),
         json.dumps(data.get('benchmarks', [])),
         session.get('user_id'))
    )
    conn.commit()
    plan_id = cursor.lastrowid
    conn.close()
    return jsonify({'ok': True, 'id': plan_id, 'plan_name': plan_name}), 201


@app.route('/api/schedule/plans')
@api_login_required
def api_schedule_list_plans():
    """List saved backwards plans for a job."""
    job_id = request.args.get('job_id')
    if not job_id:
        return jsonify({'error': 'job_id required'}), 400
    conn = get_db()
    rows = conn.execute(
        '''SELECT sp.*, u.display_name as created_by_name FROM schedule_plans sp
           LEFT JOIN users u ON sp.created_by = u.id
           WHERE sp.job_id = ? ORDER BY sp.created_at DESC''',
        (job_id,)
    ).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d['plan_data'] = json.loads(d['plan_data']) if d['plan_data'] else []
        d['summary_data'] = json.loads(d['summary_data']) if d['summary_data'] else {}
        d['weather_data'] = json.loads(d['weather_data']) if d['weather_data'] else []
        d['benchmarks_data'] = json.loads(d['benchmarks_data']) if d.get('benchmarks_data') else []
        result.append(d)
    return jsonify(result)


@app.route('/api/schedule/plans/<int:plan_id>', methods=['DELETE'])
@api_role_required('owner', 'admin')
def api_schedule_delete_plan(plan_id):
    """Delete a saved backwards plan (owner only)."""
    conn = get_db()
    row = conn.execute('SELECT id FROM schedule_plans WHERE id = ?', (plan_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Plan not found'}), 404
    conn.execute('DELETE FROM schedule_plans WHERE id = ?', (plan_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/schedule/backwards-plan/generate-pdf', methods=['POST'])
@api_login_required
def api_schedule_plan_pdf():
    """Generate a PDF of a backwards plan."""
    data = request.get_json(force=True)
    job_id = data.get('job_id')
    if not job_id:
        return jsonify({'error': 'job_id required'}), 400

    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    conn.close()
    if not job:
        return jsonify({'error': 'Job not found'}), 404

    job = dict(job)
    logo_path = os.path.abspath(os.path.join(app.static_folder, 'logo.jpg'))

    html = render_template('schedule/plan_pdf.html',
        job=job,
        plan=data.get('plan', []),
        summary=data.get('summary', {}),
        crew_recommendation=data.get('crew_recommendation', ''),
        override_impact=data.get('override_impact'),
        weather=data.get('weather', []),
        benchmarks=data.get('benchmarks', []),
        today=datetime.now().strftime('%B %d, %Y'),
        logo_path='file://' + logo_path
    )

    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    os.makedirs(proposals_dir, exist_ok=True)

    safe_name = ''.join(c if c.isalnum() or c in ' -_' else '' for c in (job.get('name') or 'Job')).strip()
    filename = f"Schedule_Plan_{safe_name}_{job_id}.pdf"
    filepath = os.path.join(proposals_dir, filename)

    try:
        wp = weasyprint.HTML(string=html, base_url=os.path.dirname(__file__))
        wp.write_pdf(filepath)
    except Exception as e:
        return jsonify({'error': f'PDF generation failed: {str(e)[:200]}'}), 500

    return jsonify({'ok': True, 'filename': filename, 'path': f'/api/schedule/plan-pdf/{filename}'})


@app.route('/api/schedule/plan-pdf/<filename>')
@api_login_required
def api_download_schedule_plan_pdf(filename):
    """Download a generated schedule plan PDF."""
    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    filepath = os.path.join(proposals_dir, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    if request.args.get('download'):
        return send_file(filepath, as_attachment=True, download_name=filename)
    return send_file(filepath, mimetype='application/pdf')


@app.route('/api/schedule/backwards-plan/email', methods=['POST'])
@api_login_required
def api_schedule_plan_email():
    """Email a schedule plan PDF to recipients."""
    data = request.get_json(force=True)
    recipients = [e.strip() for e in data.get('recipients', []) if e.strip()]
    subject = data.get('subject', 'Schedule Plan')
    body_text = data.get('body', '')

    if not recipients:
        return jsonify({'error': 'No recipients specified'}), 400

    # Load saved SMTP settings
    saved_settings = {}
    settings_path = os.path.join(os.path.dirname(__file__), 'data', 'email_settings.json')
    if os.path.exists(settings_path):
        with open(settings_path) as f:
            saved_settings = json.load(f)

    smtp_host = saved_settings.get('smtp_host', '')
    smtp_port = int(saved_settings.get('smtp_port', 587) or 587)
    smtp_user = saved_settings.get('smtp_user', '')
    smtp_pass = saved_settings.get('smtp_pass', '')
    from_email = saved_settings.get('from_email', '') or smtp_user

    if not smtp_host or not smtp_user:
        return jsonify({'error': 'SMTP settings required. Configure in Settings.'}), 400

    # Find the PDF
    pdf_filename = data.get('pdf_filename', '')
    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    pdf_path = os.path.join(proposals_dir, pdf_filename)
    if not pdf_filename or not os.path.exists(pdf_path):
        return jsonify({'error': 'No PDF found. Generate the PDF first.'}), 404

    try:
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = subject

        msg.attach(MIMEText(body_text, 'plain'))

        with open(pdf_path, 'rb') as f:
            part = MIMEBase('application', 'pdf')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{pdf_filename}"')
            msg.attach(part)

        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)

        save_email_autocomplete(recipients)
        return jsonify({'ok': True, 'sent_to': recipients})

    except Exception as e:
        return jsonify({'error': f'Email failed: {str(e)}'}), 500


# ─── Recurring Expenses ──────────────────────────────────────────

@app.route('/expenses')
@role_required('owner', 'admin')
def expenses_page():
    return render_template('expenses/overview.html')

@app.route('/api/expenses/recurring')
@api_role_required('owner', 'admin')
def api_recurring_expenses():
    conn = get_db()
    rows = conn.execute('SELECT * FROM recurring_expenses ORDER BY next_due_date ASC, category').fetchall()
    result = [dict(r) for r in rows]
    conn.close()
    # Check for upcoming due notifications
    _check_expense_notifications(result)
    return jsonify(result)

def _check_expense_notifications(expenses):
    from datetime import timedelta
    now = datetime.now()
    week_ahead = now + timedelta(days=7)
    conn = get_db()
    owners = conn.execute("SELECT id FROM users WHERE role IN ('owner','admin') AND is_active = 1").fetchall()
    for exp in expenses:
        if not exp.get('next_due_date') or not exp.get('is_active'):
            continue
        try:
            due_dt = datetime.strptime(exp['next_due_date'], '%Y-%m-%d')
        except (ValueError, TypeError):
            continue
        if due_dt <= week_ahead:
            today_str = now.strftime('%Y-%m-%d')
            for owner in owners:
                existing = conn.execute(
                    "SELECT id FROM notifications WHERE user_id = ? AND type = 'expense' AND message LIKE ? AND created_at >= ?",
                    (owner['id'], f'%expense #{exp["id"]}%', today_str)
                ).fetchone()
                if not existing:
                    status = 'OVERDUE' if due_dt < now else 'due soon'
                    conn.execute(
                        'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
                        (owner['id'], 'expense', f'Expense: {status.title()}',
                         f'{exp["category"]} - {exp["vendor"]} ${exp["amount"]:.2f} {status} ({exp["next_due_date"]}) (expense #{exp["id"]})',
                         '/expenses')
                    )
    conn.commit()
    conn.close()

@app.route('/api/expenses/recurring', methods=['POST'])
@api_role_required('owner', 'admin')
def api_create_recurring_expense():
    data = request.get_json()
    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO recurring_expenses (category, vendor, description, amount, frequency,
           due_day, start_date, end_date, is_active, next_due_date, created_by)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
        (data.get('category', ''), data.get('vendor', ''), data.get('description', ''),
         float(data.get('amount', 0)), data.get('frequency', 'Monthly'),
         int(data.get('due_day', 1)), data.get('start_date', ''), data.get('end_date', ''),
         1 if data.get('is_active', True) else 0,
         data.get('next_due_date', ''), session.get('user_id'))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cursor.lastrowid}), 201

@app.route('/api/expenses/recurring/<int:eid>', methods=['PUT'])
@api_role_required('owner', 'admin')
def api_update_recurring_expense(eid):
    data = request.get_json()
    conn = get_db()
    fields = []
    values = []
    for f in ('category', 'vendor', 'description', 'amount', 'frequency', 'due_day',
              'start_date', 'end_date', 'is_active', 'next_due_date', 'last_paid_date'):
        if f in data:
            fields.append(f'{f} = ?')
            values.append(data[f])
    if fields:
        values.append(eid)
        conn.execute(f"UPDATE recurring_expenses SET {', '.join(fields)} WHERE id = ?", values)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/expenses/recurring/<int:eid>', methods=['DELETE'])
@api_role_required('owner', 'admin')
def api_delete_recurring_expense(eid):
    conn = get_db()
    conn.execute('DELETE FROM recurring_expenses WHERE id = ?', (eid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/expenses/recurring/<int:eid>/pay', methods=['POST'])
@api_role_required('owner', 'admin')
def api_record_expense_payment(eid):
    data = request.get_json()
    conn = get_db()
    conn.execute(
        '''INSERT INTO recurring_expense_payments (recurring_expense_id, amount_paid, payment_date,
           payment_method, reference_number, notes, created_by)
           VALUES (?,?,?,?,?,?,?)''',
        (eid, float(data.get('amount_paid', 0)), data.get('payment_date', datetime.now().strftime('%Y-%m-%d')),
         data.get('payment_method', ''), data.get('reference_number', ''),
         data.get('notes', ''), session.get('user_id'))
    )
    # Update last_paid_date and next_due_date
    if data.get('next_due_date'):
        conn.execute(
            'UPDATE recurring_expenses SET last_paid_date = ?, next_due_date = ? WHERE id = ?',
            (data.get('payment_date', datetime.now().strftime('%Y-%m-%d')), data['next_due_date'], eid)
        )
    else:
        conn.execute(
            'UPDATE recurring_expenses SET last_paid_date = ? WHERE id = ?',
            (data.get('payment_date', datetime.now().strftime('%Y-%m-%d')), eid)
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/expenses/recurring/<int:eid>/payments')
@api_role_required('owner', 'admin')
def api_expense_payments(eid):
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM recurring_expense_payments WHERE recurring_expense_id = ? ORDER BY payment_date DESC',
        (eid,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ─── Licenses ────────────────────────────────────────────────────

LICENSES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'licenses')
os.makedirs(LICENSES_DIR, exist_ok=True)

@app.route('/licenses')
@role_required('owner')
def licenses_page():
    return render_template('licenses/list.html')

@app.route('/api/licenses')
@api_role_required('owner')
def api_licenses_list():
    conn = get_db()
    rows = conn.execute('SELECT * FROM licenses ORDER BY expiration_date ASC').fetchall()
    result = []
    now = datetime.now()
    owners = conn.execute("SELECT id FROM users WHERE role = 'owner' AND is_active = 1").fetchall()
    for r in rows:
        d = dict(r)
        # Auto-update status based on expiration
        if d.get('expiration_date'):
            try:
                exp_dt = datetime.strptime(d['expiration_date'], '%Y-%m-%d')
                days_left = (exp_dt - now).days
                if days_left < 0:
                    d['status'] = 'Expired'
                elif days_left <= 60:
                    d['status'] = 'Expiring Soon'
                else:
                    if d['status'] not in ('Pending Renewal',):
                        d['status'] = 'Active'
                d['days_until_expiry'] = days_left
            except (ValueError, TypeError):
                d['days_until_expiry'] = None
        # Update status in DB if changed
        if d['status'] != r['status']:
            conn.execute('UPDATE licenses SET status = ?, updated_at = datetime("now","localtime") WHERE id = ?',
                         (d['status'], d['id']))
        result.append(d)
    conn.commit()
    # Check for notifications (within 60 days)
    _check_license_notifications(result, owners, conn)
    conn.close()
    return jsonify(result)

def _check_license_notifications(licenses, owners, conn):
    now = datetime.now()
    today_str = now.strftime('%Y-%m-%d')
    from datetime import timedelta
    week_ago = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    for lic in licenses:
        days = lic.get('days_until_expiry')
        if days is not None and days <= 60:
            for owner in owners:
                existing = conn.execute(
                    "SELECT id FROM notifications WHERE user_id = ? AND type = 'license' AND message LIKE ? AND created_at >= ?",
                    (owner['id'], f'%license #{lic["id"]}%', week_ago)
                ).fetchone()
                if not existing:
                    status = 'EXPIRED' if days < 0 else f'expires in {days} days'
                    conn.execute(
                        'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
                        (owner['id'], 'license', f'License: {lic["license_name"]}',
                         f'{lic["license_name"]} ({lic["license_number"]}) {status} (license #{lic["id"]})',
                         '/licenses')
                    )
    conn.commit()

@app.route('/api/licenses', methods=['POST'])
@api_role_required('owner')
def api_create_license():
    conn = get_db()
    if request.content_type and 'multipart' in request.content_type:
        data = request.form
        file = request.files.get('file')
        file_path = ''
        file_hash = ''
        if file and file.filename:
            from werkzeug.utils import secure_filename
            content = file.read()
            file_hash = hashlib.md5(content).hexdigest()
            fname = secure_filename(file.filename)
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.seek(0)
            file.save(os.path.join(LICENSES_DIR, fname))
            file_path = fname
        conn.execute(
            '''INSERT INTO licenses (license_type, license_name, license_number, issuing_body,
               holder_name, issue_date, expiration_date, renewal_cost, status, notes, file_path, file_hash, created_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (data.get('license_type', ''), data.get('license_name', ''), data.get('license_number', ''),
             data.get('issuing_body', ''), data.get('holder_name', ''),
             data.get('issue_date', ''), data.get('expiration_date', ''),
             float(data.get('renewal_cost', 0) or 0), data.get('status', 'Active'),
             data.get('notes', ''), file_path, file_hash, session.get('user_id'))
        )
    else:
        data = request.get_json(force=True)
        conn.execute(
            '''INSERT INTO licenses (license_type, license_name, license_number, issuing_body,
               holder_name, issue_date, expiration_date, renewal_cost, status, notes, created_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
            (data.get('license_type', ''), data.get('license_name', ''), data.get('license_number', ''),
             data.get('issuing_body', ''), data.get('holder_name', ''),
             data.get('issue_date', ''), data.get('expiration_date', ''),
             float(data.get('renewal_cost', 0) or 0), data.get('status', 'Active'),
             data.get('notes', ''), session.get('user_id'))
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/licenses/<int:lid>', methods=['PUT'])
@api_role_required('owner')
def api_update_license(lid):
    conn = get_db()
    if request.content_type and 'multipart' in request.content_type:
        data = request.form
        file = request.files.get('file')
        fields = []
        values = []
        for f in ('license_type', 'license_name', 'license_number', 'issuing_body',
                   'holder_name', 'issue_date', 'expiration_date', 'renewal_cost', 'status', 'notes'):
            if f in data:
                fields.append(f'{f} = ?')
                values.append(float(data[f]) if f == 'renewal_cost' else data[f])
        if file and file.filename:
            from werkzeug.utils import secure_filename
            content = file.read()
            fh = hashlib.md5(content).hexdigest()
            fname = secure_filename(file.filename)
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.seek(0)
            file.save(os.path.join(LICENSES_DIR, fname))
            fields.append('file_path = ?')
            values.append(fname)
            fields.append('file_hash = ?')
            values.append(fh)
        if fields:
            fields.append("updated_at = datetime('now','localtime')")
            values.append(lid)
            conn.execute(f"UPDATE licenses SET {', '.join(fields)} WHERE id = ?", values)
            conn.commit()
    else:
        data = request.get_json(force=True)
        fields = []
        values = []
        for f in ('license_type', 'license_name', 'license_number', 'issuing_body',
                   'holder_name', 'issue_date', 'expiration_date', 'renewal_cost', 'status', 'notes'):
            if f in data:
                fields.append(f'{f} = ?')
                values.append(data[f])
        if fields:
            fields.append("updated_at = datetime('now','localtime')")
            values.append(lid)
            conn.execute(f"UPDATE licenses SET {', '.join(fields)} WHERE id = ?", values)
            conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/licenses/<int:lid>', methods=['DELETE'])
@api_role_required('owner')
def api_delete_license(lid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM licenses WHERE id = ?', (lid,)).fetchone()
    if row and row['file_path']:
        fpath = os.path.join(LICENSES_DIR, row['file_path'])
        if os.path.exists(fpath):
            os.remove(fpath)
    conn.execute('DELETE FROM licenses WHERE id = ?', (lid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/licenses/<int:lid>/file')
@api_role_required('owner')
def api_license_file(lid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM licenses WHERE id = ?', (lid,)).fetchone()
    conn.close()
    if not row or not row['file_path']:
        return 'Not found', 404
    fpath = os.path.join(LICENSES_DIR, row['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    return send_file(fpath, mimetype='application/pdf')

# ─── RFIs ────────────────────────────────────────────────────────

@app.route('/rfis')
@role_required('owner', 'admin', 'project_manager')
def rfis_page():
    return render_template('rfis/list.html')

@app.route('/api/rfis')
@api_role_required('owner', 'admin', 'project_manager')
def api_rfis_list():
    job_id = request.args.get('job_id', '')
    status = request.args.get('status', '')
    conn = get_db()
    sql = '''SELECT r.*, j.name as job_name FROM rfis r
             LEFT JOIN jobs j ON r.job_id = j.id WHERE 1=1'''
    params = []
    if job_id:
        sql += ' AND r.job_id = ?'
        params.append(job_id)
    if status:
        sql += ' AND r.status = ?'
        params.append(status)
    sql += ' ORDER BY r.date_submitted DESC, r.rfi_number DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/rfis', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_create_rfi():
    data = request.get_json()
    conn = get_db()
    # Auto-assign RFI number per job
    max_num = conn.execute('SELECT MAX(rfi_number) FROM rfis WHERE job_id = ?',
                           (data['job_id'],)).fetchone()[0] or 0
    cursor = conn.execute(
        '''INSERT INTO rfis (job_id, rfi_number, subject, question, requested_by,
           assigned_to, date_required, created_by)
           VALUES (?,?,?,?,?,?,?,?)''',
        (data['job_id'], max_num + 1, data.get('subject', ''), data.get('question', ''),
         data.get('requested_by', ''), data.get('assigned_to') or None,
         data.get('date_required', ''), session.get('user_id'))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cursor.lastrowid}), 201

@app.route('/api/rfis/<int:rid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_rfi(rid):
    data = request.get_json()
    conn = get_db()
    fields = []
    values = []
    for f in ('subject', 'question', 'answer', 'requested_by', 'assigned_to',
              'status', 'date_required', 'date_answered'):
        if f in data:
            fields.append(f'{f} = ?')
            values.append(data[f] if data[f] != '' else (None if f == 'assigned_to' else data[f]))
    if fields:
        fields.append("updated_at = datetime('now','localtime')")
        values.append(rid)
        conn.execute(f"UPDATE rfis SET {', '.join(fields)} WHERE id = ?", values)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/rfis/<int:rid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_rfi(rid):
    conn = get_db()
    conn.execute('DELETE FROM rfis WHERE id = ?', (rid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Change Orders ───────────────────────────────────────────────

@app.route('/change-orders')
@role_required('owner', 'admin', 'project_manager')
def change_orders_page():
    return render_template('change_orders/list.html')

@app.route('/api/change-orders')
@api_role_required('owner', 'admin', 'project_manager')
def api_change_orders_list():
    job_id = request.args.get('job_id', '')
    status = request.args.get('status', '')
    conn = get_db()
    sql = '''SELECT co.*, j.name as job_name FROM change_orders co
             LEFT JOIN jobs j ON co.job_id = j.id WHERE 1=1'''
    params = []
    if job_id:
        sql += ' AND co.job_id = ?'
        params.append(job_id)
    if status:
        sql += ' AND co.status = ?'
        params.append(status)
    sql += ' ORDER BY co.created_at DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/change-orders/<int:coid>')
@api_role_required('owner', 'admin', 'project_manager')
def api_get_change_order(coid):
    conn = get_db()
    co = conn.execute(
        'SELECT co.*, j.name as job_name FROM change_orders co LEFT JOIN jobs j ON co.job_id = j.id WHERE co.id = ?',
        (coid,)
    ).fetchone()
    if not co:
        conn.close()
        return jsonify({'error': 'Change order not found'}), 404
    result = dict(co)
    # Find which building section this CO's SOV line belongs to
    if co['sov_item_id'] and co['pay_app_contract_id']:
        sov_line = conn.execute('SELECT sort_order FROM pay_app_sov_items WHERE id = ?', (co['sov_item_id'],)).fetchone()
        if sov_line:
            header = conn.execute(
                'SELECT id FROM pay_app_sov_items WHERE contract_id = ? AND is_header = 1 AND sort_order < ? ORDER BY sort_order DESC LIMIT 1',
                (co['pay_app_contract_id'], sov_line['sort_order'])
            ).fetchone()
            result['sov_building_id'] = header['id'] if header else None
    conn.close()
    return jsonify(result)

@app.route('/api/change-orders', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_create_change_order():
    data = request.get_json()
    conn = get_db()
    max_num = conn.execute('SELECT MAX(co_number) FROM change_orders WHERE job_id = ?',
                           (data['job_id'],)).fetchone()[0] or 0
    co_number = max_num + 1
    co_title = data.get('title', '')
    co_amount = float(data.get('amount', 0))
    cursor = conn.execute(
        '''INSERT INTO change_orders (job_id, co_number, title, scope_description, reason,
           amount, gc_name, created_by)
           VALUES (?,?,?,?,?,?,?,?)''',
        (data['job_id'], co_number, co_title, data.get('scope_description', ''),
         data.get('reason', ''), co_amount,
         data.get('gc_name', ''), session.get('user_id'))
    )
    co_id = cursor.lastrowid

    # Immediately create SOV line item so CO is billable right away
    contract = conn.execute(
        'SELECT id FROM pay_app_contracts WHERE job_id = ? ORDER BY id LIMIT 1',
        (data['job_id'],)
    ).fetchone()

    sov_item_id = None
    sov_section_id = data.get('sov_section_id')  # header item ID for building placement
    if contract:
        cid = contract['id']

        # Determine insert position
        if sov_section_id:
            # Find the last item in this building section (everything after the header
            # until the next header or end of list)
            section_header = conn.execute(
                'SELECT sort_order FROM pay_app_sov_items WHERE id = ? AND contract_id = ?',
                (sov_section_id, cid)
            ).fetchone()
            if section_header:
                # Find next header after this one
                next_header = conn.execute(
                    'SELECT MIN(sort_order) FROM pay_app_sov_items WHERE contract_id = ? AND is_header = 1 AND sort_order > ?',
                    (cid, section_header['sort_order'])
                ).fetchone()[0]
                if next_header is not None:
                    insert_sort = next_header  # insert before the next header
                else:
                    # Last section — insert at end
                    insert_sort = conn.execute('SELECT COALESCE(MAX(sort_order), -1) + 1 FROM pay_app_sov_items WHERE contract_id = ?', (cid,)).fetchone()[0]
                # Shift items down
                conn.execute('UPDATE pay_app_sov_items SET sort_order = sort_order + 1 WHERE contract_id = ? AND sort_order >= ?', (cid, insert_sort))
            else:
                insert_sort = conn.execute('SELECT COALESCE(MAX(sort_order), -1) + 1 FROM pay_app_sov_items WHERE contract_id = ?', (cid,)).fetchone()[0]
        else:
            insert_sort = conn.execute('SELECT COALESCE(MAX(sort_order), -1) + 1 FROM pay_app_sov_items WHERE contract_id = ?', (cid,)).fetchone()[0]

        sov_cursor = conn.execute(
            '''INSERT INTO pay_app_sov_items (contract_id, item_number, description, scheduled_value, sort_order)
               VALUES (?,?,?,?,?)''',
            (cid, 0, f"CO #{co_number}: {co_title}", co_amount, insert_sort)
        )
        sov_item_id = sov_cursor.lastrowid

        # Renumber all item_numbers sequentially
        rows = conn.execute('SELECT id FROM pay_app_sov_items WHERE contract_id = ? ORDER BY sort_order, id', (cid,)).fetchall()
        for idx, row in enumerate(rows):
            conn.execute('UPDATE pay_app_sov_items SET item_number = ? WHERE id = ?', (idx + 1, row['id']))

        # Backfill existing pay apps with 0 entries
        existing_apps = conn.execute(
            'SELECT id FROM pay_applications WHERE contract_id = ?', (cid,)
        ).fetchall()
        for pa in existing_apps:
            conn.execute(
                'INSERT OR IGNORE INTO pay_app_line_entries (pay_app_id, sov_item_id, work_this_period, materials_stored) VALUES (?,?,0,0)',
                (pa['id'], sov_item_id)
            )

        conn.execute(
            'UPDATE change_orders SET pay_app_contract_id = ?, sov_item_id = ? WHERE id = ?',
            (cid, sov_item_id, co_id)
        )

    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': co_id, 'sov_item_id': sov_item_id}), 201

@app.route('/api/change-orders/<int:coid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_change_order(coid):
    data = request.get_json()
    conn = get_db()
    fields = []
    values = []
    for f in ('title', 'scope_description', 'reason', 'amount', 'status',
              'gc_name', 'submitted_date', 'approved_date', 'approved_by'):
        if f in data:
            fields.append(f'{f} = ?')
            values.append(data[f])
    if fields:
        fields.append("updated_at = datetime('now','localtime')")
        values.append(coid)
        conn.execute(f"UPDATE change_orders SET {', '.join(fields)} WHERE id = ?", values)

        # Sync SOV line item if amount or title changed
        if 'amount' in data or 'title' in data:
            co = conn.execute('SELECT co_number, title, amount, sov_item_id FROM change_orders WHERE id = ?', (coid,)).fetchone()
            if co and co['sov_item_id']:
                conn.execute(
                    'UPDATE pay_app_sov_items SET description = ?, scheduled_value = ? WHERE id = ?',
                    (f"CO #{co['co_number']}: {co['title']}", co['amount'], co['sov_item_id'])
                )

        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/change-orders/<int:coid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_change_order(coid):
    conn = get_db()
    # Remove linked SOV line item and its line entries
    co = conn.execute('SELECT sov_item_id FROM change_orders WHERE id = ?', (coid,)).fetchone()
    if co and co['sov_item_id']:
        sov_id = co['sov_item_id']
        # Clear FK reference before deleting SOV item
        conn.execute('UPDATE change_orders SET sov_item_id = NULL, pay_app_contract_id = NULL WHERE id = ?', (coid,))
        conn.execute('DELETE FROM pay_app_line_entries WHERE sov_item_id = ?', (sov_id,))
        conn.execute('DELETE FROM pay_app_sov_items WHERE id = ?', (sov_id,))
    conn.execute('DELETE FROM change_orders WHERE id = ?', (coid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/change-orders/<int:coid>/generate-proposal', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_generate_co_proposal(coid):
    conn = get_db()
    co = conn.execute(
        '''SELECT co.*, j.name as job_name, j.address as job_address, j.city as job_city, j.state as job_state
           FROM change_orders co LEFT JOIN jobs j ON co.job_id = j.id WHERE co.id = ?''',
        (coid,)
    ).fetchone()
    conn.close()
    if not co:
        return jsonify({'error': 'Change order not found'}), 404

    co = dict(co)
    today = datetime.now().strftime('%B %d, %Y')
    logo_path = os.path.abspath(os.path.join(app.static_folder, 'logo.jpg'))

    html = render_template('change_orders/proposal_pdf.html', co=co, today=today, logo_path='file://' + logo_path)

    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    os.makedirs(proposals_dir, exist_ok=True)

    safe_title = ''.join(c if c.isalnum() or c in ' -_' else '' for c in (co.get('title') or 'CO')).strip()
    filename = f"CO_{co['job_id']}_{co['co_number']}_{safe_title}.pdf"
    filepath = os.path.join(proposals_dir, filename)

    try:
        wp = weasyprint.HTML(string=html, base_url=os.path.dirname(__file__))
        wp.write_pdf(filepath)
    except Exception as e:
        return jsonify({'error': f'PDF generation failed: {str(e)[:200]}'}), 500

    # Store filename on CO
    conn2 = get_db()
    conn2.execute('UPDATE change_orders SET proposal_file = ? WHERE id = ?', (filename, coid))
    conn2.commit()
    conn2.close()

    return jsonify({'ok': True, 'filename': filename, 'path': f'/api/change-orders/{coid}/proposal/{filename}'})

@app.route('/api/change-orders/<int:coid>/proposal/<filename>')
@api_role_required('owner', 'admin', 'project_manager')
def api_download_co_proposal(coid, filename):
    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    filepath = os.path.join(proposals_dir, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    if request.args.get('download'):
        return send_file(filepath, as_attachment=True, download_name=filename)
    return send_file(filepath, mimetype='application/pdf')

@app.route('/api/change-orders/<int:coid>/approve', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_approve_change_order(coid):
    conn = get_db()
    co = conn.execute('SELECT * FROM change_orders WHERE id = ?', (coid,)).fetchone()
    if not co:
        conn.close()
        return jsonify({'error': 'Change order not found'}), 404

    today = datetime.now().strftime('%Y-%m-%d')
    conn.execute(
        "UPDATE change_orders SET status = 'Approved', approved_date = ?, approved_by = ?, updated_at = datetime('now','localtime') WHERE id = ?",
        (today, session.get('display_name', session.get('username', '')), coid)
    )

    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Submittals ──────────────────────────────────────────────────

SUBMITTALS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'submittals')
os.makedirs(SUBMITTALS_DIR, exist_ok=True)

@app.route('/submittals')
@role_required('owner', 'admin', 'project_manager')
def submittals_page():
    return render_template('submittals/list.html')

@app.route('/api/submittals')
@api_role_required('owner', 'admin', 'project_manager')
def api_submittals_list():
    job_id = request.args.get('job_id', '')
    status = request.args.get('status', '')
    conn = get_db()
    sql = '''SELECT s.*, j.name as job_name, sf.title as library_title, sf.id as library_file_id
             FROM submittals s
             LEFT JOIN jobs j ON s.job_id = j.id
             LEFT JOIN submittal_files sf ON s.submittal_file_id = sf.id
             WHERE 1=1'''
    params = []
    if job_id:
        sql += ' AND s.job_id = ?'
        params.append(job_id)
    if status:
        sql += ' AND s.status = ?'
        params.append(status)
    sql += ' ORDER BY s.created_at DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/submittals', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_create_submittal():
    conn = get_db()
    if request.content_type and 'multipart' in request.content_type:
        data = request.form
        file = request.files.get('file')
        file_path = ''
        if file and file.filename:
            from werkzeug.utils import secure_filename
            fname = secure_filename(file.filename)
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.save(os.path.join(SUBMITTALS_DIR, fname))
            file_path = fname
        job_id = data.get('job_id')
        max_num = conn.execute('SELECT MAX(submittal_number) FROM submittals WHERE job_id = ?',
                               (job_id,)).fetchone()[0] or 0
        conn.execute(
            '''INSERT INTO submittals (job_id, submittal_number, spec_section, description, vendor,
               date_required, file_path, created_by)
               VALUES (?,?,?,?,?,?,?,?)''',
            (job_id, max_num + 1, data.get('spec_section', ''), data.get('description', ''),
             data.get('vendor', ''), data.get('date_required', ''),
             file_path, session.get('user_id'))
        )
    else:
        data = request.get_json(force=True)
        job_id = data.get('job_id')
        max_num = conn.execute('SELECT MAX(submittal_number) FROM submittals WHERE job_id = ?',
                               (job_id,)).fetchone()[0] or 0
        conn.execute(
            '''INSERT INTO submittals (job_id, submittal_number, spec_section, description, vendor,
               date_required, created_by)
               VALUES (?,?,?,?,?,?,?)''',
            (job_id, max_num + 1, data.get('spec_section', ''), data.get('description', ''),
             data.get('vendor', ''), data.get('date_required', ''), session.get('user_id'))
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/submittals/<int:sid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_submittal(sid):
    conn = get_db()
    if request.content_type and 'multipart' in request.content_type:
        data = request.form
        file = request.files.get('file')
        fields = []
        values = []
        for f in ('spec_section', 'description', 'vendor', 'status', 'revision_number',
                   'date_submitted', 'date_required', 'date_returned', 'reviewer', 'reviewer_comments'):
            if f in data:
                fields.append(f'{f} = ?')
                values.append(int(data[f]) if f == 'revision_number' else data[f])
        if file and file.filename:
            from werkzeug.utils import secure_filename
            fname = secure_filename(file.filename)
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.save(os.path.join(SUBMITTALS_DIR, fname))
            fields.append('file_path = ?')
            values.append(fname)
        if fields:
            fields.append("updated_at = datetime('now','localtime')")
            values.append(sid)
            conn.execute(f"UPDATE submittals SET {', '.join(fields)} WHERE id = ?", values)
            conn.commit()
    else:
        data = request.get_json(force=True)
        fields = []
        values = []
        for f in ('spec_section', 'description', 'vendor', 'status', 'revision_number',
                   'date_submitted', 'date_required', 'date_returned', 'reviewer', 'reviewer_comments'):
            if f in data:
                fields.append(f'{f} = ?')
                values.append(data[f])
        if fields:
            fields.append("updated_at = datetime('now','localtime')")
            values.append(sid)
            conn.execute(f"UPDATE submittals SET {', '.join(fields)} WHERE id = ?", values)
            conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/submittals/<int:sid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_submittal(sid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM submittals WHERE id = ?', (sid,)).fetchone()
    if row and row['file_path']:
        fpath = os.path.join(SUBMITTALS_DIR, row['file_path'])
        if os.path.exists(fpath):
            os.remove(fpath)
    conn.execute('DELETE FROM submittals WHERE id = ?', (sid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/submittals/<int:sid>/file')
@api_role_required('owner', 'admin', 'project_manager')
def api_submittal_file(sid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM submittals WHERE id = ?', (sid,)).fetchone()
    conn.close()
    if not row or not row['file_path']:
        return 'Not found', 404
    fpath = os.path.join(SUBMITTALS_DIR, row['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    return send_file(fpath, mimetype='application/pdf')

# ─── Submittal Library ────────────────────────────────────────

@app.route('/api/submittal-library')
@api_role_required('owner', 'admin', 'project_manager')
def api_submittal_library_list():
    conn = get_db()
    rows = conn.execute(
        '''SELECT sf.*, COUNT(s.id) as usage_count
           FROM submittal_files sf
           LEFT JOIN submittals s ON s.submittal_file_id = sf.id
           GROUP BY sf.id ORDER BY sf.title'''
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/submittal-library', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_create_submittal_library():
    conn = get_db()
    if request.content_type and 'multipart' in request.content_type:
        data = request.form
        file = request.files.get('file')
        file_path = ''
        file_hash = ''
        if file and file.filename:
            import hashlib
            from werkzeug.utils import secure_filename
            content = file.read()
            file_hash = hashlib.md5(content).hexdigest()
            # Check for duplicate
            existing = conn.execute('SELECT id, title FROM submittal_files WHERE file_hash = ?', (file_hash,)).fetchone()
            if existing:
                conn.close()
                return jsonify({'ok': False, 'duplicate': True, 'existing_id': existing['id'],
                                'existing_title': existing['title']}), 409
            fname = secure_filename(file.filename)
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.seek(0)
            file.save(os.path.join(SUBMITTALS_DIR, fname))
            file_path = fname
        cursor = conn.execute(
            '''INSERT INTO submittal_files (title, file_path, file_hash, vendor, category, description, keywords)
               VALUES (?,?,?,?,?,?,?)''',
            (data.get('title', ''), file_path, file_hash,
             data.get('vendor', ''), data.get('category', ''), data.get('description', ''),
             data.get('keywords', ''))
        )
        lib_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'id': lib_id}), 201
    else:
        data = request.get_json(force=True)
        cursor = conn.execute(
            '''INSERT INTO submittal_files (title, file_path, file_hash, vendor, category, description, keywords)
               VALUES (?,?,?,?,?,?,?)''',
            (data.get('title', ''), data.get('file_path', ''), data.get('file_hash', ''),
             data.get('vendor', ''), data.get('category', ''), data.get('description', ''),
             data.get('keywords', ''))
        )
        lib_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'id': lib_id}), 201

@app.route('/api/submittal-library/<int:lid>/file')
@api_role_required('owner', 'admin', 'project_manager')
def api_submittal_library_file(lid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM submittal_files WHERE id = ?', (lid,)).fetchone()
    conn.close()
    if not row or not row['file_path']:
        return 'Not found', 404
    fpath = os.path.join(SUBMITTALS_DIR, row['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    return send_file(fpath, mimetype='application/pdf')

@app.route('/api/submittals/<int:sid>/link-library', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_link_submittal_to_library(sid):
    data = request.get_json(force=True)
    lib_id = data.get('submittal_file_id')
    if not lib_id:
        return jsonify({'error': 'submittal_file_id required'}), 400
    conn = get_db()
    lib = conn.execute('SELECT id, file_path FROM submittal_files WHERE id = ?', (lib_id,)).fetchone()
    if not lib:
        conn.close()
        return jsonify({'error': 'Library file not found'}), 404
    conn.execute(
        "UPDATE submittals SET submittal_file_id = ?, updated_at = datetime('now','localtime') WHERE id = ?",
        (lib_id, sid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/submittal-library/<int:lid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_submittal_library(lid):
    data = request.get_json(force=True)
    conn = get_db()
    fields = []
    values = []
    for f in ('title', 'vendor', 'category', 'description', 'keywords'):
        if f in data:
            fields.append(f'{f} = ?')
            values.append(data[f])
    if fields:
        values.append(lid)
        conn.execute(f"UPDATE submittal_files SET {', '.join(fields)} WHERE id = ?", values)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/submittal-library/<int:lid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_submittal_library(lid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM submittal_files WHERE id = ?', (lid,)).fetchone()
    if row and row['file_path']:
        fpath = os.path.join(SUBMITTALS_DIR, row['file_path'])
        if os.path.exists(fpath):
            os.remove(fpath)
    # Unlink any submittals that reference this library file
    conn.execute('UPDATE submittals SET submittal_file_id = NULL WHERE submittal_file_id = ?', (lid,))
    conn.execute('DELETE FROM submittal_files WHERE id = ?', (lid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Documents (Closeout) ───────────────────────────────────────

CLOSEOUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'closeout')
os.makedirs(CLOSEOUT_DIR, exist_ok=True)

@app.route('/documents')
@role_required('owner', 'admin', 'project_manager')
def documents_page():
    return render_template('documents/list.html')

@app.route('/documents/job/<int:job_id>')
@role_required('owner', 'admin', 'project_manager')
def documents_job_page(job_id):
    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    conn.close()
    if not job:
        return 'Job not found', 404
    return render_template('documents/job.html', job=job)

@app.route('/api/documents/checklist')
@api_role_required('owner', 'admin', 'project_manager')
def api_checklist_list():
    job_id = request.args.get('job_id', '')
    conn = get_db()
    if job_id:
        rows = conn.execute('SELECT * FROM closeout_checklists WHERE job_id = ? ORDER BY sort_order, id', (job_id,)).fetchall()
    else:
        rows = conn.execute('SELECT * FROM closeout_checklists ORDER BY job_id, sort_order, id').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/documents/checklist', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_checklist_create():
    conn = get_db()
    if request.content_type and 'multipart' in request.content_type:
        data = request.form
        file = request.files.get('file')
        file_path = ''
        file_hash = ''
        if file and file.filename:
            from werkzeug.utils import secure_filename
            content = file.read()
            file_hash = hashlib.md5(content).hexdigest()
            fname = secure_filename(file.filename)
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.seek(0)
            file.save(os.path.join(CLOSEOUT_DIR, fname))
            file_path = fname
        conn.execute(
            '''INSERT INTO closeout_checklists (job_id, item_name, item_type, status, file_path, file_hash, notes, sort_order, created_by)
               VALUES (?,?,?,?,?,?,?,?,?)''',
            (data.get('job_id'), data.get('item_name', ''), data.get('item_type', 'Other'),
             data.get('status', 'Not Started'), file_path, file_hash, data.get('notes', ''),
             int(data.get('sort_order', 0)), session.get('user_id'))
        )
    else:
        data = request.get_json(force=True)
        conn.execute(
            '''INSERT INTO closeout_checklists (job_id, item_name, item_type, status, notes, sort_order, created_by)
               VALUES (?,?,?,?,?,?,?)''',
            (data.get('job_id'), data.get('item_name', ''), data.get('item_type', 'Other'),
             data.get('status', 'Not Started'), data.get('notes', ''),
             int(data.get('sort_order', 0)), session.get('user_id'))
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/documents/checklist/<int:cid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_checklist_update(cid):
    conn = get_db()
    if request.content_type and 'multipart' in request.content_type:
        data = request.form
        file = request.files.get('file')
        fields = []
        values = []
        for f in ('item_name', 'item_type', 'status', 'notes', 'sort_order'):
            if f in data:
                fields.append(f'{f} = ?')
                values.append(int(data[f]) if f == 'sort_order' else data[f])
        if file and file.filename:
            from werkzeug.utils import secure_filename
            content = file.read()
            fh = hashlib.md5(content).hexdigest()
            fname = secure_filename(file.filename)
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.seek(0)
            file.save(os.path.join(CLOSEOUT_DIR, fname))
            fields.append('file_path = ?')
            values.append(fname)
            fields.append('file_hash = ?')
            values.append(fh)
        if fields:
            fields.append("updated_at = datetime('now','localtime')")
            values.append(cid)
            conn.execute(f"UPDATE closeout_checklists SET {', '.join(fields)} WHERE id = ?", values)
            conn.commit()
    else:
        data = request.get_json(force=True)
        fields = []
        values = []
        for f in ('item_name', 'item_type', 'status', 'notes', 'sort_order'):
            if f in data:
                fields.append(f'{f} = ?')
                values.append(data[f])
        if fields:
            fields.append("updated_at = datetime('now','localtime')")
            values.append(cid)
            conn.execute(f"UPDATE closeout_checklists SET {', '.join(fields)} WHERE id = ?", values)
            conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/documents/checklist/<int:cid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_checklist_delete(cid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM closeout_checklists WHERE id = ?', (cid,)).fetchone()
    if row and row['file_path']:
        fpath = os.path.join(CLOSEOUT_DIR, row['file_path'])
        if os.path.exists(fpath):
            os.remove(fpath)
    conn.execute('DELETE FROM closeout_checklists WHERE id = ?', (cid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/documents/checklist/<int:cid>/file')
@api_role_required('owner', 'admin', 'project_manager')
def api_checklist_file(cid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM closeout_checklists WHERE id = ?', (cid,)).fetchone()
    conn.close()
    if not row or not row['file_path']:
        return 'Not found', 404
    fpath = os.path.join(CLOSEOUT_DIR, row['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    return send_file(fpath, mimetype='application/pdf')

@app.route('/api/documents/checklist/defaults', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_checklist_add_defaults():
    data = request.get_json(force=True)
    job_id = data.get('job_id')
    if not job_id:
        return jsonify({'error': 'job_id required'}), 400

    defaults = [
        ('O&M Manual', 'O&M Manual'),
        ('Warranty Letter', 'Warranty Letter'),
        ('As-Built Drawings', 'As-Built'),
        ('Start-Up Report', 'Start-Up Report'),
        ('Balancing Report', 'Balancing Report'),
        ('Test Report', 'Test Report'),
        ('Lien Waiver', 'Lien Waiver'),
        ('Certificate of Completion', 'Certificate of Completion'),
        ('Permit Closeout', 'Permit'),
    ]
    conn = get_db()
    existing = conn.execute('SELECT item_name FROM closeout_checklists WHERE job_id = ?', (job_id,)).fetchall()
    existing_names = {r['item_name'] for r in existing}
    sort_idx = len(existing)
    for name, item_type in defaults:
        if name not in existing_names:
            conn.execute(
                '''INSERT INTO closeout_checklists (job_id, item_name, item_type, sort_order, created_by)
                   VALUES (?,?,?,?,?)''',
                (job_id, name, item_type, sort_idx, session.get('user_id'))
            )
            sort_idx += 1
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/documents/transmittals')
@api_role_required('owner', 'admin', 'project_manager')
def api_transmittals_list():
    job_id = request.args.get('job_id', '')
    conn = get_db()
    if job_id:
        rows = conn.execute('SELECT * FROM transmittals WHERE job_id = ? ORDER BY transmittal_number DESC', (job_id,)).fetchall()
    else:
        rows = conn.execute('SELECT * FROM transmittals ORDER BY created_at DESC').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/documents/transmittals', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_transmittal_create():
    data = request.get_json(force=True)
    conn = get_db()
    job_id = data.get('job_id')
    max_num = conn.execute('SELECT MAX(transmittal_number) FROM transmittals WHERE job_id = ?',
                           (job_id,)).fetchone()[0] or 0
    cursor = conn.execute(
        '''INSERT INTO transmittals (job_id, transmittal_number, to_company, to_attention,
           subject, notes, sent_date, sent_via, created_by)
           VALUES (?,?,?,?,?,?,?,?,?)''',
        (job_id, max_num + 1, data.get('to_company', ''), data.get('to_attention', ''),
         data.get('subject', ''), data.get('notes', ''), data.get('sent_date', ''),
         data.get('sent_via', 'Email'), session.get('user_id'))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cursor.lastrowid}), 201

@app.route('/api/documents/transmittals/<int:tid>/generate', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_generate_transmittal(tid):
    conn = get_db()
    trans = conn.execute('SELECT * FROM transmittals WHERE id = ?', (tid,)).fetchone()
    if not trans:
        conn.close()
        return jsonify({'error': 'Transmittal not found'}), 404

    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (trans['job_id'],)).fetchone()
    # Get completed checklist items for this job
    items = conn.execute(
        "SELECT * FROM closeout_checklists WHERE job_id = ? AND status = 'Complete' ORDER BY sort_order",
        (trans['job_id'],)
    ).fetchall()
    conn.close()

    trans = dict(trans)
    job = dict(job) if job else {}
    items = [dict(i) for i in items]
    today = datetime.now().strftime('%B %d, %Y')
    logo_path = os.path.abspath(os.path.join(app.static_folder, 'logo.jpg'))

    html = render_template('documents/transmittal_pdf.html',
                           transmittal=trans, job=job, documents=items,
                           today=today, logo_path='file://' + logo_path)

    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    os.makedirs(proposals_dir, exist_ok=True)

    filename = f"Transmittal_{trans['job_id']}_{trans['transmittal_number']}.pdf"
    filepath = os.path.join(proposals_dir, filename)

    try:
        wp = weasyprint.HTML(string=html, base_url=os.path.dirname(__file__))
        wp.write_pdf(filepath)
    except Exception as e:
        return jsonify({'error': f'PDF generation failed: {str(e)[:200]}'}), 500

    # Store filename
    conn2 = get_db()
    conn2.execute('UPDATE transmittals SET proposal_file = ? WHERE id = ?', (filename, tid))
    conn2.commit()
    conn2.close()

    return jsonify({'ok': True, 'filename': filename, 'path': f'/api/documents/transmittals/{tid}/pdf/{filename}'})

@app.route('/api/documents/transmittals/<int:tid>/pdf/<filename>')
@api_role_required('owner', 'admin', 'project_manager')
def api_download_transmittal(tid, filename):
    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    filepath = os.path.join(proposals_dir, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    if request.args.get('download'):
        return send_file(filepath, as_attachment=True, download_name=filename)
    return send_file(filepath, mimetype='application/pdf')


# ─── Contracts ───────────────────────────────────────────────────

CONTRACTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'contracts')
os.makedirs(CONTRACTS_DIR, exist_ok=True)

@app.route('/contracts')
@role_required('owner', 'admin', 'project_manager')
def contracts_page():
    return render_template('contracts/list.html')

@app.route('/api/contracts')
@api_role_required('owner', 'admin', 'project_manager')
def api_contracts_list():
    job_id = request.args.get('job_id', type=int)
    status = request.args.get('status', '')
    ctype = request.args.get('type', '')
    conn = get_db()
    sql = '''SELECT c.*, j.name as job_name FROM contracts c
             JOIN jobs j ON c.job_id = j.id WHERE 1=1'''
    params = []
    if job_id:
        sql += ' AND c.job_id = ?'
        params.append(job_id)
    if status:
        sql += ' AND c.status = ?'
        params.append(status)
    if ctype:
        sql += ' AND c.contract_type = ?'
        params.append(ctype)
    sql += ' ORDER BY c.created_at DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d['type'] = d.get('contract_type', '')
        d['has_file'] = bool(d.get('file_path'))
        d['has_review'] = bool(d.get('ai_review'))
        result.append(d)
    return jsonify(result)

@app.route('/api/contracts', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_contracts_create():
    conn = get_db()
    file_path = ''
    file_hash = ''
    if request.content_type and 'multipart' in request.content_type:
        job_id = request.form.get('job_id', type=int)
        title = request.form.get('title', '')
        contractor = request.form.get('contractor', '')
        contract_type = request.form.get('contract_type', 'Prime')
        value = float(request.form.get('value', 0))
        status = request.form.get('status', 'Draft')
        notes = request.form.get('notes', '')
        contract_date = request.form.get('contract_date', '')
        file = request.files.get('file')
        if file and file.filename:
            from werkzeug.utils import secure_filename
            content = file.read()
            file_hash = hashlib.md5(content).hexdigest()
            fname = secure_filename(file.filename)
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.seek(0)
            file.save(os.path.join(CONTRACTS_DIR, fname))
            file_path = fname
    else:
        data = request.get_json(force=True)
        job_id = data.get('job_id')
        title = data.get('title', '')
        contractor = data.get('contractor', '')
        contract_type = data.get('contract_type', 'Prime')
        value = float(data.get('value', 0))
        status = data.get('status', 'Draft')
        notes = data.get('notes', '')
        contract_date = data.get('contract_date', '')

    cursor = conn.execute(
        '''INSERT INTO contracts (job_id, title, contractor, contract_type, file_path, file_hash,
           upload_date, value, status, notes, created_by, contract_date)
           VALUES (?,?,?,?,?,?,date('now','localtime'),?,?,?,?,?)''',
        (job_id, title, contractor, contract_type, file_path, file_hash, value, status, notes, session['user_id'], contract_date)
    )
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return jsonify({'ok': True, 'id': new_id}), 201

@app.route('/api/contracts/<int:cid>')
@api_role_required('owner', 'admin', 'project_manager')
def api_contracts_get(cid):
    conn = get_db()
    row = conn.execute(
        '''SELECT c.*, j.name as job_name FROM contracts c
           JOIN jobs j ON c.job_id = j.id WHERE c.id = ?''', (cid,)
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Contract not found'}), 404
    d = dict(row)
    d['type'] = d.get('contract_type', '')
    d['has_file'] = bool(d.get('file_path'))
    d['has_review'] = bool(d.get('ai_review'))
    if d.get('ai_review'):
        try:
            d['ai_review'] = json.loads(d['ai_review'])
        except (json.JSONDecodeError, TypeError):
            pass
    return jsonify(d)

@app.route('/api/contracts/<int:cid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_contracts_update(cid):
    conn = get_db()
    file_path_update = None
    if request.content_type and 'multipart' in request.content_type:
        title = request.form.get('title', '')
        contractor = request.form.get('contractor', '')
        contract_type = request.form.get('contract_type', 'Prime')
        value = float(request.form.get('value', 0))
        status = request.form.get('status', 'Draft')
        notes = request.form.get('notes', '')
        contract_date = request.form.get('contract_date', '')
        file = request.files.get('file')
        if file and file.filename:
            from werkzeug.utils import secure_filename
            fname = secure_filename(file.filename)
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.save(os.path.join(CONTRACTS_DIR, fname))
            file_path_update = fname
    else:
        data = request.get_json(force=True)
        title = data.get('title', '')
        contractor = data.get('contractor', '')
        contract_type = data.get('contract_type', 'Prime')
        value = float(data.get('value', 0))
        status = data.get('status', 'Draft')
        notes = data.get('notes', '')
        contract_date = data.get('contract_date', '')

    if file_path_update:
        conn.execute(
            '''UPDATE contracts SET title=?, contractor=?, contract_type=?, value=?,
               status=?, notes=?, contract_date=?, file_path=?, upload_date=date('now','localtime'),
               updated_at=datetime('now','localtime') WHERE id=?''',
            (title, contractor, contract_type, value, status, notes, contract_date, file_path_update, cid)
        )
    else:
        conn.execute(
            '''UPDATE contracts SET title=?, contractor=?, contract_type=?, value=?,
               status=?, notes=?, contract_date=?, updated_at=datetime('now','localtime') WHERE id=?''',
            (title, contractor, contract_type, value, status, notes, contract_date, cid)
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/contracts/<int:cid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_contracts_delete(cid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM contracts WHERE id = ?', (cid,)).fetchone()
    if row and row['file_path']:
        fpath = os.path.join(CONTRACTS_DIR, row['file_path'])
        if os.path.exists(fpath):
            os.remove(fpath)
    conn.execute('DELETE FROM contracts WHERE id = ?', (cid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/contracts/<int:cid>/file')
@api_role_required('owner', 'admin', 'project_manager')
def api_contracts_file(cid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM contracts WHERE id = ?', (cid,)).fetchone()
    conn.close()
    if not row or not row['file_path']:
        return 'Not found', 404
    fpath = os.path.join(CONTRACTS_DIR, row['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    return send_file(fpath, mimetype='application/pdf')

@app.route('/api/contracts/<int:cid>/review', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_contracts_review(cid):
    """AI review of a contract — compare against job bid data."""
    conn = get_db()
    contract = conn.execute('SELECT * FROM contracts WHERE id = ?', (cid,)).fetchone()
    if not contract:
        conn.close()
        return jsonify({'error': 'Contract not found'}), 404

    # Get bid data for comparison
    bid = conn.execute(
        'SELECT * FROM bids WHERE job_id = ? ORDER BY id DESC LIMIT 1', (contract['job_id'],)
    ).fetchone()

    # Try to extract text from the PDF
    pdf_text = ''
    if contract['file_path']:
        fpath = os.path.join(CONTRACTS_DIR, contract['file_path'])
        if os.path.exists(fpath):
            try:
                import pdfplumber
                with pdfplumber.open(fpath) as pdf:
                    for page in pdf.pages[:20]:
                        pdf_text += (page.extract_text() or '') + '\n'
            except ImportError:
                pdf_text = '[PDF text extraction unavailable - install pdfplumber]'
            except Exception as e:
                pdf_text = f'[Error extracting PDF text: {str(e)}]'

    # Build structured review
    findings = []
    concerns = []
    recommendations = []
    key_terms = []

    # ── Value comparison against bid ──
    bid_total = 0
    if bid:
        bid_total = bid['total_price'] if 'total_price' in bid.keys() else 0
        if contract['value'] and bid_total:
            diff = contract['value'] - bid_total
            pct = (diff / bid_total * 100) if bid_total else 0
            key_terms.append(f"Contract Value: ${contract['value']:,.2f}")
            key_terms.append(f"Bid Total: ${bid_total:,.2f}")
            if abs(pct) > 10:
                concerns.append(f"Contract value differs from bid total by {pct:+.1f}% — significant variance requires review")
                findings.append({'type': 'warning', 'category': 'Value Comparison',
                    'message': f"Contract value (${contract['value']:,.2f}) differs from bid total (${bid_total:,.2f}) by {pct:+.1f}%"})
            elif abs(pct) > 5:
                concerns.append(f"Contract value differs from bid total by {pct:+.1f}%")
                findings.append({'type': 'warning', 'category': 'Value Comparison',
                    'message': f"Contract value (${contract['value']:,.2f}) differs from bid total (${bid_total:,.2f}) by {pct:+.1f}%"})
            else:
                findings.append({'type': 'ok', 'category': 'Value Comparison',
                    'message': f"Contract value (${contract['value']:,.2f}) aligns with bid total (${bid_total:,.2f}) within 5%"})
        elif contract['value']:
            key_terms.append(f"Contract Value: ${contract['value']:,.2f}")
    else:
        recommendations.append("No bid found for this job — add bid data for value comparison")
        findings.append({'type': 'info', 'category': 'Bid Data',
            'message': 'No bid found for this job — cannot compare contract value to bid'})

    # ── Contract metadata checks ──
    if contract['contractor']:
        key_terms.append(f"Contractor: {contract['contractor']}")
    if contract['contract_type']:
        key_terms.append(f"Type: {contract['contract_type']}")
    if contract['status']:
        key_terms.append(f"Status: {contract['status']}")

    # ── PDF clause analysis ──
    clause_checks = [
        ('scope of work', 'Scope of Work', 'Defines the specific work to be performed'),
        ('payment terms', 'Payment Terms', 'Outlines how and when payments are made'),
        ('change order', 'Change Order Process', 'Defines procedure for contract modifications'),
        ('termination', 'Termination Clause', 'Conditions under which contract can be ended'),
        ('insurance', 'Insurance Requirements', 'Required coverage types and amounts'),
        ('warranty', 'Warranty Provisions', 'Post-completion warranty obligations'),
        ('liquidated damages', 'Liquidated Damages', 'Penalties for delayed completion'),
        ('retainage', 'Retainage', 'Percentage withheld until project completion'),
        ('indemnif', 'Indemnification', 'Liability protection and hold-harmless clauses'),
        ('dispute', 'Dispute Resolution', 'Process for resolving disagreements'),
        ('lien', 'Lien Waiver', 'Mechanic\'s lien release requirements'),
        ('safety', 'Safety Requirements', 'Jobsite safety obligations'),
    ]

    clauses_found = []
    clauses_missing = []

    if pdf_text and not pdf_text.startswith('['):
        text_lower = pdf_text.lower()
        for keyword, label, desc in clause_checks:
            if keyword in text_lower:
                clauses_found.append(label)
                findings.append({'type': 'ok', 'category': 'Clause Check', 'message': f'{label} — {desc}'})
            else:
                clauses_missing.append(label)
                findings.append({'type': 'warning', 'category': 'Clause Check', 'message': f'{label} NOT found — {desc}'})

        if clauses_missing:
            for clause in clauses_missing[:3]:
                concerns.append(f"Missing clause: {clause}")
            if len(clauses_missing) > 3:
                concerns.append(f"...and {len(clauses_missing) - 3} more missing clauses")
            recommendations.append("Review contract for missing standard clauses listed above")
        if clauses_found:
            recommendations.append(f"{len(clauses_found)} of {len(clause_checks)} standard clauses found in contract")
    elif not contract['file_path']:
        recommendations.append("Upload a contract PDF for full clause analysis")
        findings.append({'type': 'info', 'category': 'Document', 'message': 'No PDF uploaded — upload a contract document for full review'})
    elif pdf_text.startswith('['):
        recommendations.append(pdf_text.strip('[]'))

    # ── Risk assessment ──
    risk_score = len(concerns)
    if risk_score == 0:
        risk_level = 'Low'
    elif risk_score <= 2:
        risk_level = 'Medium'
    else:
        risk_level = 'High'

    # ── Summary ──
    summary_parts = []
    summary_parts.append(f"Contract \"{contract['title']}\" for {contract['contractor'] or 'unknown contractor'}")
    if contract['value']:
        summary_parts.append(f"valued at ${contract['value']:,.2f}")
    if clauses_found:
        summary_parts.append(f"{len(clauses_found)}/{len(clause_checks)} standard clauses detected")
    if concerns:
        summary_parts.append(f"{len(concerns)} concern(s) identified")
    else:
        summary_parts.append("no major concerns identified")

    review = {
        'reviewed_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'contract_title': contract['title'],
        'contract_value': contract['value'],
        'bid_total': bid_total,
        'pdf_extracted': bool(pdf_text and not pdf_text.startswith('[')),
        'summary': ' — '.join(summary_parts) + '.',
        'risk_level': risk_level,
        'key_terms': key_terms,
        'concerns': concerns,
        'recommendations': recommendations,
        'findings': findings,
    }

    if not findings:
        findings.append({'type': 'info', 'category': 'General', 'message': 'Review complete — no issues found'})

    # Save review to contract
    conn.execute('UPDATE contracts SET ai_review = ? WHERE id = ?', (json.dumps(review), cid))
    conn.commit()
    conn.close()

    return jsonify(review)


# ─── Lien Waivers ────────────────────────────────────────────────

LIEN_WAIVERS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'lien_waivers')
os.makedirs(LIEN_WAIVERS_DIR, exist_ok=True)

@app.route('/lien-waivers')
@role_required('owner', 'admin', 'project_manager')
def lien_waivers_page():
    return render_template('lien_waivers/list.html')

@app.route('/api/lien-waivers')
@api_role_required('owner', 'admin', 'project_manager')
def api_lien_waivers_list():
    job_id = request.args.get('job_id', type=int)
    status = request.args.get('status', '')
    wtype = request.args.get('type', '')
    conn = get_db()
    sql = '''SELECT lw.*, j.name as job_name,
             pa.application_number as pay_app_number, pa.period_to as pay_app_period
             FROM lien_waivers lw
             JOIN jobs j ON lw.job_id = j.id
             LEFT JOIN pay_applications pa ON lw.pay_app_id = pa.id
             WHERE 1=1'''
    params = []
    if job_id:
        sql += ' AND lw.job_id = ?'
        params.append(job_id)
    if status:
        sql += ' AND lw.status = ?'
        params.append(status)
    if wtype:
        sql += ' AND lw.waiver_type = ?'
        params.append(wtype)
    sql += ' ORDER BY lw.created_at DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d['has_file'] = bool(d.get('file_path'))
        d['has_pdf'] = bool(d.get('proposal_file'))
        result.append(d)
    return jsonify(result)

@app.route('/api/lien-waivers', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_lien_waivers_create():
    data = request.get_json(force=True)
    job_id = data.get('job_id')
    if not job_id:
        return jsonify({'error': 'job_id is required'}), 400

    conn = get_db()

    # Auto-number per job
    max_num = conn.execute(
        'SELECT MAX(waiver_number) FROM lien_waivers WHERE job_id = ?', (job_id,)
    ).fetchone()[0]
    waiver_number = (max_num or 0) + 1

    # Auto-fill from job and contract data
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    contract = conn.execute(
        'SELECT * FROM contracts WHERE job_id = ? AND status = ? ORDER BY value DESC LIMIT 1',
        (job_id, 'Active')
    ).fetchone()
    customer = None
    if job and job['customer_id']:
        customer = conn.execute('SELECT * FROM customers WHERE id = ?', (job['customer_id'],)).fetchone()

    contract_amount = float(data.get('contract_amount', 0))
    if not contract_amount and contract:
        contract_amount = contract['value'] or 0

    against_company = data.get('against_company', '')
    if not against_company and customer:
        against_company = customer['company_name'] or ''

    premises = data.get('premises_description', '')
    if not premises and job:
        premises = job['name'] or ''

    state_val = data.get('state', '')
    if not state_val and job:
        state_val = job['state'] or ''

    # Auto-calculate previous payments from executed waivers
    prev_row = conn.execute(
        'SELECT COALESCE(SUM(current_payment), 0) FROM lien_waivers WHERE job_id = ? AND status = ?',
        (job_id, 'Executed')
    ).fetchone()
    auto_previous = prev_row[0] if prev_row else 0
    previous_payments = float(data.get('previous_payments', 0)) or auto_previous

    current_payment = float(data.get('current_payment', 0))
    contract_balance = contract_amount - previous_payments - current_payment

    pay_app_id = data.get('pay_app_id') or None

    cursor = conn.execute(
        '''INSERT INTO lien_waivers (job_id, waiver_number, waiver_type, waiver_date,
           title_company, file_number, state, county, contract_amount, previous_payments,
           current_payment, contract_balance, claimant, against_company, premises_description,
           through_date, signer_name, signer_title, status, notes, created_by, pay_app_id)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (job_id, waiver_number,
         data.get('waiver_type', 'Conditional Progress'),
         data.get('waiver_date', datetime.now().strftime('%Y-%m-%d')),
         data.get('title_company', ''),
         data.get('file_number', ''),
         state_val,
         data.get('county', ''),
         contract_amount,
         previous_payments,
         current_payment,
         contract_balance,
         data.get('claimant', 'LGHVAC Mechanical, LLC'),
         against_company,
         premises,
         data.get('through_date', ''),
         data.get('signer_name', ''),
         data.get('signer_title', ''),
         'Draft',
         data.get('notes', ''),
         session['user_id'],
         pay_app_id)
    )
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return jsonify({'ok': True, 'id': new_id, 'waiver_number': waiver_number}), 201

@app.route('/api/lien-waivers/<int:wid>')
@api_role_required('owner', 'admin', 'project_manager')
def api_lien_waivers_get(wid):
    conn = get_db()
    row = conn.execute(
        '''SELECT lw.*, j.name as job_name,
           pa.application_number as pay_app_number, pa.period_to as pay_app_period
           FROM lien_waivers lw
           JOIN jobs j ON lw.job_id = j.id
           LEFT JOIN pay_applications pa ON lw.pay_app_id = pa.id
           WHERE lw.id = ?''', (wid,)
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Lien waiver not found'}), 404
    d = dict(row)
    d['has_file'] = bool(d.get('file_path'))
    d['has_pdf'] = bool(d.get('proposal_file'))
    return jsonify(d)

@app.route('/api/lien-waivers/<int:wid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_lien_waivers_update(wid):
    data = request.get_json(force=True)
    conn = get_db()

    contract_amount = float(data.get('contract_amount', 0))
    previous_payments = float(data.get('previous_payments', 0))
    current_payment = float(data.get('current_payment', 0))
    contract_balance = contract_amount - previous_payments - current_payment

    pay_app_id = data.get('pay_app_id') or None

    conn.execute(
        '''UPDATE lien_waivers SET waiver_type=?, waiver_date=?, title_company=?,
           file_number=?, state=?, county=?, contract_amount=?, previous_payments=?,
           current_payment=?, contract_balance=?, claimant=?, against_company=?,
           premises_description=?, through_date=?, signer_name=?, signer_title=?,
           status=?, notes=?, pay_app_id=?, updated_at=datetime('now','localtime') WHERE id=?''',
        (data.get('waiver_type', 'Conditional Progress'),
         data.get('waiver_date', ''),
         data.get('title_company', ''),
         data.get('file_number', ''),
         data.get('state', ''),
         data.get('county', ''),
         contract_amount,
         previous_payments,
         current_payment,
         contract_balance,
         data.get('claimant', 'LGHVAC Mechanical, LLC'),
         data.get('against_company', ''),
         data.get('premises_description', ''),
         data.get('through_date', ''),
         data.get('signer_name', ''),
         data.get('signer_title', ''),
         data.get('status', 'Draft'),
         data.get('notes', ''),
         pay_app_id,
         wid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/lien-waivers/<int:wid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_lien_waivers_delete(wid):
    conn = get_db()
    row = conn.execute('SELECT status, file_path, proposal_file FROM lien_waivers WHERE id = ?', (wid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if row['status'] != 'Draft':
        conn.close()
        return jsonify({'error': 'Only Draft waivers can be deleted'}), 400
    # Clean up files
    if row['file_path']:
        fpath = os.path.join(LIEN_WAIVERS_DIR, row['file_path'])
        if os.path.exists(fpath):
            os.remove(fpath)
    if row['proposal_file']:
        proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
        fpath = os.path.join(proposals_dir, row['proposal_file'])
        if os.path.exists(fpath):
            os.remove(fpath)
    conn.execute('DELETE FROM lien_waivers WHERE id = ?', (wid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/lien-waivers/<int:wid>/generate', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_lien_waivers_generate(wid):
    conn = get_db()
    waiver = conn.execute('SELECT * FROM lien_waivers WHERE id = ?', (wid,)).fetchone()
    if not waiver:
        conn.close()
        return jsonify({'error': 'Lien waiver not found'}), 404

    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (waiver['job_id'],)).fetchone()
    conn.close()

    waiver = dict(waiver)
    job = dict(job) if job else {}
    logo_path = os.path.abspath(os.path.join(app.static_folder, 'logo.jpg'))

    html = render_template('lien_waivers/waiver_pdf.html',
                           waiver=waiver, job=job,
                           logo_path='file://' + logo_path)

    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    os.makedirs(proposals_dir, exist_ok=True)

    filename = f"LienWaiver_{waiver['job_id']}_{waiver['waiver_number']}.pdf"
    filepath = os.path.join(proposals_dir, filename)

    try:
        wp = weasyprint.HTML(string=html, base_url=os.path.dirname(__file__))
        wp.write_pdf(filepath)
    except Exception as e:
        return jsonify({'error': f'PDF generation failed: {str(e)[:200]}'}), 500

    # Store filename
    conn2 = get_db()
    conn2.execute('UPDATE lien_waivers SET proposal_file = ? WHERE id = ?', (filename, wid))
    conn2.commit()
    conn2.close()

    return jsonify({'ok': True, 'filename': filename,
                    'path': f'/api/lien-waivers/{wid}/pdf/{filename}'})

@app.route('/api/lien-waivers/<int:wid>/pdf/<filename>')
@api_role_required('owner', 'admin', 'project_manager')
def api_lien_waivers_pdf(wid, filename):
    proposals_dir = os.path.join(os.path.dirname(__file__), 'data', 'proposals')
    filepath = os.path.join(proposals_dir, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    if request.args.get('download'):
        return send_file(filepath, as_attachment=True, download_name=filename)
    return send_file(filepath, mimetype='application/pdf')

@app.route('/api/lien-waivers/<int:wid>/upload', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_lien_waivers_upload(wid):
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'No file provided'}), 400
    from werkzeug.utils import secure_filename
    fname = secure_filename(file.filename)
    fname = f"{int(datetime.now().timestamp())}_{fname}"
    original_name = file.filename
    file.save(os.path.join(LIEN_WAIVERS_DIR, fname))
    conn = get_db()
    conn.execute('UPDATE lien_waivers SET file_path = ?, original_filename = ?, updated_at = datetime(\'now\',\'localtime\') WHERE id = ?',
                 (fname, original_name, wid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'filename': fname})

@app.route('/api/lien-waivers/upload-new', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_lien_waivers_upload_new():
    """Create a draft lien waiver and attach an uploaded file in one step (for drag-and-drop)."""
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'No file provided'}), 400
    job_id = request.form.get('job_id')
    if not job_id:
        return jsonify({'error': 'Job is required'}), 400
    conn = get_db()
    job = conn.execute('SELECT name FROM jobs WHERE id = ?', (job_id,)).fetchone()
    job_name = job['name'] if job else ''
    pay_app_id = request.form.get('pay_app_id') or None
    waiver_type = request.form.get('waiver_type', 'Conditional Progress')
    if waiver_type not in ('Conditional Progress', 'Unconditional Progress', 'Conditional Final', 'Unconditional Final'):
        waiver_type = 'Conditional Progress'
    max_num = conn.execute('SELECT MAX(waiver_number) FROM lien_waivers WHERE job_id = ?', (job_id,)).fetchone()[0] or 0
    today = datetime.now().strftime('%Y-%m-%d')
    cursor = conn.execute(
        '''INSERT INTO lien_waivers (job_id, waiver_number, waiver_type, waiver_date, claimant,
           premises_description, status, created_by, pay_app_id, created_at) VALUES (?,?,?,?,?,?,?,?,?,datetime('now','localtime'))''',
        (job_id, max_num + 1, waiver_type, today, 'LGHVAC Mechanical, LLC',
         job_name, 'Draft', session['user_id'], pay_app_id)
    )
    wid = cursor.lastrowid
    from werkzeug.utils import secure_filename
    fname = secure_filename(file.filename)
    original_name = file.filename
    fname = f"{int(datetime.now().timestamp())}_{fname}"
    file.save(os.path.join(LIEN_WAIVERS_DIR, fname))
    conn.execute('UPDATE lien_waivers SET file_path = ?, original_filename = ? WHERE id = ?', (fname, original_name, wid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': wid})

@app.route('/api/lien-waivers/pay-apps-for-job/<int:job_id>')
@api_role_required('owner', 'admin', 'project_manager')
def api_lien_waivers_payapps_for_job(job_id):
    """Get all pay applications for a job (across all contracts) for lien waiver dropdown."""
    conn = get_db()
    rows = conn.execute(
        '''SELECT pa.id, pa.application_number, pa.period_to, pa.status,
           pac.gc_name, pac.project_name
           FROM pay_applications pa
           JOIN pay_app_contracts pac ON pa.contract_id = pac.id
           WHERE pac.job_id = ?
           ORDER BY pa.application_number''',
        (job_id,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/lien-waivers/<int:wid>/file')
@api_role_required('owner', 'admin', 'project_manager')
def api_lien_waivers_file(wid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM lien_waivers WHERE id = ?', (wid,)).fetchone()
    conn.close()
    if not row or not row['file_path']:
        return 'Not found', 404
    fpath = os.path.join(LIEN_WAIVERS_DIR, row['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    return send_file(fpath, mimetype='application/pdf')


# ─── Workflow Dashboard ─────────────────────────────────────────

@app.route('/workflow')
@role_required('owner', 'admin', 'project_manager')
def workflow_page():
    return render_template('workflow/dashboard.html')

@app.route('/api/workflow/summary')
@api_role_required('owner', 'admin', 'project_manager')
def api_workflow_summary():
    job_id = request.args.get('job_id', type=int)
    conn = get_db()

    # Get jobs
    if job_id:
        jobs = conn.execute('SELECT id, name, status FROM jobs WHERE id = ?', (job_id,)).fetchall()
    else:
        jobs = conn.execute("SELECT id, name, status FROM jobs WHERE status NOT IN ('Complete','Cancelled') ORDER BY name").fetchall()

    job_ids = [j['id'] for j in jobs]
    if not job_ids:
        conn.close()
        return jsonify({'summary': {}, 'jobs': []})

    placeholders = ','.join('?' * len(job_ids))

    # Aggregate RFIs
    rfis = conn.execute(
        f'SELECT job_id, status, COUNT(*) as cnt FROM rfis WHERE job_id IN ({placeholders}) GROUP BY job_id, status',
        job_ids
    ).fetchall()

    # Aggregate Submittals
    submittals = conn.execute(
        f'SELECT job_id, status, COUNT(*) as cnt FROM submittals WHERE job_id IN ({placeholders}) GROUP BY job_id, status',
        job_ids
    ).fetchall()

    # Aggregate Change Orders
    change_orders = conn.execute(
        f'SELECT job_id, status, COUNT(*) as cnt FROM change_orders WHERE job_id IN ({placeholders}) GROUP BY job_id, status',
        job_ids
    ).fetchall()

    # Aggregate Schedule Events
    schedule = conn.execute(
        f'SELECT job_id, status, COUNT(*) as cnt FROM job_schedule_events WHERE job_id IN ({placeholders}) GROUP BY job_id, status',
        job_ids
    ).fetchall()

    # Aggregate Pay Apps (through contracts)
    pay_apps = conn.execute(
        f'''SELECT pac.job_id, pa.status, COUNT(*) as cnt
            FROM pay_applications pa
            JOIN pay_app_contracts pac ON pa.contract_id = pac.id
            WHERE pac.job_id IN ({placeholders})
            GROUP BY pac.job_id, pa.status''',
        job_ids
    ).fetchall()

    # Aggregate Closeout Checklists
    closeout = conn.execute(
        f'SELECT job_id, status, COUNT(*) as cnt FROM closeout_checklists WHERE job_id IN ({placeholders}) GROUP BY job_id, status',
        job_ids
    ).fetchall()

    # Aggregate Contracts
    contracts = conn.execute(
        f'SELECT job_id, status, COUNT(*) as cnt FROM contracts WHERE job_id IN ({placeholders}) GROUP BY job_id, status',
        job_ids
    ).fetchall()

    conn.close()

    # Build per-job and summary dicts
    def build_status_dict(rows):
        result = {}
        for row in rows:
            jid = row['job_id']
            if jid not in result:
                result[jid] = {}
            result[jid][row['status']] = row['cnt']
        return result

    rfi_map = build_status_dict(rfis)
    sub_map = build_status_dict(submittals)
    co_map = build_status_dict(change_orders)
    sched_map = build_status_dict(schedule)
    pa_map = build_status_dict(pay_apps)
    cl_map = build_status_dict(closeout)
    ct_map = build_status_dict(contracts)

    # Build summary totals
    def merge_totals(status_map):
        totals = {}
        for jid, statuses in status_map.items():
            for s, c in statuses.items():
                totals[s] = totals.get(s, 0) + c
        return totals

    summary = {
        'rfis': merge_totals(rfi_map),
        'submittals': merge_totals(sub_map),
        'change_orders': merge_totals(co_map),
        'schedule': merge_totals(sched_map),
        'pay_apps': merge_totals(pa_map),
        'closeout': merge_totals(cl_map),
        'contracts': merge_totals(ct_map),
    }

    # Build per-job data
    job_data = []
    for j in jobs:
        jid = j['id']
        job_data.append({
            'id': jid,
            'name': j['name'],
            'status': j['status'],
            'rfis': rfi_map.get(jid, {}),
            'submittals': sub_map.get(jid, {}),
            'change_orders': co_map.get(jid, {}),
            'schedule': sched_map.get(jid, {}),
            'pay_apps': pa_map.get(jid, {}),
            'closeout': cl_map.get(jid, {}),
            'contracts': ct_map.get(jid, {}),
        })

    return jsonify({'summary': summary, 'jobs': job_data})


# ─── BillTrust Integration ──────────────────────────────────────

@app.route('/api/billtrust/config')
@api_role_required('owner', 'admin')
def api_billtrust_config():
    conn = get_db()
    configs = conn.execute('SELECT id, supplier_name, is_active, use_mock, last_sync_at FROM billtrust_config ORDER BY supplier_name').fetchall()
    conn.close()
    return jsonify([dict(c) for c in configs])

@app.route('/api/billtrust/config/<int:cid>', methods=['PUT'])
@api_role_required('owner', 'admin')
def api_billtrust_config_update(cid):
    data = request.get_json()
    conn = get_db()
    conn.execute(
        '''UPDATE billtrust_config SET client_id=?, client_secret=?, is_active=?, use_mock=?,
           updated_at=datetime('now','localtime') WHERE id=?''',
        (data.get('client_id', ''), data.get('client_secret', ''),
         data.get('is_active', 1), data.get('use_mock', 1), cid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/billtrust/sync/<int:config_id>', methods=['POST'])
@api_role_required('owner', 'admin')
def api_billtrust_sync(config_id):
    conn = get_db()
    try:
        from billtrust import get_client_for_supplier
        client = get_client_for_supplier(conn, config_id)
        if not client:
            conn.close()
            return jsonify({'error': 'Supplier config not found'}), 404

        result = client.sync_invoices(conn, config_id)
        conn.execute(
            "UPDATE billtrust_config SET last_sync_at = datetime('now','localtime') WHERE id = ?",
            (config_id,)
        )
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

    conn.close()
    return jsonify(result)

@app.route('/api/billtrust/invoices')
@api_role_required('owner', 'admin')
def api_billtrust_invoices():
    supplier_id = request.args.get('supplier_id', type=int)
    status = request.args.get('status', '')
    conn = get_db()
    sql = '''SELECT si.*, bc.supplier_name FROM supplier_invoices si
             JOIN billtrust_config bc ON si.supplier_config_id = bc.id WHERE 1=1'''
    params = []
    if supplier_id:
        sql += ' AND si.supplier_config_id = ?'
        params.append(supplier_id)
    if status:
        sql += ' AND LOWER(si.status) = LOWER(?)'
        params.append(status)
    sql += ' ORDER BY si.invoice_date DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/billtrust/invoices/<int:iid>/link', methods=['PUT'])
@api_role_required('owner', 'admin')
def api_billtrust_invoice_link(iid):
    """Link a supplier invoice to a job."""
    data = request.get_json()
    conn = get_db()
    conn.execute('UPDATE supplier_invoices SET job_id = ? WHERE id = ?', (data.get('job_id'), iid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/billtrust/test/<int:config_id>', methods=['POST'])
@api_role_required('owner', 'admin')
def api_billtrust_test(config_id):
    conn = get_db()
    try:
        from billtrust import test_connection
        result = test_connection(conn, config_id)
        conn.close()
        return jsonify(result)
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/billtrust/detect-pdf', methods=['POST'])
@api_role_required('owner', 'admin')
def api_billtrust_detect_pdf():
    """Check if a PDF is a BillTrust supplier invoice by reading page 1."""
    pdf_file = request.files.get('pdf_file')
    if not pdf_file:
        return jsonify({'is_billtrust': False})
    try:
        import pdfplumber, io
        content = pdf_file.read()
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            if not pdf.pages:
                return jsonify({'is_billtrust': False})
            text = (pdf.pages[0].extract_text() or '').upper()
        # Check for known supplier names
        if 'LOCKE SUPPLY' in text:
            return jsonify({'is_billtrust': True, 'supplier_name': 'Locke Supply'})
        if 'PLUMB SUPPLY' in text:
            return jsonify({'is_billtrust': True, 'supplier_name': 'Plumb Supply'})
        return jsonify({'is_billtrust': False})
    except Exception:
        return jsonify({'is_billtrust': False})

@app.route('/api/billtrust/import', methods=['POST'])
@api_role_required('owner', 'admin')
def api_billtrust_import():
    """Import invoices from BillTrust CSV + PDF export files."""
    csv_file = request.files.get('csv_file')
    pdf_file = request.files.get('pdf_file')
    supplier_name = request.form.get('supplier_name', 'Locke Supply')

    if not csv_file and not pdf_file:
        return jsonify({'error': 'CSV or PDF file is required'}), 400

    job_id = request.form.get('job_id', '').strip() or None
    if job_id:
        try:
            job_id = int(job_id)
        except ValueError:
            job_id = None

    csv_content = csv_file.read() if csv_file else None
    pdf_content = pdf_file.read() if pdf_file else None

    conn = get_db()
    try:
        # Look up supplier config
        config = conn.execute(
            'SELECT id FROM billtrust_config WHERE supplier_name = ?', (supplier_name,)
        ).fetchone()
        if not config:
            conn.close()
            return jsonify({'error': f'Supplier "{supplier_name}" not found in config'}), 404

        supplier_config_id = config['id']

        from invoice_import import import_billtrust_files
        result = import_billtrust_files(csv_content, pdf_content, supplier_config_id, conn, job_id=job_id)

        # Log to sync log
        stats = result.get('stats', {})
        conn.execute('''
            INSERT INTO billtrust_sync_log (config_id, sync_type, invoices_found, invoices_new, invoices_updated, errors)
            VALUES (?, 'csv_pdf_import', ?, ?, ?, ?)
        ''', (supplier_config_id, stats.get('total', 0), stats.get('new', 0),
              stats.get('updated', 0), json.dumps([])))
        conn.commit()

        # Notify other owners/admins
        current_user = session.get('user_id')
        admins = conn.execute(
            "SELECT id FROM users WHERE role IN ('owner','admin') AND id != ?",
            (current_user,)
        ).fetchall()
        conn.close()

        for admin in admins:
            create_notification(
                admin['id'], 'invoice',
                f'BillTrust Import: {stats.get("new",0)} new, {stats.get("updated",0)} updated',
                f'{supplier_name} — {stats.get("total",0)} invoices imported via CSV/PDF',
                '/invoices'
            )

        return jsonify(result)
    except Exception as e:
        conn.close()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ─── Plans ─────────────────────────────────────────────────────

PLANS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'plans')
os.makedirs(PLANS_DIR, exist_ok=True)

# ── Takeoff Template (matches Takeoff_Template.xlsm "Program" tab) ──
TAKEOFF_TEMPLATE = {
    "sections": [
        {
            "name": "Rough - In",
            "items": [
                # ── CRDs / Bath Exhaust ──
                {"part": "12x14 CRD", "sku": "I-CRD50", "category": "CRD", "price": 43.00, "default_use": True},
                {"part": "12x14 Fire/Smoke Radiation Damper", "sku": "FSD-111-1", "category": "CRD", "price": 608.00, "default_use": False},
                {"part": "6x12x8 CRD Boot", "sku": "50CRD", "category": "CRD", "price": 35.00, "default_use": True},
                {"part": "4\" Round CRD", "sku": "55CRD 4\"", "category": "CRD", "price": 32.00, "default_use": True},
                {"part": "6\" Round CRD", "sku": "55CRD 6\"", "category": "CRD", "price": 38.00, "default_use": False},
                {"part": "80CFM CRD Fan", "sku": "QTXE080-22016845", "category": "CRD", "price": 100.50, "default_use": True},
                {"part": "Broan 688 Exhaust Fan", "sku": "688-22019236", "category": "Exhaust Fan", "price": 18.45, "default_use": False},
                # ── Round Pipe (workbook) ──
                {"part": "4\" Adjustable 90", "sku": "L0121", "category": "Round Pipe", "price": 4.50, "default_use": False},
                {"part": "3\" Adjustable 90", "sku": "L0120", "category": "Round Pipe", "price": 3.75, "default_use": False},
                {"part": "4x3 Reducer", "sku": "L0285", "category": "Round Pipe", "price": 3.50, "default_use": False},
                {"part": "3\" Conductor Pipe", "sku": "L0463", "category": "Round Pipe", "price": 8.00, "default_use": False},
                {"part": "4\" Conductor Pipe", "sku": "L0464", "category": "Round Pipe", "price": 10.00, "default_use": False},
                # ── Boots ──
                {"part": "8\" Foam Boot", "sku": "L7001", "category": "Boots", "price": 16.06, "default_use": True},
                {"part": "6\" Foam Boot", "sku": "608 R8 6x6x4", "category": "Boots", "price": 12.00, "default_use": True},
                {"part": "8x6 Reducer", "sku": "L0292", "category": "Duct Adapter", "price": 5.80, "default_use": False},
                # ── Finger Taps (1 per drop, sized to match) ──
                {"part": "6\" Finger Taps", "sku": "L0090", "category": "Duct Adapter", "price": 1.30, "default_use": True},
                {"part": "8\" Finger Taps", "sku": "L0092", "category": "Duct Adapter", "price": 1.65, "default_use": True},
                {"part": "10\" Finger Taps", "sku": "L0094", "category": "Duct Adapter", "price": 2.10, "default_use": False},
                # ── Supply Boots (1 per drop, sized to match) ──
                {"part": "6\" Supply Boot", "sku": "L0095", "category": "Duct Adapter", "price": 4.50, "default_use": True},
                {"part": "8\" Supply Boot", "sku": "L0096", "category": "Duct Adapter", "price": 5.75, "default_use": True},
                {"part": "10\" Supply Boot", "sku": "L0097", "category": "Duct Adapter", "price": 7.25, "default_use": False},
                # ── Vent Boxes (workbook XVENT items) ──
                {"part": "4\" Triple Brick Vent Box", "sku": "XVENT-4TB", "category": "Vent Box", "price": 18.00, "default_use": False},
                {"part": "4\" Single Brick Vent Box", "sku": "XVENT-4SB", "category": "Vent Box", "price": 12.00, "default_use": False},
                {"part": "4\" Single Siding Vent Box", "sku": "XVENT-4SS", "category": "Vent Box", "price": 10.00, "default_use": False},
                {"part": "4\" Single Soffit Vent Box", "sku": "XVENT-4SF", "category": "Vent Box", "price": 10.00, "default_use": False},
                {"part": "6\" Single Brick Vent Box", "sku": "XVENT-6SB", "category": "Vent Box", "price": 14.00, "default_use": False},
                {"part": "6\" Single Siding Vent Box", "sku": "XVENT-6SS", "category": "Vent Box", "price": 12.00, "default_use": False},
                {"part": "6\" Single Soffit Vent Box", "sku": "XVENT-6SF", "category": "Vent Box", "price": 12.00, "default_use": False},
                {"part": "4\" Triple Siding Vent Box", "sku": "XVENT-4TS", "category": "Vent Box", "price": 16.00, "default_use": False},
                {"part": "4\" Triple Soffit Vent Box", "sku": "XVENT-4TF", "category": "Vent Box", "price": 16.00, "default_use": False},
                # ── Flex Duct R6 (lower floors) ──
                {"part": "6\" R6 Flex (bags)", "sku": "L1972", "category": "Flex", "price": 31.46, "default_use": True},
                {"part": "8\" R6 Flex (bags)", "sku": "L1974", "category": "Flex", "price": 37.56, "default_use": True},
                {"part": "10\" R6 Flex (bags)", "sku": "L1976", "category": "Flex", "price": 48.50, "default_use": False},
                {"part": "12\" R6 Flex (bags)", "sku": "L1978", "category": "Flex", "price": 55.20, "default_use": False},
                # ── Flex Duct R8 (top floor - IRC code) ──
                {"part": "6\" R8 Flex (bags)", "sku": "L1939", "category": "Flex", "price": 38.50, "default_use": False},
                {"part": "8\" R8 Flex (bags)", "sku": "L1941", "category": "Flex", "price": 45.00, "default_use": False},
                {"part": "10\" R8 Flex (bags)", "sku": "L1986", "category": "Flex", "price": 56.00, "default_use": False},
                {"part": "12\" R8 Flex (bags)", "sku": "L1988", "category": "Flex", "price": 63.50, "default_use": False},
                # ── Line Sets ──
                {"part": "3/8 Line Set (in feet)", "sku": "H0716", "category": "Line Set", "price": 57.00, "default_use": True},
                {"part": "3/4 Line Set (in feet)", "sku": "H0719", "category": "Line Set", "price": 130.00, "default_use": True},
                {"part": "7/8 Line Set (in feet)", "sku": "H0720", "category": "Line Set", "price": 165.00, "default_use": False},
                {"part": "1-1/8 Line Set (in feet)", "sku": "H0722", "category": "Line Set", "price": 210.00, "default_use": False},
                {"part": "Mini Split Line Set", "sku": "", "category": "Line Set", "price": 85.00, "default_use": False},
                # ── Line Set Insulation ──
                {"part": "3/4\" Armaflex (6ft stick)", "sku": "L0484", "category": "Insulation", "price": 8.50, "default_use": True},
                {"part": "7/8\" Armaflex (6ft stick)", "sku": "H0752", "category": "Insulation", "price": 10.25, "default_use": False},
                {"part": "1-1/8\" Armaflex (6ft stick)", "sku": "H0754", "category": "Insulation", "price": 13.50, "default_use": False},
                {"part": "Armaflex Glue (qt)", "sku": "H0760", "category": "Insulation", "price": 22.00, "default_use": True},
                {"part": "3/4\" Fiberglass Pipe Wrap (roll)", "sku": "H0770", "category": "Insulation", "price": 14.00, "default_use": False},
                {"part": "7/8\" Fiberglass Pipe Wrap (roll)", "sku": "H0772", "category": "Insulation", "price": 16.50, "default_use": False},
                {"part": "1-1/8\" Fiberglass Pipe Wrap (roll)", "sku": "H0774", "category": "Insulation", "price": 19.00, "default_use": False},
                {"part": "Duct Wrap", "sku": "", "category": "Insulation", "price": 45.00, "default_use": False},
                # ── Condensate ──
                {"part": "3/4\" PVC (10ft stick)", "sku": "R0033", "category": "Condensate", "price": 4.50, "default_use": True},
                {"part": "3/4\" CPVC (10ft stick)", "sku": "P0342", "category": "Condensate", "price": 6.25, "default_use": False},
                {"part": "3/4\" PVC Fittings (bag)", "sku": "P0345", "category": "Condensate", "price": 12.00, "default_use": True},
                {"part": "3/4\" CPVC Fittings (bag)", "sku": "P0347", "category": "Condensate", "price": 15.00, "default_use": False},
                {"part": "PVC Cement & Primer Kit", "sku": "P0350", "category": "Condensate", "price": 8.75, "default_use": True},
                {"part": "CPVC Cement & Primer Kit", "sku": "P0352", "category": "Condensate", "price": 11.50, "default_use": False},
                {"part": "Condensate Trap", "sku": "P0360", "category": "Condensate", "price": 6.50, "default_use": True},
                # ── Hangers & Hardware ──
                {"part": "Metal Hanger Strap (100ft roll)", "sku": "L0200", "category": "Hangers", "price": 18.50, "default_use": True},
                {"part": "1\" Metal Screws (box)", "sku": "L0210", "category": "Hangers", "price": 12.00, "default_use": True},
                {"part": "Threaded Rod 3/8x36\"", "sku": "L0220", "category": "Hangers", "price": 3.25, "default_use": True},
                {"part": "Beam Clamps (box of 25)", "sku": "L0225", "category": "Hangers", "price": 28.00, "default_use": False},
                {"part": "3/4 Screws", "sku": "M0044", "category": "Hardware", "price": 8.00, "default_use": False},
                {"part": "Rails", "sku": "L0158", "category": "Hardware", "price": 12.00, "default_use": False},
                {"part": "Plumber Strap", "sku": "M0091", "category": "Hardware", "price": 6.50, "default_use": False},
                {"part": "4.25\" Dryer Box", "sku": "L3526", "category": "Hardware", "price": 14.00, "default_use": False},
                {"part": "16\" Boca Plate", "sku": "M5075", "category": "Hardware", "price": 3.50, "default_use": False},
                # ── Sealant / Tape ──
                {"part": "Duct Mastic (gal)", "sku": "M6003", "category": "Sealant", "price": 15.00, "default_use": True},
                {"part": "Mastic Tape (roll)", "sku": "L0305", "category": "Sealant", "price": 8.50, "default_use": True},
                {"part": "Foil Tape (roll)", "sku": "L0494", "category": "Sealant", "price": 7.25, "default_use": True},
                {"part": "Black Duct Tape", "sku": "L0444", "category": "Sealant", "price": 6.00, "default_use": False},
                {"part": "Flex Fix Tape", "sku": "", "category": "Sealant", "price": 8.00, "default_use": False},
                {"part": "Silicone", "sku": "M1842", "category": "Sealant", "price": 7.50, "default_use": False},
                {"part": "Fire Caulk 5 Gal", "sku": "M1346", "category": "Sealant", "price": 85.00, "default_use": False},
                {"part": "Paint Brush", "sku": "M0725", "category": "Sealant", "price": 3.00, "default_use": False},
                # ── Electrical (workbook items first) ──
                {"part": "14/4 S.O.Cable", "sku": "P3833", "category": "Electrical", "price": 2.50, "default_use": False},
                {"part": "18/8 T-stat Wire per foot", "sku": "P2240", "category": "Electrical", "price": 0.35, "default_use": False},
                {"part": "18/5 Thermostat Wire (250ft)", "sku": "E0050", "category": "Electrical", "price": 42.00, "default_use": True},
                {"part": "18/8 Thermostat Wire (250ft)", "sku": "E0055", "category": "Electrical", "price": 58.00, "default_use": False},
                {"part": "Low Voltage Wire 18/2 (250ft)", "sku": "E0100", "category": "Electrical", "price": 35.00, "default_use": True},
                {"part": "Disconnect 60A Non-Fused", "sku": "E0200", "category": "Electrical", "price": 18.50, "default_use": True},
                {"part": "Disconnect 60A Fused", "sku": "E0205", "category": "Electrical", "price": 28.00, "default_use": False},
                {"part": "Whip 3/4\" x 6ft", "sku": "E0210", "category": "Electrical", "price": 22.00, "default_use": True},
                {"part": "Whip 1\" x 6ft", "sku": "E0215", "category": "Electrical", "price": 32.00, "default_use": False},
                {"part": "Wire Nuts / Connectors (box)", "sku": "P1491", "category": "Electrical", "price": 8.00, "default_use": True},
                {"part": "10-2 Romex (250ft)", "sku": "E0300", "category": "Electrical", "price": 145.00, "default_use": False},
                {"part": "6-2 Romex (125ft)", "sku": "E0310", "category": "Electrical", "price": 185.00, "default_use": False},
                {"part": "8-3 Romex (125ft)", "sku": "E0320", "category": "Electrical", "price": 165.00, "default_use": False},
                # ── Return Air ──
                {"part": "14x8 Return Air Boot", "sku": "L0400", "category": "Return Air", "price": 18.00, "default_use": True},
                {"part": "16x8 Return Air Boot", "sku": "L0402", "category": "Return Air", "price": 20.00, "default_use": False},
                {"part": "20x8 Return Air Boot", "sku": "L0404", "category": "Return Air", "price": 24.00, "default_use": False},
                {"part": "20x20 Return Air Box", "sku": "L0410", "category": "Return Air", "price": 35.00, "default_use": True},
                {"part": "20x25 Return Air Box", "sku": "L0412", "category": "Return Air", "price": 40.00, "default_use": False},
                {"part": "24x24 Return Air Box", "sku": "L0414", "category": "Return Air", "price": 45.00, "default_use": False},
                {"part": "14x20 Return Grille", "sku": "L0418", "category": "Return Air", "price": 16.00, "default_use": False},
                {"part": "20x20 Return Grille", "sku": "L0419", "category": "Return Air", "price": 18.00, "default_use": False},
                {"part": "20x25 Filter Grille", "sku": "L0420", "category": "Return Air", "price": 22.00, "default_use": True},
                {"part": "16x25 Filter Grille", "sku": "L0422", "category": "Return Air", "price": 20.00, "default_use": False},
                {"part": "20x20 Filter Grille", "sku": "L0424", "category": "Return Air", "price": 20.00, "default_use": False},
                {"part": "24x24 Filter Grille", "sku": "L0426", "category": "Return Air", "price": 26.00, "default_use": False},
                # ── Plenum / Ductboard ──
                {"part": "Plenum Board (4x8 sheet)", "sku": "L0500", "category": "Plenum", "price": 32.00, "default_use": True},
                {"part": "Plenum Clips (box)", "sku": "L0505", "category": "Plenum", "price": 6.50, "default_use": True},
                {"part": "R8 Duct Board (4x10 sheet)", "sku": "L0470", "category": "Plenum", "price": 52.00, "default_use": False},
                {"part": "R6 Duct Board (4x10 sheet)", "sku": "L0512", "category": "Plenum", "price": 45.00, "default_use": False},
                {"part": "Ductboard Staples (box)", "sku": "L0515", "category": "Plenum", "price": 8.00, "default_use": False},
                # ── Gas ──
                {"part": "Gas Flex 1/2\" (per ft)", "sku": "G0100", "category": "Gas", "price": 3.50, "default_use": False},
                {"part": "Gas Flex 3/4\" (per ft)", "sku": "G0105", "category": "Gas", "price": 5.25, "default_use": False},
                {"part": "Gas Valve 1/2\"", "sku": "G0110", "category": "Gas", "price": 12.00, "default_use": False},
                {"part": "Gas Valve 3/4\"", "sku": "G0112", "category": "Gas", "price": 16.00, "default_use": False},
                {"part": "Gas Drip Leg Kit", "sku": "G0120", "category": "Gas", "price": 8.50, "default_use": False},
                {"part": "Gas Connector 1/2\" (36\")", "sku": "G0130", "category": "Gas", "price": 18.00, "default_use": False},
                {"part": "Gas Connector 3/4\" (36\")", "sku": "G0135", "category": "Gas", "price": 24.00, "default_use": False},
                {"part": "Gas Sediment Trap", "sku": "G0140", "category": "Gas", "price": 6.00, "default_use": False},
                # ── Fire Protection ──
                {"part": "Fire Wrap (1.5\" x 24\" x 25ft roll)", "sku": "FW0100", "category": "Fire Protection", "price": 85.00, "default_use": False},
                {"part": "SA Dryer Fire Wrap 16\"", "sku": "", "category": "Fire Protection", "price": 25.00, "default_use": False},
                {"part": "Fire Damper 6\"", "sku": "FW0110", "category": "Fire Protection", "price": 32.00, "default_use": False},
                {"part": "Fire Damper 8\"", "sku": "FW0112", "category": "Fire Protection", "price": 38.00, "default_use": False},
                {"part": "Fire Damper 10\"", "sku": "FW0114", "category": "Fire Protection", "price": 45.00, "default_use": False},
                {"part": "Passthrough Sleeve 6\"", "sku": "FW0120", "category": "Fire Protection", "price": 12.00, "default_use": False},
                {"part": "Passthrough Sleeve 8\"", "sku": "FW0122", "category": "Fire Protection", "price": 14.00, "default_use": False},
                {"part": "Passthrough Sleeve 10\"", "sku": "FW0124", "category": "Fire Protection", "price": 18.00, "default_use": False},
                {"part": "Firestop Caulk (tube)", "sku": "FW0130", "category": "Fire Protection", "price": 12.00, "default_use": False},
                {"part": "Firestop Putty Pad (box)", "sku": "FW0135", "category": "Fire Protection", "price": 35.00, "default_use": False},
                {"part": "Smoke Damper (round)", "sku": "FW0140", "category": "Fire Protection", "price": 65.00, "default_use": False},
                # ── Misc Rough-In (workbook) ──
                {"part": "3\" Gray Duct Strap", "sku": "", "category": "Misc Supplies", "price": 5.00, "default_use": False},
            ]
        },
        {
            "name": "Trim Out",
            "items": [
                # ── Shorts & Smalls ──
                {"part": "3\" Pump Ups", "sku": "L0681", "category": "Shorts & Smalls", "price": 3.80, "default_use": True},
                {"part": "6\" Pump Ups", "sku": "L3311", "category": "Shorts & Smalls", "price": 5.50, "default_use": False},
                {"part": "Safe T Switch", "sku": "L2507", "category": "Shorts & Smalls", "price": 17.70, "default_use": True},
                {"part": "Float Switch", "sku": "L2510", "category": "Shorts & Smalls", "price": 14.50, "default_use": True},
                {"part": "Drain Pan (plastic)", "sku": "L0700", "category": "Shorts & Smalls", "price": 18.00, "default_use": True},
                {"part": "Drain Pan (metal)", "sku": "L0705", "category": "Shorts & Smalls", "price": 25.00, "default_use": False},
                {"part": "30x30 Drain Pan", "sku": "", "category": "Shorts & Smalls", "price": 32.00, "default_use": False},
                {"part": "30x60 Drain Pan", "sku": "", "category": "Shorts & Smalls", "price": 55.00, "default_use": False},
                {"part": "P-Trap 3/4\" PVC", "sku": "L0680", "category": "Shorts & Smalls", "price": 3.50, "default_use": True},
                # ── PVC Fittings (workbook individual items) ──
                {"part": "3/4 PVC 90", "sku": "R0311", "category": "PVC Fittings", "price": 0.75, "default_use": False},
                {"part": "3/4 PVC Coupling", "sku": "R0341", "category": "PVC Fittings", "price": 0.50, "default_use": False},
                {"part": "3/4 PVC Male Adapter", "sku": "R0351", "category": "PVC Fittings", "price": 0.65, "default_use": False},
                {"part": "3/4 PVC Tee", "sku": "R0331", "category": "PVC Fittings", "price": 0.85, "default_use": False},
                {"part": "1QT PVC Cement", "sku": "R0042", "category": "PVC Fittings", "price": 8.00, "default_use": False},
                # ── Registers & Grilles (workbook items first) ──
                {"part": "8\" Supply Register", "sku": "L1731", "category": "Registers & Grilles", "price": 8.00, "default_use": False},
                {"part": "6\" Supply Register", "sku": "L1730", "category": "Registers & Grilles", "price": 6.50, "default_use": False},
                {"part": "12x6 Stamped Supply", "sku": "L1736", "category": "Registers & Grilles", "price": 7.25, "default_use": False},
                {"part": "16x8 Hart Cooley Pass Through", "sku": "L2280", "category": "Registers & Grilles", "price": 22.00, "default_use": False},
                {"part": "24x12 Stamped Return", "sku": "L1778", "category": "Registers & Grilles", "price": 14.00, "default_use": False},
                {"part": "30x14 Stamped Return", "sku": "L1791", "category": "Registers & Grilles", "price": 18.00, "default_use": False},
                {"part": "30x20 Stamped Return", "sku": "L9989", "category": "Registers & Grilles", "price": 22.00, "default_use": False},
                {"part": "24x24x8 Drop-ins", "sku": "L1928", "category": "Registers & Grilles", "price": 28.00, "default_use": False},
                {"part": "4\" Venthood Covers", "sku": "L3516", "category": "Registers & Grilles", "price": 8.00, "default_use": False},
                {"part": "6x6 Register (white)", "sku": "R0100", "category": "Registers & Grilles", "price": 5.50, "default_use": True},
                {"part": "8x4 Register (white)", "sku": "R0105", "category": "Registers & Grilles", "price": 5.50, "default_use": True},
                {"part": "10x4 Register (white)", "sku": "R0110", "category": "Registers & Grilles", "price": 6.25, "default_use": True},
                {"part": "10x6 Register (white)", "sku": "R0115", "category": "Registers & Grilles", "price": 6.75, "default_use": True},
                {"part": "12x4 Register (white)", "sku": "R0120", "category": "Registers & Grilles", "price": 6.50, "default_use": True},
                {"part": "12x6 Register (white)", "sku": "R0125", "category": "Registers & Grilles", "price": 7.25, "default_use": True},
                {"part": "14x6 Register (white)", "sku": "R0130", "category": "Registers & Grilles", "price": 8.00, "default_use": False},
                {"part": "14x8 Register (white)", "sku": "R0135", "category": "Registers & Grilles", "price": 9.00, "default_use": False},
                {"part": "6\" Round Ceiling Diffuser", "sku": "R0200", "category": "Registers & Grilles", "price": 9.50, "default_use": False},
                {"part": "8\" Round Ceiling Diffuser", "sku": "R0205", "category": "Registers & Grilles", "price": 11.00, "default_use": False},
                {"part": "10\" Round Ceiling Diffuser", "sku": "R0210", "category": "Registers & Grilles", "price": 13.00, "default_use": False},
                # ── Filters ──
                {"part": "20x25x1 Filter (12-pack)", "sku": "F0100", "category": "Filters", "price": 42.00, "default_use": True},
                {"part": "16x25x1 Filter (12-pack)", "sku": "F0105", "category": "Filters", "price": 38.00, "default_use": False},
                {"part": "20x20x1 Filter (12-pack)", "sku": "F0110", "category": "Filters", "price": 38.00, "default_use": False},
                {"part": "16x20x1 Filter (12-pack)", "sku": "F0115", "category": "Filters", "price": 36.00, "default_use": False},
                {"part": "24x24x1 Filter (12-pack)", "sku": "F0120", "category": "Filters", "price": 45.00, "default_use": False},
                # ── Mounting / Pads ──
                {"part": "Cork Pads (set of 4)", "sku": "L0800", "category": "Mounting", "price": 12.00, "default_use": True},
                {"part": "Condenser Pad (plastic)", "sku": "L0810", "category": "Mounting", "price": 28.00, "default_use": True},
                {"part": "Condenser Pad (concrete)", "sku": "L0812", "category": "Mounting", "price": 35.00, "default_use": False},
                {"part": "Wall Brackets (pair)", "sku": "L0820", "category": "Mounting", "price": 45.00, "default_use": False},
                {"part": "Rooftop Curb Adapter", "sku": "L0830", "category": "Mounting", "price": 125.00, "default_use": False},
                {"part": "Equipment Stand (28\" H)", "sku": "L0835", "category": "Mounting", "price": 85.00, "default_use": False},
                {"part": "1\" Anchor Kit", "sku": "M0051", "category": "Mounting", "price": 15.00, "default_use": False},
                # ── Refrigerant ──
                {"part": "Refrigerant R-410A (25lb)", "sku": "H1000", "category": "Refrigerant", "price": 185.00, "default_use": True},
                {"part": "Refrigerant R-410A (50lb)", "sku": "H1005", "category": "Refrigerant", "price": 310.00, "default_use": False},
                {"part": "R454B Refrigerant", "sku": "", "category": "Refrigerant", "price": 225.00, "default_use": False},
                {"part": "Nitrogen (tank rental + gas)", "sku": "H1020", "category": "Refrigerant", "price": 75.00, "default_use": True},
                # ── Brazing / Soldering ──
                {"part": "Silver Brazing Rods (pkg)", "sku": "H1030", "category": "Brazing", "price": 55.00, "default_use": True},
                {"part": "Stay-Brite #8 Solder (1lb)", "sku": "H1035", "category": "Brazing", "price": 42.00, "default_use": True},
                {"part": "Solder", "sku": "M0434", "category": "Brazing", "price": 12.00, "default_use": False},
                {"part": "Flux Paste", "sku": "H1040", "category": "Brazing", "price": 12.00, "default_use": True},
                {"part": "MAP Gas Cylinder", "sku": "H1045", "category": "Brazing", "price": 14.00, "default_use": True},
                {"part": "Silver Locking Caps", "sku": "L3011", "category": "Brazing", "price": 4.50, "default_use": False},
                # ── Consumables (workbook) ──
                {"part": "Acetylene Refill", "sku": "M3511", "category": "Consumables", "price": 65.00, "default_use": False},
                {"part": "Oxygen Refill", "sku": "M3513", "category": "Consumables", "price": 45.00, "default_use": False},
                # ── Tools (consumable) ──
                {"part": "12\" SawZaw Blade", "sku": "T0951", "category": "Tools (consumable)", "price": 8.00, "default_use": False},
                {"part": "4-3/8\" Hole Saw", "sku": "T0996", "category": "Tools (consumable)", "price": 18.00, "default_use": False},
                {"part": "Pipe Cutter 1/4-1-5/8\"", "sku": "T0100", "category": "Tools (consumable)", "price": 28.00, "default_use": False},
                {"part": "Flare Tool Kit", "sku": "T0105", "category": "Tools (consumable)", "price": 65.00, "default_use": False},
                {"part": "Torque Wrench (refrigerant)", "sku": "T0110", "category": "Tools (consumable)", "price": 45.00, "default_use": False},
                {"part": "Duct Knife / Blade (pkg)", "sku": "T0115", "category": "Tools (consumable)", "price": 15.00, "default_use": True},
                {"part": "Hacksaw Blades (pkg)", "sku": "T0120", "category": "Tools (consumable)", "price": 8.50, "default_use": True},
                {"part": "Hole Saw Kit (HVAC sizes)", "sku": "T0125", "category": "Tools (consumable)", "price": 55.00, "default_use": False},
                # ── Misc Supplies ──
                {"part": "Zip Ties 11\" (bag of 100)", "sku": "L0900", "category": "Misc Supplies", "price": 6.50, "default_use": True},
                {"part": "Zip Ties 14\" (bag of 100)", "sku": "L0902", "category": "Misc Supplies", "price": 8.00, "default_use": False},
                {"part": "Caulk / Firestop (tube)", "sku": "L0910", "category": "Misc Supplies", "price": 8.00, "default_use": True},
                {"part": "Metal Tape Measure 25ft", "sku": "T0200", "category": "Misc Supplies", "price": 14.00, "default_use": False},
                {"part": "Spray Paint (marking)", "sku": "L0920", "category": "Misc Supplies", "price": 5.50, "default_use": True},
                {"part": "Duct Seal Putty (1lb)", "sku": "L0925", "category": "Misc Supplies", "price": 4.50, "default_use": True},
                {"part": "Electrical Tape (10-pk)", "sku": "L0930", "category": "Misc Supplies", "price": 12.00, "default_use": True},
                {"part": "Tie Wire (roll)", "sku": "L0935", "category": "Misc Supplies", "price": 6.00, "default_use": True},
                # ── Condensate Pumps & IAQ ──
                {"part": "Condensate Pump (mini)", "sku": "P0400", "category": "Condensate", "price": 85.00, "default_use": False},
                {"part": "Condensate Pump (full size)", "sku": "P0405", "category": "Condensate", "price": 125.00, "default_use": False},
                {"part": "UV Light Kit", "sku": "L0950", "category": "IAQ", "price": 145.00, "default_use": False},
                {"part": "Media Filter Cabinet", "sku": "L0960", "category": "IAQ", "price": 165.00, "default_use": False},
            ]
        },
        {
            "name": "Equipment",
            "items": [
                # ── Thermostats (workbook items first) ──
                {"part": "Heat Pump Thermostat", "sku": "L0741", "category": "Thermostat", "price": 65.00, "default_use": False},
                {"part": "TH8 Vision Pro Honeywell", "sku": "", "category": "Thermostat", "price": 145.00, "default_use": False},
                {"part": "Programmable Thermostat", "sku": "L7170", "category": "Thermostat", "price": 52.50, "default_use": True},
                {"part": "WiFi Thermostat", "sku": "L7175", "category": "Thermostat", "price": 125.00, "default_use": False},
                {"part": "Smart Thermostat (Ecobee/Nest)", "sku": "", "category": "Thermostat", "price": 185.00, "default_use": False},
                # ── Air Handlers ──
                {"part": "1.5 Ton Air Handler", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 500, "default_use": False},
                {"part": "2 Ton Air Handler", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 506, "default_use": False},
                {"part": "2.5 Ton Air Handler", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 575, "default_use": False},
                {"part": "3 Ton Air Handler", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 650, "default_use": False},
                {"part": "3.5 Ton Air Handler", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 720, "default_use": False},
                {"part": "4 Ton Air Handler", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 800, "default_use": False},
                {"part": "5 Ton Air Handler", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 950, "default_use": False},
                # ── Corridor Air Handlers (workbook) ──
                {"part": "1.5 Ton Corridor Air Handler", "sku": "", "category": "Corridor AH", "price": 550, "default_use": False},
                {"part": "2 Ton Corridor Air Handler", "sku": "", "category": "Corridor AH", "price": 560, "default_use": False},
                {"part": "2.5 Ton Corridor Air Handler", "sku": "", "category": "Corridor AH", "price": 625, "default_use": False},
                {"part": "3 Ton Corridor Air Handler", "sku": "", "category": "Corridor AH", "price": 700, "default_use": False},
                {"part": "3.5 Ton Corridor Air Handler", "sku": "", "category": "Corridor AH", "price": 780, "default_use": False},
                {"part": "4 Ton Corridor Air Handler", "sku": "", "category": "Corridor AH", "price": 860, "default_use": False},
                {"part": "5 Ton Corridor Air Handler", "sku": "", "category": "Corridor AH", "price": 1020, "default_use": False},
                # ── Corridor Heat Kits (workbook) ──
                {"part": "1.5 Ton Corridor Heat Kit", "sku": "", "category": "Corridor AH", "price": 85, "default_use": False},
                {"part": "2 Ton Corridor Heat Kit", "sku": "", "category": "Corridor AH", "price": 95, "default_use": False},
                {"part": "2.5 Ton Corridor Heat Kit", "sku": "", "category": "Corridor AH", "price": 105, "default_use": False},
                {"part": "3 Ton Corridor Heat Kit", "sku": "", "category": "Corridor AH", "price": 115, "default_use": False},
                {"part": "3.5 Ton Corridor Heat Kit", "sku": "", "category": "Corridor AH", "price": 130, "default_use": False},
                {"part": "4 Ton Corridor Heat Kit", "sku": "", "category": "Corridor AH", "price": 145, "default_use": False},
                {"part": "5 Ton Corridor Heat Kit", "sku": "", "category": "Corridor AH", "price": 170, "default_use": False},
                # ── Heat Strips ──
                {"part": "5kW Heat Strip", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 85, "default_use": False},
                {"part": "8kW Heat Strip", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 105, "default_use": False},
                {"part": "10kW Heat Strip", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 125, "default_use": False},
                {"part": "15kW Heat Strip", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 155, "default_use": False},
                {"part": "20kW Heat Strip", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 185, "default_use": False},
                # ── Gas Furnaces ──
                {"part": "40K BTU Gas Furnace 96%", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 950, "default_use": False},
                {"part": "60K BTU Gas Furnace 96%", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 1100, "default_use": False},
                {"part": "80K BTU Gas Furnace 96%", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 1250, "default_use": False},
                {"part": "100K BTU Gas Furnace 96%", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 1400, "default_use": False},
                {"part": "120K BTU Gas Furnace 96%", "sku": "", "category": "Furnace/Air Handler/Heat Strip", "price": 1600, "default_use": False},
                # ── Condensers ──
                {"part": "1.5 Ton Condenser", "sku": "", "category": "Heat Pump/Condenser", "price": 820, "default_use": False},
                {"part": "2 Ton Condenser", "sku": "", "category": "Heat Pump/Condenser", "price": 950, "default_use": False},
                {"part": "2.5 Ton Condenser", "sku": "", "category": "Heat Pump/Condenser", "price": 1100, "default_use": False},
                {"part": "3 Ton Condenser", "sku": "", "category": "Heat Pump/Condenser", "price": 1300, "default_use": False},
                {"part": "3.5 Ton Condenser", "sku": "", "category": "Heat Pump/Condenser", "price": 1450, "default_use": False},
                {"part": "4 Ton Condenser", "sku": "", "category": "Heat Pump/Condenser", "price": 1650, "default_use": False},
                {"part": "5 Ton Condenser", "sku": "", "category": "Heat Pump/Condenser", "price": 1950, "default_use": False},
                # ── Heat Pumps ──
                {"part": "1.5 Ton Heat Pump", "sku": "", "category": "Heat Pump/Condenser", "price": 1050, "default_use": False},
                {"part": "2 Ton Heat Pump", "sku": "", "category": "Heat Pump/Condenser", "price": 1200, "default_use": False},
                {"part": "2.5 Ton Heat Pump", "sku": "", "category": "Heat Pump/Condenser", "price": 1380, "default_use": False},
                {"part": "3 Ton Heat Pump", "sku": "", "category": "Heat Pump/Condenser", "price": 1580, "default_use": False},
                {"part": "3.5 Ton Heat Pump", "sku": "", "category": "Heat Pump/Condenser", "price": 1750, "default_use": False},
                {"part": "4 Ton Heat Pump", "sku": "", "category": "Heat Pump/Condenser", "price": 1980, "default_use": False},
                {"part": "5 Ton Heat Pump", "sku": "", "category": "Heat Pump/Condenser", "price": 2350, "default_use": False},
                # ── Mini Splits (workbook: broken into Indoor/Outdoor) ──
                {"part": "Mini Split 9K Indoor", "sku": "L9078", "category": "Mini Split", "price": 350, "default_use": False},
                {"part": "Mini Split 9K Outdoor", "sku": "L9071", "category": "Mini Split", "price": 500, "default_use": False},
                {"part": "Mini Split 12K Indoor", "sku": "L9080", "category": "Mini Split", "price": 400, "default_use": False},
                {"part": "Mini Split 12K Outdoor", "sku": "L9073", "category": "Mini Split", "price": 550, "default_use": False},
                {"part": "Mini Split 18K Cassette", "sku": "L6455", "category": "Mini Split", "price": 650, "default_use": False},
                {"part": "Mini Split 18K Indoor", "sku": "L6448", "category": "Mini Split", "price": 475, "default_use": False},
                {"part": "Mini Split 18K Outdoor HP", "sku": "L6432", "category": "Mini Split", "price": 675, "default_use": False},
                {"part": "Mini Split 24K Indoor", "sku": "L9082", "category": "Mini Split", "price": 525, "default_use": False},
                {"part": "Mini Split 24K Outdoor", "sku": "L9075", "category": "Mini Split", "price": 825, "default_use": False},
                # ── Mini Splits (legacy combined for calc compatibility) ──
                {"part": "Mini Split 9K BTU", "sku": "", "category": "Mini Split", "price": 850, "default_use": False},
                {"part": "Mini Split 12K BTU", "sku": "", "category": "Mini Split", "price": 950, "default_use": False},
                {"part": "Mini Split 18K BTU", "sku": "", "category": "Mini Split", "price": 1150, "default_use": False},
                {"part": "Mini Split 24K BTU", "sku": "", "category": "Mini Split", "price": 1350, "default_use": False},
                {"part": "Mini Split 36K BTU", "sku": "", "category": "Mini Split", "price": 1650, "default_use": False},
                {"part": "Mini Split Multi-Zone Outdoor 2-zone", "sku": "", "category": "Mini Split", "price": 1800, "default_use": False},
                {"part": "Mini Split Multi-Zone Outdoor 3-zone", "sku": "", "category": "Mini Split", "price": 2400, "default_use": False},
                {"part": "Mini Split Multi-Zone Outdoor 4-zone", "sku": "", "category": "Mini Split", "price": 3000, "default_use": False},
                # ── Ventilation / Exhaust ──
                {"part": "ERV Unit", "sku": "", "category": "Ventilation", "price": 650, "default_use": False},
                {"part": "HRV Unit", "sku": "", "category": "Ventilation", "price": 580, "default_use": False},
                {"part": "Inline Exhaust Fan 6\"", "sku": "", "category": "Ventilation", "price": 125, "default_use": False},
                {"part": "Inline Exhaust Fan 8\"", "sku": "", "category": "Ventilation", "price": 165, "default_use": False},
                {"part": "Bath Exhaust Fan 50CFM", "sku": "", "category": "Ventilation", "price": 75, "default_use": False},
                {"part": "Bath Exhaust Fan 80CFM", "sku": "", "category": "Ventilation", "price": 95, "default_use": False},
                {"part": "Bath Exhaust Fan 110CFM", "sku": "", "category": "Ventilation", "price": 125, "default_use": False},
                {"part": "Range Hood Vent Kit", "sku": "", "category": "Ventilation", "price": 85, "default_use": False},
                {"part": "Dryer Vent Kit", "sku": "", "category": "Ventilation", "price": 45, "default_use": False},
                {"part": "Dryer Vent Elbow (4\")", "sku": "", "category": "Ventilation", "price": 6, "default_use": False},
                {"part": "Dryer Vent Hose (4\" x 8ft)", "sku": "", "category": "Ventilation", "price": 12, "default_use": False},
                {"part": "Dryer Vent Wall Cap", "sku": "", "category": "Ventilation", "price": 10, "default_use": False},
                # ── Zoning ──
                {"part": "Zoning Panel (2 zone)", "sku": "", "category": "Zoning", "price": 350, "default_use": False},
                {"part": "Zoning Panel (3 zone)", "sku": "", "category": "Zoning", "price": 450, "default_use": False},
                {"part": "Zone Damper (round)", "sku": "", "category": "Zoning", "price": 85, "default_use": False},
                {"part": "Zone Damper (rect)", "sku": "", "category": "Zoning", "price": 110, "default_use": False},
                {"part": "Zone Thermostat Sensor", "sku": "", "category": "Zoning", "price": 35, "default_use": False},
            ]
        },
        {
            "name": "Freight",
            "items": [
                {"part": "Freight to jobsite", "sku": "", "category": "Shipping", "price": 3000, "default_use": False},
                {"part": "Crane / Lift Rental", "sku": "", "category": "Shipping", "price": 1500, "default_use": False},
                {"part": "Dumpster Rental", "sku": "", "category": "Shipping", "price": 500, "default_use": False},
                {"part": "Scissor Lift Rental", "sku": "", "category": "Shipping", "price": 800, "default_use": False},
                {"part": "License", "sku": "", "category": "Permits & Fees", "price": 0, "default_use": False},
                {"part": "Permits", "sku": "", "category": "Permits & Fees", "price": 0, "default_use": False},
            ]
        }
    ],
    "waste_factor": 0.075,
    "tax_rate": 0.0885
}


@app.route('/plans')
@role_required('owner', 'admin', 'project_manager')
def plans_page():
    return render_template('plans/list.html')


@app.route('/api/plans')
@api_role_required('owner', 'admin', 'project_manager')
def api_plans_list():
    job_id = request.args.get('job_id', type=int)
    status = request.args.get('status', '')
    plan_type = request.args.get('plan_type', '')
    conn = get_db()
    sql = '''SELECT p.*, j.name as job_name FROM plans p
             JOIN jobs j ON p.job_id = j.id WHERE 1=1'''
    params = []
    if job_id:
        sql += ' AND p.job_id = ?'
        params.append(job_id)
    if status:
        sql += ' AND p.status = ?'
        params.append(status)
    if plan_type:
        sql += ' AND p.plan_type = ?'
        params.append(plan_type)
    sql += ' ORDER BY p.created_at DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d['has_file'] = bool(d.get('file_path'))
        d['has_review'] = bool(d.get('ai_review'))
        d['has_takeoff'] = bool(d.get('takeoff_data'))
        result.append(d)
    return jsonify(result)


@app.route('/api/plans', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_plans_create():
    conn = get_db()
    file_path = ''
    file_hash = ''
    page_count = 0
    if request.content_type and 'multipart' in request.content_type:
        job_id = request.form.get('job_id', type=int)
        title = request.form.get('title', '')
        plan_type = request.form.get('plan_type', 'Mechanical')
        notes = request.form.get('notes', '')
        file = request.files.get('file')
        if file and file.filename:
            from werkzeug.utils import secure_filename
            content = file.read()
            file_hash = hashlib.md5(content).hexdigest()
            fname = secure_filename(file.filename)
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.seek(0)
            file.save(os.path.join(PLANS_DIR, fname))
            file_path = fname
            # Count pages
            try:
                import pdfplumber
                with pdfplumber.open(os.path.join(PLANS_DIR, fname)) as pdf:
                    page_count = len(pdf.pages)
            except Exception:
                pass
    else:
        data = request.get_json(force=True)
        job_id = data.get('job_id')
        title = data.get('title', '')
        plan_type = data.get('plan_type', 'Mechanical')
        notes = data.get('notes', '')

    cursor = conn.execute(
        '''INSERT INTO plans (job_id, title, plan_type, file_path, file_hash, upload_date,
           notes, page_count, created_by)
           VALUES (?,?,?,?,?,date('now','localtime'),?,?,?)''',
        (job_id, title, plan_type, file_path, file_hash, notes, page_count, session['user_id'])
    )
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return jsonify({'ok': True, 'id': new_id}), 201


@app.route('/api/plans/<int:pid>')
@api_role_required('owner', 'admin', 'project_manager')
def api_plans_get(pid):
    conn = get_db()
    row = conn.execute(
        '''SELECT p.*, j.name as job_name FROM plans p
           JOIN jobs j ON p.job_id = j.id WHERE p.id = ?''', (pid,)
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Plan not found'}), 404
    d = dict(row)
    d['has_file'] = bool(d.get('file_path'))
    d['has_review'] = bool(d.get('ai_review'))
    d['has_takeoff'] = bool(d.get('takeoff_data'))
    if d.get('ai_review'):
        try:
            d['ai_review'] = json.loads(d['ai_review'])
        except (json.JSONDecodeError, TypeError):
            pass
    if d.get('takeoff_data'):
        try:
            d['takeoff_data'] = json.loads(d['takeoff_data'])
        except (json.JSONDecodeError, TypeError):
            pass
    return jsonify(d)


@app.route('/api/plans/<int:pid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_plans_update(pid):
    conn = get_db()
    file_path_update = None
    page_count_update = None
    if request.content_type and 'multipart' in request.content_type:
        title = request.form.get('title', '')
        plan_type = request.form.get('plan_type', 'Mechanical')
        status = request.form.get('status', 'Uploaded')
        notes = request.form.get('notes', '')
        takeoff_data = request.form.get('takeoff_data', '')
        file = request.files.get('file')
        if file and file.filename:
            from werkzeug.utils import secure_filename
            fname = secure_filename(file.filename)
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.save(os.path.join(PLANS_DIR, fname))
            file_path_update = fname
            try:
                import pdfplumber
                with pdfplumber.open(os.path.join(PLANS_DIR, fname)) as pdf:
                    page_count_update = len(pdf.pages)
            except Exception:
                page_count_update = 0
    else:
        data = request.get_json(force=True)
        title = data.get('title', '')
        plan_type = data.get('plan_type', 'Mechanical')
        status = data.get('status', 'Uploaded')
        notes = data.get('notes', '')
        takeoff_data = data.get('takeoff_data', '')

    if file_path_update is not None:
        conn.execute(
            '''UPDATE plans SET title=?, plan_type=?, status=?, notes=?, file_path=?,
               upload_date=date('now','localtime'), page_count=?,
               updated_at=datetime('now','localtime') WHERE id=?''',
            (title, plan_type, status, notes, file_path_update, page_count_update, pid)
        )
    elif takeoff_data:
        conn.execute(
            '''UPDATE plans SET title=?, plan_type=?, status=?, notes=?, takeoff_data=?,
               updated_at=datetime('now','localtime') WHERE id=?''',
            (title, plan_type, status, notes, takeoff_data if isinstance(takeoff_data, str) else json.dumps(takeoff_data), pid)
        )
    else:
        conn.execute(
            '''UPDATE plans SET title=?, plan_type=?, status=?, notes=?,
               updated_at=datetime('now','localtime') WHERE id=?''',
            (title, plan_type, status, notes, pid)
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/plans/<int:pid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_plans_delete(pid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM plans WHERE id = ?', (pid,)).fetchone()
    if row and row['file_path']:
        fpath = os.path.join(PLANS_DIR, row['file_path'])
        if os.path.exists(fpath):
            os.remove(fpath)
    conn.execute('DELETE FROM plans WHERE id = ?', (pid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/plans/<int:pid>/file')
@api_role_required('owner', 'admin', 'project_manager')
def api_plans_file(pid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM plans WHERE id = ?', (pid,)).fetchone()
    conn.close()
    if not row or not row['file_path']:
        return 'Not found', 404
    fpath = os.path.join(PLANS_DIR, row['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    return send_file(fpath, mimetype='application/pdf')


_review_progress = {}  # pid -> {step, message, pct, done, error, result}


@app.route('/api/plans/<int:pid>/review-estimate', methods=['GET'])
@api_role_required('owner', 'admin', 'project_manager')
def api_plans_review_estimate(pid):
    """Quick pre-scan: count mechanical sheets and estimate API cost."""
    conn = get_db()
    plan = conn.execute('SELECT * FROM plans WHERE id = ?', (pid,)).fetchone()
    if not plan:
        conn.close()
        return jsonify({'error': 'Plan not found'}), 404
    if not plan['file_path']:
        conn.close()
        return jsonify({'error': 'No file attached'}), 400

    fpath = os.path.join(PLANS_DIR, plan['file_path'])
    if not os.path.exists(fpath):
        conn.close()
        return jsonify({'error': 'File not found'}), 404
    conn.close()

    try:
        import fitz
    except ImportError:
        return jsonify({'error': 'PyMuPDF not installed'}), 500

    import re as _re
    from collections import Counter

    doc = fitz.open(fpath)
    page_count = len(doc)

    # Quick two-pass sheet identification (same logic as review)
    id_freq = Counter()
    raw = []
    for page in doc:
        text = (page.get_text() or '').strip()
        ids = set()
        for line in text.split('\n'):
            t = line.strip()
            if 2 <= len(t) <= 5:
                if _re.match(r'^ME\d{1,2}$', t): ids.add(t)
                elif _re.match(r'^M\d{1,2}$', t): ids.add(t)
                elif _re.match(r'^[PSEGAC]\d{1,3}$', t): ids.add(t)
        for s in ids:
            id_freq[s] += 1
        raw.append({'text': text, 'ids': ids, 'chars': len(text)})
    doc.close()

    noise = {s for s, c in id_freq.items() if c > 4}

    HVAC_KW = ['hvac plan','hvac equipment','mechanical plan','unit hvac',
        'condensing unit','fan coil','mini split','duct siz','equipment schedule',
        'grille','register schedule','fire damper schedule','exhaust fan','hvac',
        'refrigerant','tonnage','btu','seer','cfm','mep plan','mep','roof hvac',
        'roof mep','mechanical schedule','condensing','unit plan']

    mech_sheets = []
    hvac_pages = []
    spec_chars = 0

    for i, r in enumerate(raw):
        clean_ids = r['ids'] - noise
        text_lower = r['text'].lower()
        m_ids = [s for s in clean_ids if s.startswith('M')]
        non_m_ids = [s for s in clean_ids if not s.startswith('M')]
        is_mech = len(m_ids) > 0 and len(non_m_ids) == 0
        is_drawing = r['chars'] < 2000
        hvac_strong = any(kw in text_lower for kw in HVAC_KW)

        if is_mech:
            mech_sheets.append({'page': i+1, 'ids': sorted(clean_ids)})
        elif hvac_strong and is_drawing:
            hvac_pages.append(i+1)
        elif r['chars'] > 500 and not is_drawing:
            spec_chars += r['chars']

    image_count = min(len(mech_sheets) + len(hvac_pages), 25)
    # Cost estimate: ~1600 tokens per image, ~1 token per 4 chars of spec text
    # Sonnet pricing: $3/M input, $15/M output
    input_tokens = (image_count * 1600) + (min(spec_chars, 60000) // 4) + 2000  # prompt overhead
    output_tokens = 4000  # typical response
    est_cost = (input_tokens * 3.0 / 1_000_000) + (output_tokens * 15.0 / 1_000_000)

    return jsonify({
        'page_count': page_count,
        'mech_sheets': len(mech_sheets),
        'mech_sheet_list': [f"Page {s['page']} ({', '.join(s['ids'])})" for s in mech_sheets],
        'hvac_keyword_pages': len(hvac_pages),
        'total_images': image_count,
        'spec_pages': sum(1 for r in raw if r['chars'] > 500 and r['chars'] >= 2000),
        'spec_chars': spec_chars,
        'estimated_input_tokens': input_tokens,
        'estimated_cost': round(est_cost, 2)
    })


@app.route('/api/plans/<int:pid>/review-status', methods=['GET'])
@api_role_required('owner', 'admin', 'project_manager')
def api_plans_review_status(pid):
    """Poll endpoint for review progress."""
    prog = _review_progress.get(pid)
    if not prog:
        return jsonify({'step': 0, 'message': 'Not started', 'pct': 0, 'done': False})
    return jsonify(prog)

@app.route('/api/plans/<int:pid>/review', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_plans_review(pid):
    """Kick off AI HVAC review in background thread. Returns immediately."""
    import anthropic

    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return jsonify({'error': 'ANTHROPIC_API_KEY not configured. Add it to your .env file.'}), 500

    conn = get_db()
    plan = conn.execute('SELECT * FROM plans WHERE id = ?', (pid,)).fetchone()
    if not plan:
        conn.close()
        return jsonify({'error': 'Plan not found'}), 404

    if not plan['file_path']:
        conn.close()
        return jsonify({'error': 'No file attached to this plan.'}), 400

    fpath = os.path.join(PLANS_DIR, plan['file_path'])
    if not os.path.exists(fpath):
        conn.close()
        return jsonify({'error': 'Plan file not found on disk.'}), 404

    # Check if already running
    prog = _review_progress.get(pid)
    if prog and not prog.get('done') and not prog.get('error'):
        conn.close()
        return jsonify({'status': 'already_running', 'message': 'Review already in progress.'})

    # Get job context before closing conn
    job = conn.execute('SELECT name, address, city, state FROM jobs WHERE id = ?', (plan['job_id'],)).fetchone()
    job_context = f"Job: {job['name']}" if job else ""
    if job and job['city']:
        job_context += f" ({job['city']}, {job['state'] or ''})"

    plan_data = dict(plan)
    conn.execute("UPDATE plans SET status='Reviewing', updated_at=datetime('now','localtime') WHERE id=?", (pid,))
    conn.commit()
    conn.close()

    # Initialize progress
    _review_progress[pid] = {'step': 1, 'message': 'Opening PDF...', 'pct': 5, 'done': False}

    import threading
    def run_review():
        _do_plan_review(pid, fpath, plan_data, job_context, api_key)
    t = threading.Thread(target=run_review, daemon=True)
    t.start()

    return jsonify({'status': 'started', 'message': 'Review started. Poll /review-status for progress.'})


def _do_plan_review(pid, fpath, plan_data, job_context, api_key):
    """Background worker for plan review."""
    import anthropic, base64
    prog = _review_progress[pid]

    def update(step, msg, pct):
        prog['step'] = step
        prog['message'] = msg
        prog['pct'] = pct

    try:
        import fitz
    except ImportError:
        prog.update({'error': 'PyMuPDF not installed', 'done': True})
        return

    try:
        doc = fitz.open(fpath)
    except Exception as e:
        prog.update({'error': f'Could not open PDF: {e}', 'done': True})
        return

    page_count = len(doc)
    update(1, f'Scanning {page_count} pages...', 10)

    # ── Phase 1: Two-pass page classification ─────────────────────
    # Pass 1: Extract all potential sheet IDs and count frequency
    # Pass 2: Filter out building names and section totals (appear on many pages)
    import re as _re
    from collections import Counter

    raw_page_data = []
    id_frequency = Counter()

    for i, page in enumerate(doc):
        if i % 20 == 0:
            update(1, f'Scanning page {i+1} of {page_count}...', 10 + int(20 * i / max(page_count, 1)))
        text = page.get_text() or ''
        clean = text.strip()

        # Extract all potential sheet IDs
        potential_ids = set()
        for line in clean.split('\n'):
            token = line.strip()
            if 2 <= len(token) <= 5:
                if _re.match(r'^ME\d{1,2}$', token):
                    potential_ids.add(token)
                elif _re.match(r'^M\d{1,2}$', token):
                    potential_ids.add(token)
                elif _re.match(r'^[PSEGAC]\d{1,3}$', token):
                    potential_ids.add(token)

        for sid in potential_ids:
            id_frequency[sid] += 1

        raw_page_data.append({'text': clean, 'potential_ids': potential_ids})

    # Pass 2: IDs appearing on >4 pages are building names or section totals, not sheet IDs
    # e.g. "M8" = building name (all pages), "M7" = "sheet X of 7" (many pages)
    noise_ids = {sid for sid, count in id_frequency.items() if count > 4}

    page_info = []
    for i, rpd in enumerate(raw_page_data):
        clean = rpd['text']
        char_count = len(clean)
        text_lower = clean.lower()
        is_drawing = char_count < 2000

        # Filter to real sheet IDs (low frequency = actual sheet identifier)
        sheet_ids = rpd['potential_ids'] - noise_ids

        # Determine if this page IS a mechanical sheet
        # Must have M/ME IDs and NOT have non-mechanical IDs (A, S, P, E, G)
        # to avoid false positives from cross-reference pages
        m_ids = [s for s in sheet_ids if (s.startswith('M') and not s.startswith('ME')) or s.startswith('ME')]
        non_m_ids = [s for s in sheet_ids if not s.startswith('M')]
        is_mech_sheet = len(m_ids) > 0 and len(non_m_ids) == 0

        # Check text content for strong HVAC signals
        hvac_strong = any(kw in text_lower for kw in [
            'hvac plan', 'hvac equipment', 'mechanical plan', 'unit hvac',
            'condensing unit', 'fan coil', 'mini split', 'duct siz',
            'equipment schedule', 'grille', 'register schedule', 'fire damper schedule',
            'exhaust fan', 'hvac', 'refrigerant', 'tonnage', 'btu', 'seer', 'cfm',
            'mep plan', 'mep', 'roof hvac', 'roof mep', 'mechanical schedule',
            'condensing', 'unit plan'])

        page_info.append({
            'num': i + 1, 'text': clean, 'char_count': char_count,
            'is_drawing': is_drawing, 'sheet_ids': sheet_ids,
            'is_mech_sheet': is_mech_sheet, 'hvac_strong': hvac_strong
        })

    update(2, 'Identifying mechanical sheets...', 30)

    # ── Phase 2: Select pages — prioritize actual mechanical sheets ──
    # Tier 1: Pages with M or ME sheet IDs (the actual mechanical drawings)
    # Tier 2: Drawing pages with strong HVAC keywords but no M/ME ID
    # Tier 3: Cover/index page
    tier1 = []   # M/ME sheets — MUST include all
    tier2 = []   # HVAC-relevant drawings without sheet ID
    tier3 = []   # cover page
    text_pages = []

    for pi in page_info:
        # Mechanical sheets (M/ME) are ALWAYS sent as images, even with >2000 chars
        # (they're CAD drawings with many detail labels, not spec text)
        if pi['is_mech_sheet']:
            tier1.append(pi)
        elif not pi['is_drawing'] and pi['char_count'] > 500:
            text_pages.append(pi)
        elif pi['hvac_strong'] and pi['is_drawing']:
            tier2.append(pi)
        elif pi['num'] == 1 and pi['is_drawing']:
            tier3.append(pi)

    # Build image list: all tier 1 + fill with tier 2 and 3
    MAX_IMAGE_PAGES = 25
    image_pages = list(tier1)  # always include ALL mechanical sheets
    remaining = MAX_IMAGE_PAGES - len(image_pages)
    if remaining > 0:
        image_pages.extend(tier2[:remaining])
        remaining = MAX_IMAGE_PAGES - len(image_pages)
    if remaining > 0:
        image_pages.extend(tier3[:remaining])
    image_pages.sort(key=lambda p: p['num'])

    mech_count = len(tier1)
    update(2, f'Found {mech_count} mechanical sheets (M/ME series), sending {len(image_pages)} pages total...', 35)

    update(3, f'Rendering {len(image_pages)} drawing pages as images...', 35)

    # ── Phase 3: Render pages as JPEG images ──────────────────────
    MAX_PX = 1999
    images_b64 = []
    for idx, pi in enumerate(image_pages):
        update(3, f'Rendering page {pi["num"]} ({idx+1}/{len(image_pages)})...', 35 + int(30 * idx / max(len(image_pages), 1)))
        try:
            page = doc[pi['num'] - 1]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            if pix.width > MAX_PX or pix.height > MAX_PX:
                scale = min(MAX_PX / pix.width, MAX_PX / pix.height)
                zoom = 2.0 * scale
                pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
            img_bytes = pix.tobytes("jpeg")
            b64 = base64.standard_b64encode(img_bytes).decode('ascii')
            sheet_label = ', '.join(sorted(pi.get('sheet_ids', set()))) or ''
            label = f"Page {pi['num']}"
            if sheet_label:
                label += f" — Sheet {sheet_label}"
            images_b64.append({
                'page_num': pi['num'], 'b64': b64, 'label': label
            })
        except Exception:
            continue

    doc.close()

    # ── Phase 4: Build spec text ──────────────────────────────────
    update(4, 'Preparing specification text...', 65)
    spec_text = ''
    MAX_TEXT_CHARS = 60000
    mech_text_pages = [p for p in text_pages if p.get('hvac_strong') or p.get('is_mech_sheet')]
    other_text_pages = [p for p in text_pages if p not in mech_text_pages]
    for pi in mech_text_pages + other_text_pages:
        entry = f'--- Page {pi["num"]} (specs) ---\n{pi["text"]}\n\n'
        if len(spec_text) + len(entry) > MAX_TEXT_CHARS:
            break
        spec_text += entry

    if not images_b64 and not spec_text.strip():
        prog['error'] = 'Could not extract any content from this PDF.'
        prog['done'] = True
        conn2 = get_db()
        conn2.execute("UPDATE plans SET status='Uploaded', updated_at=datetime('now','localtime') WHERE id=?", (pid,))
        conn2.commit()
        conn2.close()
        return

    # ── Phase 5: Send to Claude Vision ────────────────────────────
    update(5, f'Sending {len(images_b64)} images to AI for analysis...', 70)

    plan_type = plan_data.get('plan_type') or 'construction'
    plan_title = plan_data.get('title') or ''

    system_prompt = f"""You are an expert HVAC mechanical contractor extracting data from construction plans for LGHVAC Mechanical, LLC. {job_context}

This is a {plan_type} plan titled "{plan_title}" with {page_count} total pages. You are being shown {len(images_b64)} page images from the mechanical drawings plus extracted text from specification pages.

YOUR JOB: Extract EVERY piece of HVAC data visible in these drawings. Read every table cell, every callout, every label, every annotation. These are CAD-generated construction documents — the data (model numbers, tonnage, CFM, duct dimensions, equipment schedules) is drawn in the images.

EXTRACT ALL OF THE FOLLOWING — report ONLY what you can actually see. Do NOT guess or infer missing data:

1. EQUIPMENT SCHEDULE — Read every row: Mark/Tag, Type, Manufacturer, Model #, Tonnage/BTU, Heating capacity, Cooling capacity, CFM, Voltage/Phase, FLA/MCA/MOCP, Location/Serving
2. DUCT SIZES — Every size annotation visible on floor plans: supply mains (rectangular), branch ducts (round), flex runs, return ducts. Include the sheet reference.
3. DIFFUSER/REGISTER/GRILLE SCHEDULE — If visible: Type, Size, CFM, Neck size, Location
4. EXHAUST FAN SCHEDULE — Model, CFM, HP, RPM, Voltage, Location
5. REFRIGERANT LINES — Suction line size, liquid line size for each unit/mark
6. UNIT COUNTS — How many of each unit type (apartment types, number of each equipment mark)
7. BUILDING INFO — Total floors, building type, unit type names (A, B, C, etc.), total apartment count
8. ELECTRICAL — Disconnect sizes, circuit requirements, wire gauges for HVAC equipment
9. NOTES & DETAILS — Any general notes, installation details, or special requirements visible on the sheets

Respond with valid JSON only (no markdown, no code fences):
{{
    "summary": "2-3 sentence factual overview: building type, total units, number of equipment types, key system info",
    "building_info": {{
        "type": "e.g. 4-story multifamily apartment",
        "total_floors": 0,
        "unit_types": [{{"name": "Type A", "beds": 0, "baths": 0, "count": 0, "hvac_mark": "C1"}}],
        "total_apartments": 0,
        "location": "city/state if shown",
        "notes": "any building-level notes"
    }},
    "equipment_schedule": [
        {{"mark": "C1", "type": "Fan Coil/Heat Pump/etc", "manufacturer": "", "model": "", "tonnage": 0, "heating_btu": 0, "cooling_btu": 0, "cfm": 0, "voltage": "208/1/60", "fla": 0, "mca": 0, "mocp": 0, "location": "where it serves", "count": 1, "notes": ""}}
    ],
    "duct_sizes": [
        {{"size": "12x8", "type": "supply_main|supply_branch|return_main|return_branch|flex|exhaust", "location": "sheet and area", "count": 1}}
    ],
    "diffusers": [
        {{"type": "supply register|return grille|ceiling diffuser", "size": "10x6", "cfm": 0, "count": 1, "location": "sheet ref"}}
    ],
    "exhaust_fans": [
        {{"mark": "EF1", "model": "", "cfm": 0, "hp": 0, "voltage": "", "type": "bath|kitchen|general", "count": 1}}
    ],
    "refrigerant_lines": [
        {{"mark": "C1", "suction_size": "7/8", "liquid_size": "3/8", "notes": ""}}
    ],
    "electrical": [
        {{"mark": "C1", "voltage": "208/1/60", "mca": 0, "mocp": 0, "wire_size": "", "disconnect": ""}}
    ],
    "notes_and_details": ["any general notes, special requirements, installation details, code references visible on sheets"],
    "sheets_analyzed": ["list of sheet IDs you can identify: M1, ME2, etc."]
}}

CRITICAL RULES:
- Report ONLY what you see. If a field is not visible, use 0 or empty string.
- Read every row of every schedule table — do not skip rows or summarize.
- Include the sheet reference for every item so the contractor can find it.
- If you see partial/unclear text, report what you can read and note it's unclear.
- Do NOT make judgments about what's missing or what should be different."""

    content_parts = []
    for img in images_b64:
        content_parts.append({"type": "text", "text": f"[{img['label']}]"})
        content_parts.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/jpeg", "data": img['b64']}
        })
    if spec_text.strip():
        content_parts.append({
            "type": "text",
            "text": f"\n\nEXTRACTED SPECIFICATION TEXT ({len(text_pages)} pages):\n{spec_text}"
        })
    content_parts.append({
        "type": "text",
        "text": "\nNow analyze all the images and text above. Return your findings as JSON."
    })

    try:
        client = anthropic.Anthropic(api_key=api_key)
        update(5, 'Extracting data from plans (this takes 30-90 seconds)...', 75)

        response = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=8192,
            system=system_prompt,
            messages=[{'role': 'user', 'content': content_parts}]
        )

        update(6, 'Processing AI response...', 90)
        response_text = response.content[0].text.strip()

        if response_text.startswith('```'):
            response_text = response_text.split('\n', 1)[1] if '\n' in response_text else response_text[3:]
            if response_text.endswith('```'):
                response_text = response_text[:-3]
            response_text = response_text.strip()

        review = json.loads(response_text)

        # Track actual token usage and cost
        usage = response.usage
        input_tokens = getattr(usage, 'input_tokens', 0)
        output_tokens = getattr(usage, 'output_tokens', 0)
        # Sonnet pricing: $3/M input, $15/M output
        actual_cost = (input_tokens * 3.0 / 1_000_000) + (output_tokens * 15.0 / 1_000_000)

        review['reviewed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        review['page_count'] = page_count
        review['pages_analyzed'] = len(images_b64)
        review['spec_pages_read'] = len(text_pages)
        review['tokens_used'] = {'input': input_tokens, 'output': output_tokens}
        review['actual_cost'] = round(actual_cost, 4)

        review.setdefault('summary', '')
        review.setdefault('building_info', {})
        review.setdefault('equipment_schedule', [])
        review.setdefault('duct_sizes', [])
        review.setdefault('diffusers', [])
        review.setdefault('exhaust_fans', [])
        review.setdefault('refrigerant_lines', [])
        review.setdefault('electrical', [])
        review.setdefault('notes_and_details', [])
        review.setdefault('sheets_analyzed', [])
        # Backwards compat: build findings from extracted data for the frontend
        review.setdefault('findings', [])
        review.setdefault('equipment_found', review.get('equipment_schedule', []))

    except json.JSONDecodeError:
        review = {
            "reviewed_at": datetime.now().strftime('%Y-%m-%d %H:%M'),
            "page_count": page_count,
            "summary": response_text[:500] if response_text else "Review completed but response could not be parsed.",
            "building_info": {}, "equipment_schedule": [], "duct_sizes": [],
            "diffusers": [], "exhaust_fans": [], "refrigerant_lines": [],
            "electrical": [], "notes_and_details": [], "sheets_analyzed": [],
            "findings": [{"type": "info", "category": "Review", "message": "AI review completed — see summary for details."}],
        }
    except Exception as e:
        prog['error'] = str(e)
        prog['done'] = True
        conn2 = get_db()
        conn2.execute("UPDATE plans SET status='Uploaded', updated_at=datetime('now','localtime') WHERE id=?", (pid,))
        conn2.commit()
        conn2.close()
        return

    # Save to DB
    update(6, 'Saving review...', 95)
    conn2 = get_db()
    conn2.execute(
        "UPDATE plans SET ai_review=?, status='Reviewed', page_count=?, updated_at=datetime('now','localtime') WHERE id=?",
        (json.dumps(review), page_count, pid)
    )
    conn2.commit()
    conn2.close()

    prog['pct'] = 100
    prog['message'] = 'Review complete!'
    prog['done'] = True
    prog['result'] = review


def _resolve_takeoff_prices(conn, job_id):
    """Resolve prices for takeoff items from supplier quotes and invoices.

    Returns {SKU: {price, source, source_type}} where source_type is:
      - 'quote'      — from a supplier quote for this job (Tier 1)
      - 'historical'  — from quotes/invoices on other jobs (Tier 2)
      - 'default'     — template price used as fallback (Tier 3)
    """
    prices = {}

    # ── Tier 1: Job quotes (prefer is_baseline) ──
    rows = conn.execute('''
        SELECT sqi.takeoff_sku, sqi.unit_price, sqi.notes,
               sq.supplier_name, sq.quote_number, sq.is_baseline
        FROM supplier_quote_items sqi
        JOIN supplier_quotes sq ON sqi.quote_id = sq.id
        WHERE sq.job_id = ? AND sqi.takeoff_sku != ''
        ORDER BY sq.is_baseline DESC, sq.quote_date DESC
    ''', (job_id,)).fetchall()
    for r in rows:
        sku = r['takeoff_sku']
        if sku in prices:
            continue  # first match wins (baseline first due to ORDER BY)
        unit_price = r['unit_price'] or 0
        # Normalize per-C / per-M pricing from notes
        notes_lower = (r['notes'] or '').lower()
        if 'per-c' in notes_lower or 'per c' in notes_lower:
            unit_price = unit_price / 100.0
        elif 'per-m' in notes_lower or 'per m' in notes_lower:
            unit_price = unit_price / 1000.0
        prices[sku] = {
            'price': round(unit_price, 4),
            'source': f"{r['supplier_name']} Q#{r['quote_number']}",
            'source_type': 'quote',
        }

    # ── Tier 2: Historical quotes from other jobs ──
    unmatched_skus_query = '''
        SELECT DISTINCT sqi.takeoff_sku
        FROM supplier_quote_items sqi
        JOIN supplier_quotes sq ON sqi.quote_id = sq.id
        WHERE sq.job_id != ? AND sqi.takeoff_sku != ''
        ORDER BY sqi.takeoff_sku
    '''
    hist_rows = conn.execute('''
        SELECT sqi.takeoff_sku, sqi.unit_price, sqi.notes,
               sq.supplier_name, sq.quote_number, sq.quote_date
        FROM supplier_quote_items sqi
        JOIN supplier_quotes sq ON sqi.quote_id = sq.id
        WHERE sq.job_id != ? AND sqi.takeoff_sku != ''
        ORDER BY sq.quote_date DESC
    ''', (job_id,)).fetchall()
    for r in hist_rows:
        sku = r['takeoff_sku']
        if sku in prices:
            continue
        unit_price = r['unit_price'] or 0
        notes_lower = (r['notes'] or '').lower()
        if 'per-c' in notes_lower or 'per c' in notes_lower:
            unit_price = unit_price / 100.0
        elif 'per-m' in notes_lower or 'per m' in notes_lower:
            unit_price = unit_price / 1000.0
        prices[sku] = {
            'price': round(unit_price, 4),
            'source': f"{r['supplier_name']} Q#{r['quote_number']} ({r['quote_date']})",
            'source_type': 'historical',
        }

    # ── Tier 2b: Historical from invoices ──
    try:
        inv_rows = conn.execute('''
            SELECT si.line_items, si.invoice_date,
                   bc.supplier_name
            FROM supplier_invoices si
            JOIN billtrust_config bc ON si.supplier_config_id = bc.id
            ORDER BY si.invoice_date DESC
        ''').fetchall()
        import json as _json
        for inv in inv_rows:
            try:
                items = _json.loads(inv['line_items'] or '[]')
            except (ValueError, TypeError):
                continue
            for li in items:
                pc = (li.get('product_code') or li.get('sku') or '').strip()
                if not pc or pc in prices:
                    continue
                unit_price = float(li.get('unit_price', 0) or 0)
                uom = (li.get('uom') or li.get('unit_of_measure') or '').upper()
                if uom in ('C', 'PER C', 'PERC'):
                    unit_price = unit_price / 100.0
                elif uom in ('M', 'PER M', 'PERM'):
                    unit_price = unit_price / 1000.0
                prices[pc] = {
                    'price': round(unit_price, 4),
                    'source': f"{inv['supplier_name']} Inv ({inv['invoice_date']})",
                    'source_type': 'historical',
                }
    except Exception:
        pass  # invoices table may not have data

    return prices


@app.route('/api/plans/<int:pid>/takeoff', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_plans_takeoff(pid):
    """Generate material takeoff from calculator questions (multi-unit-type).

    Comprehensive HVAC material calculation engine.
    Calculates every line item from unit-type inputs + building-level questions.
    Supports: heat pump, condenser, gas furnace, mini split systems.
    Handles: R6/R8 flex split by floor, per-tonnage line sets, per-tonnage
    filter/return sizing, fire wrap, CRDs, zoning, outside air, and more.
    """
    import copy, math

    conn = get_db()
    plan = conn.execute('SELECT * FROM plans WHERE id = ?', (pid,)).fetchone()
    if not plan:
        conn.close()
        return jsonify({'error': 'Plan not found'}), 404

    data = request.get_json() or {}

    # ── Read unit types array ──
    unit_types = data.get('unit_types', [])
    if not unit_types:
        conn.close()
        return jsonify({'error': 'No unit types provided'}), 400

    # ── Read building-level inputs ──
    system_type      = data.get('system_type', 'heat_pump')
    thermostat       = data.get('thermostat', 'programmable')
    stories          = int(data.get('stories', 3))
    line_set_ft      = float(data.get('line_set_length', 25))
    flex_per_drop    = float(data.get('flex_per_drop', 1))
    outdoor_loc      = data.get('outdoor_loc', 'ground')        # ground | roof
    orientation      = data.get('orientation', 'horizontal')     # horizontal | vertical
    install_loc      = data.get('install_loc', 'attic')          # attic | closet
    mounting         = data.get('mounting', 'hanger')             # hanger | shelf
    exhaust_type     = data.get('exhaust_type', 'ceiling')        # ceiling | sidewall
    ductboard        = data.get('ductboard', False)
    ductboard_per_u  = float(data.get('ductboard_per_unit', 2))
    zoning           = data.get('zoning', False)
    fire_wrap        = data.get('fire_wrap', False)
    crds             = data.get('crds', True)
    crd_kitchen      = data.get('crd_kitchen', False)
    fresh_air_per_u  = float(data.get('fresh_air_per_unit', 1))
    drain_pans       = data.get('drain_pans', True)
    outside_air      = data.get('outside_air', False)
    wrap_dry_boots   = data.get('wrap_dryers_boots', False)
    passthroughs     = data.get('passthroughs', False)
    include_freight  = data.get('include_freight', True)
    # New building-level inputs
    r8_floors        = int(data.get('r8_floors', 1))              # how many top floors get R8 (IRC code)
    tstat_wire_gauge = data.get('tstat_wire_gauge', '18/5')       # 18/5 | 18/8
    condensate_mat   = data.get('condensate_material', 'pvc')     # pvc | cpvc
    line_insul_type  = data.get('line_insul_type', 'armaflex')    # armaflex | fiberglass
    gas_line_size    = data.get('gas_line_size', '1/2')           # 1/2 | 3/4
    corridor_units   = int(data.get('corridor_units', 0))         # separate HVAC for corridors/common areas

    # Job tax rate
    job = conn.execute('SELECT tax_rate FROM jobs WHERE id = ?', (plan['job_id'],)).fetchone()
    job_tax_rate = (job['tax_rate'] / 100.0) if job and job['tax_rate'] else 0.0885

    # ── Build takeoff ──
    takeoff = copy.deepcopy(TAKEOFF_TEMPLATE)
    takeoff['tax_rate'] = job_tax_rate
    takeoff['inputs'] = {
        'unit_types': unit_types, 'system_type': system_type, 'thermostat': thermostat,
        'stories': stories, 'line_set_length': line_set_ft, 'flex_per_drop': flex_per_drop,
        'outdoor_loc': outdoor_loc, 'orientation': orientation, 'install_loc': install_loc,
        'mounting': mounting, 'exhaust_type': exhaust_type, 'ductboard': ductboard,
        'ductboard_per_unit': ductboard_per_u, 'zoning': zoning, 'fire_wrap': fire_wrap,
        'crds': crds, 'crd_kitchen': crd_kitchen, 'fresh_air_per_unit': fresh_air_per_u,
        'drain_pans': drain_pans, 'outside_air': outside_air,
        'wrap_dryers_boots': wrap_dry_boots, 'passthroughs': passthroughs,
        'include_freight': include_freight, 'r8_floors': r8_floors,
        'tstat_wire_gauge': tstat_wire_gauge, 'condensate_material': condensate_mat,
        'line_insul_type': line_insul_type, 'gas_line_size': gas_line_size,
        'corridor_units': corridor_units,
    }

    # ── Aggregate across unit types ──
    U = sum(ut['qty'] for ut in unit_types)   # total units
    total_beds      = sum(ut['beds'] * ut['qty'] for ut in unit_types)
    total_baths     = sum(ut['baths'] * ut['qty'] for ut in unit_types)
    total_drops_6   = sum(ut.get('drops_6', 0) * ut['qty'] for ut in unit_types)
    total_drops_8   = sum(ut.get('drops_8', 0) * ut['qty'] for ut in unit_types)
    total_drops_10  = sum(ut.get('drops_10', 0) * ut['qty'] for ut in unit_types)
    total_returns   = sum(ut.get('returns', 1) * ut['qty'] for ut in unit_types)
    total_drops     = total_drops_6 + total_drops_8 + total_drops_10

    # ── Floor-based R6/R8 split (IRC code: top floor(s) = R8, lower = R6) ──
    # r8_floors = number of top floors requiring R8 flex
    # If building has 3 stories and r8_floors=1, top 1/3 of units get R8
    r8_fraction = min(r8_floors / max(stories, 1), 1.0)
    r6_fraction = 1.0 - r8_fraction
    # Apply to each drop size
    r8_drops_6  = math.ceil(total_drops_6 * r8_fraction)
    r6_drops_6  = total_drops_6 - r8_drops_6
    r8_drops_8  = math.ceil(total_drops_8 * r8_fraction)
    r6_drops_8  = total_drops_8 - r8_drops_8
    r8_drops_10 = math.ceil(total_drops_10 * r8_fraction)
    r6_drops_10 = total_drops_10 - r8_drops_10

    # Equipment counts per tonnage
    tonnage_counts = {}   # {tonnage_str: qty}
    hs_counts = {}        # {kw_str: qty}
    for ut in unit_types:
        t = ut['tonnage']
        ts = str(t) if t != int(t) else str(int(t))
        tonnage_counts[ts] = tonnage_counts.get(ts, 0) + ut['qty']
        hs = ut.get('heat_strip', 0)
        if hs > 0:
            hs_counts[str(hs)] = hs_counts.get(str(hs), 0) + ut['qty']

    # ── Per-tonnage line set sizing ──
    # Different tonnages may need different suction line sizes
    # liquid line is always 3/8; suction: <=3T=3/4", 3.5-4T=7/8", 5T=1-1/8"
    ls_ft_by_size = {'3/8': 0, '3/4': 0, '7/8': 0, '1-1/8': 0}
    for ut in unit_types:
        t = ut['tonnage']
        ut_ls = ut.get('line_set_override', line_set_ft)  # per-unit-type override
        ft = ut_ls * ut['qty']
        ls_ft_by_size['3/8'] += ft                        # liquid line always 3/8
        if t <= 3:
            ls_ft_by_size['3/4'] += ft
        elif t <= 4:
            ls_ft_by_size['7/8'] += ft
        else:
            ls_ft_by_size['1-1/8'] += ft
    total_ls_ft = ls_ft_by_size['3/8']   # same as sum of all unit line set lengths

    is_gas  = system_type == 'gas_furnace'
    is_hp   = system_type == 'heat_pump'
    is_ac   = system_type == 'condenser'
    is_mini = system_type == 'mini_split'
    above_living = install_loc == 'attic'  # unit above living space

    # Max tonnage (for whip sizing)
    max_tonnage = max(ut['tonnage'] for ut in unit_types)

    # ── Flex bags (R6 for lower floors, R8 for top floors) ──
    bags_6_r6  = math.ceil(r6_drops_6  * flex_per_drop) if r6_drops_6 > 0 else 0
    bags_6_r8  = math.ceil(r8_drops_6  * flex_per_drop) if r8_drops_6 > 0 else 0
    bags_8_r6  = math.ceil(r6_drops_8  * flex_per_drop) if r6_drops_8 > 0 else 0
    bags_8_r8  = math.ceil(r8_drops_8  * flex_per_drop) if r8_drops_8 > 0 else 0
    bags_10_r6 = math.ceil(r6_drops_10 * flex_per_drop) if r6_drops_10 > 0 else 0
    bags_10_r8 = math.ceil(r8_drops_10 * flex_per_drop) if r8_drops_10 > 0 else 0

    bath_ceil = math.ceil(total_baths)

    # ── Quantity map: (qty, use_flag) ──
    calc = {}

    # =====================================================================
    # ROUGH-IN SECTION
    # =====================================================================

    # ── CRDs / Bath Exhaust ──
    # 12x14 CRD = 1 per system/unit
    # 6x12x8 CRD Boot = 1 per supply drop (connects CRD to duct)
    # 4" or 6" Round CRD = per fresh air intake per unit
    # 80CFM CRD Fan = per bathroom (+ per kitchen if crd_kitchen checked)
    crd_fan_count = bath_ceil + (U if crd_kitchen else 0) if crds else 0
    fresh_air_total = math.ceil(U * fresh_air_per_u) if crds else 0
    calc['12x14 CRD']        = (U if crds else 0, crds)
    calc['12x14 Fire/Smoke Radiation Damper'] = (0, False)  # manual add for fire-rated buildings
    calc['6x12x8 CRD Boot']  = (total_drops if crds else 0, crds)
    calc['4\" Round CRD']     = (fresh_air_total if crds else 0, crds and fresh_air_total > 0)
    calc['6\" Round CRD']     = (0, False)  # alternative size, manual switch
    calc['80CFM CRD Fan']    = (crd_fan_count, crds)
    calc['Broan 688 Exhaust Fan'] = (0, False)  # alternative to CRD fan, manual add

    # ── Boots (per drop, sized to match) ──
    calc['8\" Foam Boot']    = (total_drops_8, total_drops_8 > 0)
    calc['6\" Foam Boot']    = (total_drops_6, total_drops_6 > 0)
    calc['8x6 Reducer']      = (0, False)  # manual add

    # ── Finger Taps: 1 per drop, sized to match drop diameter ──
    calc['6\" Finger Taps']  = (total_drops_6,  total_drops_6 > 0)
    calc['8\" Finger Taps']  = (total_drops_8,  total_drops_8 > 0)
    calc['10\" Finger Taps'] = (total_drops_10, total_drops_10 > 0)

    # ── Supply Boots: 1 per drop, sized to match ──
    calc['6\" Supply Boot']  = (total_drops_6,  total_drops_6 > 0)
    calc['8\" Supply Boot']  = (total_drops_8,  total_drops_8 > 0)
    calc['10\" Supply Boot'] = (total_drops_10, total_drops_10 > 0)

    # ── Flex Duct R6 (lower floors) ──
    calc['6\" R6 Flex (bags)']  = (bags_6_r6,  bags_6_r6 > 0)
    calc['8\" R6 Flex (bags)']  = (bags_8_r6,  bags_8_r6 > 0)
    calc['10\" R6 Flex (bags)'] = (bags_10_r6, bags_10_r6 > 0)
    calc['12\" R6 Flex (bags)'] = (0, False)

    # ── Flex Duct R8 (top floor - IRC code requirement) ──
    calc['6\" R8 Flex (bags)']  = (bags_6_r8,  bags_6_r8 > 0)
    calc['8\" R8 Flex (bags)']  = (bags_8_r8,  bags_8_r8 > 0)
    calc['10\" R8 Flex (bags)'] = (bags_10_r8, bags_10_r8 > 0)
    calc['12\" R8 Flex (bags)'] = (0, False)

    # ── Line Sets (per-tonnage sizing, aggregated by line size) ──
    # Sold in 50ft rolls; calculate rolls needed
    ls_rolls_38   = math.ceil(ls_ft_by_size['3/8']   / 50) if ls_ft_by_size['3/8'] > 0 else 0
    ls_rolls_34   = math.ceil(ls_ft_by_size['3/4']   / 50) if ls_ft_by_size['3/4'] > 0 else 0
    ls_rolls_78   = math.ceil(ls_ft_by_size['7/8']   / 50) if ls_ft_by_size['7/8'] > 0 else 0
    ls_rolls_118  = math.ceil(ls_ft_by_size['1-1/8'] / 50) if ls_ft_by_size['1-1/8'] > 0 else 0
    calc['3/8 Line Set (in feet)']   = (ls_rolls_38,  ls_rolls_38 > 0 and not is_mini)
    calc['3/4 Line Set (in feet)']   = (ls_rolls_34,  ls_rolls_34 > 0 and not is_mini)
    calc['7/8 Line Set (in feet)']   = (ls_rolls_78,  ls_rolls_78 > 0 and not is_mini)
    calc['1-1/8 Line Set (in feet)'] = (ls_rolls_118, ls_rolls_118 > 0 and not is_mini)

    # ── Line Set Insulation (Armaflex or Fiberglass, sized to match suction line) ──
    use_armaflex = line_insul_type == 'armaflex'
    use_fiberglass = line_insul_type == 'fiberglass'
    af_sticks_34  = math.ceil(ls_ft_by_size['3/4'] / 6)   if ls_ft_by_size['3/4'] > 0 else 0
    af_sticks_78  = math.ceil(ls_ft_by_size['7/8'] / 6)   if ls_ft_by_size['7/8'] > 0 else 0
    af_sticks_118 = math.ceil(ls_ft_by_size['1-1/8'] / 6) if ls_ft_by_size['1-1/8'] > 0 else 0
    calc['3/4\" Armaflex (6ft stick)']   = (af_sticks_34  if use_armaflex else 0, use_armaflex and af_sticks_34 > 0 and not is_mini)
    calc['7/8\" Armaflex (6ft stick)']   = (af_sticks_78  if use_armaflex else 0, use_armaflex and af_sticks_78 > 0 and not is_mini)
    calc['1-1/8\" Armaflex (6ft stick)'] = (af_sticks_118 if use_armaflex else 0, use_armaflex and af_sticks_118 > 0 and not is_mini)
    calc['Armaflex Glue (qt)']           = (max(1, math.ceil(U / 4)) if use_armaflex else 0, use_armaflex and not is_mini)
    # Fiberglass pipe wrap alternative (rolls cover ~50ft)
    fg_rolls_34  = math.ceil(ls_ft_by_size['3/4'] / 50)   if ls_ft_by_size['3/4'] > 0 else 0
    fg_rolls_78  = math.ceil(ls_ft_by_size['7/8'] / 50)   if ls_ft_by_size['7/8'] > 0 else 0
    fg_rolls_118 = math.ceil(ls_ft_by_size['1-1/8'] / 50) if ls_ft_by_size['1-1/8'] > 0 else 0
    calc['3/4\" Fiberglass Pipe Wrap (roll)']   = (fg_rolls_34  if use_fiberglass else 0, use_fiberglass and fg_rolls_34 > 0 and not is_mini)
    calc['7/8\" Fiberglass Pipe Wrap (roll)']   = (fg_rolls_78  if use_fiberglass else 0, use_fiberglass and fg_rolls_78 > 0 and not is_mini)
    calc['1-1/8\" Fiberglass Pipe Wrap (roll)'] = (fg_rolls_118 if use_fiberglass else 0, use_fiberglass and fg_rolls_118 > 0 and not is_mini)

    # ── Condensate (PVC vs CPVC based on condensate_material selection) ──
    use_pvc  = condensate_mat == 'pvc'
    use_cpvc = condensate_mat == 'cpvc'
    calc['3/4\" PVC (10ft stick)']      = (U if use_pvc else 0,  use_pvc)
    calc['3/4\" CPVC (10ft stick)']     = (U if use_cpvc else 0, use_cpvc)
    calc['3/4\" PVC Fittings (bag)']    = (max(1, math.ceil(U / 3)) if use_pvc else 0,  use_pvc)
    calc['3/4\" CPVC Fittings (bag)']   = (max(1, math.ceil(U / 3)) if use_cpvc else 0, use_cpvc)
    calc['PVC Cement & Primer Kit']     = (max(1, math.ceil(U / 6)) if use_pvc else 0,  use_pvc)
    calc['CPVC Cement & Primer Kit']    = (max(1, math.ceil(U / 6)) if use_cpvc else 0, use_cpvc)
    calc['Condensate Trap']             = (U, True)

    # ── Hangers & Hardware ──
    # Metal strap: 2ft per drop for hanger / 100ft per roll
    calc['Metal Hanger Strap (100ft roll)'] = (max(1, math.ceil(total_drops * 2 / 100)), True)
    calc['1\" Metal Screws (box)']          = (max(1, math.ceil(U / 5)), True)
    calc['Threaded Rod 3/8x36\"']           = (total_drops, total_drops > 0)
    calc['Beam Clamps (box of 25)']         = (0, False)  # manual add for steel framing

    # ── Sealant / Tape ──
    calc['Duct Mastic (gal)']  = (max(1, math.ceil(U / 3)), True)
    calc['Mastic Tape (roll)'] = (max(1, math.ceil(U / 4)), True)
    calc['Foil Tape (roll)']   = (max(1, math.ceil(U / 4)), True)

    # ── Return Air ──
    # Return boot/box sizing based on tonnage: <=2T=14x8, 2.5-3T=16x8, 3.5T+=20x8
    ret_14x8 = 0; ret_16x8 = 0; ret_20x8 = 0
    ret_box_20x20 = 0; ret_box_20x25 = 0; ret_box_24x24 = 0
    filt_20x25 = 0; filt_16x25 = 0; filt_20x20 = 0; filt_16x20 = 0; filt_24x24 = 0
    for ut in unit_types:
        t = ut['tonnage']
        r = ut.get('returns', 1) * ut['qty']
        if t <= 2:
            ret_14x8 += r;  ret_box_20x20 += r; filt_20x20 += r
        elif t <= 3:
            ret_16x8 += r;  ret_box_20x25 += r; filt_20x25 += r
        else:
            ret_20x8 += r;  ret_box_24x24 += r; filt_20x25 += r  # larger tonnage = 20x25 filter
    calc['14x8 Return Air Boot']  = (ret_14x8, ret_14x8 > 0)
    calc['16x8 Return Air Boot']  = (ret_16x8, ret_16x8 > 0)
    calc['20x8 Return Air Boot']  = (ret_20x8, ret_20x8 > 0)
    calc['20x20 Return Air Box']  = (ret_box_20x20, ret_box_20x20 > 0)
    calc['20x25 Return Air Box']  = (ret_box_20x25, ret_box_20x25 > 0)
    calc['24x24 Return Air Box']  = (ret_box_24x24, ret_box_24x24 > 0)
    calc['14x20 Return Grille']   = (0, False)
    calc['20x20 Return Grille']   = (0, False)
    # Filter grilles (sized by tonnage)
    calc['20x25 Filter Grille']   = (filt_20x25, filt_20x25 > 0)
    calc['16x25 Filter Grille']   = (0, False)
    calc['20x20 Filter Grille']   = (filt_20x20, filt_20x20 > 0)
    calc['24x24 Filter Grille']   = (filt_24x24, filt_24x24 > 0)

    # ── Plenum / Ductboard ──
    if ductboard:
        calc['Plenum Board (4x8 sheet)']    = (0, False)
        calc['Plenum Clips (box)']          = (0, False)
        calc['R8 Duct Board (4x10 sheet)']  = (ductboard_per_u * U, True)
        calc['R6 Duct Board (4x10 sheet)']  = (0, False)
        calc['Ductboard Staples (box)']     = (max(1, math.ceil(U / 4)), True)
    else:
        calc['Plenum Board (4x8 sheet)']    = (U, True)
        calc['Plenum Clips (box)']          = (max(1, math.ceil(U / 4)), True)
        calc['R8 Duct Board (4x10 sheet)']  = (0, False)
        calc['R6 Duct Board (4x10 sheet)']  = (0, False)
        calc['Ductboard Staples (box)']     = (0, False)

    # ── Electrical: Thermostat Wire ──
    # 250ft rolls, ~25ft per unit run => 1 roll per ~10 units
    tstat_rolls = max(1, math.ceil(U / 10))
    calc['18/5 Thermostat Wire (250ft)']  = (tstat_rolls if tstat_wire_gauge == '18/5' else 0, tstat_wire_gauge == '18/5')
    calc['18/8 Thermostat Wire (250ft)']  = (tstat_rolls if tstat_wire_gauge == '18/8' else 0, tstat_wire_gauge == '18/8')
    calc['Low Voltage Wire 18/2 (250ft)'] = (max(1, math.ceil(U / 10)), True)

    # ── Electrical: Disconnects & Whips ──
    calc['Disconnect 60A Non-Fused'] = (U, not is_mini)
    calc['Disconnect 60A Fused']     = (0, False)
    # Whip sizing: <=3T=3/4", larger=1"
    whip_small = sum(ut['qty'] for ut in unit_types if ut['tonnage'] <= 3)
    whip_large = sum(ut['qty'] for ut in unit_types if ut['tonnage'] > 3)
    calc['Whip 3/4\" x 6ft']            = (whip_small if not is_mini else 0, whip_small > 0 and not is_mini)
    calc['Whip 1\" x 6ft']              = (whip_large if not is_mini else 0, whip_large > 0 and not is_mini)
    calc['Wire Nuts / Connectors (box)'] = (max(1, math.ceil(U / 8)), True)
    calc['10-2 Romex (250ft)']           = (0, False)  # user adds manually if needed
    calc['6-2 Romex (125ft)']            = (0, False)
    calc['8-3 Romex (125ft)']            = (0, False)

    # ── Gas (only if gas furnace system) ──
    gas_sm = gas_line_size == '1/2'
    gas_lg = gas_line_size == '3/4'
    calc['Gas Flex 1/2\" (per ft)']   = (15 * U if is_gas and gas_sm else 0, is_gas and gas_sm)
    calc['Gas Flex 3/4\" (per ft)']   = (15 * U if is_gas and gas_lg else 0, is_gas and gas_lg)
    calc['Gas Valve 1/2\"']           = (U if is_gas and gas_sm else 0, is_gas and gas_sm)
    calc['Gas Valve 3/4\"']           = (U if is_gas and gas_lg else 0, is_gas and gas_lg)
    calc['Gas Drip Leg Kit']          = (U if is_gas else 0, is_gas)
    calc['Gas Connector 1/2\" (36\")'] = (U if is_gas and gas_sm else 0, is_gas and gas_sm)
    calc['Gas Connector 3/4\" (36\")'] = (U if is_gas and gas_lg else 0, is_gas and gas_lg)
    calc['Gas Sediment Trap']          = (U if is_gas else 0, is_gas)

    # ── Fire Protection ──
    # Fire wrap: rolls per rated penetration; estimate drops * stories for fire-rated assemblies
    fire_penetrations = total_drops * (stories - 1) if fire_wrap else 0
    fire_rolls = math.ceil(fire_penetrations / 8) if fire_penetrations > 0 else 0   # ~8 penetrations per roll
    calc['Fire Wrap (1.5\" x 24\" x 25ft roll)'] = (fire_rolls, fire_wrap)
    # Fire dampers at rated wall/floor penetrations
    fd_6  = math.ceil(total_drops_6  * (stories - 1) * 0.5) if fire_wrap else 0  # ~50% of penetrations need damper
    fd_8  = math.ceil(total_drops_8  * (stories - 1) * 0.5) if fire_wrap else 0
    fd_10 = math.ceil(total_drops_10 * (stories - 1) * 0.5) if fire_wrap else 0
    calc['Fire Damper 6\"']  = (fd_6,  fire_wrap and fd_6 > 0)
    calc['Fire Damper 8\"']  = (fd_8,  fire_wrap and fd_8 > 0)
    calc['Fire Damper 10\"'] = (fd_10, fire_wrap and fd_10 > 0)
    # Passthrough sleeves (fire-rated wall penetrations)
    pt_per_unit = 2 if passthroughs else 0
    pt_6  = math.ceil(total_drops_6  / total_drops * U * pt_per_unit) if total_drops > 0 and passthroughs else 0
    pt_8  = math.ceil(total_drops_8  / total_drops * U * pt_per_unit) if total_drops > 0 and passthroughs else 0
    pt_10 = math.ceil(total_drops_10 / total_drops * U * pt_per_unit) if total_drops > 0 and passthroughs else 0
    calc['Passthrough Sleeve 6\"']  = (pt_6,  passthroughs and pt_6 > 0)
    calc['Passthrough Sleeve 8\"']  = (pt_8,  passthroughs and pt_8 > 0)
    calc['Passthrough Sleeve 10\"'] = (pt_10, passthroughs and pt_10 > 0)
    # Firestop caulk: 1 per 3 units baseline + 1 per unit for fire wrap builds
    calc['Firestop Caulk (tube)'] = (max(1, math.ceil(U / 3)) + (U if fire_wrap else 0), True)
    calc['Firestop Putty Pad (box)'] = (max(1, math.ceil(U / 6)) if fire_wrap else 0, fire_wrap)
    calc['Smoke Damper (round)'] = (0, False)  # manual add

    # =====================================================================
    # TRIM OUT SECTION
    # =====================================================================

    # ── Shorts & Smalls ──
    calc['3\" Pump Ups']        = (3 * U, True)                         # 3 per unit
    calc['Safe T Switch']       = (U if above_living else 0, above_living)  # required above living space
    calc['Float Switch']        = (U, True)                              # 1 per unit always
    calc['Drain Pan (plastic)'] = (U if (drain_pans or above_living) else 0, drain_pans or above_living)
    calc['Drain Pan (metal)']   = (0, False)
    calc['P-Trap 3/4\" PVC']    = (U, True)                              # 1 per unit condensate

    # ── Registers (Supply) — sized from drop size ──
    # 6" drops: mix of 8x4 and 10x4 registers
    # 8" drops: mix of 10x6 and 12x4 registers
    # 10" drops: mix of 12x6 and 14x6 registers
    calc['6x6 Register (white)']  = (0, False)
    calc['8x4 Register (white)']  = (math.ceil(total_drops_6 * 0.5), total_drops_6 > 0)
    calc['10x4 Register (white)'] = (total_drops_6 - math.ceil(total_drops_6 * 0.5), total_drops_6 > 0)
    calc['10x6 Register (white)'] = (math.ceil(total_drops_8 * 0.5), total_drops_8 > 0)
    calc['12x4 Register (white)'] = (total_drops_8 - math.ceil(total_drops_8 * 0.5), total_drops_8 > 0)
    calc['12x6 Register (white)'] = (math.ceil(total_drops_10 * 0.5), total_drops_10 > 0)
    calc['14x6 Register (white)'] = (total_drops_10 - math.ceil(total_drops_10 * 0.5), total_drops_10 > 0)
    calc['14x8 Register (white)'] = (0, False)
    calc['6\" Round Ceiling Diffuser']  = (0, False)
    calc['8\" Round Ceiling Diffuser']  = (0, False)
    calc['10\" Round Ceiling Diffuser'] = (0, False)

    # ── Filters (sized by tonnage, 12-packs for initial stock) ──
    calc['20x25x1 Filter (12-pack)'] = (max(1, math.ceil(filt_20x25 / 12)) if filt_20x25 > 0 else 0, filt_20x25 > 0)
    calc['16x25x1 Filter (12-pack)'] = (0, False)
    calc['20x20x1 Filter (12-pack)'] = (max(1, math.ceil(filt_20x20 / 12)) if filt_20x20 > 0 else 0, filt_20x20 > 0)
    calc['16x20x1 Filter (12-pack)'] = (max(1, math.ceil(filt_16x20 / 12)) if filt_16x20 > 0 else 0, filt_16x20 > 0)
    calc['24x24x1 Filter (12-pack)'] = (max(1, math.ceil(filt_24x24 / 12)) if filt_24x24 > 0 else 0, filt_24x24 > 0)

    # ── Mounting / Pads ──
    calc['Cork Pads (set of 4)']       = (U, True)                         # vibration isolation, 1 set per indoor unit
    calc['Condenser Pad (plastic)']    = (U if outdoor_loc == 'ground' and not is_mini else 0, outdoor_loc == 'ground' and not is_mini)
    calc['Condenser Pad (concrete)']   = (0, False)
    calc['Wall Brackets (pair)']       = (U if is_mini else 0, is_mini)
    calc['Rooftop Curb Adapter']       = (U if outdoor_loc == 'roof' else 0, outdoor_loc == 'roof' and not is_mini)
    calc['Equipment Stand (28\" H)']   = (0, False)  # manual add

    # ── Refrigerant ──
    calc['Refrigerant R-410A (25lb)']    = (max(1, math.ceil(U / 4)), not is_mini)
    calc['Refrigerant R-410A (50lb)']    = (0, False)
    calc['Nitrogen (tank rental + gas)'] = (max(1, math.ceil(U / 10)), True)

    # ── Brazing / Soldering ──
    calc['Silver Brazing Rods (pkg)']  = (max(1, math.ceil(U / 5)), not is_mini)
    calc['Stay-Brite #8 Solder (1lb)'] = (max(1, math.ceil(U / 8)), not is_mini)
    calc['Flux Paste']                  = (max(1, math.ceil(U / 10)), not is_mini)
    calc['MAP Gas Cylinder']            = (max(1, math.ceil(U / 6)), not is_mini)

    # ── Tools (consumable) ──
    calc['Pipe Cutter 1/4-1-5/8\"']    = (0, False)
    calc['Flare Tool Kit']              = (0, False)
    calc['Torque Wrench (refrigerant)'] = (0, False)
    calc['Duct Knife / Blade (pkg)']    = (max(1, math.ceil(U / 10)), True)
    calc['Hacksaw Blades (pkg)']        = (max(1, math.ceil(U / 15)), True)
    calc['Hole Saw Kit (HVAC sizes)']   = (0, False)

    # ── Misc Supplies ──
    calc['Zip Ties 11\" (bag of 100)']  = (max(1, math.ceil(U / 8)), True)
    calc['Zip Ties 14\" (bag of 100)']  = (0, False)
    calc['Caulk / Firestop (tube)']     = (max(1, math.ceil(U / 3)) + (U if fire_wrap else 0), True)
    calc['Metal Tape Measure 25ft']     = (0, False)
    calc['Spray Paint (marking)']       = (max(1, math.ceil(U / 10)), True)
    calc['Duct Seal Putty (1lb)']       = (max(1, math.ceil(U / 6)), True)
    calc['Electrical Tape (10-pk)']     = (max(1, math.ceil(U / 15)), True)
    calc['Tie Wire (roll)']             = (max(1, math.ceil(U / 10)), True)
    calc['Condensate Pump (mini)']      = (0, False)
    calc['Condensate Pump (full size)'] = (0, False)
    calc['UV Light Kit']                = (0, False)
    calc['Media Filter Cabinet']        = (0, False)

    # =====================================================================
    # EQUIPMENT SECTION
    # =====================================================================

    # ── Air Handlers — aggregate per tonnage ──
    for t in ['1.5', '2', '2.5', '3', '3.5', '4', '5']:
        qty = tonnage_counts.get(t, 0)
        calc[f'{t} Ton Air Handler'] = (qty if not is_mini and not is_gas and qty > 0 else 0, not is_mini and not is_gas and qty > 0)
        calc[f'{t} Ton Condenser']   = (qty if (is_ac or is_gas) and qty > 0 else 0, (is_ac or is_gas) and qty > 0)
        calc[f'{t} Ton Heat Pump']   = (qty if is_hp and qty > 0 else 0, is_hp and qty > 0)

    # ── Heat Strips — aggregate per kW ──
    for kw in ['5', '8', '10', '15', '20']:
        qty = hs_counts.get(kw, 0)
        calc[f'{kw}kW Heat Strip'] = (qty if not is_mini and not is_gas and qty > 0 else 0, not is_mini and not is_gas and qty > 0)

    # ── Gas Furnaces — sized by tonnage ──
    for btu in ['40K', '60K', '80K', '100K', '120K']:
        calc[f'{btu} BTU Gas Furnace 96%'] = (0, False)
    if is_gas:
        for ut in unit_types:
            t = ut['tonnage']
            if t >= 5:
                calc['120K BTU Gas Furnace 96%'] = (calc.get('120K BTU Gas Furnace 96%', (0, False))[0] + ut['qty'], True)
            elif t >= 4:
                calc['100K BTU Gas Furnace 96%'] = (calc.get('100K BTU Gas Furnace 96%', (0, False))[0] + ut['qty'], True)
            elif t >= 2.5:
                calc['80K BTU Gas Furnace 96%'] = (calc.get('80K BTU Gas Furnace 96%', (0, False))[0] + ut['qty'], True)
            elif t >= 1.5:
                calc['60K BTU Gas Furnace 96%'] = (calc.get('60K BTU Gas Furnace 96%', (0, False))[0] + ut['qty'], True)
            else:
                calc['40K BTU Gas Furnace 96%'] = (calc.get('40K BTU Gas Furnace 96%', (0, False))[0] + ut['qty'], True)

    # ── Mini Splits ──
    for sz in ['9K', '12K', '18K', '24K', '36K']:
        calc[f'Mini Split {sz} BTU'] = (0, False)
    for z in ['2-zone', '3-zone', '4-zone']:
        calc[f'Mini Split Multi-Zone Outdoor {z}'] = (0, False)
    if is_mini:
        mini_sizes = {'9K': 9000, '12K': 12000, '18K': 18000, '24K': 24000, '36K': 36000}
        for ut in unit_types:
            btu = ut['tonnage'] * 12000
            best = min(mini_sizes.keys(), key=lambda k: abs(mini_sizes[k] - btu))
            key = f'Mini Split {best} BTU'
            calc[key] = (calc.get(key, (0, False))[0] + ut['qty'], True)

    # ── Thermostats — 1 per unit ──
    calc['Programmable Thermostat']        = (U if thermostat == 'programmable' else 0, thermostat == 'programmable')
    calc['WiFi Thermostat']                = (U if thermostat == 'wifi' else 0, thermostat == 'wifi')
    calc['Smart Thermostat (Ecobee/Nest)'] = (U if thermostat == 'smart' else 0, thermostat == 'smart')

    # ── Ventilation / Exhaust ──
    # ERV/HRV: if outside_air, 1 per unit (or per corridor if corridors specified)
    erv_qty = corridor_units if corridor_units > 0 and outside_air else (U if outside_air else 0)
    calc['ERV Unit']               = (erv_qty, outside_air)
    calc['HRV Unit']               = (0, False)
    calc['Inline Exhaust Fan 6\"'] = (0, False)
    calc['Inline Exhaust Fan 8\"'] = (0, False)
    # Bath exhaust fans: only if NOT using CRDs (CRDs replace bath fans)
    bath_fan_qty = bath_ceil if not crds else 0
    calc['Bath Exhaust Fan 50CFM']  = (0, False)
    calc['Bath Exhaust Fan 80CFM']  = (bath_fan_qty if not crds and exhaust_type == 'ceiling' else 0, not crds and exhaust_type == 'ceiling')
    calc['Bath Exhaust Fan 110CFM'] = (bath_fan_qty if not crds and exhaust_type == 'sidewall' else 0, not crds and exhaust_type == 'sidewall')
    calc['Range Hood Vent Kit']     = (0, False)
    # Dryer vent: if wrap_dryers_boots, full kit + components per unit
    calc['Dryer Vent Kit']          = (U if wrap_dry_boots else 0, wrap_dry_boots)
    calc['Dryer Vent Elbow (4\")']  = (2 * U if wrap_dry_boots else 0, wrap_dry_boots)   # ~2 elbows per run
    calc['Dryer Vent Hose (4\" x 8ft)'] = (U if wrap_dry_boots else 0, wrap_dry_boots)
    calc['Dryer Vent Wall Cap']     = (U if wrap_dry_boots else 0, wrap_dry_boots)

    # ── Zoning ──
    if zoning:
        calc['Zoning Panel (2 zone)']   = (U, True)
        calc['Zone Damper (round)']     = (total_drops, True)
        calc['Zone Thermostat Sensor']  = (U, True)
    else:
        calc['Zoning Panel (2 zone)']   = (0, False)
        calc['Zone Damper (round)']     = (0, False)
        calc['Zone Thermostat Sensor']  = (0, False)
    calc['Zoning Panel (3 zone)'] = (0, False)
    calc['Zone Damper (rect)']    = (0, False)

    # =====================================================================
    # FREIGHT SECTION
    # =====================================================================
    calc['Freight to jobsite']  = (1 if include_freight else 0, include_freight)
    calc['Crane / Lift Rental'] = (1 if outdoor_loc == 'roof' else 0, outdoor_loc == 'roof')
    calc['Dumpster Rental']     = (1 if U >= 20 else 0, U >= 20)  # auto-add dumpster for large jobs
    calc['Scissor Lift Rental'] = (0, False)
    calc['License']             = (0, False)
    calc['Permits']             = (0, False)

    # ── New workbook items (default = manual add) ──
    # Rough-In: Round Pipe
    calc['4\" Adjustable 90']   = (0, False)
    calc['3\" Adjustable 90']   = (0, False)
    calc['4x3 Reducer']         = (0, False)
    calc['3\" Conductor Pipe']  = (0, False)
    calc['4\" Conductor Pipe']  = (0, False)
    # Rough-In: Vent Boxes
    for vb in ['4\" Triple Brick Vent Box', '4\" Single Brick Vent Box',
               '4\" Single Siding Vent Box', '4\" Single Soffit Vent Box',
               '6\" Single Brick Vent Box', '6\" Single Siding Vent Box',
               '6\" Single Soffit Vent Box', '4\" Triple Siding Vent Box',
               '4\" Triple Soffit Vent Box']:
        calc[vb] = (0, False)
    # Rough-In: Electrical
    calc['14/4 S.O.Cable']         = (0, False)
    calc['18/8 T-stat Wire per foot'] = (0, False)
    # Rough-In: Hardware
    calc['3/4 Screws']             = (0, False)
    calc['Rails']                  = (0, False)
    calc['Plumber Strap']          = (0, False)
    calc['4.25\" Dryer Box']       = (0, False)
    calc['16\" Boca Plate']        = (0, False)
    # Rough-In: Tape/Sealant
    calc['Black Duct Tape']        = (0, False)
    calc['Flex Fix Tape']          = (0, False)
    calc['Silicone']               = (0, False)
    calc['Fire Caulk 5 Gal']      = (0, False)
    calc['Paint Brush']            = (0, False)
    # Rough-In: Misc
    calc['SA Dryer Fire Wrap 16\"'] = (0, False)
    calc['Duct Wrap']              = (0, False)
    calc['3\" Gray Duct Strap']    = (0, False)
    calc['Mini Split Line Set']    = (0, False)
    # Trim Out: PVC Fittings
    calc['3/4 PVC 90']             = (0, False)
    calc['3/4 PVC Coupling']       = (0, False)
    calc['3/4 PVC Male Adapter']   = (0, False)
    calc['3/4 PVC Tee']            = (0, False)
    calc['1QT PVC Cement']         = (0, False)
    # Trim Out: Shorts & Smalls
    calc['6\" Pump Ups']           = (0, False)
    calc['30x30 Drain Pan']        = (0, False)
    calc['30x60 Drain Pan']        = (0, False)
    # Trim Out: Registers (workbook)
    calc['8\" Supply Register']    = (0, False)
    calc['6\" Supply Register']    = (0, False)
    calc['12x6 Stamped Supply']    = (0, False)
    calc['16x8 Hart Cooley Pass Through'] = (0, False)
    calc['24x12 Stamped Return']   = (0, False)
    calc['30x14 Stamped Return']   = (0, False)
    calc['30x20 Stamped Return']   = (0, False)
    calc['24x24x8 Drop-ins']       = (0, False)
    calc['4\" Venthood Covers']    = (0, False)
    # Trim Out: Mounting
    calc['1\" Anchor Kit']         = (0, False)
    # Trim Out: Refrigerant
    calc['R454B Refrigerant']      = (0, False)
    # Trim Out: Brazing
    calc['Solder']                 = (0, False)
    calc['Silver Locking Caps']    = (0, False)
    # Trim Out: Consumables
    calc['Acetylene Refill']       = (0, False)
    calc['Oxygen Refill']          = (0, False)
    # Trim Out: Tools
    calc['12\" SawZaw Blade']      = (0, False)
    calc['4-3/8\" Hole Saw']       = (0, False)
    # Equipment: Thermostats
    calc['Heat Pump Thermostat']   = (U if is_hp and thermostat == 'programmable' else 0, is_hp and thermostat == 'programmable')
    calc['TH8 Vision Pro Honeywell'] = (0, False)
    # Equipment: Corridor AH
    for t in ['1.5', '2', '2.5', '3', '3.5', '4', '5']:
        calc[f'{t} Ton Corridor Air Handler'] = (0, False)
        calc[f'{t} Ton Corridor Heat Kit']    = (0, False)
    # Equipment: Mini Splits (workbook indoor/outdoor)
    calc['Mini Split 9K Indoor']     = (0, False)
    calc['Mini Split 9K Outdoor']    = (0, False)
    calc['Mini Split 12K Indoor']    = (0, False)
    calc['Mini Split 12K Outdoor']   = (0, False)
    calc['Mini Split 18K Cassette']  = (0, False)
    calc['Mini Split 18K Indoor']    = (0, False)
    calc['Mini Split 18K Outdoor HP'] = (0, False)
    calc['Mini Split 24K Indoor']    = (0, False)
    calc['Mini Split 24K Outdoor']   = (0, False)

    # ── Resolve dynamic prices from supplier quotes/invoices ──
    resolved_prices = _resolve_takeoff_prices(conn, plan['job_id'])
    price_counts = {'quote': 0, 'historical': 0, 'default': 0}

    # ── Apply to template ──
    for section in takeoff['sections']:
        for item in section['items']:
            if item['part'] in calc:
                qty, use = calc[item['part']]
                item['quantity'] = qty
                item['use'] = use
            else:
                item['quantity'] = 0
                item['use'] = item.get('default_use', False)

            # Dynamic price override from supplier data
            sku = item.get('sku', '')
            if sku and sku in resolved_prices:
                rp = resolved_prices[sku]
                item['price'] = rp['price']
                item['price_source'] = rp['source']
                item['price_source_type'] = rp['source_type']
                price_counts[rp['source_type']] += 1
            else:
                item['price_source'] = 'Template default'
                item['price_source_type'] = 'default'
                price_counts['default'] += 1

            q = item['quantity']
            waste = q * takeoff['waste_factor'] if item['use'] and q > 0 else 0
            item['total_with_waste'] = round((q + waste) * item['price'], 2) if item['use'] else 0

    # ── Summary ──
    subtotal = sum(
        item['total_with_waste']
        for section in takeoff['sections']
        for item in section['items']
        if item['use']
    )
    tax = round(subtotal * job_tax_rate, 2)

    # Per-apartment = grand_total / number of apartments (U = total units/apartments)
    grand_total = round(subtotal + tax, 2)
    price_per_apt = round(grand_total / U, 2) if U > 0 else 0
    # Number of HVAC systems (units + corridor units if any)
    num_systems = U + corridor_units
    price_per_system = round(grand_total / num_systems, 2) if num_systems > 0 else 0

    takeoff['summary'] = {
        'subtotal': round(subtotal, 2),
        'tax_rate': job_tax_rate,
        'tax': tax,
        'grand_total': grand_total,
        'total_units': U,
        'total_bedrooms': total_beds,
        'total_bathrooms': math.ceil(total_baths),
        'total_drops': total_drops,
        'total_systems': num_systems,
        'price_per_apartment': price_per_apt,
        'price_per_system': price_per_system,
        'price_sources': price_counts,
    }
    takeoff['generated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    # Save
    conn.execute(
        "UPDATE plans SET takeoff_data=?, status='Takeoff Complete', updated_at=datetime('now','localtime') WHERE id=?",
        (json.dumps(takeoff), pid)
    )
    conn.commit()
    conn.close()
    return jsonify(takeoff)


# ─── Supplier Quotes (Phase 2) ──────────────────────────────────

SUPPLIER_QUOTES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'supplier_quotes')
os.makedirs(SUPPLIER_QUOTES_DIR, exist_ok=True)

@app.route('/supplier-quotes')
@role_required('owner', 'admin', 'project_manager')
def supplier_quotes_list():
    return render_template('supplier_quotes/list.html')

@app.route('/supplier-quotes/compare')
@role_required('owner', 'admin', 'project_manager')
def supplier_quotes_compare():
    return render_template('supplier_quotes/compare.html')

@app.route('/supplier-quotes/<int:qid>')
@role_required('owner', 'admin', 'project_manager')
def supplier_quotes_detail(qid):
    return render_template('supplier_quotes/detail.html', quote_id=qid)

@app.route('/api/supplier-quotes')
@api_role_required('owner', 'admin', 'project_manager')
def api_supplier_quotes():
    job_id = request.args.get('job_id', '')
    status = request.args.get('status', '')
    supplier = request.args.get('supplier', '')
    conn = get_db()
    query = '''SELECT sq.*, j.name as job_name FROM supplier_quotes sq
               LEFT JOIN jobs j ON sq.job_id = j.id WHERE 1=1'''
    params = []
    if job_id:
        query += ' AND sq.job_id = ?'
        params.append(job_id)
    if status:
        query += ' AND sq.status = ?'
        params.append(status)
    if supplier:
        query += ' AND sq.supplier_name LIKE ?'
        params.append(f'%{supplier}%')
    query += ' ORDER BY sq.updated_at DESC'
    quotes = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(q) for q in quotes])

@app.route('/api/supplier-quotes', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_create_supplier_quote():
    conn = get_db()
    if request.content_type and 'multipart' in request.content_type:
        data = request.form
        file = request.files.get('file')
        file_path = ''
        file_hash = ''
        if file and file.filename:
            from werkzeug.utils import secure_filename
            content = file.read()
            file_hash = hashlib.md5(content).hexdigest()
            fname = secure_filename(file.filename)
            fname = f"{int(datetime.now().timestamp())}_{fname}"
            file.seek(0)
            file.save(os.path.join(SUPPLIER_QUOTES_DIR, fname))
            file_path = fname
        cursor = conn.execute(
            '''INSERT INTO supplier_quotes (job_id, plan_id, supplier_name, supplier_config_id,
               quote_number, quote_date, expiration_date, status, notes, file_path, file_hash, created_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (data.get('job_id'), data.get('plan_id'), data.get('supplier_name', ''),
             data.get('supplier_config_id'), data.get('quote_number', ''),
             data.get('quote_date', ''), data.get('expiration_date', ''),
             data.get('status', 'Received'), data.get('notes', ''),
             file_path, file_hash, session.get('user_id'))
        )
    else:
        data = request.get_json()
        cursor = conn.execute(
            '''INSERT INTO supplier_quotes (job_id, plan_id, supplier_name, supplier_config_id,
               quote_number, quote_date, expiration_date, status, notes, created_by)
               VALUES (?,?,?,?,?,?,?,?,?,?)''',
            (data.get('job_id'), data.get('plan_id'), data.get('supplier_name', ''),
             data.get('supplier_config_id'), data.get('quote_number', ''),
             data.get('quote_date', ''), data.get('expiration_date', ''),
             data.get('status', 'Requested'), data.get('notes', ''), session.get('user_id'))
        )
    qid = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': qid}), 201

@app.route('/api/supplier-quotes/<int:qid>')
@api_role_required('owner', 'admin', 'project_manager')
def api_supplier_quote_detail(qid):
    conn = get_db()
    q = conn.execute(
        '''SELECT sq.*, j.name as job_name FROM supplier_quotes sq
           LEFT JOIN jobs j ON sq.job_id = j.id WHERE sq.id = ?''', (qid,)
    ).fetchone()
    if not q:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    items = conn.execute('SELECT * FROM supplier_quote_items WHERE quote_id = ? ORDER BY line_number', (qid,)).fetchall()
    conn.close()
    result = dict(q)
    result['items'] = [dict(i) for i in items]
    return jsonify(result)

@app.route('/api/supplier-quotes/<int:qid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_supplier_quote(qid):
    data = request.get_json()
    conn = get_db()
    conn.execute(
        '''UPDATE supplier_quotes SET supplier_name=?, quote_number=?, quote_date=?,
           expiration_date=?, status=?, subtotal=?, tax_amount=?, freight=?, total=?,
           notes=?, file_path=?, updated_at=datetime('now','localtime') WHERE id=?''',
        (data.get('supplier_name',''), data.get('quote_number',''), data.get('quote_date',''),
         data.get('expiration_date',''), data.get('status','Requested'),
         float(data.get('subtotal',0) or 0), float(data.get('tax_amount',0) or 0),
         float(data.get('freight',0) or 0), float(data.get('total',0) or 0),
         data.get('notes',''), data.get('file_path',''), qid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/supplier-quotes/<int:qid>/items', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_save_supplier_quote_items(qid):
    data = request.get_json()
    items = data.get('items', [])
    conn = get_db()
    conn.execute('DELETE FROM supplier_quote_items WHERE quote_id = ?', (qid,))
    subtotal = 0
    for i, item in enumerate(items):
        ext = float(item.get('quantity',0) or 0) * float(item.get('unit_price',0) or 0)
        conn.execute(
            '''INSERT INTO supplier_quote_items (quote_id, line_number, sku, description,
               quantity, unit_price, extended_price, takeoff_sku, notes, requires_submittal, submittal_file_id)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
            (qid, i+1, item.get('sku',''), item.get('description',''),
             float(item.get('quantity',0) or 0), float(item.get('unit_price',0) or 0),
             ext, item.get('takeoff_sku',''), item.get('notes',''),
             1 if item.get('requires_submittal') else 0,
             int(item['submittal_file_id']) if item.get('submittal_file_id') else None)
        )
        subtotal += ext
    conn.execute('UPDATE supplier_quotes SET subtotal=?, updated_at=datetime("now","localtime") WHERE id=?', (subtotal, qid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'subtotal': subtotal})

@app.route('/api/supplier-quotes/<int:qid>/upload', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_upload_supplier_quote_file(qid):
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'No file uploaded'}), 400
    from werkzeug.utils import secure_filename
    content = file.read()
    file_hash = hashlib.md5(content).hexdigest()
    fname = secure_filename(file.filename)
    fname = f"{int(datetime.now().timestamp())}_{fname}"
    file.seek(0)
    file.save(os.path.join(SUPPLIER_QUOTES_DIR, fname))
    conn = get_db()
    conn.execute("UPDATE supplier_quotes SET file_path=?, file_hash=?, updated_at=datetime('now','localtime') WHERE id=?", (fname, file_hash, qid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'file_path': fname})

@app.route('/api/supplier-quotes/<int:qid>/file')
@api_role_required('owner', 'admin', 'project_manager')
def api_supplier_quote_file(qid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM supplier_quotes WHERE id = ?', (qid,)).fetchone()
    conn.close()
    if not row or not row['file_path']:
        return jsonify({'error': 'No file'}), 404
    fpath = os.path.join(SUPPLIER_QUOTES_DIR, row['file_path'])
    if not os.path.exists(fpath):
        return jsonify({'error': 'File not found'}), 404
    return send_file(fpath)

@app.route('/api/supplier-quotes/<int:qid>/select', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_select_supplier_quote(qid):
    conn = get_db()
    q = conn.execute('SELECT job_id FROM supplier_quotes WHERE id = ?', (qid,)).fetchone()
    if not q:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    conn.execute('UPDATE supplier_quotes SET is_baseline = 0 WHERE job_id = ?', (q['job_id'],))
    conn.execute("UPDATE supplier_quotes SET is_baseline = 1, status = 'Selected', updated_at=datetime('now','localtime') WHERE id = ?", (qid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/supplier-quotes/compare')
@api_role_required('owner', 'admin', 'project_manager')
def api_supplier_quote_compare():
    job_id = request.args.get('job_id')
    if not job_id:
        return jsonify({'error': 'job_id required'}), 400
    conn = get_db()
    quotes = conn.execute(
        "SELECT * FROM supplier_quotes WHERE job_id = ? AND status IN ('Received','Reviewing','Selected') ORDER BY supplier_name",
        (job_id,)
    ).fetchall()
    result = []
    for q in quotes:
        items = conn.execute('SELECT * FROM supplier_quote_items WHERE quote_id = ? ORDER BY line_number', (q['id'],)).fetchall()
        d = dict(q)
        d['items'] = [dict(i) for i in items]
        result.append(d)
    conn.close()
    return jsonify(result)

@app.route('/api/supplier-quotes/<int:qid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_supplier_quote(qid):
    conn = get_db()
    row = conn.execute('SELECT file_path FROM supplier_quotes WHERE id = ?', (qid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if row['file_path']:
        fpath = os.path.join(SUPPLIER_QUOTES_DIR, row['file_path'])
        if os.path.exists(fpath):
            os.remove(fpath)
    conn.execute('DELETE FROM supplier_quotes WHERE id = ?', (qid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/supplier-quotes/<int:qid>/toggle-submittal', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_toggle_quote_item_submittal(qid):
    """Toggle requires_submittal on a quote line item."""
    data = request.get_json(force=True)
    item_id = data.get('item_id')
    requires = 1 if data.get('requires_submittal') else 0
    conn = get_db()
    conn.execute(
        "UPDATE supplier_quote_items SET requires_submittal = ? WHERE id = ? AND quote_id = ?",
        (requires, item_id, qid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/supplier-quotes/<int:qid>/generate-submittals', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_generate_submittals_from_quote(qid):
    """Generate submittals for all quote items marked requires_submittal.
    Matches against submittal_files library by description/vendor.
    Returns which items were matched, created, and which are missing."""
    conn = get_db()
    quote = conn.execute(
        'SELECT sq.*, j.name as job_name FROM supplier_quotes sq LEFT JOIN jobs j ON sq.job_id = j.id WHERE sq.id = ?',
        (qid,)
    ).fetchone()
    if not quote:
        conn.close()
        return jsonify({'error': 'Quote not found'}), 404

    job_id = quote['job_id']
    if not job_id:
        conn.close()
        return jsonify({'error': 'Quote has no job assigned'}), 400

    # Get items that require submittals
    items = conn.execute(
        'SELECT * FROM supplier_quote_items WHERE quote_id = ? AND requires_submittal = 1 ORDER BY line_number',
        (qid,)
    ).fetchall()

    if not items:
        conn.close()
        return jsonify({'error': 'No line items marked as requiring submittals'}), 400

    # Get all library files for matching
    library = conn.execute('SELECT * FROM submittal_files ORDER BY title').fetchall()

    # Get existing submittals for this job to avoid duplicates
    existing = conn.execute(
        'SELECT description, vendor, submittal_file_id FROM submittals WHERE job_id = ?',
        (job_id,)
    ).fetchall()
    existing_descs = set((r['description'].lower().strip(), (r['vendor'] or '').lower().strip()) for r in existing)

    created = []
    matched = []
    missing = []

    for item in items:
        desc = (item['description'] or '').strip()
        sku = (item['sku'] or '').strip()
        vendor = (quote['supplier_name'] or '').strip()

        # Skip if already exists in this job's submittals
        if (desc.lower(), vendor.lower()) in existing_descs:
            matched.append({'item_id': item['id'], 'description': desc, 'status': 'already_exists'})
            continue

        # Try to find a match in the library
        lib_match = None

        # If item already has a linked library file, use that
        if item['submittal_file_id']:
            lib_match = conn.execute('SELECT * FROM submittal_files WHERE id = ?', (item['submittal_file_id'],)).fetchone()

        # Search library by description keywords or SKU
        if not lib_match:
            for lib in library:
                lib_title = (lib['title'] or '').lower()
                lib_desc = (lib['description'] or '').lower()
                lib_keywords = (lib['keywords'] or '').lower() if 'keywords' in lib.keys() else ''
                lib_searchable = lib_title + ' ' + lib_desc + ' ' + lib_keywords
                item_desc_lower = desc.lower()

                # Match by SKU if available
                if sku and (sku.lower() in lib_title or sku.lower() in lib_keywords):
                    lib_match = lib
                    break
                # Match by description keywords (at least 3 significant words match)
                desc_words = [w for w in item_desc_lower.split() if len(w) > 3]
                if desc_words:
                    title_matches = sum(1 for w in desc_words if w in lib_searchable)
                    if title_matches >= min(3, len(desc_words)):
                        lib_match = lib
                        break

        # Create the submittal entry
        max_num = conn.execute('SELECT MAX(submittal_number) FROM submittals WHERE job_id = ?',
                               (job_id,)).fetchone()[0] or 0

        submittal_file_id = lib_match['id'] if lib_match else None
        file_path = ''
        if lib_match and lib_match['file_path']:
            file_path = lib_match['file_path']

        conn.execute(
            '''INSERT INTO submittals (job_id, submittal_number, description, vendor,
               submittal_file_id, file_path, status, created_by)
               VALUES (?,?,?,?,?,?,?,?)''',
            (job_id, max_num + 1, desc, vendor, submittal_file_id,
             file_path, 'Pending', session.get('user_id'))
        )

        # Update the quote item with the library link if found
        if lib_match and not item['submittal_file_id']:
            conn.execute(
                'UPDATE supplier_quote_items SET submittal_file_id = ? WHERE id = ?',
                (lib_match['id'], item['id'])
            )

        if lib_match:
            matched.append({'item_id': item['id'], 'description': desc, 'status': 'matched',
                           'library_title': lib_match['title']})
        else:
            missing.append({'item_id': item['id'], 'description': desc, 'sku': sku, 'vendor': vendor})

        created.append({'description': desc, 'has_library_file': bool(lib_match)})

    conn.commit()
    conn.close()

    return jsonify({
        'ok': True,
        'created': len(created),
        'matched': len([m for m in matched if m['status'] == 'matched']),
        'already_existed': len([m for m in matched if m['status'] == 'already_exists']),
        'missing': missing,
        'details': {'created': created, 'matched': matched, 'missing': missing}
    })

@app.route('/api/supplier-quotes/<int:qid>/parse', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_parse_supplier_quote(qid):
    """Parse line items from a Locke Supply PDF quote using pdfplumber."""
    conn = get_db()
    q = conn.execute('SELECT file_path, supplier_name FROM supplier_quotes WHERE id = ?', (qid,)).fetchone()
    if not q or not q['file_path']:
        conn.close()
        return jsonify({'error': 'No PDF file found for this quote'}), 404
    fpath = os.path.join(SUPPLIER_QUOTES_DIR, q['file_path'])
    if not os.path.exists(fpath):
        conn.close()
        return jsonify({'error': 'PDF file not found on disk'}), 404
    try:
        import pdfplumber
        import re
        items = []
        # Locke Supply PDFs: last table on each page has line items in a single cell
        # Format: "LineNo SKU [- Reference] Qty BO Shipped UM UnitPrice ExtPrice\nDescription..."
        # The end pattern (qty BO shipped UM price amount) is consistent
        line_header_pattern = re.compile(
            r'^(\d+)\s+(.+?)\s+(\d[\d,]*)\s+\d+\s+\d+\s+\w+\s+([\d,.]+)\s+([\d,.]+)\s*$'
        )
        with pdfplumber.open(fpath) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if not tables:
                    continue
                # Last table contains the line items
                last_table = tables[-1]
                for row in last_table:
                    if not row:
                        continue
                    # All item data is in the first cell
                    cell = str(row[0] or '')
                    if not cell.strip():
                        continue
                    # Split cell into lines and parse each item block
                    lines = cell.split('\n')
                    i = 0
                    while i < len(lines):
                        m = line_header_pattern.match(lines[i].strip())
                        if m:
                            line_no = int(m.group(1))
                            sku_ref = m.group(2).strip()
                            # Split SKU from reference (e.g. "PTC073J35AXXX PNFE647695" or "MECHANICAL SALES")
                            sku_parts = sku_ref.split()
                            sku = sku_parts[0] if sku_parts else sku_ref
                            # Remove trailing dash/reference
                            if ' - ' in sku_ref:
                                sku = sku_ref.split(' - ')[0].strip()
                            qty = float(m.group(3).replace(',', ''))
                            unit_price = float(m.group(4).replace(',', ''))
                            ext_price = float(m.group(5).replace(',', ''))
                            # Next line(s) are description until another line item or end
                            desc_parts = []
                            i += 1
                            while i < len(lines):
                                next_line = lines[i].strip()
                                if line_header_pattern.match(next_line):
                                    break
                                if next_line and not next_line.startswith('***'):
                                    desc_parts.append(next_line)
                                i += 1
                            desc = desc_parts[0] if desc_parts else sku
                            items.append({
                                'line_number': line_no, 'sku': sku,
                                'description': desc[:200],
                                'quantity': qty, 'unit_price': unit_price,
                                'extended_price': ext_price
                            })
                        else:
                            i += 1
        # Save parsed items
        conn.execute('DELETE FROM supplier_quote_items WHERE quote_id = ?', (qid,))
        subtotal = 0
        for item in items:
            conn.execute(
                '''INSERT INTO supplier_quote_items (quote_id, line_number, sku, description,
                   quantity, unit_price, extended_price) VALUES (?,?,?,?,?,?,?)''',
                (qid, item['line_number'], item['sku'], item['description'],
                 item['quantity'], item['unit_price'], item['extended_price'])
            )
            subtotal += item['extended_price']
        conn.execute(
            "UPDATE supplier_quotes SET subtotal=?, updated_at=datetime('now','localtime') WHERE id=?",
            (subtotal, qid)
        )
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'items_parsed': len(items), 'subtotal': subtotal})
    except ImportError:
        conn.close()
        return jsonify({'error': 'pdfplumber not installed'}), 500
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


# ─── AI Price Check Helpers ─────────────────────────────────────

def _build_historical_prices(conn, items):
    """Build historical pricing data from past quotes and invoices."""
    historical = {}
    _STOPWORDS = {'the','a','an','and','or','for','with','in','of','to','from','by','at','on','x','w','per'}

    for item in items:
        sku = (item.get('sku') or '').strip().upper()
        desc = (item.get('description') or '').strip()
        item_key = f"{sku}|{desc}"
        sources = []

        # Source 1: Past supplier_quote_items
        if sku and len(sku) >= 3:
            # Exact match
            rows = conn.execute('''
                SELECT sqi.unit_price, sqi.quantity, sqi.extended_price, sqi.description,
                       sq.supplier_name, sq.quote_date, sq.id as quote_id,
                       j.name as job_name
                FROM supplier_quote_items sqi
                JOIN supplier_quotes sq ON sq.id = sqi.quote_id
                LEFT JOIN jobs j ON j.id = sq.job_id
                WHERE UPPER(TRIM(sqi.sku)) = ? AND sq.id != ?
                ORDER BY sq.quote_date DESC
            ''', (sku, item.get('_quote_id', 0))).fetchall()
            for r in rows:
                # Detect per-C/per-M pricing
                qty = float(r['quantity'] or 0)
                up = float(r['unit_price'] or 0)
                ext = float(r['extended_price'] or 0)
                norm_price = up  # default: per-each
                pricing_unit = 'each'
                if qty > 0 and ext > 0 and up > 0:
                    each_ext = qty * up
                    per_c_ext = qty * up / 100
                    per_m_ext = qty * up / 1000
                    if abs(per_c_ext - ext) < abs(each_ext - ext) * 0.5:
                        norm_price = up / 100
                        pricing_unit = 'per-C'
                    elif abs(per_m_ext - ext) < abs(each_ext - ext) * 0.5:
                        norm_price = up / 1000
                        pricing_unit = 'per-M'
                sources.append({
                    'type': 'quote', 'price': norm_price, 'pricing_unit': pricing_unit,
                    'supplier': r['supplier_name'], 'date': r['quote_date'] or '',
                    'job': r['job_name'] or '', 'ref': f"Quote #{r['quote_id']}"
                })

            # Prefix match (6+ chars) if no exact matches
            if not sources and len(sku) >= 6:
                rows = conn.execute('''
                    SELECT sqi.unit_price, sqi.quantity, sqi.extended_price, sqi.description,
                           sq.supplier_name, sq.quote_date, sq.id as quote_id,
                           j.name as job_name
                    FROM supplier_quote_items sqi
                    JOIN supplier_quotes sq ON sq.id = sqi.quote_id
                    LEFT JOIN jobs j ON j.id = sq.job_id
                    WHERE UPPER(TRIM(sqi.sku)) LIKE ? AND sq.id != ?
                    ORDER BY sq.quote_date DESC LIMIT 20
                ''', (sku[:6] + '%', item.get('_quote_id', 0))).fetchall()
                for r in rows:
                    qty = float(r['quantity'] or 0)
                    up = float(r['unit_price'] or 0)
                    ext = float(r['extended_price'] or 0)
                    norm_price = up
                    pricing_unit = 'each'
                    if qty > 0 and ext > 0 and up > 0:
                        each_ext = qty * up
                        per_c_ext = qty * up / 100
                        if abs(per_c_ext - ext) < abs(each_ext - ext) * 0.5:
                            norm_price = up / 100
                            pricing_unit = 'per-C'
                    sources.append({
                        'type': 'quote', 'price': norm_price, 'pricing_unit': pricing_unit,
                        'supplier': r['supplier_name'], 'date': r['quote_date'] or '',
                        'job': r['job_name'] or '', 'ref': f"Quote #{r['quote_id']} (prefix match)"
                    })

        # Source 2: supplier_invoices line_items JSON
        if sku and len(sku) >= 3:
            inv_rows = conn.execute('''
                SELECT si.line_items, si.invoice_number, si.invoice_date,
                       bc.supplier_name, j.name as job_name
                FROM supplier_invoices si
                JOIN billtrust_config bc ON bc.id = si.supplier_config_id
                LEFT JOIN jobs j ON j.id = si.job_id
                WHERE si.line_items IS NOT NULL AND si.line_items != '[]'
                ORDER BY si.invoice_date DESC LIMIT 200
            ''').fetchall()
            for inv in inv_rows:
                try:
                    inv_items = json.loads(inv['line_items'] or '[]')
                except (json.JSONDecodeError, TypeError):
                    continue
                for li in inv_items:
                    li_code = (li.get('product_code') or '').strip().upper()
                    li_desc = (li.get('description') or '').upper()
                    # Check product_code match or SKU in description
                    if li_code == sku or (len(sku) >= 6 and li_code and li_code.startswith(sku[:6])) or sku in li_desc:
                        up = float(li.get('unit_price') or 0)
                        qty = float(li.get('qty_shipped') or li.get('qty_ordered') or li.get('quantity') or 0)
                        ext = float(li.get('extended_price') or 0)
                        norm_price = up
                        pricing_unit = 'each'
                        if qty > 0 and ext > 0 and up > 0:
                            each_ext = qty * up
                            per_c_ext = qty * up / 100
                            if abs(per_c_ext - ext) < abs(each_ext - ext) * 0.5:
                                norm_price = up / 100
                                pricing_unit = 'per-C'
                        sources.append({
                            'type': 'invoice', 'price': norm_price, 'pricing_unit': pricing_unit,
                            'supplier': inv['supplier_name'], 'date': inv['invoice_date'] or '',
                            'job': inv['job_name'] or '', 'ref': f"Invoice {inv['invoice_number']}"
                        })

        # Description word overlap fallback
        if not sources and desc and len(desc) >= 10:
            desc_words = set(w.lower() for w in desc.split() if len(w) > 2 and w.lower() not in _STOPWORDS)
            if len(desc_words) >= 3:
                rows = conn.execute('''
                    SELECT sqi.unit_price, sqi.quantity, sqi.extended_price, sqi.description,
                           sq.supplier_name, sq.quote_date, sq.id as quote_id,
                           j.name as job_name
                    FROM supplier_quote_items sqi
                    JOIN supplier_quotes sq ON sq.id = sqi.quote_id
                    LEFT JOIN jobs j ON j.id = sq.job_id
                    WHERE sq.id != ?
                    ORDER BY sq.quote_date DESC LIMIT 500
                ''', (item.get('_quote_id', 0),)).fetchall()
                for r in rows:
                    r_words = set(w.lower() for w in (r['description'] or '').split() if len(w) > 2 and w.lower() not in _STOPWORDS)
                    if r_words and len(desc_words & r_words) / max(len(desc_words), 1) >= 0.6:
                        up = float(r['unit_price'] or 0)
                        qty = float(r['quantity'] or 0)
                        ext = float(r['extended_price'] or 0)
                        norm_price = up
                        pricing_unit = 'each'
                        if qty > 0 and ext > 0 and up > 0:
                            each_ext = qty * up
                            per_c_ext = qty * up / 100
                            if abs(per_c_ext - ext) < abs(each_ext - ext) * 0.5:
                                norm_price = up / 100
                                pricing_unit = 'per-C'
                        sources.append({
                            'type': 'quote', 'price': norm_price, 'pricing_unit': pricing_unit,
                            'supplier': r['supplier_name'], 'date': r['quote_date'] or '',
                            'job': r['job_name'] or '', 'ref': f"Quote #{r['quote_id']} (desc match)"
                        })

        if sources:
            prices = [s['price'] for s in sources if s['price'] > 0]
            if prices:
                prices_sorted = sorted(prices)
                avg_price = sum(prices) / len(prices)
                trend = 'stable'
                dated = sorted([s for s in sources if s.get('date')], key=lambda x: x['date'])
                if len(dated) >= 2:
                    recent = [s['price'] for s in dated[-3:]]
                    older = [s['price'] for s in dated[:max(1, len(dated)//2)]]
                    recent_avg = sum(recent) / len(recent)
                    older_avg = sum(older) / len(older)
                    if recent_avg > older_avg * 1.05:
                        trend = 'increasing'
                    elif recent_avg < older_avg * 0.95:
                        trend = 'decreasing'
                historical[item_key] = {
                    'avg_price': round(avg_price, 4),
                    'min_price': round(prices_sorted[0], 4),
                    'max_price': round(prices_sorted[-1], 4),
                    'price_count': len(prices),
                    'last_purchased': dated[-1]['date'] if dated else '',
                    'trend': trend,
                    'sources': sources[:10]
                }
    return historical


def _search_web_prices(items, max_searches=10):
    """Search DuckDuckGo Lite for public HVAC pricing on top-value items."""
    import re as _re
    import time as _time
    try:
        from urllib.request import urlopen, Request
        from urllib.parse import quote_plus
        from urllib.error import URLError
    except ImportError:
        return {}

    # Sort by extended_price descending, only items >= $100
    valued = [it for it in items if float(it.get('extended_price') or 0) >= 100]
    valued.sort(key=lambda x: float(x.get('extended_price') or 0), reverse=True)
    valued = valued[:max_searches]

    web_data = {}
    price_pattern = _re.compile(r'\$\s?(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)')

    for item in valued:
        sku = (item.get('sku') or '').strip()
        desc = (item.get('description') or '').strip()
        item_key = f"{sku}|{desc}"

        query_term = sku if sku and len(sku) >= 4 else desc[:60]
        query = f"{query_term} HVAC supply price buy"
        url = f"https://lite.duckduckgo.com/lite/?q={quote_plus(query)}"

        try:
            req = Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            resp = urlopen(req, timeout=8)
            html = resp.read().decode('utf-8', errors='ignore')
            found_prices = []
            for match in price_pattern.finditer(html):
                val = float(match.group(1).replace(',', ''))
                if 1 <= val <= 50000:
                    found_prices.append(val)

            if found_prices:
                # Filter outliers (>3x or <1/3 of median)
                found_prices.sort()
                median = found_prices[len(found_prices) // 2]
                filtered = [p for p in found_prices if median / 3 <= p <= median * 3]
                if filtered:
                    web_data[item_key] = {
                        'found': True,
                        'range_low': round(min(filtered), 2),
                        'range_high': round(max(filtered), 2),
                        'median': round(median, 2),
                        'sources': [{'query': query, 'prices_found': len(filtered)}],
                        'note': 'Web prices are typically retail; distributor pricing is usually 20-40% lower.'
                    }
            else:
                web_data[item_key] = {'found': False, 'note': 'No web pricing found for this item.'}
        except Exception:
            web_data[item_key] = {'found': False, 'note': 'Web search failed for this item.'}

        _time.sleep(0.5)

    return web_data


def _ai_analyze_pricing(items, historical, web_data, competitor_data, quote_info):
    """Use Claude Haiku to analyze pricing across all data sources."""
    import anthropic
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return {'error': 'No ANTHROPIC_API_KEY configured', 'items': [], 'summary': '', 'recommendations': []}

    # Build per-item context
    items_context = []
    for item in items:
        sku = (item.get('sku') or '').strip()
        desc = (item.get('description') or '').strip()
        item_key = f"{sku}|{desc}"
        entry = {
            'sku': sku, 'description': desc,
            'quantity': item.get('quantity', 0),
            'unit_price': item.get('unit_price', 0),
            'extended_price': item.get('extended_price', 0),
        }
        hist = historical.get(item_key)
        if hist:
            entry['historical'] = {
                'avg_price': hist['avg_price'], 'min_price': hist['min_price'],
                'max_price': hist['max_price'], 'price_count': hist['price_count'],
                'trend': hist['trend']
            }
        web = web_data.get(item_key)
        if web and web.get('found'):
            entry['web_pricing'] = {
                'range_low': web['range_low'], 'range_high': web['range_high']
            }
        items_context.append(entry)

    prompt = f"""You are an HVAC construction pricing analyst for LGHVAC LLC, an HVAC contractor in Oklahoma.

Analyze this supplier quote and determine if prices are competitive.

IMPORTANT HVAC industry context:
- "Per-C" pricing = price per 100 units (common for pipe, flex duct, collars, fittings)
- "Per-M" pricing = price per 1000 units
- Distributor prices (Locke Supply, Plumb Supply) are typically 20-40% below retail/web prices
- Seasonal trends: equipment prices rise in spring/summer
- Major brands: Carrier, Trane, Lennox, Rheem, Goodman, Daikin

Quote from: {quote_info.get('supplier_name', 'Unknown')}
Quote total: ${quote_info.get('total', 0):,.2f}

LINE ITEMS WITH AVAILABLE DATA:
{json.dumps(items_context, indent=2)}

"""
    if competitor_data:
        prompt += f"""COMPETITOR PRICING DATA:
{json.dumps(competitor_data, indent=2)}

"""

    prompt += """For EACH line item, provide:
1. assessment: "excellent" (great deal), "fair" (market rate), "above_average" (slightly high), or "high" (overpriced)
2. savings_low: conservative estimated savings in dollars (0 if no savings)
3. savings_high: optimistic estimated savings in dollars (0 if no savings)
4. ai_note: brief explanation (1 sentence)

Also provide:
- summary: 2-3 sentence overall assessment
- recommendations: list of 3-5 actionable negotiation points

Return ONLY valid JSON in this format:
{
  "items": [{"sku": "...", "assessment": "...", "savings_low": 0, "savings_high": 0, "ai_note": "..."}],
  "summary": "...",
  "recommendations": ["...", "..."]
}"""

    try:
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=4000,
            messages=[{'role': 'user', 'content': prompt}]
        )
        raw = resp.content[0].text
        import re
        match = re.search(r'\{[\s\S]*\}', raw)
        if match:
            result = json.loads(match.group())
            return result
        return {'error': 'Could not parse AI response', 'items': [], 'summary': raw[:500], 'recommendations': []}
    except Exception as e:
        return {'error': str(e), 'items': [], 'summary': '', 'recommendations': []}


# ─── AI Price Check Endpoints ───────────────────────────────────

@app.route('/api/supplier-quotes/<int:qid>/price-check', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_run_price_check(qid):
    """Run AI-powered pricing analysis on a supplier quote."""
    conn = get_db()
    quote = conn.execute('''SELECT sq.*, j.name as job_name
        FROM supplier_quotes sq LEFT JOIN jobs j ON j.id = sq.job_id
        WHERE sq.id = ?''', (qid,)).fetchone()
    if not quote:
        conn.close()
        return jsonify({'error': 'Quote not found'}), 404

    items_rows = conn.execute(
        'SELECT * FROM supplier_quote_items WHERE quote_id = ? ORDER BY line_number', (qid,)
    ).fetchall()
    if not items_rows:
        conn.close()
        return jsonify({'error': 'No line items to analyze'}), 400

    items = []
    for r in items_rows:
        items.append({
            'sku': r['sku'], 'description': r['description'],
            'quantity': r['quantity'], 'unit_price': r['unit_price'],
            'extended_price': r['extended_price'], '_quote_id': qid
        })

    # Get request data (competitor text/items)
    data = request.get_json(silent=True) or {}
    competitor_text = data.get('competitor_text', '').strip()
    competitor_items = data.get('competitor_items', [])
    competitor_data = {}
    if competitor_text:
        competitor_data['raw_text'] = competitor_text
    if competitor_items:
        competitor_data['items'] = competitor_items

    # Run the 3 analysis steps
    historical = _build_historical_prices(conn, items)
    web_data = _search_web_prices(items)
    quote_info = {
        'supplier_name': quote['supplier_name'],
        'total': quote['total'] or 0,
        'job_name': quote['job_name'] or ''
    }
    ai_result = _ai_analyze_pricing(items, historical, web_data, competitor_data, quote_info)

    # Build per-item report
    item_reports = []
    total_savings_low = 0
    total_savings_high = 0
    items_with_savings = 0
    for i, item in enumerate(items):
        sku = (item.get('sku') or '').strip()
        desc = (item.get('description') or '').strip()
        item_key = f"{sku}|{desc}"
        hist = historical.get(item_key, {})
        web = web_data.get(item_key, {})

        # Match AI item by index or SKU
        ai_item = {}
        ai_items = ai_result.get('items', [])
        if i < len(ai_items):
            ai_item = ai_items[i]
        else:
            for ai in ai_items:
                if ai.get('sku', '').upper() == sku.upper():
                    ai_item = ai
                    break

        # Find competitor price for this item
        comp_price = None
        if competitor_items:
            for ci in competitor_items:
                ci_sku = (ci.get('sku') or '').strip().upper()
                if ci_sku and ci_sku == sku.upper():
                    comp_price = float(ci.get('unit_price') or 0)
                    break

        savings_low = float(ai_item.get('savings_low') or 0)
        savings_high = float(ai_item.get('savings_high') or 0)
        total_savings_low += savings_low
        total_savings_high += savings_high
        if savings_high > 0:
            items_with_savings += 1

        item_reports.append({
            'sku': sku, 'description': desc,
            'quantity': item['quantity'], 'unit_price': item['unit_price'],
            'extended_price': item['extended_price'],
            'historical_avg': hist.get('avg_price'),
            'historical_min': hist.get('min_price'),
            'historical_max': hist.get('max_price'),
            'historical_count': hist.get('price_count', 0),
            'historical_trend': hist.get('trend'),
            'web_low': web.get('range_low') if web.get('found') else None,
            'web_high': web.get('range_high') if web.get('found') else None,
            'competitor_price': comp_price,
            'assessment': ai_item.get('assessment', 'fair'),
            'savings_low': savings_low,
            'savings_high': savings_high,
            'ai_note': ai_item.get('ai_note', '')
        })

    review = {
        'quote_id': qid,
        'supplier_name': quote['supplier_name'],
        'quote_total': quote['total'] or 0,
        'items_reviewed': len(item_reports),
        'items_with_savings': items_with_savings,
        'total_savings_low': round(total_savings_low, 2),
        'total_savings_high': round(total_savings_high, 2),
        'summary': ai_result.get('summary', ''),
        'recommendations': ai_result.get('recommendations', []),
        'items': item_reports,
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'has_competitor_data': bool(competitor_text or competitor_items),
        'has_historical_data': bool(historical),
        'has_web_data': any(w.get('found') for w in web_data.values())
    }

    # Cache in database
    conn.execute(
        '''INSERT INTO pricing_reviews (quote_id, review_data, total_savings_low, total_savings_high,
           items_reviewed, items_with_savings, created_by)
           VALUES (?, ?, ?, ?, ?, ?, ?)''',
        (qid, json.dumps(review), total_savings_low, total_savings_high,
         len(item_reports), items_with_savings, session.get('user_id'))
    )
    conn.commit()
    conn.close()
    return jsonify(review)


@app.route('/api/supplier-quotes/<int:qid>/price-check', methods=['GET'])
@api_role_required('owner', 'admin', 'project_manager')
def api_get_price_check(qid):
    """Return the most recent cached pricing review for a quote."""
    conn = get_db()
    row = conn.execute(
        'SELECT review_data, created_at FROM pricing_reviews WHERE quote_id = ? ORDER BY created_at DESC LIMIT 1',
        (qid,)
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({'cached': False})
    try:
        review = json.loads(row['review_data'])
        review['cached'] = True
        review['cached_at'] = row['created_at']
        return jsonify(review)
    except (json.JSONDecodeError, TypeError):
        return jsonify({'cached': False})


@app.route('/api/supplier-quotes/<int:qid>/price-check/competitor', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_extract_competitor_pdf(qid):
    """Extract line items from a competitor's quote PDF using pdfplumber + AI."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Only PDF files are supported'}), 400

    try:
        import pdfplumber
        import tempfile

        # Save to temp file
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        file.save(tmp.name)
        tmp.close()

        # Extract text
        text_parts = []
        with pdfplumber.open(tmp.name) as pdf:
            for page in pdf.pages[:10]:
                text_parts.append(page.extract_text() or '')
        os.unlink(tmp.name)
        pdf_text = '\n'.join(text_parts)

        if len(pdf_text.strip()) < 20:
            return jsonify({'error': 'Could not extract text from PDF'}), 400

        # Use Claude to extract line items
        import anthropic
        api_key = os.environ.get('ANTHROPIC_API_KEY', '')
        if not api_key:
            return jsonify({'error': 'ANTHROPIC_API_KEY not configured'}), 500

        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=4000,
            messages=[{'role': 'user', 'content': f"""Extract line items from this supplier quote PDF text.

Return ONLY a valid JSON array of objects with these fields:
- sku: product code/SKU (string)
- description: item description (string)
- quantity: number
- unit_price: price per unit (number)
- extended_price: total for this line (number)
- supplier_name: supplier name if visible in the document (string)

If a field is not found, use empty string for text or 0 for numbers.

PDF TEXT:
{pdf_text[:6000]}"""}]
        )
        raw = resp.content[0].text
        import re
        match = re.search(r'\[[\s\S]*\]', raw)
        if match:
            extracted = json.loads(match.group())
            # Ensure supplier_name from doc or filename
            supplier = ''
            for it in extracted:
                if it.get('supplier_name'):
                    supplier = it['supplier_name']
                    break
            if not supplier:
                supplier = file.filename.rsplit('.', 1)[0][:50]
            for it in extracted:
                it['supplier_name'] = it.get('supplier_name') or supplier
            return jsonify({'ok': True, 'items': extracted, 'supplier_name': supplier})
        return jsonify({'error': 'Could not parse items from PDF'}), 400
    except ImportError:
        return jsonify({'error': 'pdfplumber not installed'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── Inventory (Phase 3) ────────────────────────────────────────

@app.route('/inventory')
@role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def inventory_list():
    return render_template('inventory/list.html')

@app.route('/inventory/<int:iid>')
@role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def inventory_detail(iid):
    return render_template('inventory/detail.html', item_id=iid)

@app.route('/api/inventory')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def api_inventory():
    search = request.args.get('q', '').strip()
    category = request.args.get('category', '')
    conn = get_db()
    query = 'SELECT * FROM inventory_items WHERE 1=1'
    params = []
    if search:
        query += ' AND (sku LIKE ? OR description LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])
    if category:
        query += ' AND category = ?'
        params.append(category)
    query += ' ORDER BY description'
    items = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(i) for i in items])

@app.route('/api/inventory', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_create_inventory_item():
    data = request.get_json()
    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO inventory_items (sku, description, category, quantity_on_hand, unit,
           location, reorder_point, notes, created_by)
           VALUES (?,?,?,?,?,?,?,?,?)''',
        (data.get('sku',''), data.get('description',''), data.get('category',''),
         float(data.get('quantity_on_hand',0) or 0), data.get('unit','each'),
         data.get('location','Warehouse'), float(data.get('reorder_point',0) or 0),
         data.get('notes',''), session.get('user_id'))
    )
    iid = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': iid}), 201

@app.route('/api/inventory/<int:iid>')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def api_inventory_detail(iid):
    conn = get_db()
    item = conn.execute('SELECT * FROM inventory_items WHERE id = ?', (iid,)).fetchone()
    if not item:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    transactions = conn.execute(
        '''SELECT it.*, j.name as job_name, u.display_name as user_name
           FROM inventory_transactions it
           LEFT JOIN jobs j ON it.job_id = j.id
           LEFT JOIN users u ON it.created_by = u.id
           WHERE it.inventory_item_id = ? ORDER BY it.created_at DESC LIMIT 100''', (iid,)
    ).fetchall()
    conn.close()
    result = dict(item)
    result['transactions'] = [dict(t) for t in transactions]
    return jsonify(result)

@app.route('/api/inventory/<int:iid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_update_inventory_item(iid):
    data = request.get_json()
    conn = get_db()
    conn.execute(
        '''UPDATE inventory_items SET sku=?, description=?, category=?, unit=?,
           location=?, reorder_point=?, notes=?, updated_at=datetime('now','localtime') WHERE id=?''',
        (data.get('sku',''), data.get('description',''), data.get('category',''),
         data.get('unit','each'), data.get('location','Warehouse'),
         float(data.get('reorder_point',0) or 0), data.get('notes',''), iid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/inventory/<int:iid>/adjust', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_inventory_adjust(iid):
    data = request.get_json()
    ttype = data.get('transaction_type', 'adjust')
    qty = float(data.get('quantity', 0) or 0)
    if qty == 0:
        return jsonify({'error': 'Quantity required'}), 400
    conn = get_db()
    conn.execute(
        '''INSERT INTO inventory_transactions (inventory_item_id, transaction_type, quantity,
           job_id, reference, notes, created_by) VALUES (?,?,?,?,?,?,?)''',
        (iid, ttype, qty, data.get('job_id') or None, data.get('reference',''),
         data.get('notes',''), session.get('user_id'))
    )
    if ttype in ('receive', 'return'):
        conn.execute('UPDATE inventory_items SET quantity_on_hand = quantity_on_hand + ?, updated_at=datetime("now","localtime") WHERE id=?', (abs(qty), iid))
    elif ttype == 'issue':
        conn.execute('UPDATE inventory_items SET quantity_on_hand = MAX(0, quantity_on_hand - ?), updated_at=datetime("now","localtime") WHERE id=?', (abs(qty), iid))
    elif ttype == 'adjust':
        conn.execute('UPDATE inventory_items SET quantity_on_hand = quantity_on_hand + ?, updated_at=datetime("now","localtime") WHERE id=?', (qty, iid))
    elif ttype == 'count':
        conn.execute('UPDATE inventory_items SET quantity_on_hand = ?, last_count_date=date("now","localtime"), updated_at=datetime("now","localtime") WHERE id=?', (abs(qty), iid))
    conn.commit()
    check_reorder_alerts(conn)
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/inventory/count', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_inventory_bulk_count():
    data = request.get_json()
    counts = data.get('counts', [])
    conn = get_db()
    for c in counts:
        iid = c.get('id')
        qty = float(c.get('quantity', 0) or 0)
        conn.execute(
            '''INSERT INTO inventory_transactions (inventory_item_id, transaction_type, quantity, notes, created_by)
               VALUES (?,?,?,?,?)''',
            (iid, 'count', qty, 'Physical count', session.get('user_id'))
        )
        conn.execute(
            'UPDATE inventory_items SET quantity_on_hand=?, last_count_date=date("now","localtime"), updated_at=datetime("now","localtime") WHERE id=?',
            (qty, iid)
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/inventory/check-availability')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def api_inventory_check_availability():
    job_id = request.args.get('job_id')
    if not job_id:
        return jsonify({'error': 'job_id required'}), 400
    conn = get_db()
    plan = conn.execute("SELECT takeoff_data FROM plans WHERE job_id = ? AND status = 'Takeoff Complete' ORDER BY updated_at DESC LIMIT 1", (job_id,)).fetchone()
    if not plan or not plan['takeoff_data']:
        conn.close()
        return jsonify({'error': 'No takeoff data found'}), 404
    import json as _json
    takeoff = _json.loads(plan['takeoff_data']) if isinstance(plan['takeoff_data'], str) else plan['takeoff_data']
    result = []
    for item in takeoff:
        sku = item.get('sku', '')
        needed = float(item.get('quantity', 0) or 0)
        inv = conn.execute('SELECT quantity_on_hand FROM inventory_items WHERE sku = ?', (sku,)).fetchone()
        in_stock = inv['quantity_on_hand'] if inv else 0
        result.append({
            'sku': sku, 'description': item.get('description', ''),
            'needed': needed, 'in_stock': in_stock,
            'to_order': max(0, needed - in_stock)
        })
    conn.close()
    return jsonify(result)


# ─── Phase 5: Invoices & Duplicate Detection ─────────────────────

@app.route('/invoices')
@login_required
def invoices_page():
    return render_template('invoices/list.html')

@app.route('/api/invoices')
@api_login_required
def api_list_invoices():
    conn = get_db()
    rows = conn.execute('''
        SELECT si.*, bc.supplier_name,
               j.name as job_name,
               sq.quote_number as matched_quote_number,
               sq.id as supplier_quote_id
        FROM supplier_invoices si
        LEFT JOIN billtrust_config bc ON si.supplier_config_id = bc.id
        LEFT JOIN jobs j ON si.job_id = j.id
        LEFT JOIN supplier_quotes sq ON si.supplier_quote_id = sq.id
        ORDER BY si.invoice_date DESC
    ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/invoices/<int:iid>')
@api_login_required
def api_get_invoice(iid):
    conn = get_db()
    inv = conn.execute('''
        SELECT si.*, bc.supplier_name,
               j.name as job_name
        FROM supplier_invoices si
        LEFT JOIN billtrust_config bc ON si.supplier_config_id = bc.id
        LEFT JOIN jobs j ON si.job_id = j.id
        WHERE si.id = ?
    ''', (iid,)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    result = dict(inv)
    # Get line items if they exist
    items = conn.execute('''
        SELECT * FROM supplier_invoice_items WHERE invoice_id = ?
        ORDER BY line_number
    ''', (iid,)).fetchall()
    result['items'] = [dict(i) for i in items]
    conn.close()
    return jsonify(result)

@app.route('/api/invoices/duplicates')
@api_login_required
def api_list_duplicate_invoices():
    conn = get_db()
    rows = conn.execute('''
        SELECT si.*, bc.supplier_name, j.name as job_name
        FROM supplier_invoices si
        LEFT JOIN billtrust_config bc ON si.supplier_config_id = bc.id
        LEFT JOIN jobs j ON si.job_id = j.id
        WHERE si.is_duplicate = 1
        ORDER BY si.invoice_date DESC
    ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/invoices/<int:iid>/resolve-duplicate', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_resolve_duplicate(iid):
    data = request.get_json() or {}
    action = data.get('action', 'dismiss')  # 'dismiss' or 'confirm'
    conn = get_db()
    if action == 'dismiss':
        conn.execute('UPDATE supplier_invoices SET is_duplicate = 0, duplicate_of_id = NULL WHERE id = ?', (iid,))
    elif action == 'confirm':
        # Mark as confirmed duplicate — could archive or delete
        conn.execute("UPDATE supplier_invoices SET status = 'void', notes = COALESCE(notes,'') || ' [Confirmed duplicate]' WHERE id = ?", (iid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/invoices/<int:iid>/match-quote', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_match_invoice_to_quote(iid):
    data = request.get_json() or {}
    quote_id = data.get('supplier_quote_id')
    if not quote_id:
        return jsonify({'error': 'supplier_quote_id required'}), 400
    conn = get_db()
    conn.execute('UPDATE supplier_invoices SET supplier_quote_id = ? WHERE id = ?', (quote_id, iid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/jobs/<int:job_id>/invoice-summary')
@api_login_required
def api_job_invoice_summary(job_id):
    conn = get_db()
    # Total invoiced
    inv_row = conn.execute('''
        SELECT COUNT(*) as count, COALESCE(SUM(total),0) as total_invoiced,
               COALESCE(SUM(balance_due),0) as total_balance
        FROM supplier_invoices WHERE job_id = ? AND is_duplicate = 0
    ''', (job_id,)).fetchone()
    # Total quoted (baseline)
    quote_row = conn.execute('''
        SELECT COALESCE(SUM(total),0) as total_quoted
        FROM supplier_quotes WHERE job_id = ? AND is_baseline = 1
    ''', (job_id,)).fetchone()
    total_invoiced = inv_row['total_invoiced']
    total_quoted = quote_row['total_quoted']
    variance = total_invoiced - total_quoted if total_quoted > 0 else 0
    conn.close()
    return jsonify({
        'invoice_count': inv_row['count'],
        'total_invoiced': total_invoiced,
        'total_balance': inv_row['total_balance'],
        'total_quoted': total_quoted,
        'variance': variance,
        'pct_of_quote': round((total_invoiced / total_quoted * 100), 1) if total_quoted > 0 else 0
    })

@app.route('/api/billtrust/sync-log')
@api_login_required
def api_billtrust_sync_log():
    conn = get_db()
    rows = conn.execute('''
        SELECT sl.*, bc.supplier_name
        FROM billtrust_sync_log sl
        LEFT JOIN billtrust_config bc ON sl.config_id = bc.id
        ORDER BY sl.created_at DESC LIMIT 50
    ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# ─── Invoice Review Flags & Detail ───────────────────────────────

@app.route('/api/invoice-flags')
@api_login_required
def api_list_invoice_flags():
    job_id = request.args.get('job_id')
    severity = request.args.get('severity')
    resolved = request.args.get('resolved')
    limit = request.args.get('limit', 200, type=int)
    conn = get_db()
    sql = '''
        SELECT irf.*, j.name as job_name
        FROM invoice_review_flags irf
        LEFT JOIN jobs j ON irf.job_id = j.id
        WHERE 1=1
    '''
    params = []
    if job_id:
        sql += ' AND irf.job_id = ?'
        params.append(job_id)
    if severity:
        sql += ' AND irf.severity = ?'
        params.append(severity)
    if resolved is not None and resolved != '':
        sql += ' AND irf.resolved = ?'
        params.append(int(resolved))
    sql += ' ORDER BY irf.created_at DESC LIMIT ?'
    params.append(limit)
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/invoice-flags/<int:fid>/resolve', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_resolve_invoice_flag(fid):
    conn = get_db()
    conn.execute('''
        UPDATE invoice_review_flags
        SET resolved = 1, resolved_by = ?, resolved_at = datetime('now','localtime')
        WHERE id = ?
    ''', (session['user_id'], fid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/invoices/reports')
@role_required('owner', 'admin')
def invoice_reports_page():
    return render_template('invoices/reports.html')

@app.route('/invoices/<int:iid>')
@login_required
def invoice_detail_page(iid):
    return render_template('invoices/detail.html', invoice_id=iid)


# ─── Phase 6: Material Requests ──────────────────────────────────

@app.route('/api/jobs/<int:job_id>/line-items-for-request')
@api_login_required
def api_line_items_for_request(job_id):
    phase = request.args.get('phase', '')
    grouped = request.args.get('grouped', '')
    conn = get_db()
    line_items = conn.execute('''
        SELECT id, line_number, sku, description, qty_ordered, stock_ns
        FROM line_items WHERE job_id = ? ORDER BY line_number
    ''', (job_id,)).fetchall()

    # Batch-query inventory for all SKUs
    skus = list(set((li['sku'] or '').strip() for li in line_items if (li['sku'] or '').strip()))
    inv_map = {}
    if skus:
        placeholders = ','.join('?' * len(skus))
        inv_rows = conn.execute(f'''
            SELECT id, sku, quantity_on_hand, reorder_point FROM inventory_items WHERE sku IN ({placeholders})
        ''', skus).fetchall()
        for row in inv_rows:
            inv_map[row['sku']] = row

    # Batch-query already-requested totals
    li_ids = [li['id'] for li in line_items]
    requested_map = {}
    if li_ids:
        placeholders = ','.join('?' * len(li_ids))
        req_rows = conn.execute(f'''
            SELECT mri.line_item_id, COALESCE(SUM(mri.quantity_requested), 0) as total
            FROM material_request_items mri
            JOIN material_requests mr ON mri.request_id = mr.id
            WHERE mri.line_item_id IN ({placeholders}) AND mr.status != 'Cancelled'
            GROUP BY mri.line_item_id
        ''', li_ids).fetchall()
        for row in req_rows:
            requested_map[row['line_item_id']] = row['total']

    # Batch-query phase assignments for grouped mode
    phase_map = {}  # line_item_id -> set of phases
    if grouped == '1' and li_ids:
        placeholders = ','.join('?' * len(li_ids))
        phase_rows = conn.execute(f'''
            SELECT msi.line_item_id, ms.phase
            FROM material_shipment_items msi
            JOIN material_shipments ms ON msi.shipment_id = ms.id
            WHERE ms.job_id = ? AND msi.line_item_id IN ({placeholders})
        ''', [job_id] + li_ids).fetchall()
        for row in phase_rows:
            phase_map.setdefault(row['line_item_id'], set()).add(row['phase'])

    results = []
    for li in line_items:
        sku = (li['sku'] or '').strip()
        inv = inv_map.get(sku)
        stock_on_hand = inv['quantity_on_hand'] if inv else 0
        reorder_point = inv['reorder_point'] if inv else 0
        inventory_item_id = inv['id'] if inv else None
        already_requested = requested_map.get(li['id'], 0)
        qty_still_needed = max(0, (li['qty_ordered'] or 0) - already_requested)

        # Stock status determination
        if stock_on_hand <= 0:
            stock_status = 'Out of Stock'
        elif stock_on_hand <= reorder_point:
            stock_status = 'Low Stock'
        else:
            stock_status = 'In Stock'

        has_phase_shipment = False
        if phase and not grouped:
            phases_for_item = phase_map.get(li['id'], set()) if phase_map else set()
            if not phases_for_item:
                ps = conn.execute('''
                    SELECT COUNT(*) as cnt FROM material_shipment_items msi
                    JOIN material_shipments ms ON msi.shipment_id = ms.id
                    WHERE msi.line_item_id = ? AND ms.phase = ?
                ''', (li['id'], phase)).fetchone()
                has_phase_shipment = ps['cnt'] > 0
            else:
                has_phase_shipment = phase in phases_for_item

        in_stock = stock_on_hand >= qty_still_needed and qty_still_needed > 0
        item_data = {
            'line_item_id': li['id'],
            'line_number': li['line_number'],
            'sku': sku,
            'description': li['description'] or '',
            'qty_ordered': li['qty_ordered'] or 0,
            'qty_still_needed': qty_still_needed,
            'stock_on_hand': stock_on_hand,
            'reorder_point': reorder_point,
            'stock_status': stock_status,
            'in_stock': in_stock,
            'inventory_item_id': inventory_item_id,
            'has_phase_shipment': has_phase_shipment
        }
        results.append(item_data)

    conn.close()

    # Grouped response: organize by phase
    if grouped == '1':
        phase_order = ['Rough-In', 'Trim Out', 'Equipment', 'Startup', 'Unassigned']
        phase_buckets = {p: [] for p in phase_order}
        for item in results:
            item_phases = phase_map.get(item['line_item_id'], set())
            if not item_phases:
                phase_buckets['Unassigned'].append(item)
            else:
                for p in item_phases:
                    if p in phase_buckets:
                        phase_buckets[p].append(item)
                    else:
                        phase_buckets.setdefault(p, []).append(item)
        phases_out = []
        for p in phase_order:
            items = phase_buckets.get(p, [])
            if not items:
                continue
            in_stock_count = sum(1 for i in items if i['stock_status'] == 'In Stock')
            need_ordering = len(items) - in_stock_count
            phases_out.append({
                'phase': p,
                'items': items,
                'summary': {
                    'total_items': len(items),
                    'in_stock': in_stock_count,
                    'need_ordering': need_ordering
                }
            })
        return jsonify({'phases': phases_out})

    return jsonify(results)


@app.route('/api/material-requests/parse-file', methods=['POST'])
@api_login_required
def api_material_requests_parse_file():
    """Parse a CSV or Excel file and cross-reference against inventory."""
    import csv as csv_module
    import io

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'No file provided'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    rows = []

    # Column alias matching
    header_aliases = {
        'sku': ['sku', 'part number', 'part #', 'item number', 'product code'],
        'description': ['description', 'desc', 'item description', 'material', 'name'],
        'quantity': ['quantity', 'qty', 'amount', 'count'],
        'unit': ['unit', 'uom', 'unit of measure'],
    }

    if ext == 'csv':
        try:
            content = file.read().decode('utf-8-sig')
            reader = csv_module.DictReader(io.StringIO(content))
            raw_headers = {h.strip().lower(): h for h in (reader.fieldnames or [])}
            col_map = {}
            for field, aliases in header_aliases.items():
                for alias in aliases:
                    if alias in raw_headers:
                        col_map[field] = raw_headers[alias]
                        break
            for row in reader:
                rows.append({
                    'sku': (row.get(col_map.get('sku', ''), '') or '').strip(),
                    'description': (row.get(col_map.get('description', ''), '') or '').strip(),
                    'quantity': row.get(col_map.get('quantity', ''), '') or '0',
                    'unit': (row.get(col_map.get('unit', ''), '') or 'each').strip(),
                })
        except Exception as e:
            return jsonify({'error': f'Could not read CSV: {str(e)[:200]}'}), 400

    elif ext in ('xlsx', 'xls'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            headers = {}
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers[str(val).strip().lower()] = col
            col_map = {}
            for field, aliases in header_aliases.items():
                for alias in aliases:
                    if alias in headers:
                        col_map[field] = headers[alias]
                        break
            for row_num in range(2, ws.max_row + 1):
                sku_val = ws.cell(row=row_num, column=col_map.get('sku', 0)).value if 'sku' in col_map else ''
                desc_val = ws.cell(row=row_num, column=col_map.get('description', 0)).value if 'description' in col_map else ''
                qty_val = ws.cell(row=row_num, column=col_map.get('quantity', 0)).value if 'quantity' in col_map else 0
                unit_val = ws.cell(row=row_num, column=col_map.get('unit', 0)).value if 'unit' in col_map else 'each'
                if not sku_val and not desc_val:
                    continue
                rows.append({
                    'sku': str(sku_val or '').strip(),
                    'description': str(desc_val or '').strip(),
                    'quantity': str(qty_val or '0'),
                    'unit': str(unit_val or 'each').strip(),
                })
        except Exception as e:
            return jsonify({'error': f'Could not read Excel file: {str(e)[:200]}'}), 400
    else:
        return jsonify({'error': 'Unsupported file type. Please upload .csv or .xlsx'}), 400

    # Cross-reference against inventory
    conn = get_db()
    skus = list(set(r['sku'] for r in rows if r['sku']))
    inv_map = {}
    if skus:
        placeholders = ','.join('?' * len(skus))
        inv_rows = conn.execute(f'''
            SELECT id, sku, quantity_on_hand, reorder_point FROM inventory_items WHERE sku IN ({placeholders})
        ''', skus).fetchall()
        for row in inv_rows:
            inv_map[row['sku']] = row
    conn.close()

    items = []
    for r in rows:
        qty = 0
        try:
            qty = float(r['quantity'])
        except (ValueError, TypeError):
            pass
        if qty <= 0 and not r['sku'] and not r['description']:
            continue
        inv = inv_map.get(r['sku'])
        stock_on_hand = inv['quantity_on_hand'] if inv else 0
        reorder_point = inv['reorder_point'] if inv else 0
        inventory_item_id = inv['id'] if inv else None
        if stock_on_hand <= 0:
            stock_status = 'Out of Stock'
        elif stock_on_hand <= reorder_point:
            stock_status = 'Low Stock'
        else:
            stock_status = 'In Stock'
        in_stock = stock_on_hand >= qty and qty > 0
        items.append({
            'sku': r['sku'],
            'description': r['description'],
            'quantity': qty,
            'unit': r['unit'] or 'each',
            'stock_on_hand': stock_on_hand,
            'reorder_point': reorder_point,
            'stock_status': stock_status,
            'in_stock': in_stock,
            'inventory_item_id': inventory_item_id,
            'order_needed': not in_stock
        })

    return jsonify({'items': items, 'count': len(items)})

@app.route('/material-requests')
@login_required
def material_requests_page():
    return render_template('material_requests/list.html')

@app.route('/api/material-requests')
@api_login_required
def api_list_material_requests():
    job_id = request.args.get('job_id')
    status = request.args.get('status')
    conn = get_db()
    sql = '''
        SELECT mr.*, j.name as job_name,
               u.display_name as requester_name
        FROM material_requests mr
        LEFT JOIN jobs j ON mr.job_id = j.id
        LEFT JOIN users u ON mr.requested_by = u.id
        WHERE 1=1
    '''
    params = []
    if job_id:
        sql += ' AND mr.job_id = ?'
        params.append(job_id)
    if status:
        sql += ' AND mr.status = ?'
        params.append(status)
    sql += ' ORDER BY mr.created_at DESC'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/material-requests', methods=['POST'])
@api_login_required
def api_create_material_request():
    data = request.get_json() or {}
    job_id = data.get('job_id')
    if not job_id:
        return jsonify({'error': 'job_id required'}), 400
    conn = get_db()
    cursor = conn.execute('''
        INSERT INTO material_requests (job_id, requested_by, priority, needed_by, notes, status, phase)
        VALUES (?, ?, ?, ?, ?, 'Pending', ?)
    ''', (job_id, session.get('user_id'), data.get('priority', 'Normal'),
          data.get('needed_by', ''), data.get('notes', ''), data.get('phase', '')))
    rid = cursor.lastrowid
    items = data.get('items', [])
    for item in items:
        conn.execute('''
            INSERT INTO material_request_items (request_id, sku, description, quantity_requested, unit,
                line_item_id, order_needed, inventory_item_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (rid, item.get('sku', ''), item.get('description', ''),
              float(item.get('quantity_requested', 0)), item.get('unit', 'each'),
              item.get('line_item_id'), int(item.get('order_needed', 0)),
              item.get('inventory_item_id')))
    conn.commit()
    # Notify PMs and owners
    job = conn.execute('SELECT name FROM jobs WHERE id = ?', (job_id,)).fetchone()
    requester = conn.execute('SELECT display_name FROM users WHERE id = ?', (session.get('user_id'),)).fetchone()
    pms = conn.execute("SELECT id FROM users WHERE role IN ('owner','admin','project_manager') AND is_active = 1").fetchall()
    for pm in pms:
        if pm['id'] != session.get('user_id'):
            create_notification(pm['id'], 'material_request',
                f"New Material Request for {job['name'] if job else 'Unknown'}",
                f"{requester['display_name'] if requester else 'Someone'} requested materials ({len(items)} items)",
                f"/material-requests")
    conn.close()
    return jsonify({'ok': True, 'id': rid}), 201

@app.route('/api/material-requests/<int:rid>')
@api_login_required
def api_get_material_request(rid):
    conn = get_db()
    req = conn.execute('''
        SELECT mr.*, j.name as job_name, u.display_name as requester_name
        FROM material_requests mr
        LEFT JOIN jobs j ON mr.job_id = j.id
        LEFT JOIN users u ON mr.requested_by = u.id
        WHERE mr.id = ?
    ''', (rid,)).fetchone()
    if not req:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    result = dict(req)
    items = conn.execute('''
        SELECT mri.*, ii.quantity_on_hand as stock_on_hand
        FROM material_request_items mri
        LEFT JOIN inventory_items ii ON mri.inventory_item_id = ii.id
        WHERE mri.request_id = ?
    ''', (rid,)).fetchall()
    result['items'] = [dict(i) for i in items]
    conn.close()
    return jsonify(result)

@app.route('/api/material-requests/<int:rid>', methods=['PUT'])
@api_login_required
def api_update_material_request(rid):
    data = request.get_json() or {}
    conn = get_db()
    conn.execute('''
        UPDATE material_requests
        SET priority = ?, needed_by = ?, notes = ?, updated_at = datetime('now','localtime')
        WHERE id = ?
    ''', (data.get('priority', 'Normal'), data.get('needed_by', ''),
          data.get('notes', ''), rid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/material-requests/<int:rid>/approve', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_approve_material_request(rid):
    data = request.get_json() or {}
    conn = get_db()
    items = data.get('items', [])
    for item in items:
        conn.execute('''
            UPDATE material_request_items
            SET quantity_approved = ?
            WHERE id = ?
        ''', (float(item.get('quantity_approved', 0)), item['id']))
    conn.execute('''
        UPDATE material_requests
        SET status = 'Approved', approved_by = ?, approved_at = datetime('now','localtime'),
            updated_at = datetime('now','localtime')
        WHERE id = ?
    ''', (session.get('user_id'), rid))
    conn.commit()
    # Notify requester
    req = conn.execute('SELECT requested_by, job_id FROM material_requests WHERE id = ?', (rid,)).fetchone()
    if req and req['requested_by']:
        job = conn.execute('SELECT name FROM jobs WHERE id = ?', (req['job_id'],)).fetchone()
        create_notification(req['requested_by'], 'material_request',
            f"Material Request #{rid} Approved",
            f"Your material request for {job['name'] if job else 'Unknown'} has been approved",
            f"/material-requests")
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/material-requests/<int:rid>/fulfill', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_fulfill_material_request(rid):
    data = request.get_json() or {}
    conn = get_db()
    items = data.get('items', [])
    req = conn.execute('SELECT job_id FROM material_requests WHERE id = ?', (rid,)).fetchone()
    all_fulfilled = True
    for item in items:
        qty_fulfilled = float(item.get('quantity_fulfilled', 0))
        qty_requested = float(item.get('quantity_requested', 0))
        if qty_fulfilled < qty_requested:
            all_fulfilled = False
        conn.execute('''
            UPDATE material_request_items
            SET quantity_fulfilled = ?
            WHERE id = ?
        ''', (qty_fulfilled, item['id']))
        # Decrement inventory if inventory_item_id is linked
        inv_id = item.get('inventory_item_id')
        if inv_id and qty_fulfilled > 0:
            conn.execute('''
                UPDATE inventory_items
                SET quantity_on_hand = MAX(0, quantity_on_hand - ?),
                    updated_at = datetime('now','localtime')
                WHERE id = ?
            ''', (qty_fulfilled, inv_id))
            conn.execute('''
                INSERT INTO inventory_transactions
                (inventory_item_id, transaction_type, quantity, job_id, reference, notes, created_by)
                VALUES (?, 'issue', ?, ?, ?, ?, ?)
            ''', (inv_id, -qty_fulfilled, req['job_id'] if req else None,
                  f"Material Request #{rid}", f"Fulfilled from MR #{rid}",
                  session.get('user_id')))
    new_status = 'Fulfilled' if all_fulfilled else 'Partial'
    conn.execute('''
        UPDATE material_requests
        SET status = ?, fulfilled_by = ?, fulfilled_at = datetime('now','localtime'),
            updated_at = datetime('now','localtime')
        WHERE id = ?
    ''', (new_status, session.get('user_id'), rid))
    conn.commit()
    # Notify requester
    mr = conn.execute('SELECT requested_by, job_id FROM material_requests WHERE id = ?', (rid,)).fetchone()
    if mr and mr['requested_by']:
        job = conn.execute('SELECT name FROM jobs WHERE id = ?', (mr['job_id'],)).fetchone()
        create_notification(mr['requested_by'], 'material_request',
            f"Material Request #{rid} {new_status}",
            f"Your material request for {job['name'] if job else 'Unknown'} has been {new_status.lower()}",
            f"/material-requests")
    conn.close()
    return jsonify({'ok': True, 'status': new_status})


# ─── Material Orders ─────────────────────────────────────────────

@app.route('/orders')
@role_required('owner', 'admin', 'project_manager', 'warehouse')
def orders_list():
    return render_template('orders/list.html')

@app.route('/orders/<int:oid>')
@role_required('owner', 'admin', 'project_manager', 'warehouse')
def orders_detail(oid):
    return render_template('orders/detail.html', order_id=oid)

@app.route('/api/orders')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_orders_list():
    job_id = request.args.get('job_id', '')
    status = request.args.get('status', '')
    conn = get_db()
    query = '''SELECT mo.*, j.name as job_name,
               (SELECT COUNT(*) FROM material_order_items WHERE order_id = mo.id) as item_count
               FROM material_orders mo
               LEFT JOIN jobs j ON mo.job_id = j.id WHERE 1=1'''
    params = []
    if job_id:
        query += ' AND mo.job_id = ?'
        params.append(job_id)
    if status:
        query += ' AND mo.status = ?'
        params.append(status)
    query += ' ORDER BY mo.updated_at DESC'
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/orders/prepare')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_orders_prepare():
    job_id = request.args.get('job_id')
    takeoff_type = request.args.get('takeoff_type', 'residential')
    if not job_id:
        return jsonify({'error': 'job_id required'}), 400
    conn = get_db()
    job = conn.execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    if not job:
        conn.close()
        return jsonify({'error': 'Job not found'}), 404

    # Find baseline quote for this job
    quote = conn.execute('''SELECT * FROM supplier_quotes
                            WHERE job_id = ? AND is_baseline = 1
                            ORDER BY updated_at DESC LIMIT 1''', (job_id,)).fetchone()
    quote_items = []
    quote_id = None
    supplier_name = ''
    quote_number = ''
    if quote:
        quote_id = quote['id']
        supplier_name = quote['supplier_name'] or ''
        quote_number = quote['quote_number'] or ''
        quote_items = [dict(r) for r in conn.execute(
            'SELECT * FROM supplier_quote_items WHERE quote_id = ? ORDER BY line_number',
            (quote['id'],)).fetchall()]

    # Find bid for this job
    bid = conn.execute('SELECT * FROM bids WHERE job_id = ? ORDER BY updated_at DESC LIMIT 1',
                       (job_id,)).fetchone()
    bid_id = None
    takeoff_items = []
    unit_types = []
    systems = []
    config = {}

    if bid:
        bid_id = bid['id']
        if takeoff_type == 'residential':
            takeoff_items = [dict(r) for r in conn.execute(
                'SELECT * FROM bid_takeoff_items WHERE bid_id = ? AND enabled = 1 ORDER BY phase, sort_order',
                (bid['id'],)).fetchall()]
            unit_types = [dict(r) for r in conn.execute(
                'SELECT * FROM bid_takeoff_unit_types WHERE bid_id = ? ORDER BY sort_order',
                (bid['id'],)).fetchall()]
            try:
                config = json.loads(bid['takeoff_config'] or '{}')
            except Exception:
                config = {}
        else:
            takeoff_items = [dict(r) for r in conn.execute(
                'SELECT * FROM bid_commercial_takeoff_items WHERE bid_id = ? AND enabled = 1 ORDER BY phase, sort_order',
                (bid['id'],)).fetchall()]
            systems = [dict(r) for r in conn.execute(
                'SELECT * FROM bid_commercial_takeoff_systems WHERE bid_id = ? ORDER BY sort_order',
                (bid['id'],)).fetchall()]
            try:
                config = json.loads(bid['commercial_takeoff_config'] or '{}')
            except Exception:
                config = {}

    conn.close()
    return jsonify({
        'job_name': job['name'],
        'quote_id': quote_id,
        'bid_id': bid_id,
        'supplier_name': supplier_name,
        'quote_number': quote_number,
        'quote_items': quote_items,
        'takeoff_items': takeoff_items,
        'unit_types': unit_types,
        'systems': systems,
        'config': config
    })

@app.route('/api/orders', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_create_order():
    d = request.get_json()
    if not d or not d.get('job_id'):
        return jsonify({'error': 'job_id required'}), 400
    conn = get_db()

    # Generate order number
    count = conn.execute('SELECT COUNT(*) FROM material_orders').fetchone()[0]
    order_number = f'ORD-{count + 1:04d}'

    cur = conn.execute('''INSERT INTO material_orders
        (job_id, quote_id, bid_id, takeoff_type, order_number, supplier_name,
         status, subtotal, tax_amount, freight, total, notes, created_by)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (d['job_id'], d.get('quote_id'), d.get('bid_id'),
         d.get('takeoff_type', 'residential'), order_number,
         d.get('supplier_name', ''), 'Draft',
         d.get('subtotal', 0), d.get('tax_amount', 0), d.get('freight', 0),
         d.get('total', 0), d.get('notes', ''),
         session.get('user_id')))
    order_id = cur.lastrowid

    for item in d.get('items', []):
        conn.execute('''INSERT INTO material_order_items
            (order_id, line_number, sku, description, quote_qty, takeoff_qty,
             order_qty, unit_price, extended_price, takeoff_sku, source, discrepancy, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (order_id, item.get('line_number', 0), item.get('sku', ''),
             item.get('description', ''), item.get('quote_qty', 0),
             item.get('takeoff_qty', 0), item.get('order_qty', 0),
             item.get('unit_price', 0), item.get('extended_price', 0),
             item.get('takeoff_sku', ''), item.get('source', 'manual'),
             item.get('discrepancy', ''), item.get('notes', '')))

    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': order_id}), 201

@app.route('/api/orders/<int:oid>')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_order_detail(oid):
    conn = get_db()
    order = conn.execute('''SELECT mo.*, j.name as job_name
                            FROM material_orders mo
                            LEFT JOIN jobs j ON mo.job_id = j.id
                            WHERE mo.id = ?''', (oid,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Order not found'}), 404
    items = conn.execute('SELECT * FROM material_order_items WHERE order_id = ? ORDER BY line_number',
                         (oid,)).fetchall()
    conn.close()
    result = dict(order)
    result['items'] = [dict(i) for i in items]
    return jsonify(result)

@app.route('/api/orders/<int:oid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_update_order(oid):
    d = request.get_json()
    conn = get_db()
    order = conn.execute('SELECT * FROM material_orders WHERE id = ?', (oid,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Order not found'}), 404

    # Update order fields
    subtotal = 0
    items_data = d.get('items')
    if items_data is not None and order['status'] == 'Draft':
        conn.execute('DELETE FROM material_order_items WHERE order_id = ?', (oid,))
        for item in items_data:
            ext = (item.get('order_qty', 0) or 0) * (item.get('unit_price', 0) or 0)
            subtotal += ext
            conn.execute('''INSERT INTO material_order_items
                (order_id, line_number, sku, description, quote_qty, takeoff_qty,
                 order_qty, unit_price, extended_price, takeoff_sku, source, discrepancy, notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (oid, item.get('line_number', 0), item.get('sku', ''),
                 item.get('description', ''), item.get('quote_qty', 0),
                 item.get('takeoff_qty', 0), item.get('order_qty', 0),
                 item.get('unit_price', 0), ext,
                 item.get('takeoff_sku', ''), item.get('source', 'manual'),
                 item.get('discrepancy', ''), item.get('notes', '')))
    else:
        subtotal = order['subtotal'] or 0

    tax = d.get('tax_amount', order['tax_amount'] or 0)
    freight = d.get('freight', order['freight'] or 0)
    total = subtotal + tax + freight

    conn.execute('''UPDATE material_orders SET
        order_number = ?, expected_delivery = ?, notes = ?,
        subtotal = ?, tax_amount = ?, freight = ?, total = ?,
        updated_at = datetime('now','localtime')
        WHERE id = ?''',
        (d.get('order_number', order['order_number']),
         d.get('expected_delivery', order['expected_delivery']),
         d.get('notes', order['notes']),
         subtotal, tax, freight, total, oid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/orders/<int:oid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_delete_order(oid):
    conn = get_db()
    order = conn.execute('SELECT status FROM material_orders WHERE id = ?', (oid,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Order not found'}), 404
    if order['status'] != 'Draft':
        conn.close()
        return jsonify({'error': 'Only draft orders can be deleted'}), 400
    conn.execute('DELETE FROM material_order_items WHERE order_id = ?', (oid,))
    conn.execute('DELETE FROM material_orders WHERE id = ?', (oid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/orders/<int:oid>/submit', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_submit_order(oid):
    conn = get_db()
    order = conn.execute('SELECT * FROM material_orders WHERE id = ?', (oid,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Order not found'}), 404
    if order['status'] != 'Draft':
        conn.close()
        return jsonify({'error': 'Only draft orders can be submitted'}), 400
    conn.execute('''UPDATE material_orders SET status = 'Submitted',
        submitted_date = datetime('now','localtime'),
        updated_at = datetime('now','localtime') WHERE id = ?''', (oid,))
    conn.commit()

    # Notify owners/admins
    users = conn.execute("SELECT id FROM users WHERE role IN ('owner','admin') AND is_active = 1").fetchall()
    for u in users:
        create_notification(u['id'], 'order', 'Order Submitted',
            f"Order {order['order_number'] or 'ORD-'+str(oid)} has been submitted.",
            f"/orders/{oid}")
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/orders/<int:oid>/confirm', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_confirm_order(oid):
    conn = get_db()
    order = conn.execute('SELECT * FROM material_orders WHERE id = ?', (oid,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Order not found'}), 404
    if order['status'] != 'Submitted':
        conn.close()
        return jsonify({'error': 'Only submitted orders can be confirmed'}), 400
    conn.execute('''UPDATE material_orders SET status = 'Confirmed',
        confirmed_date = datetime('now','localtime'),
        updated_at = datetime('now','localtime') WHERE id = ?''', (oid,))
    conn.commit()

    if order['created_by']:
        create_notification(order['created_by'], 'order', 'Order Confirmed',
            f"Order {order['order_number'] or 'ORD-'+str(oid)} has been confirmed.",
            f"/orders/{oid}")
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/orders/<int:oid>/receive', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_receive_order(oid):
    d = request.get_json()
    conn = get_db()
    order = conn.execute('SELECT * FROM material_orders WHERE id = ?', (oid,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Order not found'}), 404
    if order['status'] not in ('Confirmed', 'Partial'):
        conn.close()
        return jsonify({'error': 'Order must be confirmed or partial to receive'}), 400

    for ri in d.get('items', []):
        item_id = ri.get('item_id')
        rcv_qty = ri.get('received_qty', 0)
        if item_id and rcv_qty > 0:
            conn.execute('''UPDATE material_order_items
                SET received_qty = received_qty + ?
                WHERE id = ? AND order_id = ?''',
                (rcv_qty, item_id, oid))

    # Check if all items fully received
    items = conn.execute('''SELECT order_qty, received_qty FROM material_order_items
                            WHERE order_id = ?''', (oid,)).fetchall()
    all_received = all(
        (row['received_qty'] or 0) >= (row['order_qty'] or 0)
        for row in items
    )
    new_status = 'Received' if all_received else 'Partial'
    update_fields = "status = ?, updated_at = datetime('now','localtime')"
    params = [new_status]
    if all_received:
        update_fields += ", received_date = datetime('now','localtime')"
    params.append(oid)
    conn.execute(f'UPDATE material_orders SET {update_fields} WHERE id = ?', params)
    conn.commit()

    if order['created_by']:
        msg = 'fully received' if all_received else 'partially received'
        create_notification(order['created_by'], 'order', f'Order {msg.title()}',
            f"Order {order['order_number'] or 'ORD-'+str(oid)} has been {msg}.",
            f"/orders/{oid}")
    conn.close()
    return jsonify({'ok': True, 'status': new_status})


# ─── Duplicate File Check ────────────────────────────────────────

@app.route('/api/files/check-duplicate', methods=['POST'])
@api_login_required
def api_check_duplicate():
    file = request.files.get('file')
    doc_type = request.form.get('doc_type', '')
    if not file or not file.filename:
        return jsonify({'error': 'No file'}), 400
    content = file.read()
    file.seek(0)  # Reset for potential subsequent read
    result = check_duplicate(content, doc_type, file.filename)
    return jsonify(result)


# ─── Universal Table Export (Excel + PDF) ────────────────────────

@app.route('/api/export/excel', methods=['POST'])
@api_login_required
def api_export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    payload = request.get_json(force=True)
    title = payload.get('title', 'Export')
    headers = payload.get('headers', [])
    rows = payload.get('rows', [])
    filename = payload.get('filename', title)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    money_format = '#,##0.00'

    # Header row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            # Try to convert money strings to numbers
            clean = val
            is_money = False
            if isinstance(val, str) and val.startswith('$'):
                is_money = True
                try:
                    clean = float(val.replace('$', '').replace(',', ''))
                except ValueError:
                    clean = val
            cell = ws.cell(row=r, column=c, value=clean)
            cell.border = thin_border
            if is_money and isinstance(clean, float):
                cell.number_format = money_format

    # Auto-width
    for c in range(1, len(headers) + 1):
        max_len = len(str(headers[c - 1]))
        for r in range(2, len(rows) + 2):
            val = ws.cell(row=r, column=c).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 3, 50)

    export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'exports')
    os.makedirs(export_dir, exist_ok=True)
    safe = "".join(ch if ch.isalnum() or ch in ' -_' else '' for ch in filename)
    filepath = os.path.join(export_dir, f'{safe}.xlsx')
    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name=f'{safe}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/export/pdf', methods=['POST'])
@api_login_required
def api_export_pdf():
    payload = request.get_json(force=True)
    title = payload.get('title', 'Export')
    headers = payload.get('headers', [])
    rows = payload.get('rows', [])
    filename = payload.get('filename', title)

    generated_at = datetime.now().strftime('%B %d, %Y %I:%M %p')
    html = render_template('export_table_pdf.html', title=title, headers=headers,
                           rows=rows, generated_at=generated_at)

    export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'exports')
    os.makedirs(export_dir, exist_ok=True)
    safe = "".join(ch if ch.isalnum() or ch in ' -_' else '' for ch in filename)
    filepath = os.path.join(export_dir, f'{safe}.pdf')

    try:
        wp = weasyprint.HTML(string=html, base_url=os.path.dirname(os.path.abspath(__file__)))
        wp.write_pdf(filepath)
    except Exception as e:
        return jsonify({'error': f'PDF generation failed: {str(e)[:200]}'}), 500

    return send_file(filepath, as_attachment=True, download_name=f'{safe}.pdf',
                     mimetype='application/pdf')


# ─── Submittal AI Analysis + Duplicate Detection ────────────────

@app.route('/api/submittals/analyze', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_analyze_submittal():
    import anthropic

    file = request.files.get('file')
    job_id = request.form.get('job_id')
    if not file or not file.filename or not job_id:
        return jsonify({'error': 'File and job_id required'}), 400

    # Extract text from PDF
    pdf_text = ''
    try:
        import pdfplumber
        file_bytes = file.read()
        file.seek(0)
        import io
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages[:5]:  # First 5 pages max
                pdf_text += (page.extract_text() or '') + '\n'
    except Exception:
        pdf_text = ''

    if not pdf_text.strip():
        return jsonify({'product_name': '', 'manufacturer': '', 'description': '', 'duplicate_of': None})

    # Call Claude to extract product info
    client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY', ''))
    try:
        extract_resp = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=300,
            messages=[{
                'role': 'user',
                'content': f"""Extract from this HVAC submittal document:
- product_name: the primary product name/model
- manufacturer: the manufacturer or brand
- description: a short one-line description suitable for a submittal log

Return ONLY valid JSON: {{"product_name":"...","manufacturer":"...","description":"..."}}

Document text:
{pdf_text[:3000]}"""
            }]
        )
        import re
        raw = extract_resp.content[0].text
        match = re.search(r'\{[^}]+\}', raw, re.DOTALL)
        info = json.loads(match.group()) if match else {}
    except Exception:
        info = {}

    product_name = info.get('product_name', '')
    manufacturer = info.get('manufacturer', '')
    description = info.get('description', '')

    # Check for duplicates among existing submittals for this job
    duplicate_of = None
    conn = get_db()
    existing = conn.execute(
        'SELECT id, submittal_number, description, vendor FROM submittals WHERE job_id = ?',
        (job_id,)
    ).fetchall()
    conn.close()

    if existing and (product_name or description):
        existing_list = [dict(r) for r in existing]
        try:
            dup_resp = client.messages.create(
                model='claude-haiku-4-5-20251001',
                max_tokens=200,
                messages=[{
                    'role': 'user',
                    'content': f"""I'm uploading a submittal for: {product_name} by {manufacturer} — {description}

Existing submittals for this job:
{json.dumps(existing_list, default=str)}

Does this new submittal match any existing one (same product/model)?
Return ONLY valid JSON: {{"duplicate_of_id": <id or null>, "confidence": "high"/"medium"/"low"}}"""
                }]
            )
            raw2 = dup_resp.content[0].text
            match2 = re.search(r'\{[^}]+\}', raw2, re.DOTALL)
            dup_info = json.loads(match2.group()) if match2 else {}
            dup_id = dup_info.get('duplicate_of_id')
            confidence = dup_info.get('confidence', 'low')
            if dup_id and confidence in ('high', 'medium'):
                dup_row = next((e for e in existing_list if e['id'] == dup_id), None)
                if dup_row:
                    duplicate_of = {
                        'id': dup_row['id'],
                        'submittal_number': dup_row['submittal_number'],
                        'description': dup_row['description'] or '',
                    }
        except Exception:
            pass

    return jsonify({
        'product_name': product_name,
        'manufacturer': manufacturer,
        'description': description,
        'duplicate_of': duplicate_of,
    })


# ─── Feedback / Feature Requests ─────────────────────────────────

@app.route('/feedback')
@login_required
def feedback_page():
    return render_template('feedback/list.html')

@app.route('/api/feedback')
@api_login_required
def api_feedback_list():
    conn = get_db()
    rows = conn.execute(
        '''SELECT f.*, u.display_name as submitter_name,
           (SELECT COUNT(*) FROM feedback_upvotes WHERE feedback_id = f.id) as upvote_count
           FROM feedback_requests f
           LEFT JOIN users u ON f.submitted_by = u.id
           ORDER BY f.created_at DESC'''
    ).fetchall()
    result = [dict(r) for r in rows]
    # Check if current user has upvoted each item
    user_id = session['user_id']
    user_upvotes = set()
    uv_rows = conn.execute('SELECT feedback_id FROM feedback_upvotes WHERE user_id = ?', (user_id,)).fetchall()
    for uv in uv_rows:
        user_upvotes.add(uv['feedback_id'])
    for r in result:
        r['user_upvoted'] = r['id'] in user_upvotes
        r['upvotes'] = r['upvote_count']
    conn.close()
    return jsonify(result)

@app.route('/api/feedback', methods=['POST'])
@api_login_required
def api_feedback_create():
    data = request.get_json(force=True)
    title = data.get('title', '').strip()
    if not title:
        return jsonify({'error': 'Title is required'}), 400
    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO feedback_requests (title, description, category, priority, submitted_by)
           VALUES (?,?,?,?,?)''',
        (title, data.get('description', ''), data.get('category', 'Feature'),
         data.get('priority', 'Medium'), session['user_id'])
    )
    fb_id = cursor.lastrowid
    conn.commit()
    # Notify all owners
    owners = conn.execute("SELECT id FROM users WHERE role = 'owner'").fetchall()
    for owner in owners:
        create_notification(
            owner['id'], 'feedback', 'New Feedback Request',
            f'{title}', '/feedback'
        )
    conn.close()
    return jsonify({'ok': True, 'id': fb_id})

@app.route('/api/feedback/<int:fid>', methods=['PUT'])
@api_login_required
def api_feedback_update(fid):
    data = request.get_json(force=True)
    conn = get_db()
    fb = conn.execute('SELECT * FROM feedback_requests WHERE id = ?', (fid,)).fetchone()
    if not fb:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    user_id = session['user_id']
    user_role = session.get('role', '')

    # Owners can update status and response; submitter can update title/description
    if user_role == 'owner':
        status = data.get('status', fb['status'])
        owner_response = data.get('owner_response', fb['owner_response'])
        title = data.get('title', fb['title'])
        description = data.get('description', fb['description'])
        category = data.get('category', fb['category'])
        priority = data.get('priority', fb['priority'])
    elif user_id == fb['submitted_by']:
        status = fb['status']
        owner_response = fb['owner_response']
        title = data.get('title', fb['title'])
        description = data.get('description', fb['description'])
        category = data.get('category', fb['category'])
        priority = data.get('priority', fb['priority'])
    else:
        conn.close()
        return jsonify({'error': 'Unauthorized'}), 403

    old_status = fb['status']

    conn.execute(
        '''UPDATE feedback_requests SET title=?, description=?, category=?, priority=?,
           status=?, owner_response=?, updated_at=datetime('now','localtime') WHERE id=?''',
        (title, description, category, priority, status, owner_response, fid)
    )
    conn.commit()

    # Send notification to submitter when status changes
    if status != old_status and fb['submitted_by'] != user_id:
        submitter_id = fb['submitted_by']
        if status == 'Completed':
            create_notification(
                submitter_id, 'feedback',
                'Feedback Completed',
                f'Your feedback "{fb["title"]}" has been marked as completed.',
                '/feedback'
            )
        else:
            create_notification(
                submitter_id, 'feedback',
                'Feedback Status Updated',
                f'Your feedback "{fb["title"]}" status changed to {status}.',
                '/feedback'
            )

    conn.close()
    return jsonify({'ok': True})

@app.route('/api/feedback/<int:fid>/upvote', methods=['POST'])
@api_login_required
def api_feedback_upvote(fid):
    user_id = session['user_id']
    conn = get_db()
    existing = conn.execute(
        'SELECT id FROM feedback_upvotes WHERE feedback_id = ? AND user_id = ?',
        (fid, user_id)
    ).fetchone()
    if existing:
        conn.execute('DELETE FROM feedback_upvotes WHERE id = ?', (existing['id'],))
        conn.execute('UPDATE feedback_requests SET upvotes = MAX(0, upvotes - 1) WHERE id = ?', (fid,))
    else:
        conn.execute('INSERT INTO feedback_upvotes (feedback_id, user_id) VALUES (?,?)', (fid, user_id))
        conn.execute('UPDATE feedback_requests SET upvotes = upvotes + 1 WHERE id = ?', (fid,))
    conn.commit()
    new_count = conn.execute('SELECT upvotes FROM feedback_requests WHERE id = ?', (fid,)).fetchone()
    conn.close()
    return jsonify({'ok': True, 'upvotes': new_count['upvotes'] if new_count else 0, 'toggled': not existing})

@app.route('/api/feedback/<int:fid>', methods=['DELETE'])
@api_role_required('owner')
def api_feedback_delete(fid):
    conn = get_db()
    conn.execute('DELETE FROM feedback_requests WHERE id = ?', (fid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ─── Job Pipeline (32-Step Workflow) ──────────────────────────────

PIPELINE_STEPS = [
    (1, 'Receive ITB / Invitation to Bid', 'bidding', 'bids'),
    (2, 'Review Plans & Specifications', 'bidding', 'plans'),
    (3, 'Perform Takeoff', 'bidding', 'plans'),
    (4, 'Request Supplier Quotes', 'bidding', 'supplier_quotes'),
    (5, 'Build Bid / Estimate', 'bidding', 'bids'),
    (6, 'Create Proposal', 'bidding', 'bids'),
    (7, 'Submit Bid', 'bidding', 'bids'),
    (8, 'Bid Follow-up', 'bidding', 'bid_followups'),
    (9, 'Contract Execution', 'contract', 'contracts'),
    (10, 'Bonds & Insurance (COI)', 'contract', 'coi'),
    (11, 'Notice to Proceed', 'contract', 'contracts'),
    (12, 'Set Up Pay App / SOV', 'contract', 'payapps'),
    (13, 'Permits & Inspections', 'contract', 'permits'),
    (14, 'Assign Project Manager', 'contract', 'projects'),
    (15, 'Submittal Preparation', 'preconstruction', 'submittals'),
    (16, 'Submittal Submission', 'preconstruction', 'submittals'),
    (17, 'Submittal Approval', 'preconstruction', 'submittals'),
    (18, 'Pre-Construction Meeting', 'preconstruction', 'precon'),
    (19, 'RFI Resolution', 'preconstruction', 'rfis'),
    (20, 'Material Ordering / Delivery Schedule', 'materials', 'materials'),
    (21, 'Material Receiving', 'materials', 'receiving'),
    (22, 'Material Shortage Check', 'materials', 'inventory'),
    (23, 'Invoice Verification', 'finance', 'invoices'),
    (24, 'Pay Application Submission', 'finance', 'payapps'),
    (25, 'Material Shipping by Phase', 'construction', 'shipments'),
    (26, 'Rough-In', 'construction', 'schedule'),
    (27, 'Trim Out', 'construction', 'schedule'),
    (28, 'Equipment Start-Up', 'construction', 'schedule'),
    (29, 'Job Photos / Documentation', 'construction', 'photos'),
    (30, 'Punch List', 'construction', 'projects'),
    (31, 'Closeout / O&M / Warranty', 'closeout', 'documents'),
    (32, 'Final Billing / Lien Waiver / COI', 'closeout', 'billing'),
]

MODULE_LINKS = {
    'bids': '/bids', 'plans': '/plans', 'supplier_quotes': '/supplier-quotes',
    'bid_followups': '/bids', 'contracts': '/contracts', 'coi': '/coi',
    'payapps': '/payapps', 'projects': '/projects/{job_id}',
    'permits': '/permits', 'submittals': '/submittals',
    'precon': '/projects/{job_id}?tab=precon',
    'rfis': '/rfis', 'materials': '/materials/job/{job_id}',
    'receiving': '/receiving', 'inventory': '/inventory',
    'invoices': '/invoices', 'shipments': '/material-shipments',
    'schedule': '/schedule', 'photos': '/photos',
    'documents': '/documents', 'billing': '/projects/{job_id}',
}

def seed_pipeline_steps(conn, job_id):
    """Idempotent seed of 32 pipeline steps for a job."""
    existing = conn.execute('SELECT step_number FROM job_pipeline_steps WHERE job_id = ?', (job_id,)).fetchall()
    existing_nums = {r['step_number'] for r in existing}
    for step_num, name, category, module in PIPELINE_STEPS:
        if step_num not in existing_nums:
            conn.execute(
                '''INSERT INTO job_pipeline_steps (job_id, step_number, step_name, step_category, linked_module)
                   VALUES (?,?,?,?,?)''',
                (job_id, step_num, name, category, module)
            )

def auto_detect_pipeline(conn, job_id):
    """Auto-detect step completion from linked modules."""
    updates = {}
    # Step 1: bid exists
    bid = conn.execute('SELECT id, status FROM bids WHERE job_id = ?', (job_id,)).fetchone()
    if bid:
        updates[1] = 'complete'
        # Step 5: bid has values
        if bid['status'] in ('Draft', 'Submitted', 'Accepted', 'Rejected'):
            updates[5] = 'complete'
        # Step 6: proposal exists (bid with status beyond draft)
        if bid['status'] in ('Submitted', 'Accepted', 'Rejected'):
            updates[6] = 'complete'
        # Step 7: bid submitted
        if bid['status'] in ('Submitted', 'Accepted', 'Rejected'):
            updates[7] = 'complete'
        # Step 8: followups
        followup = conn.execute('SELECT id FROM bid_followups WHERE bid_id = ?', (bid['id'],)).fetchone()
        if followup:
            updates[8] = 'complete'

    # Step 2-3: plans reviewed / takeoff
    plan_reviewed = conn.execute("SELECT id FROM plans WHERE job_id = ? AND status = 'Reviewed'", (job_id,)).fetchone()
    plan_takeoff = conn.execute("SELECT id FROM plans WHERE job_id = ? AND status = 'Takeoff Complete'", (job_id,)).fetchone()
    if plan_reviewed or plan_takeoff:
        updates[2] = 'complete'
    if plan_takeoff:
        updates[3] = 'complete'

    # Step 4: supplier quotes
    sq = conn.execute("SELECT id FROM supplier_quotes WHERE job_id = ? AND status IN ('Received','Reviewing','Selected')", (job_id,)).fetchone()
    if sq:
        updates[4] = 'complete'

    # Step 9: contract exists
    contract = conn.execute('SELECT id FROM contracts WHERE job_id = ?', (job_id,)).fetchone()
    if contract:
        updates[9] = 'complete'

    # Step 10: COI exists
    coi = conn.execute('SELECT id FROM certificates_of_insurance WHERE job_id = ?', (job_id,)).fetchone()
    if coi:
        updates[10] = 'complete'

    # Step 12: pay app contract exists
    pac = conn.execute('SELECT id FROM pay_app_contracts WHERE job_id = ?', (job_id,)).fetchone()
    if pac:
        updates[12] = 'complete'

    # Step 13: permits — all permits approved or N/A
    permits = conn.execute('SELECT status FROM permits WHERE job_id = ?', (job_id,)).fetchall()
    if permits and all(p['status'] in ('Approved', 'N/A') for p in permits):
        updates[13] = 'complete'

    # Step 14: PM assigned
    job = conn.execute('SELECT project_manager_id FROM jobs WHERE id = ?', (job_id,)).fetchone()
    if job and job['project_manager_id']:
        updates[14] = 'complete'

    # Steps 15-17: submittals
    sub_submitted = conn.execute("SELECT id FROM submittals WHERE job_id = ? AND status IN ('Submitted','Approved','Approved as Noted')", (job_id,)).fetchone()
    sub_approved = conn.execute("SELECT id FROM submittals WHERE job_id = ? AND status IN ('Approved','Approved as Noted')", (job_id,)).fetchone()
    sub_any = conn.execute("SELECT id FROM submittals WHERE job_id = ?", (job_id,)).fetchone()
    if sub_any:
        updates[15] = 'complete'
    if sub_submitted:
        updates[16] = 'complete'
    if sub_approved:
        updates[17] = 'complete'

    # Step 18: precon meeting
    precon = conn.execute("SELECT id FROM precon_meetings WHERE job_id = ? AND status = 'Completed'", (job_id,)).fetchone()
    if precon:
        updates[18] = 'complete'

    # Step 19: RFIs
    rfi_closed = conn.execute("SELECT id FROM rfis WHERE job_id = ? AND status = 'Closed'", (job_id,)).fetchone()
    if rfi_closed:
        updates[19] = 'complete'

    # Step 20: materials ordered
    mat_ordered = conn.execute("SELECT id FROM line_items WHERE job_id = ? AND qty_ordered > 0", (job_id,)).fetchone()
    if mat_ordered:
        updates[20] = 'complete'

    # Step 24: pay app submitted
    if pac:
        pa = conn.execute("SELECT id FROM pay_applications WHERE contract_id = ? AND status IN ('Submitted','Approved','Paid')", (pac['id'],)).fetchone()
        if pa:
            updates[24] = 'complete'

    # Step 26-28: schedule phases
    for step, phase in [(26, 'Rough-In'), (27, 'Trim Out'), (28, 'Startup')]:
        phase_done = conn.execute("SELECT id FROM job_schedule_events WHERE job_id = ? AND phase_name = ? AND status = 'Complete'", (job_id, phase)).fetchone()
        if phase_done:
            updates[step] = 'complete'

    # Step 31: closeout items complete
    closeout = conn.execute("SELECT COUNT(*) as total, SUM(CASE WHEN status='Complete' THEN 1 ELSE 0 END) as done FROM closeout_checklists WHERE job_id = ?", (job_id,)).fetchone()
    if closeout and closeout['total'] > 0 and closeout['done'] == closeout['total']:
        updates[31] = 'complete'

    # Step 32: lien waiver executed
    lw = conn.execute("SELECT id FROM lien_waivers WHERE job_id = ? AND status = 'Executed'", (job_id,)).fetchone()
    if lw:
        updates[32] = 'complete'

    # Apply auto-detected updates (only upgrade, never downgrade manual overrides)
    for step_num, new_status in updates.items():
        conn.execute(
            '''UPDATE job_pipeline_steps SET status = ?, updated_at = datetime('now','localtime')
               WHERE job_id = ? AND step_number = ? AND status IN ('pending','active')''',
            (new_status, job_id, step_num)
        )

@app.route('/api/jobs/<int:job_id>/pipeline/seed', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_seed_pipeline(job_id):
    conn = get_db()
    seed_pipeline_steps(conn, job_id)
    auto_detect_pipeline(conn, job_id)
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/jobs/<int:job_id>/pipeline')
@api_role_required('owner', 'admin', 'project_manager')
def api_get_pipeline(job_id):
    conn = get_db()
    # Auto-seed if empty
    existing = conn.execute('SELECT COUNT(*) FROM job_pipeline_steps WHERE job_id = ?', (job_id,)).fetchone()[0]
    if existing == 0:
        seed_pipeline_steps(conn, job_id)
        conn.commit()
    auto_detect_pipeline(conn, job_id)
    conn.commit()
    steps = conn.execute(
        'SELECT * FROM job_pipeline_steps WHERE job_id = ? ORDER BY step_number', (job_id,)
    ).fetchall()
    conn.close()
    result = []
    for s in steps:
        d = dict(s)
        link_tpl = MODULE_LINKS.get(d['linked_module'], '')
        d['module_link'] = link_tpl.replace('{job_id}', str(job_id))
        result.append(d)
    return jsonify(result)

@app.route('/api/jobs/<int:job_id>/pipeline/<int:step>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_pipeline_step(job_id, step):
    data = request.get_json() or {}
    conn = get_db()
    updates = []
    params = []
    if 'status' in data:
        updates.append('status = ?')
        params.append(data['status'])
        if data['status'] == 'complete':
            updates.append('completed_date = date("now","localtime")')
            updates.append('completed_by = ?')
            params.append(session.get('user_id'))
    if 'notes' in data:
        updates.append('notes = ?')
        params.append(data['notes'])
    updates.append('updated_at = datetime("now","localtime")')
    params.extend([job_id, step])
    conn.execute(
        f'UPDATE job_pipeline_steps SET {", ".join(updates)} WHERE job_id = ? AND step_number = ?',
        params
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/pipeline/overview')
@api_role_required('owner', 'admin', 'project_manager')
def api_pipeline_overview():
    conn = get_db()
    jobs = conn.execute("SELECT id, name, status FROM jobs WHERE status NOT IN ('Complete','Cancelled') ORDER BY name").fetchall()
    result = []
    for job in jobs:
        steps = conn.execute(
            'SELECT step_number, step_name, step_category, status FROM job_pipeline_steps WHERE job_id = ? ORDER BY step_number',
            (job['id'],)
        ).fetchall()
        if not steps:
            continue
        active_step = None
        complete_count = sum(1 for s in steps if s['status'] == 'complete')
        for s in steps:
            if s['status'] in ('pending', 'active'):
                active_step = dict(s)
                break
        result.append({
            'job_id': job['id'], 'job_name': job['name'], 'job_status': job['status'],
            'total_steps': len(steps), 'complete_steps': complete_count,
            'current_step': active_step,
            'steps': [dict(s) for s in steps]
        })
    conn.close()
    return jsonify(result)

# ─── Material Receiving ──────────────────────────────────────────

@app.route('/receiving')
@role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def receiving_page():
    return render_template('receiving/list.html')

@app.route('/api/receiving/<int:job_id>')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def api_receiving_items(job_id):
    conn = get_db()
    items = conn.execute('''
        SELECT li.id, li.sku, li.description, li.qty_ordered,
               COALESCE(SUM(re.quantity), 0) as qty_received
        FROM line_items li
        LEFT JOIN received_entries re ON re.line_item_id = li.id
        WHERE li.job_id = ?
        GROUP BY li.id
        ORDER BY li.line_number
    ''', (job_id,)).fetchall()
    conn.close()
    result = []
    for i in items:
        d = dict(i)
        d['qty_remaining'] = max(0, (d['qty_ordered'] or 0) - (d['qty_received'] or 0))
        result.append(d)
    return jsonify(result)

@app.route('/api/receiving/<int:job_id>/quick', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_quick_receive(job_id):
    data = request.get_json() or {}
    items = data.get('items', [])
    conn = get_db()
    # Create delivery receipt
    cur = conn.execute(
        '''INSERT INTO delivery_receipts (job_id, supplier_name, po_number, received_by, notes)
           VALUES (?,?,?,?,?)''',
        (job_id, data.get('supplier_name',''), data.get('po_number',''),
         session.get('user_id'), data.get('notes',''))
    )
    receipt_id = cur.lastrowid
    for item in items:
        lid = item.get('line_item_id')
        qty = float(item.get('qty', 0) or 0)
        if qty <= 0 or not lid:
            continue
        # Find next available column
        col = conn.execute(
            'SELECT COALESCE(MAX(column_number), 0) + 1 as next_col FROM received_entries WHERE line_item_id = ?',
            (lid,)
        ).fetchone()['next_col']
        if col > 15:
            col = 15
        conn.execute(
            'INSERT OR REPLACE INTO received_entries (line_item_id, column_number, quantity, entry_date) VALUES (?,?,?,date("now","localtime"))',
            (lid, col, qty)
        )
        # Also update inventory if item exists
        inv = conn.execute('SELECT id FROM inventory_items WHERE sku = (SELECT sku FROM line_items WHERE id = ?)', (lid,)).fetchone()
        if inv:
            conn.execute('UPDATE inventory_items SET quantity_on_hand = quantity_on_hand + ?, updated_at=datetime("now","localtime") WHERE id=?', (qty, inv['id']))
            conn.execute(
                'INSERT INTO inventory_transactions (inventory_item_id, transaction_type, quantity, job_id, reference, created_by) VALUES (?,?,?,?,?,?)',
                (inv['id'], 'receive', qty, job_id, f'Receipt #{receipt_id}', session.get('user_id'))
            )
    conn.commit()
    # Check for shortage alerts after receiving
    check_reorder_alerts(conn)
    conn.close()
    return jsonify({'ok': True, 'receipt_id': receipt_id})

@app.route('/api/receiving/pending')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def api_receiving_pending():
    conn = get_db()
    jobs = conn.execute('''
        SELECT j.id, j.name, COUNT(li.id) as total_items,
               SUM(CASE WHEN li.qty_ordered > COALESCE(recv.received, 0) THEN 1 ELSE 0 END) as items_pending
        FROM jobs j
        JOIN line_items li ON li.job_id = j.id AND li.qty_ordered > 0
        LEFT JOIN (
            SELECT line_item_id, SUM(quantity) as received
            FROM received_entries GROUP BY line_item_id
        ) recv ON recv.line_item_id = li.id
        WHERE j.status NOT IN ('Complete','Cancelled')
        GROUP BY j.id
        HAVING items_pending > 0
        ORDER BY j.name
    ''').fetchall()
    conn.close()
    return jsonify([dict(j) for j in jobs])

# ─── Material Shortage Notifications ──────────────────────────────

def check_reorder_alerts(conn):
    """Check for items below reorder point and create notifications."""
    low_items = conn.execute('''
        SELECT id, sku, description, quantity_on_hand, reorder_point
        FROM inventory_items
        WHERE reorder_point > 0 AND quantity_on_hand <= reorder_point
    ''').fetchall()
    if not low_items:
        return
    # Notify warehouse + admin users
    users = conn.execute("SELECT id FROM users WHERE role IN ('owner','admin','warehouse') AND is_active = 1").fetchall()
    for item in low_items:
        for u in users:
            # Check if notification already exists (avoid duplicates)
            existing = conn.execute(
                "SELECT id FROM notifications WHERE user_id = ? AND type = 'inventory' AND title LIKE ? AND is_read = 0",
                (u['id'], f'%{item["sku"]}%')
            ).fetchone()
            if not existing:
                create_notification(
                    u['id'], 'inventory',
                    f'Low Stock: {item["sku"]}',
                    f'{item["description"]} — {item["quantity_on_hand"]} on hand (reorder point: {item["reorder_point"]})',
                    '/inventory'
                )

# ─── Drag-Drop Invoice Upload ────────────────────────────────────

@app.route('/api/invoices/upload', methods=['POST'])
@api_role_required('owner', 'admin')
def api_upload_invoice():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    job_id = request.form.get('job_id') or None
    os.makedirs(os.path.join(app.root_path, 'data', 'invoices'), exist_ok=True)
    filename = f'{datetime.now().strftime("%Y%m%d_%H%M%S")}_{f.filename}'
    filepath = os.path.join(app.root_path, 'data', 'invoices', filename)
    f.save(filepath)
    conn = get_db()
    # Get or create a default supplier config
    supplier = conn.execute('SELECT id FROM billtrust_config LIMIT 1').fetchone()
    supplier_id = supplier['id'] if supplier else 1
    conn.execute(
        '''INSERT INTO supplier_invoices (supplier_config_id, invoice_number, status, job_id, notes, total)
           VALUES (?,?,?,?,?,?)''',
        (supplier_id, f.filename.replace('.pdf','').replace('.PDF',''),
         'open', job_id, f'Uploaded via drag-drop: {filepath}', 0)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'file': filename})

# ─── Invoice Verification ────────────────────────────────────────

@app.route('/api/invoices/<int:iid>/verify', methods=['POST'])
@api_role_required('owner', 'admin')
def api_verify_invoice(iid):
    conn = get_db()
    invoice = conn.execute('SELECT * FROM supplier_invoices WHERE id = ?', (iid,)).fetchone()
    if not invoice:
        conn.close()
        return jsonify({'error': 'Invoice not found'}), 404
    quote_id = invoice['supplier_quote_id']
    if not quote_id:
        conn.close()
        return jsonify({'error': 'No linked supplier quote to verify against'}), 400

    # Get quote items
    quote_items = conn.execute(
        'SELECT sku, description, quantity, unit_price, extended_price FROM supplier_quote_items WHERE quote_id = ?',
        (quote_id,)
    ).fetchall()
    quote_map = {qi['sku']: dict(qi) for qi in quote_items if qi['sku']}

    # Parse invoice line items (JSON)
    inv_lines = json.loads(invoice['line_items'] or '[]')

    flags = []
    matched = 0
    for line in inv_lines:
        inv_sku = line.get('sku', '')
        inv_qty = float(line.get('quantity', 0) or 0)
        inv_price = float(line.get('unit_price', 0) or 0)
        if inv_sku in quote_map:
            qi = quote_map[inv_sku]
            matched += 1
            if abs(inv_qty - qi['quantity']) > 0.01:
                flags.append({
                    'sku': inv_sku, 'severity': 'error', 'category': 'qty_mismatch',
                    'message': f'Qty mismatch for {inv_sku}: invoice={inv_qty}, quote={qi["quantity"]}'
                })
            if abs(inv_price - qi['unit_price']) > 0.01:
                variance = ((inv_price - qi['unit_price']) / qi['unit_price'] * 100) if qi['unit_price'] else 0
                flags.append({
                    'sku': inv_sku, 'severity': 'warning', 'category': 'price_mismatch',
                    'message': f'Price mismatch for {inv_sku}: invoice=${inv_price:.2f}, quote=${qi["unit_price"]:.2f} ({variance:+.1f}%)'
                })
        elif inv_sku:
            flags.append({
                'sku': inv_sku, 'severity': 'warning', 'category': 'not_on_quote',
                'message': f'{inv_sku} on invoice but not found in quote'
            })

    # Check for quote items missing from invoice
    inv_skus = {line.get('sku','') for line in inv_lines}
    for sku, qi in quote_map.items():
        if sku not in inv_skus:
            flags.append({
                'sku': sku, 'severity': 'info', 'category': 'missing_from_invoice',
                'message': f'{sku} on quote but not on invoice (qty={qi["quantity"]})'
            })

    verification_data = json.dumps({
        'flags': flags, 'matched': matched,
        'total_invoice_lines': len(inv_lines), 'total_quote_lines': len(quote_items)
    })
    status = 'verified' if not flags else ('issues_found' if any(f['severity'] == 'error' for f in flags) else 'reviewed')
    conn.execute(
        'UPDATE supplier_invoices SET verification_status=?, verification_data=?, updated_at=datetime("now","localtime") WHERE id=?',
        (status, verification_data, iid)
    )
    # Store flags in invoice_review_flags table
    conn.execute('DELETE FROM invoice_review_flags WHERE invoice_id = ?', (iid,))
    for flag in flags:
        conn.execute(
            'INSERT INTO invoice_review_flags (invoice_id, invoice_number, job_id, severity, category, message) VALUES (?,?,?,?,?,?)',
            (iid, invoice['invoice_number'], invoice['job_id'], flag['severity'], flag['category'], flag['message'])
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'status': status, 'flags': flags, 'matched': matched})

# ─── Job Photos ──────────────────────────────────────────────────

@app.route('/photos')
@role_required('owner', 'admin', 'project_manager', 'warehouse')
def photos_page():
    return render_template('photos/gallery.html')

@app.route('/api/photos')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_list_photos():
    job_id = request.args.get('job_id', type=int)
    category = request.args.get('category', '')
    album_id = request.args.get('album_id', '')
    conn = get_db()
    query = 'SELECT p.*, j.name as job_name, u.display_name as uploaded_by_name FROM job_photos p LEFT JOIN jobs j ON p.job_id = j.id LEFT JOIN users u ON p.uploaded_by = u.id WHERE 1=1'
    params = []
    if job_id:
        query += ' AND p.job_id = ?'
        params.append(job_id)
    if category:
        query += ' AND p.category = ?'
        params.append(category)
    if album_id == 'none':
        query += ' AND p.album_id IS NULL'
    elif album_id:
        query += ' AND p.album_id = ?'
        params.append(int(album_id))
    query += ' ORDER BY p.created_at DESC'
    photos = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(p) for p in photos])

@app.route('/api/photos', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_upload_photos():
    job_id = request.form.get('job_id')
    if not job_id:
        return jsonify({'error': 'Job required'}), 400
    category = request.form.get('category', 'General')
    caption = request.form.get('caption', '')
    album_id = request.form.get('album_id') or None
    os.makedirs(os.path.join(app.root_path, 'data', 'photos'), exist_ok=True)
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'No files'}), 400
    conn = get_db()
    ids = []
    for f in files:
        filename = f'{datetime.now().strftime("%Y%m%d_%H%M%S")}_{f.filename}'
        filepath = os.path.join(app.root_path, 'data', 'photos', filename)
        f.save(filepath)
        cur = conn.execute(
            '''INSERT INTO job_photos (job_id, file_path, caption, category, taken_date, uploaded_by, album_id)
               VALUES (?,?,?,?,date('now','localtime'),?,?)''',
            (job_id, f'data/photos/{filename}', caption, category, session.get('user_id'), album_id)
        )
        ids.append(cur.lastrowid)
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'ids': ids})

@app.route('/api/photos/<int:pid>/file')
@api_login_required
def api_photo_file(pid):
    conn = get_db()
    photo = conn.execute('SELECT file_path FROM job_photos WHERE id = ?', (pid,)).fetchone()
    conn.close()
    if not photo:
        return jsonify({'error': 'Not found'}), 404
    return send_file(os.path.join(app.root_path, photo['file_path']))

@app.route('/api/photos/<int:pid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_photo(pid):
    conn = get_db()
    photo = conn.execute('SELECT file_path FROM job_photos WHERE id = ?', (pid,)).fetchone()
    if photo:
        path = os.path.join(app.root_path, photo['file_path'])
        if os.path.exists(path):
            os.remove(path)
    conn.execute('DELETE FROM job_photos WHERE id = ?', (pid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Photo Albums ────────────────────────────────────────────────

@app.route('/api/photos/albums')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_list_albums():
    job_id = request.args.get('job_id', type=int)
    if not job_id:
        return jsonify([])
    conn = get_db()
    albums = conn.execute('''
        SELECT a.*, COUNT(p.id) as photo_count,
               (SELECT pp.id FROM job_photos pp WHERE pp.album_id = a.id ORDER BY pp.created_at DESC LIMIT 1) as cover_photo_id
        FROM photo_albums a
        LEFT JOIN job_photos p ON p.album_id = a.id
        WHERE a.job_id = ?
        GROUP BY a.id
        ORDER BY a.name
    ''', (job_id,)).fetchall()
    conn.close()
    return jsonify([dict(a) for a in albums])

@app.route('/api/photos/albums', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_create_album():
    data = request.get_json()
    job_id = data.get('job_id')
    name = (data.get('name') or '').strip()
    if not job_id or not name:
        return jsonify({'error': 'Job and name required'}), 400
    conn = get_db()
    cur = conn.execute('INSERT INTO photo_albums (job_id, name) VALUES (?, ?)', (job_id, name))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cur.lastrowid})

@app.route('/api/photos/albums/<int:aid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_album(aid):
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Name required'}), 400
    conn = get_db()
    conn.execute('UPDATE photo_albums SET name = ? WHERE id = ?', (name, aid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/photos/albums/<int:aid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_album(aid):
    conn = get_db()
    conn.execute('UPDATE job_photos SET album_id = NULL WHERE album_id = ?', (aid,))
    conn.execute('DELETE FROM photo_albums WHERE id = ?', (aid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/photos/<int:pid>/move', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_move_photo(pid):
    data = request.get_json()
    album_id = data.get('album_id')  # null = remove from album
    conn = get_db()
    conn.execute('UPDATE job_photos SET album_id = ? WHERE id = ?', (album_id, pid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Material Shipments by Phase ──────────────────────────────────

@app.route('/material-shipments')
@role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def material_shipments_page():
    return render_template('material_shipments/list.html')

@app.route('/api/material-shipments')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def api_list_shipments():
    conn = get_db()
    shipments = conn.execute('''
        SELECT ms.*, j.name as job_name, u.display_name as created_by_name
        FROM material_shipments ms
        LEFT JOIN jobs j ON ms.job_id = j.id
        LEFT JOIN users u ON ms.created_by = u.id
        ORDER BY ms.created_at DESC
    ''').fetchall()
    result = []
    for s in shipments:
        d = dict(s)
        items = conn.execute('SELECT * FROM material_shipment_items WHERE shipment_id = ?', (s['id'],)).fetchall()
        d['items'] = [dict(i) for i in items]
        result.append(d)
    conn.close()
    return jsonify(result)

@app.route('/api/material-shipments', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_create_shipment():
    data = request.get_json() or {}
    conn = get_db()
    cur = conn.execute(
        '''INSERT INTO material_shipments (job_id, phase, shipment_date, status, notes, created_by)
           VALUES (?,?,?,?,?,?)''',
        (data['job_id'], data.get('phase','Rough-In'), data.get('shipment_date',''),
         data.get('status','Draft'), data.get('notes',''), session.get('user_id'))
    )
    shipment_id = cur.lastrowid
    for item in data.get('items', []):
        conn.execute(
            '''INSERT INTO material_shipment_items (shipment_id, line_item_id, sku, description, quantity, quantity_loaded, notes)
               VALUES (?,?,?,?,?,?,?)''',
            (shipment_id, item.get('line_item_id'), item.get('sku',''), item.get('description',''),
             float(item.get('quantity', 0) or 0), float(item.get('quantity_loaded', 0) or 0), item.get('notes',''))
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': shipment_id}), 201

@app.route('/api/material-shipments/<int:sid>', methods=['GET'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def api_get_shipment(sid):
    conn = get_db()
    s = conn.execute('SELECT ms.*, j.name as job_name FROM material_shipments ms LEFT JOIN jobs j ON ms.job_id = j.id WHERE ms.id = ?', (sid,)).fetchone()
    if not s:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    d = dict(s)
    items = conn.execute('SELECT * FROM material_shipment_items WHERE shipment_id = ?', (sid,)).fetchall()
    d['items'] = [dict(i) for i in items]
    conn.close()
    return jsonify(d)

@app.route('/api/material-shipments/<int:sid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_update_shipment(sid):
    data = request.get_json() or {}
    conn = get_db()
    conn.execute(
        '''UPDATE material_shipments SET phase=?, shipment_date=?, status=?, notes=?,
           shipped_by=?, updated_at=datetime('now','localtime') WHERE id=?''',
        (data.get('phase','Rough-In'), data.get('shipment_date',''), data.get('status','Draft'),
         data.get('notes',''), data.get('shipped_by'), sid)
    )
    if 'items' in data:
        conn.execute('DELETE FROM material_shipment_items WHERE shipment_id = ?', (sid,))
        for item in data['items']:
            conn.execute(
                '''INSERT INTO material_shipment_items (shipment_id, line_item_id, sku, description, quantity, quantity_loaded, notes)
                   VALUES (?,?,?,?,?,?,?)''',
                (sid, item.get('line_item_id'), item.get('sku',''), item.get('description',''),
                 float(item.get('quantity', 0) or 0), float(item.get('quantity_loaded', 0) or 0), item.get('notes',''))
            )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/material-shipments/<int:sid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_shipment(sid):
    conn = get_db()
    conn.execute('DELETE FROM material_shipments WHERE id = ?', (sid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/jobs/<int:job_id>/available-materials')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse', 'supplier')
def api_available_materials(job_id):
    conn = get_db()
    items = conn.execute('''
        SELECT li.id, li.sku, li.description, li.qty_ordered,
               COALESCE(recv.received, 0) as qty_received,
               COALESCE(ship.shipped, 0) as qty_shipped
        FROM line_items li
        LEFT JOIN (SELECT line_item_id, SUM(quantity) as received FROM received_entries GROUP BY line_item_id) recv ON recv.line_item_id = li.id
        LEFT JOIN (SELECT line_item_id, SUM(quantity_loaded) as shipped FROM material_shipment_items GROUP BY line_item_id) ship ON ship.line_item_id = li.id
        WHERE li.job_id = ? AND li.qty_ordered > 0
        ORDER BY li.line_number
    ''', (job_id,)).fetchall()
    conn.close()
    result = []
    for i in items:
        d = dict(i)
        d['qty_available'] = max(0, (d['qty_received'] or 0) - (d['qty_shipped'] or 0))
        result.append(d)
    return jsonify(result)

# ─── Billing Schedules ───────────────────────────────────────────

@app.route('/api/jobs/<int:job_id>/pay-apps')
@api_role_required('owner', 'admin', 'project_manager')
def api_job_pay_apps(job_id):
    """Get all pay apps for a job with billing totals."""
    conn = get_db()
    contracts = conn.execute(
        'SELECT id, contract_name, original_contract_sum, retainage_work_pct FROM pay_app_contracts WHERE job_id = ?',
        (job_id,)
    ).fetchall()
    result = []
    total_contract = 0
    total_billed = 0
    total_retainage = 0
    total_this_period = 0
    for c in contracts:
        cid = c['id']
        apps = conn.execute(
            'SELECT id, app_number, period_from, period_to, status FROM pay_applications WHERE contract_id = ? ORDER BY app_number',
            (cid,)
        ).fetchall()
        for app in apps:
            lines = conn.execute(
                'SELECT COALESCE(SUM(work_this_period + materials_stored), 0) as this_period, COALESCE(SUM(total_completed), 0) as total_billed FROM pay_app_line_entries WHERE pay_app_id = ?',
                (app['id'],)
            ).fetchone()
            this_period = lines['this_period'] or 0
            billed = lines['total_billed'] or 0
            ret_pct = c['retainage_work_pct'] or 0
            retainage = round(billed * ret_pct / 100, 2)
            total_this_period += this_period
            total_billed += billed
            total_retainage += retainage
            result.append({
                'id': app['id'], 'contract_id': cid,
                'app_number': app['app_number'],
                'period_from': app['period_from'], 'period_to': app['period_to'],
                'contract_name': c['contract_name'],
                'contract_sum': c['original_contract_sum'],
                'this_period': round(this_period, 2),
                'total_billed': round(billed, 2),
                'retainage': retainage,
                'status': app['status']
            })
        total_contract += c['original_contract_sum'] or 0
    conn.close()
    return jsonify({
        'apps': result,
        'totals': {
            'contract': round(total_contract, 2),
            'billed': round(total_billed, 2),
            'retainage': round(total_retainage, 2),
            'this_period': round(total_this_period, 2),
        }
    })

@app.route('/api/jobs/<int:job_id>/billing-schedule')
@api_role_required('owner', 'admin', 'project_manager')
def api_list_billing_schedule(job_id):
    conn = get_db()
    items = conn.execute('SELECT * FROM billing_schedules WHERE job_id = ? ORDER BY billing_number', (job_id,)).fetchall()
    conn.close()
    return jsonify([dict(i) for i in items])

@app.route('/api/jobs/<int:job_id>/billing-schedule', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_create_billing_milestone(job_id):
    data = request.get_json() or {}
    conn = get_db()
    next_num = conn.execute('SELECT COALESCE(MAX(billing_number), 0) + 1 FROM billing_schedules WHERE job_id = ?', (job_id,)).fetchone()[0]
    conn.execute(
        '''INSERT INTO billing_schedules (job_id, billing_number, description, scheduled_date, amount, status, required_docs, notes, created_by)
           VALUES (?,?,?,?,?,?,?,?,?)''',
        (job_id, next_num, data.get('description',''), data.get('scheduled_date',''),
         float(data.get('amount', 0) or 0), data.get('status','Pending'),
         json.dumps(data.get('required_docs', [])), data.get('notes',''), session.get('user_id'))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/billing-schedule/<int:bid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_billing_milestone(bid):
    data = request.get_json() or {}
    conn = get_db()
    conn.execute(
        '''UPDATE billing_schedules SET description=?, scheduled_date=?, amount=?, status=?,
           required_docs=?, notes=?, updated_at=datetime('now','localtime') WHERE id=?''',
        (data.get('description',''), data.get('scheduled_date',''),
         float(data.get('amount', 0) or 0), data.get('status','Pending'),
         json.dumps(data.get('required_docs', [])), data.get('notes',''), bid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/billing-schedule/<int:bid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_billing_milestone(bid):
    conn = get_db()
    conn.execute('DELETE FROM billing_schedules WHERE id = ?', (bid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/billing-schedule/upcoming')
@api_role_required('owner', 'admin', 'project_manager')
def api_upcoming_billing():
    conn = get_db()
    items = conn.execute('''
        SELECT bs.*, j.name as job_name FROM billing_schedules bs
        LEFT JOIN jobs j ON bs.job_id = j.id
        WHERE bs.status IN ('Pending','Ready')
        ORDER BY bs.scheduled_date
    ''').fetchall()
    conn.close()
    return jsonify([dict(i) for i in items])

# ─── Project Documents (General File Upload) ────────────────────────

PROJECT_DOCS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'project_docs')
os.makedirs(PROJECT_DOCS_DIR, exist_ok=True)

@app.route('/api/projects/<int:job_id>/documents')
@api_role_required('owner', 'admin', 'project_manager')
def api_list_project_documents(job_id):
    conn = get_db()
    docs = conn.execute('''
        SELECT pd.*, u.display_name as uploader_name
        FROM project_documents pd
        LEFT JOIN users u ON pd.uploaded_by = u.id
        WHERE pd.job_id = ?
        ORDER BY pd.created_at DESC
    ''', (job_id,)).fetchall()
    conn.close()
    return jsonify([dict(d) for d in docs])

@app.route('/api/projects/<int:job_id>/documents', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_upload_project_document(job_id):
    from werkzeug.utils import secure_filename
    import mimetypes
    files = request.files.getlist('files')
    if not files or not files[0].filename:
        return jsonify({'error': 'No file provided'}), 400
    category = request.form.get('category', 'Other')
    notes = request.form.get('notes', '')
    conn = get_db()
    uploaded = []
    for file in files:
        if not file.filename:
            continue
        fname = secure_filename(file.filename)
        ts_fname = f"{int(datetime.now().timestamp())}_{fname}"
        file.save(os.path.join(PROJECT_DOCS_DIR, ts_fname))
        file_size = os.path.getsize(os.path.join(PROJECT_DOCS_DIR, ts_fname))
        mime = mimetypes.guess_type(fname)[0] or 'application/octet-stream'
        conn.execute('''
            INSERT INTO project_documents (job_id, file_name, file_path, file_size, mime_type, category, notes, uploaded_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (job_id, fname, ts_fname, file_size, mime, category, notes, session.get('user_id')))
        uploaded.append(fname)
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'uploaded': uploaded})

@app.route('/api/projects/<int:job_id>/documents/<int:doc_id>/file')
@api_role_required('owner', 'admin', 'project_manager')
def api_view_project_document(job_id, doc_id):
    conn = get_db()
    doc = conn.execute('SELECT * FROM project_documents WHERE id = ? AND job_id = ?', (doc_id, job_id)).fetchone()
    conn.close()
    if not doc or not doc['file_path']:
        return 'Not found', 404
    fpath = os.path.join(PROJECT_DOCS_DIR, doc['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    return send_file(fpath, mimetype=doc['mime_type'], download_name=doc['file_name'])

@app.route('/api/projects/<int:job_id>/documents/<int:doc_id>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_project_document(job_id, doc_id):
    conn = get_db()
    doc = conn.execute('SELECT file_path FROM project_documents WHERE id = ? AND job_id = ?', (doc_id, job_id)).fetchone()
    if doc and doc['file_path']:
        fpath = os.path.join(PROJECT_DOCS_DIR, doc['file_path'])
        if os.path.exists(fpath):
            os.remove(fpath)
    conn.execute('DELETE FROM project_documents WHERE id = ? AND job_id = ?', (doc_id, job_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Certificates of Insurance (COI) ─────────────────────────────

@app.route('/coi')
@role_required('owner', 'admin', 'project_manager')
def coi_page():
    return render_template('coi/list.html')

@app.route('/api/coi')
@api_role_required('owner', 'admin', 'project_manager')
def api_list_coi():
    conn = get_db()
    cois = conn.execute('''
        SELECT c.*, j.name as job_name FROM certificates_of_insurance c
        LEFT JOIN jobs j ON c.job_id = j.id
        ORDER BY c.expiration_date
    ''').fetchall()
    conn.close()
    today = datetime.now().strftime('%Y-%m-%d')
    soon = (datetime.now() + timedelta(days=60)).strftime('%Y-%m-%d')
    result = []
    for c in cois:
        d = dict(c)
        if d['expiration_date'] and d['expiration_date'] < today:
            d['computed_status'] = 'Expired'
        elif d['expiration_date'] and d['expiration_date'] <= soon:
            d['computed_status'] = 'Expiring Soon'
        else:
            d['computed_status'] = d['status']
        result.append(d)
    return jsonify(result)

@app.route('/api/coi', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_create_coi():
    if request.content_type and 'multipart' in request.content_type:
        data = request.form.to_dict()
        f = request.files.get('file')
    else:
        data = request.get_json() or {}
        f = None
    file_path = ''
    if f:
        os.makedirs(os.path.join(app.root_path, 'data', 'coi'), exist_ok=True)
        filename = f'{datetime.now().strftime("%Y%m%d_%H%M%S")}_{f.filename}'
        filepath = os.path.join(app.root_path, 'data', 'coi', filename)
        f.save(filepath)
        file_path = f'data/coi/{filename}'
    conn = get_db()
    conn.execute(
        '''INSERT INTO certificates_of_insurance (job_id, policy_type, carrier, policy_number,
           effective_date, expiration_date, coverage_amount, certificate_holder, file_path, status, notes, created_by)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
        (data.get('job_id') or None, data.get('policy_type','General Liability'),
         data.get('carrier',''), data.get('policy_number',''),
         data.get('effective_date',''), data.get('expiration_date',''),
         float(data.get('coverage_amount', 0) or 0), data.get('certificate_holder',''),
         file_path, data.get('status','Active'), data.get('notes',''), session.get('user_id'))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/coi/<int:cid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_update_coi(cid):
    if request.content_type and 'multipart' in request.content_type:
        data = request.form.to_dict()
        f = request.files.get('file')
    else:
        data = request.get_json() or {}
        f = None
    conn = get_db()
    file_path = data.get('file_path', '')
    if f:
        os.makedirs(os.path.join(app.root_path, 'data', 'coi'), exist_ok=True)
        filename = f'{datetime.now().strftime("%Y%m%d_%H%M%S")}_{f.filename}'
        filepath = os.path.join(app.root_path, 'data', 'coi', filename)
        f.save(filepath)
        file_path = f'data/coi/{filename}'
    conn.execute(
        '''UPDATE certificates_of_insurance SET job_id=?, policy_type=?, carrier=?, policy_number=?,
           effective_date=?, expiration_date=?, coverage_amount=?, certificate_holder=?,
           file_path=?, status=?, notes=?, updated_at=datetime('now','localtime') WHERE id=?''',
        (data.get('job_id') or None, data.get('policy_type','General Liability'),
         data.get('carrier',''), data.get('policy_number',''),
         data.get('effective_date',''), data.get('expiration_date',''),
         float(data.get('coverage_amount', 0) or 0), data.get('certificate_holder',''),
         file_path, data.get('status','Active'), data.get('notes',''), cid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/coi/<int:cid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_coi(cid):
    conn = get_db()
    conn.execute('DELETE FROM certificates_of_insurance WHERE id = ?', (cid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/coi/<int:cid>/file')
@api_login_required
def api_coi_file(cid):
    conn = get_db()
    coi = conn.execute('SELECT file_path FROM certificates_of_insurance WHERE id = ?', (cid,)).fetchone()
    conn.close()
    if not coi or not coi['file_path']:
        return jsonify({'error': 'No file'}), 404
    return send_file(os.path.join(app.root_path, coi['file_path']))

# ─── Billing Summary ─────────────────────────────────────────────

@app.route('/billing-summary')
@role_required('owner', 'admin')
def billing_summary_page():
    return render_template('billing_summary.html')

@app.route('/api/billing-summary')
@api_role_required('owner', 'admin')
def api_billing_summary():
    """Comprehensive billing summary across all projects — pay apps, invoices, retainage."""
    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')

    # Get all jobs
    jobs = conn.execute("SELECT id, name, status FROM jobs ORDER BY name").fetchall()

    projects = []
    totals = {
        'total_contract': 0, 'total_billed': 0, 'total_paid': 0,
        'total_retained': 0, 'total_outstanding': 0, 'total_cos': 0
    }

    for job in jobs:
        job_id = job['id']

        # Pay app contracts for this job
        contracts = conn.execute(
            'SELECT * FROM pay_app_contracts WHERE job_id = ?', (job_id,)
        ).fetchall()

        if not contracts:
            continue

        for contract in contracts:
            cid = contract['id']
            original_sum = contract['original_contract_sum'] or 0

            # Get all pay apps for this contract
            pay_apps = conn.execute(
                '''SELECT pa.*, GROUP_CONCAT(
                       COALESCE(ple.work_this_period, 0) + COALESCE(ple.materials_stored, 0)
                   ) as line_totals
                   FROM pay_applications pa
                   LEFT JOIN pay_app_line_entries ple ON ple.pay_app_id = pa.id
                   WHERE pa.contract_id = ?
                   GROUP BY pa.id
                   ORDER BY pa.application_number''',
                (cid,)
            ).fetchall()

            # Calculate totals from SOV entries
            total_billed = 0
            total_paid = 0
            total_retained = 0
            co_net = 0

            app_details = []
            for pa in pay_apps:
                # Sum work+materials for this pay app
                entries = conn.execute(
                    '''SELECT SUM(COALESCE(work_this_period, 0) + COALESCE(materials_stored, 0)) as period_total
                       FROM pay_app_line_entries WHERE pay_app_id = ?''',
                    (pa['id'],)
                ).fetchone()
                period_billed = entries['period_total'] or 0

                ret_pct = contract['retainage_work_pct'] or 10
                period_retained = round(period_billed * ret_pct / 100, 2)
                period_paid_net = period_billed - period_retained if pa['status'] == 'Paid' else 0

                # If approved but not paid, it's outstanding
                if pa['status'] in ('Approved', 'Submitted'):
                    period_paid_net = 0

                total_billed += period_billed
                if pa['status'] == 'Paid':
                    total_paid += period_billed - period_retained
                    total_retained += period_retained

                co_net += (pa['co_additions'] or 0) - (pa['co_deductions'] or 0)

                app_details.append({
                    'id': pa['id'],
                    'number': pa['application_number'],
                    'period_to': pa['period_to'] or '',
                    'date': pa['application_date'] or '',
                    'billed': round(period_billed, 2),
                    'retained': round(period_retained, 2),
                    'paid': round(period_paid_net, 2),
                    'status': pa['status'],
                })

            current_contract = original_sum + co_net
            outstanding = total_billed - total_paid - total_retained

            proj = {
                'job_id': job_id,
                'job_name': job['name'],
                'job_status': job['status'],
                'contract_id': cid,
                'gc_name': contract['gc_name'] or '',
                'original_contract': original_sum,
                'co_net': round(co_net, 2),
                'current_contract': round(current_contract, 2),
                'total_billed': round(total_billed, 2),
                'total_paid': round(total_paid, 2),
                'total_retained': round(total_retained, 2),
                'outstanding': round(outstanding, 2),
                'pct_billed': round(total_billed / current_contract * 100, 1) if current_contract else 0,
                'pct_collected': round(total_paid / total_billed * 100, 1) if total_billed else 0,
                'pay_apps': app_details,
                'retainage_pct': contract['retainage_work_pct'] or 10,
            }
            projects.append(proj)

            totals['total_contract'] += current_contract
            totals['total_billed'] += total_billed
            totals['total_paid'] += total_paid
            totals['total_retained'] += total_retained
            totals['total_outstanding'] += outstanding
            totals['total_cos'] += co_net

    # Service invoices (non-pay-app billing)
    svc_invoices = conn.execute('''
        SELECT ci.*, j.name as job_name FROM client_invoices ci
        LEFT JOIN jobs j ON ci.job_id = j.id
        ORDER BY ci.issue_date DESC
    ''').fetchall()

    svc_total = sum(i['amount'] or 0 for i in svc_invoices)
    svc_paid = sum(i['amount'] or 0 for i in svc_invoices if i['status'] == 'Paid')
    svc_outstanding = svc_total - svc_paid

    # Aging buckets for outstanding
    aging = {'0_30': 0, '31_60': 0, '61_90': 0, '90_plus': 0}
    for p in projects:
        if p['outstanding'] > 0 and p['pay_apps']:
            last_app = p['pay_apps'][-1]
            app_date = last_app.get('date') or last_app.get('period_to') or ''
            if app_date:
                try:
                    days = (datetime.strptime(today, '%Y-%m-%d') - datetime.strptime(app_date, '%Y-%m-%d')).days
                except (ValueError, TypeError):
                    days = 0
                if days <= 30: aging['0_30'] += p['outstanding']
                elif days <= 60: aging['31_60'] += p['outstanding']
                elif days <= 90: aging['61_90'] += p['outstanding']
                else: aging['90_plus'] += p['outstanding']

    for inv in svc_invoices:
        if inv['status'] != 'Paid' and (inv['amount'] or 0) > 0:
            if inv['issue_date']:
                try:
                    days = (datetime.strptime(today, '%Y-%m-%d') - datetime.strptime(inv['issue_date'], '%Y-%m-%d')).days
                except (ValueError, TypeError):
                    days = 0
                amt = (inv['amount'] or 0) - (inv['amount_paid'] or 0)
                if days <= 30: aging['0_30'] += amt
                elif days <= 60: aging['31_60'] += amt
                elif days <= 90: aging['61_90'] += amt
                else: aging['90_plus'] += amt

    conn.close()

    return jsonify({
        'projects': projects,
        'totals': {k: round(v, 2) for k, v in totals.items()},
        'service_invoices': [dict(i) for i in svc_invoices],
        'service_totals': {
            'total': round(svc_total, 2),
            'paid': round(svc_paid, 2),
            'outstanding': round(svc_outstanding, 2),
        },
        'aging': {k: round(v, 2) for k, v in aging.items()},
    })

# ─── Payment Days-to-Receive & Aged Receivables ──────────────────

@app.route('/api/accounting/receivables')
@api_role_required('owner', 'admin')
def api_aged_receivables():
    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    invoices = conn.execute('''
        SELECT ci.*, j.name as job_name FROM client_invoices ci
        LEFT JOIN jobs j ON ci.job_id = j.id
        WHERE ci.status != 'Paid'
        ORDER BY ci.issue_date
    ''').fetchall()
    conn.close()
    buckets = {'0_30': [], '31_60': [], '61_90': [], '90_plus': []}
    for inv in invoices:
        d = dict(inv)
        if d['issue_date']:
            days = (datetime.strptime(today, '%Y-%m-%d') - datetime.strptime(d['issue_date'], '%Y-%m-%d')).days
        else:
            days = 0
        d['days_outstanding'] = days
        if days <= 30:
            buckets['0_30'].append(d)
        elif days <= 60:
            buckets['31_60'].append(d)
        elif days <= 90:
            buckets['61_90'].append(d)
        else:
            buckets['90_plus'].append(d)
    return jsonify({
        'buckets': {k: v for k, v in buckets.items()},
        'totals': {k: sum(i['amount'] for i in v) for k, v in buckets.items()}
    })

# ─── Supplier Delivery Schedules ─────────────────────────────────

@app.route('/api/delivery-schedules')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_list_delivery_schedules():
    job_id = request.args.get('job_id', type=int)
    conn = get_db()
    query = 'SELECT ds.*, j.name as job_name FROM delivery_schedules ds LEFT JOIN jobs j ON ds.job_id = j.id WHERE 1=1'
    params = []
    if job_id:
        query += ' AND ds.job_id = ?'
        params.append(job_id)
    query += ' ORDER BY ds.expected_date'
    items = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(i) for i in items])

@app.route('/api/delivery-schedules', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_create_delivery_schedule():
    data = request.get_json() or {}
    conn = get_db()
    conn.execute(
        '''INSERT INTO delivery_schedules (job_id, supplier_name, expected_date, status,
           items_summary, tracking_number, notes, created_by) VALUES (?,?,?,?,?,?,?,?)''',
        (data['job_id'], data.get('supplier_name',''), data.get('expected_date',''),
         data.get('status','Scheduled'), data.get('items_summary',''),
         data.get('tracking_number',''), data.get('notes',''), session.get('user_id'))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/delivery-schedules/<int:did>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_update_delivery_schedule(did):
    data = request.get_json() or {}
    conn = get_db()
    conn.execute(
        '''UPDATE delivery_schedules SET supplier_name=?, expected_date=?, actual_date=?, status=?,
           items_summary=?, tracking_number=?, notes=?, updated_at=datetime('now','localtime') WHERE id=?''',
        (data.get('supplier_name',''), data.get('expected_date',''), data.get('actual_date',''),
         data.get('status','Scheduled'), data.get('items_summary',''),
         data.get('tracking_number',''), data.get('notes',''), did)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/delivery-schedules/<int:did>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_delete_delivery_schedule(did):
    conn = get_db()
    conn.execute('DELETE FROM delivery_schedules WHERE id = ?', (did,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/delivery-schedules/upcoming')
@api_role_required('owner', 'admin', 'project_manager', 'warehouse')
def api_upcoming_deliveries():
    conn = get_db()
    items = conn.execute('''
        SELECT ds.*, j.name as job_name FROM delivery_schedules ds
        LEFT JOIN jobs j ON ds.job_id = j.id
        WHERE ds.status NOT IN ('Delivered')
        ORDER BY ds.expected_date
    ''').fetchall()
    conn.close()
    return jsonify([dict(i) for i in items])

# ─── Lien Waiver Auto-Prompt (on payment) ─────────────────────────
# (Integrated into create_payment above — see modification below)

# ─── Job Import Tools ─────────────────────────────────────────────

@app.route('/jobs/import')
@role_required('owner', 'admin')
def job_import_page():
    return render_template('jobs/import.html')

@app.route('/api/jobs/import-excel', methods=['POST'])
@api_role_required('owner', 'admin')
def api_import_jobs_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    f = request.files['file']
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip().lower() for c in ws[1]]
        conn = get_db()
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, h in enumerate(headers):
                if i < len(row):
                    row_dict[h] = row[i]
            name = row_dict.get('name') or row_dict.get('job_name') or row_dict.get('project')
            if not name:
                continue
            status = row_dict.get('status', 'In Progress')
            cur = conn.execute(
                '''INSERT INTO jobs (name, status, address, city, state, zip_code)
                   VALUES (?,?,?,?,?,?)''',
                (str(name), str(status),
                 str(row_dict.get('address', '') or ''),
                 str(row_dict.get('city', '') or ''),
                 str(row_dict.get('state', '') or ''),
                 str(row_dict.get('zip', '') or row_dict.get('zip_code', '') or ''))
            )
            job_id = cur.lastrowid
            # Link customer if provided
            customer_name = row_dict.get('customer') or row_dict.get('customer_name')
            if customer_name:
                cust = conn.execute('SELECT id FROM customers WHERE name = ?', (str(customer_name),)).fetchone()
                if cust:
                    conn.execute('UPDATE jobs SET customer_id = ? WHERE id = ?', (cust['id'], job_id))
            # Seed pipeline
            seed_pipeline_steps(conn, job_id)
            # Auto-complete pipeline steps based on status
            if status in ('Awarded', 'In Progress', 'Complete'):
                for s in range(1, 9):
                    conn.execute("UPDATE job_pipeline_steps SET status='complete' WHERE job_id=? AND step_number=?", (job_id, s))
            if status in ('In Progress', 'Complete'):
                for s in range(9, 14):
                    conn.execute("UPDATE job_pipeline_steps SET status='complete' WHERE job_id=? AND step_number=?", (job_id, s))
            if status == 'Complete':
                conn.execute("UPDATE job_pipeline_steps SET status='complete' WHERE job_id=?", (job_id,))
            imported += 1
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'imported': imported})
    except ImportError:
        return jsonify({'error': 'openpyxl not installed. Run: pip install openpyxl'}), 500

@app.route('/api/jobs/quick-add', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_quick_add_job():
    data = request.get_json() or {}
    name = data.get('name')
    if not name:
        return jsonify({'error': 'Job name required'}), 400
    conn = get_db()
    cur = conn.execute(
        '''INSERT INTO jobs (name, status, address, city, state, zip_code)
           VALUES (?,?,?,?,?,?)''',
        (name, data.get('status', 'In Progress'),
         data.get('address',''), data.get('city',''), data.get('state',''), data.get('zip_code',''))
    )
    job_id = cur.lastrowid
    if data.get('customer_id'):
        conn.execute('UPDATE jobs SET customer_id = ? WHERE id = ?', (data['customer_id'], job_id))
    seed_pipeline_steps(conn, job_id)
    # Mark completed steps
    for step_num in data.get('completed_steps', []):
        conn.execute(
            "UPDATE job_pipeline_steps SET status='complete', completed_date=date('now','localtime') WHERE job_id=? AND step_number=?",
            (job_id, int(step_num))
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'job_id': job_id}), 201


# ─── Team Pay (Internal Progress Payroll) ─────────────────────────

@app.route('/team-pay')
@role_required('owner')
def team_pay_list_page():
    return render_template('team_pay/list.html')

@app.route('/team-pay/job/<int:schedule_id>')
@role_required('owner')
def team_pay_job_page(schedule_id):
    return render_template('team_pay/job.html', schedule_id=schedule_id)

@app.route('/team-pay/period/<int:period_id>')
@role_required('owner')
def team_pay_period_page(period_id):
    return render_template('team_pay/period.html', period_id=period_id)

@app.route('/api/team-pay/schedules', methods=['GET'])
@api_role_required('owner')
def api_team_pay_schedules_list():
    conn = get_db()
    rows = conn.execute('''
        SELECT s.*, j.name as job_name,
            (SELECT COUNT(*) FROM team_pay_members WHERE schedule_id = s.id) as member_count,
            (SELECT COALESCE(SUM(scheduled_amount), 0) FROM team_pay_members WHERE schedule_id = s.id) as scheduled_total,
            (SELECT COUNT(*) FROM team_pay_periods WHERE schedule_id = s.id) as period_count,
            (SELECT COALESCE(SUM(e.amount), 0) FROM team_pay_entries e
             JOIN team_pay_periods p ON e.period_id = p.id
             WHERE p.schedule_id = s.id) as total_paid
        FROM team_pay_schedules s
        JOIN jobs j ON s.job_id = j.id
        ORDER BY s.created_at DESC
    ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/team-pay/schedules', methods=['POST'])
@api_role_required('owner')
def api_team_pay_schedules_create():
    data = request.get_json()
    job_id = data.get('job_id')
    if not job_id:
        return jsonify({'error': 'Job is required'}), 400
    conn = get_db()
    existing = conn.execute('SELECT id FROM team_pay_schedules WHERE job_id = ?', (job_id,)).fetchone()
    if existing:
        conn.close()
        return jsonify({'error': 'A team pay schedule already exists for this job'}), 409
    cursor = conn.execute(
        'INSERT INTO team_pay_schedules (job_id, total_job_value, notes, created_by) VALUES (?,?,?,?)',
        (job_id, float(data.get('total_job_value', 0)), data.get('notes', ''), session.get('user_id'))
    )
    conn.commit()
    schedule_id = cursor.lastrowid
    conn.close()
    return jsonify({'ok': True, 'id': schedule_id}), 201

@app.route('/api/team-pay/schedules/<int:sid>', methods=['GET'])
@api_role_required('owner')
def api_team_pay_schedule_detail(sid):
    conn = get_db()
    row = conn.execute('''
        SELECT s.*, j.name as job_name
        FROM team_pay_schedules s
        JOIN jobs j ON s.job_id = j.id
        WHERE s.id = ?
    ''', (sid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(dict(row))

@app.route('/api/team-pay/schedules/<int:sid>', methods=['PUT'])
@api_role_required('owner')
def api_team_pay_schedule_update(sid):
    data = request.get_json()
    conn = get_db()
    conn.execute(
        "UPDATE team_pay_schedules SET total_job_value = ?, notes = ?, updated_at = datetime('now','localtime') WHERE id = ?",
        (float(data.get('total_job_value', 0)), data.get('notes', ''), sid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/team-pay/schedules/<int:sid>', methods=['DELETE'])
@api_role_required('owner')
def api_team_pay_schedule_delete(sid):
    conn = get_db()
    conn.execute('DELETE FROM team_pay_schedules WHERE id = ?', (sid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# --- Team Pay Members ---

@app.route('/api/team-pay/schedules/<int:sid>/members', methods=['GET'])
@api_role_required('owner')
def api_team_pay_members_list(sid):
    conn = get_db()
    rows = conn.execute('''
        SELECT m.*, u.display_name, u.username,
            (SELECT COALESCE(SUM(e.amount), 0) FROM team_pay_entries e
             JOIN team_pay_periods p ON e.period_id = p.id
             WHERE e.member_id = m.id) as total_paid
        FROM team_pay_members m
        JOIN users u ON m.user_id = u.id
        WHERE m.schedule_id = ?
        ORDER BY m.sort_order, m.id
    ''', (sid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/team-pay/schedules/<int:sid>/members', methods=['POST'])
@api_role_required('owner')
def api_team_pay_member_add(sid):
    data = request.get_json()
    user_id = data.get('user_id')
    if not user_id:
        return jsonify({'error': 'User is required'}), 400
    conn = get_db()
    existing = conn.execute('SELECT id FROM team_pay_members WHERE schedule_id = ? AND user_id = ?', (sid, user_id)).fetchone()
    if existing:
        conn.close()
        return jsonify({'error': 'Member already in schedule'}), 409
    max_order = conn.execute('SELECT COALESCE(MAX(sort_order), 0) FROM team_pay_members WHERE schedule_id = ?', (sid,)).fetchone()[0]
    conn.execute(
        'INSERT INTO team_pay_members (schedule_id, user_id, scheduled_amount, sort_order) VALUES (?,?,?,?)',
        (sid, user_id, float(data.get('scheduled_amount', 0)), max_order + 1)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/team-pay/members/<int:mid>', methods=['PUT'])
@api_role_required('owner')
def api_team_pay_member_update(mid):
    data = request.get_json()
    conn = get_db()
    fields, values = [], []
    if 'scheduled_amount' in data:
        fields.append('scheduled_amount = ?')
        values.append(float(data['scheduled_amount']))
    if 'sort_order' in data:
        fields.append('sort_order = ?')
        values.append(int(data['sort_order']))
    if fields:
        values.append(mid)
        conn.execute(f"UPDATE team_pay_members SET {', '.join(fields)} WHERE id = ?", values)
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/team-pay/members/<int:mid>', methods=['DELETE'])
@api_role_required('owner')
def api_team_pay_member_delete(mid):
    conn = get_db()
    conn.execute('DELETE FROM team_pay_members WHERE id = ?', (mid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# --- Team Pay Periods ---

@app.route('/api/team-pay/schedules/<int:sid>/periods', methods=['GET'])
@api_role_required('owner')
def api_team_pay_periods_list(sid):
    conn = get_db()
    rows = conn.execute('''
        SELECT p.*,
            (SELECT COALESCE(SUM(e.amount), 0) FROM team_pay_entries e WHERE e.period_id = p.id) as distributed_total
        FROM team_pay_periods p
        WHERE p.schedule_id = ?
        ORDER BY p.period_number
    ''', (sid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/team-pay/schedules/<int:sid>/periods', methods=['POST'])
@api_role_required('owner')
def api_team_pay_period_create(sid):
    conn = get_db()
    max_num = conn.execute('SELECT COALESCE(MAX(period_number), 0) FROM team_pay_periods WHERE schedule_id = ?', (sid,)).fetchone()[0]
    cursor = conn.execute(
        'INSERT INTO team_pay_periods (schedule_id, period_number, created_by) VALUES (?,?,?)',
        (sid, max_num + 1, session.get('user_id'))
    )
    period_id = cursor.lastrowid
    # Pre-create entry rows for all members
    members = conn.execute('SELECT id FROM team_pay_members WHERE schedule_id = ?', (sid,)).fetchall()
    for m in members:
        conn.execute('INSERT INTO team_pay_entries (period_id, member_id, amount) VALUES (?,?,0)', (period_id, m['id']))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': period_id}), 201

@app.route('/api/team-pay/periods/<int:pid>', methods=['GET'])
@api_role_required('owner')
def api_team_pay_period_detail(pid):
    """Main G703 endpoint — returns members with calculated previous/total/balance."""
    conn = get_db()
    period = conn.execute('SELECT * FROM team_pay_periods WHERE id = ?', (pid,)).fetchone()
    if not period:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    schedule_id = period['schedule_id']
    period_number = period['period_number']
    schedule = conn.execute('''
        SELECT s.*, j.name as job_name
        FROM team_pay_schedules s JOIN jobs j ON s.job_id = j.id
        WHERE s.id = ?
    ''', (schedule_id,)).fetchone()
    # Get all members
    members = conn.execute('''
        SELECT m.*, u.display_name, u.username
        FROM team_pay_members m JOIN users u ON m.user_id = u.id
        WHERE m.schedule_id = ?
        ORDER BY m.sort_order, m.id
    ''', (schedule_id,)).fetchall()
    # Get current period entries
    entries = conn.execute('SELECT * FROM team_pay_entries WHERE period_id = ?', (pid,)).fetchall()
    entry_map = {e['member_id']: e['amount'] for e in entries}
    # Get prior period IDs
    prior_periods = conn.execute(
        'SELECT id FROM team_pay_periods WHERE schedule_id = ? AND period_number < ?',
        (schedule_id, period_number)
    ).fetchall()
    prior_ids = [p['id'] for p in prior_periods]
    # Calculate previous payments per member
    prev_map = {}
    if prior_ids:
        placeholders = ','.join('?' * len(prior_ids))
        prev_rows = conn.execute(
            f'SELECT member_id, COALESCE(SUM(amount), 0) as prev FROM team_pay_entries WHERE period_id IN ({placeholders}) GROUP BY member_id',
            prior_ids
        ).fetchall()
        prev_map = {r['member_id']: r['prev'] for r in prev_rows}
    conn.close()
    result = dict(period)
    result['schedule_id'] = schedule_id
    result['job_name'] = schedule['job_name'] if schedule else ''
    result['total_job_value'] = schedule['total_job_value'] if schedule else 0
    result['members'] = []
    for m in members:
        prev = prev_map.get(m['id'], 0)
        this_amt = entry_map.get(m['id'], 0)
        result['members'].append({
            'member_id': m['id'],
            'user_id': m['user_id'],
            'display_name': m['display_name'],
            'username': m['username'],
            'scheduled_amount': m['scheduled_amount'],
            'previous_payments': prev,
            'this_period': this_amt,
        })
    return jsonify(result)

@app.route('/api/team-pay/periods/<int:pid>', methods=['PUT'])
@api_role_required('owner')
def api_team_pay_period_update(pid):
    data = request.get_json()
    conn = get_db()
    period = conn.execute('SELECT * FROM team_pay_periods WHERE id = ?', (pid,)).fetchone()
    if not period:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if period['status'] == 'Finalized' and data.get('status') != 'Draft':
        conn.close()
        return jsonify({'error': 'Period is finalized'}), 400
    # Update period fields
    update_fields = []
    update_vals = []
    if 'notes' in data:
        update_fields.append('notes = ?')
        update_vals.append(data['notes'])
    if 'source_amount' in data:
        update_fields.append('source_amount = ?')
        update_vals.append(float(data['source_amount']))
    if 'payment_date' in data:
        update_fields.append('payment_date = ?')
        update_vals.append(data['payment_date'])
    if 'status' in data:
        update_fields.append('status = ?')
        update_vals.append(data['status'])
    if update_fields:
        update_vals.append(pid)
        conn.execute(f"UPDATE team_pay_periods SET {', '.join(update_fields)} WHERE id = ?", update_vals)
    # Save entries
    entries = data.get('entries', [])
    for entry in entries:
        conn.execute('''
            INSERT INTO team_pay_entries (period_id, member_id, amount)
            VALUES (?, ?, ?)
            ON CONFLICT(period_id, member_id) DO UPDATE SET amount = excluded.amount
        ''', (pid, entry['member_id'], float(entry.get('amount', 0))))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/team-pay/periods/<int:pid>', methods=['DELETE'])
@api_role_required('owner')
def api_team_pay_period_delete(pid):
    conn = get_db()
    period = conn.execute('SELECT status FROM team_pay_periods WHERE id = ?', (pid,)).fetchone()
    if period and period['status'] == 'Finalized':
        conn.close()
        return jsonify({'error': 'Cannot delete a finalized period'}), 400
    conn.execute('DELETE FROM team_pay_periods WHERE id = ?', (pid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/team-pay/dashboard')
@api_role_required('owner')
def api_team_pay_dashboard():
    """Cross-job summary per member."""
    conn = get_db()
    rows = conn.execute('''
        SELECT u.id as user_id, u.display_name, u.username,
            COALESCE(SUM(m.scheduled_amount), 0) as total_scheduled,
            COALESCE((SELECT SUM(e.amount) FROM team_pay_entries e WHERE e.member_id IN
                (SELECT id FROM team_pay_members WHERE user_id = u.id)), 0) as total_paid
        FROM users u
        JOIN team_pay_members m ON m.user_id = u.id
        GROUP BY u.id
        ORDER BY u.display_name
    ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ─── Team Chat ──────────────────────────────────────────────────

TEAM_CHAT_UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'team_chat_files')

@app.route('/team-chat')
@login_required
def team_chat_page():
    return render_template('team_chat/chat.html')

@app.route('/api/team-chat/users')
@api_login_required
def api_tc_users():
    conn = get_db()
    users = conn.execute('SELECT id, username, display_name, role FROM users WHERE is_active = 1 ORDER BY display_name').fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])

@app.route('/api/team-chat/channels')
@api_login_required
def api_tc_channels():
    uid = session['user_id']
    conn = get_db()
    channels = conn.execute('''
        SELECT c.*, (
            SELECT COUNT(*) FROM tc_messages m
            WHERE m.channel_id = c.id AND m.id > COALESCE(
                (SELECT rs.last_read_message_id FROM tc_read_status rs
                 WHERE rs.user_id = ? AND rs.channel_id = c.id), 0)
        ) as unread_count
        FROM tc_channels c
        JOIN tc_channel_members cm ON cm.channel_id = c.id AND cm.user_id = ?
        ORDER BY c.name
    ''', (uid, uid)).fetchall()
    conn.close()
    return jsonify([dict(c) for c in channels])

@app.route('/api/team-chat/channels', methods=['POST'])
@api_role_required('owner', 'admin')
def api_tc_create_channel():
    data = request.get_json()
    name = (data.get('name') or '').strip().lower().replace(' ', '-')
    desc = (data.get('description') or '').strip()
    if not name:
        return jsonify({'error': 'Channel name required'}), 400

    conn = get_db()
    existing = conn.execute('SELECT id FROM tc_channels WHERE name = ?', (name,)).fetchone()
    if existing:
        conn.close()
        return jsonify({'error': 'Channel name already exists'}), 400

    cur = conn.execute(
        'INSERT INTO tc_channels (name, description, is_default, created_by) VALUES (?, ?, 0, ?)',
        (name, desc, session['user_id'])
    )
    channel_id = cur.lastrowid

    # Auto-enroll all active users
    users = conn.execute('SELECT id FROM users WHERE is_active = 1').fetchall()
    for u in users:
        conn.execute('INSERT OR IGNORE INTO tc_channel_members (channel_id, user_id) VALUES (?, ?)',
                     (channel_id, u['id']))

    conn.commit()
    conn.close()
    return jsonify({'id': channel_id, 'name': name})

@app.route('/api/team-chat/channels/<int:cid>', methods=['DELETE'])
@api_role_required('owner', 'admin')
def api_tc_delete_channel(cid):
    conn = get_db()
    ch = conn.execute('SELECT * FROM tc_channels WHERE id = ?', (cid,)).fetchone()
    if not ch:
        conn.close()
        return jsonify({'error': 'Channel not found'}), 404
    if ch['is_default']:
        conn.close()
        return jsonify({'error': 'Cannot delete default channel'}), 400
    conn.execute('DELETE FROM tc_channels WHERE id = ?', (cid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/team-chat/channels/<int:cid>/messages')
@api_login_required
def api_tc_channel_messages(cid):
    uid = session['user_id']
    before_id = request.args.get('before_id', type=int)
    limit = request.args.get('limit', 50, type=int)

    conn = get_db()
    # Verify membership
    member = conn.execute('SELECT 1 FROM tc_channel_members WHERE channel_id = ? AND user_id = ?', (cid, uid)).fetchone()
    if not member:
        conn.close()
        return jsonify({'error': 'Not a member of this channel'}), 403

    if before_id:
        msgs = conn.execute('''
            SELECT m.*, u.display_name as sender_name FROM tc_messages m
            LEFT JOIN users u ON u.id = m.sender_id
            WHERE m.channel_id = ? AND m.id < ?
            ORDER BY m.id DESC LIMIT ?
        ''', (cid, before_id, limit)).fetchall()
    else:
        msgs = conn.execute('''
            SELECT m.*, u.display_name as sender_name FROM tc_messages m
            LEFT JOIN users u ON u.id = m.sender_id
            WHERE m.channel_id = ?
            ORDER BY m.id DESC LIMIT ?
        ''', (cid, limit)).fetchall()

    conn.close()
    result = [dict(m) for m in msgs]
    result.reverse()
    return jsonify(result)

@app.route('/api/team-chat/channels/<int:cid>/messages', methods=['POST'])
@api_login_required
def api_tc_send_channel_message(cid):
    uid = session['user_id']
    content = (request.form.get('content') or '').strip()
    file = request.files.get('file')

    if not content and not file:
        return jsonify({'error': 'Message content or file required'}), 400

    conn = get_db()
    member = conn.execute('SELECT 1 FROM tc_channel_members WHERE channel_id = ? AND user_id = ?', (cid, uid)).fetchone()
    if not member:
        conn.close()
        return jsonify({'error': 'Not a member'}), 403

    file_path = file_name = file_type = ''
    if file and file.filename:
        os.makedirs(TEAM_CHAT_UPLOAD_DIR, exist_ok=True)
        import uuid
        ext = os.path.splitext(file.filename)[1]
        saved_name = f"{uuid.uuid4().hex}{ext}"
        save_path = os.path.join(TEAM_CHAT_UPLOAD_DIR, saved_name)
        file.save(save_path)
        file_path = saved_name
        file_name = file.filename
        file_type = file.content_type or ''

    cur = conn.execute(
        '''INSERT INTO tc_messages (channel_id, sender_id, content, file_path, file_name, file_type)
           VALUES (?, ?, ?, ?, ?, ?)''',
        (cid, uid, content, file_path, file_name, file_type)
    )
    msg_id = cur.lastrowid

    # Update sender's read status
    conn.execute('''INSERT INTO tc_read_status (user_id, channel_id, last_read_message_id, updated_at)
        VALUES (?, ?, ?, datetime('now','localtime'))
        ON CONFLICT(user_id, channel_id) DO UPDATE SET last_read_message_id = ?, updated_at = datetime('now','localtime')''',
        (uid, cid, msg_id, msg_id))

    # Collect notification data before closing conn
    sender_name = session.get('display_name') or session.get('username', 'Someone')
    ch = conn.execute('SELECT name FROM tc_channels WHERE id = ?', (cid,)).fetchone()
    ch_name = ch['name'] if ch else 'channel'
    member_ids = [m['user_id'] for m in conn.execute(
        'SELECT user_id FROM tc_channel_members WHERE channel_id = ? AND user_id != ?', (cid, uid)).fetchall()]
    preview = (content[:80] + '...') if len(content) > 80 else content

    conn.commit()
    conn.close()

    # Notify other channel members (create_notification opens its own conn)
    for mid in member_ids:
        create_notification(mid, 'team_chat',
            f'{sender_name} in #{ch_name}',
            preview or '(file attachment)',
            f'/team-chat?channel={cid}')

    return jsonify({'id': msg_id})

@app.route('/api/team-chat/channels/<int:cid>/read', methods=['POST'])
@api_login_required
def api_tc_channel_read(cid):
    uid = session['user_id']
    data = request.get_json() or {}
    last_id = data.get('last_message_id', 0)
    conn = get_db()
    conn.execute('''INSERT INTO tc_read_status (user_id, channel_id, last_read_message_id, updated_at)
        VALUES (?, ?, ?, datetime('now','localtime'))
        ON CONFLICT(user_id, channel_id) DO UPDATE SET
            last_read_message_id = MAX(last_read_message_id, ?),
            updated_at = datetime('now','localtime')''',
        (uid, cid, last_id, last_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Team Chat DMs ──────────────────────────────────────────

@app.route('/api/team-chat/dm/conversations')
@api_login_required
def api_tc_dm_conversations():
    uid = session['user_id']
    conn = get_db()
    # Get unique DM peers
    rows = conn.execute('''
        SELECT peer_id, u.display_name, u.username,
            (SELECT MAX(m2.id) FROM tc_messages m2
             WHERE m2.channel_id IS NULL
             AND ((m2.sender_id = ? AND m2.dm_recipient_id = peer_id)
               OR (m2.sender_id = peer_id AND m2.dm_recipient_id = ?))
            ) as last_msg_id,
            (SELECT m3.content FROM tc_messages m3 WHERE m3.id = (
                SELECT MAX(m4.id) FROM tc_messages m4
                WHERE m4.channel_id IS NULL
                AND ((m4.sender_id = ? AND m4.dm_recipient_id = peer_id)
                  OR (m4.sender_id = peer_id AND m4.dm_recipient_id = ?))
            )) as last_message,
            (SELECT COUNT(*) FROM tc_messages m5
             WHERE m5.channel_id IS NULL AND m5.sender_id = peer_id AND m5.dm_recipient_id = ?
             AND m5.id > COALESCE(
                (SELECT rs.last_read_message_id FROM tc_read_status rs
                 WHERE rs.user_id = ? AND rs.dm_peer_id = peer_id), 0)
            ) as unread_count
        FROM (
            SELECT DISTINCT CASE WHEN sender_id = ? THEN dm_recipient_id ELSE sender_id END as peer_id
            FROM tc_messages WHERE channel_id IS NULL AND (sender_id = ? OR dm_recipient_id = ?)
        ) peers
        JOIN users u ON u.id = peer_id
        ORDER BY last_msg_id DESC
    ''', (uid, uid, uid, uid, uid, uid, uid, uid, uid)).fetchall()
    conn.close()
    return jsonify([{
        'peer_id': r['peer_id'],
        'display_name': r['display_name'] or r['username'],
        'last_message': r['last_message'] or '',
        'unread_count': r['unread_count']
    } for r in rows])

@app.route('/api/team-chat/dm/<int:peer_id>/messages')
@api_login_required
def api_tc_dm_messages(peer_id):
    uid = session['user_id']
    before_id = request.args.get('before_id', type=int)
    limit = request.args.get('limit', 50, type=int)

    conn = get_db()
    if before_id:
        msgs = conn.execute('''
            SELECT m.*, u.display_name as sender_name FROM tc_messages m
            LEFT JOIN users u ON u.id = m.sender_id
            WHERE m.channel_id IS NULL AND m.id < ?
            AND ((m.sender_id = ? AND m.dm_recipient_id = ?) OR (m.sender_id = ? AND m.dm_recipient_id = ?))
            ORDER BY m.id DESC LIMIT ?
        ''', (before_id, uid, peer_id, peer_id, uid, limit)).fetchall()
    else:
        msgs = conn.execute('''
            SELECT m.*, u.display_name as sender_name FROM tc_messages m
            LEFT JOIN users u ON u.id = m.sender_id
            WHERE m.channel_id IS NULL
            AND ((m.sender_id = ? AND m.dm_recipient_id = ?) OR (m.sender_id = ? AND m.dm_recipient_id = ?))
            ORDER BY m.id DESC LIMIT ?
        ''', (uid, peer_id, peer_id, uid, limit)).fetchall()

    conn.close()
    result = [dict(m) for m in msgs]
    result.reverse()
    return jsonify(result)

@app.route('/api/team-chat/dm/<int:peer_id>/messages', methods=['POST'])
@api_login_required
def api_tc_send_dm(peer_id):
    uid = session['user_id']
    content = (request.form.get('content') or '').strip()
    file = request.files.get('file')

    if not content and not file:
        return jsonify({'error': 'Message content or file required'}), 400

    file_path = file_name = file_type = ''
    if file and file.filename:
        os.makedirs(TEAM_CHAT_UPLOAD_DIR, exist_ok=True)
        import uuid
        ext = os.path.splitext(file.filename)[1]
        saved_name = f"{uuid.uuid4().hex}{ext}"
        save_path = os.path.join(TEAM_CHAT_UPLOAD_DIR, saved_name)
        file.save(save_path)
        file_path = saved_name
        file_name = file.filename
        file_type = file.content_type or ''

    conn = get_db()
    cur = conn.execute(
        '''INSERT INTO tc_messages (sender_id, dm_recipient_id, content, file_path, file_name, file_type)
           VALUES (?, ?, ?, ?, ?, ?)''',
        (uid, peer_id, content, file_path, file_name, file_type)
    )
    msg_id = cur.lastrowid

    # Update sender's read status
    conn.execute('''INSERT INTO tc_read_status (user_id, dm_peer_id, last_read_message_id, updated_at)
        VALUES (?, ?, ?, datetime('now','localtime'))
        ON CONFLICT(user_id, dm_peer_id) DO UPDATE SET last_read_message_id = ?, updated_at = datetime('now','localtime')''',
        (uid, peer_id, msg_id, msg_id))

    sender_name = session.get('display_name') or session.get('username', 'Someone')
    preview = (content[:80] + '...') if len(content) > 80 else content

    conn.commit()
    conn.close()

    # Notify recipient (create_notification opens its own conn)
    create_notification(peer_id, 'team_chat',
        f'DM from {sender_name}',
        preview or '(file attachment)',
        f'/team-chat?dm={uid}')

    return jsonify({'id': msg_id})

@app.route('/api/team-chat/dm/<int:peer_id>/read', methods=['POST'])
@api_login_required
def api_tc_dm_read(peer_id):
    uid = session['user_id']
    data = request.get_json() or {}
    last_id = data.get('last_message_id', 0)
    conn = get_db()
    conn.execute('''INSERT INTO tc_read_status (user_id, dm_peer_id, last_read_message_id, updated_at)
        VALUES (?, ?, ?, datetime('now','localtime'))
        ON CONFLICT(user_id, dm_peer_id) DO UPDATE SET
            last_read_message_id = MAX(last_read_message_id, ?),
            updated_at = datetime('now','localtime')''',
        (uid, peer_id, last_id, last_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Team Chat Files & Unread ───────────────────────────────

@app.route('/api/team-chat/files/<int:msg_id>')
@api_login_required
def api_tc_file(msg_id):
    conn = get_db()
    msg = conn.execute('SELECT file_path, file_name, file_type FROM tc_messages WHERE id = ?', (msg_id,)).fetchone()
    conn.close()
    if not msg or not msg['file_path']:
        return 'File not found', 404
    fpath = os.path.join(TEAM_CHAT_UPLOAD_DIR, msg['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    as_attachment = not (msg['file_type'] or '').startswith('image/')
    return send_file(fpath, download_name=msg['file_name'], as_attachment=as_attachment)

@app.route('/api/team-chat/unread-total')
@api_login_required
def api_tc_unread_total():
    uid = session['user_id']
    conn = get_db()
    # Channel unreads
    ch_unread = conn.execute('''
        SELECT COALESCE(SUM(cnt), 0) as total FROM (
            SELECT COUNT(*) as cnt FROM tc_messages m
            JOIN tc_channel_members cm ON cm.channel_id = m.channel_id AND cm.user_id = ?
            WHERE m.id > COALESCE(
                (SELECT rs.last_read_message_id FROM tc_read_status rs
                 WHERE rs.user_id = ? AND rs.channel_id = m.channel_id), 0)
        )
    ''', (uid, uid)).fetchone()['total']

    # DM unreads
    dm_unread = conn.execute('''
        SELECT COUNT(*) as total FROM tc_messages m
        WHERE m.channel_id IS NULL AND m.dm_recipient_id = ?
        AND m.id > COALESCE(
            (SELECT rs.last_read_message_id FROM tc_read_status rs
             WHERE rs.user_id = ? AND rs.dm_peer_id = m.sender_id), 0)
    ''', (uid, uid)).fetchone()['total']

    conn.close()
    return jsonify({'total': ch_unread + dm_unread})

# ─── Training ────────────────────────────────────────────────────

@app.route('/training')
@login_required
def training_page():
    return render_template('training/index.html')

@app.route('/api/training/progress')
@api_login_required
def api_training_progress():
    conn = get_db()
    rows = conn.execute(
        'SELECT lesson_key FROM training_progress WHERE user_id = ?',
        (session['user_id'],)
    ).fetchall()
    conn.close()
    return jsonify({'completed': [r['lesson_key'] for r in rows]})

@app.route('/api/training/progress', methods=['POST'])
@api_login_required
def api_training_complete():
    data = request.get_json(force=True)
    lesson_key = data.get('lesson_key', '')
    if not lesson_key:
        return jsonify({'error': 'Missing lesson_key'}), 400
    conn = get_db()
    conn.execute(
        'INSERT OR IGNORE INTO training_progress (user_id, lesson_key) VALUES (?,?)',
        (session['user_id'], lesson_key)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/training/progress', methods=['DELETE'])
@api_login_required
def api_training_reset():
    conn = get_db()
    conn.execute('DELETE FROM training_progress WHERE user_id = ?', (session['user_id'],))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Reminders ────────────────────────────────────────────────────

@app.route('/reminders')
@login_required
def reminders_page():
    return render_template('reminders/list.html')

@app.route('/api/reminders')
@api_login_required
def api_reminders_list():
    conn = get_db()
    uid = session['user_id']
    rows = conn.execute(
        "SELECT * FROM reminders WHERE user_id = ? ORDER BY CASE WHEN due_date = '' THEN 1 ELSE 0 END, due_date ASC",
        (uid,)
    ).fetchall()
    result = [dict(r) for r in rows]
    _check_reminder_notifications(result, uid, conn)
    conn.close()
    return jsonify(result)

def _check_reminder_notifications(reminders, user_id, conn):
    now = datetime.now()
    today_str = now.strftime('%Y-%m-%d')
    tomorrow = (now + timedelta(days=1)).strftime('%Y-%m-%d')
    for r in reminders:
        if r['status'] != 'Active' or not r.get('due_date'):
            continue
        due = r['due_date']
        if due > tomorrow:
            continue
        # Check if notification already sent today for this reminder
        existing = conn.execute(
            "SELECT id FROM notifications WHERE user_id = ? AND type = 'reminder' AND message LIKE ? AND created_at >= ?",
            (user_id, f'%reminder#{r["id"]}%', today_str)
        ).fetchone()
        if existing:
            continue
        if due < today_str:
            label = 'is overdue'
        elif due == today_str:
            label = 'is due today'
        else:
            label = 'is due tomorrow'
        conn.execute(
            'INSERT INTO notifications (user_id, type, title, message, link) VALUES (?,?,?,?,?)',
            (user_id, 'reminder', f'Reminder: {r["title"]}',
             f'"{r["title"]}" {label} (reminder#{r["id"]})',
             '/reminders')
        )
    conn.commit()

@app.route('/api/reminders', methods=['POST'])
@api_login_required
def api_create_reminder():
    data = request.get_json(force=True)
    conn = get_db()
    conn.execute(
        'INSERT INTO reminders (user_id, title, description, due_date) VALUES (?,?,?,?)',
        (session['user_id'], data.get('title', ''), data.get('description', ''), data.get('due_date', ''))
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/reminders/<int:rid>', methods=['PUT'])
@api_login_required
def api_update_reminder(rid):
    conn = get_db()
    row = conn.execute('SELECT * FROM reminders WHERE id = ? AND user_id = ?', (rid, session['user_id'])).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    data = request.get_json(force=True)
    conn.execute(
        '''UPDATE reminders SET title = ?, description = ?, due_date = ?, status = ?,
           updated_at = datetime('now','localtime') WHERE id = ?''',
        (data.get('title', row['title']), data.get('description', row['description']),
         data.get('due_date', row['due_date']), data.get('status', row['status']), rid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/reminders/<int:rid>/complete', methods=['PUT'])
@api_login_required
def api_complete_reminder(rid):
    conn = get_db()
    row = conn.execute('SELECT * FROM reminders WHERE id = ? AND user_id = ?', (rid, session['user_id'])).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    conn.execute(
        "UPDATE reminders SET status = 'Completed', updated_at = datetime('now','localtime') WHERE id = ?",
        (rid,)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/reminders/<int:rid>', methods=['DELETE'])
@api_login_required
def api_delete_reminder(rid):
    conn = get_db()
    row = conn.execute('SELECT * FROM reminders WHERE id = ? AND user_id = ?', (rid, session['user_id'])).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    conn.execute('DELETE FROM reminders WHERE id = ?', (rid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Shared Files (Dropbox Replacement) ──────────────────────────

SHARED_FILES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'shared_files')
os.makedirs(SHARED_FILES_DIR, exist_ok=True)

@app.route('/shared-files')
@login_required
def shared_files_page():
    return render_template('shared_files/index.html')

@app.route('/api/shared-files')
@api_login_required
def api_shared_files_list():
    parent_id = request.args.get('parent_id')
    conn = get_db()
    if parent_id:
        rows = conn.execute(
            '''SELECT sf.*, u.display_name as uploader_name
               FROM shared_files sf LEFT JOIN users u ON sf.uploaded_by = u.id
               WHERE sf.parent_id = ?
               ORDER BY sf.is_folder DESC, LOWER(sf.name) ASC''', (parent_id,)
        ).fetchall()
    else:
        rows = conn.execute(
            '''SELECT sf.*, u.display_name as uploader_name
               FROM shared_files sf LEFT JOIN users u ON sf.uploaded_by = u.id
               WHERE sf.parent_id IS NULL
               ORDER BY sf.is_folder DESC, LOWER(sf.name) ASC'''
        ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/shared-files/<int:fid>/breadcrumbs')
@api_login_required
def api_shared_files_breadcrumbs(fid):
    conn = get_db()
    path = []
    current = fid
    while current:
        row = conn.execute('SELECT id, name, parent_id FROM shared_files WHERE id = ?', (current,)).fetchone()
        if not row:
            break
        path.append({'id': row['id'], 'name': row['name']})
        current = row['parent_id']
    conn.close()
    path.reverse()
    return jsonify(path)

@app.route('/api/shared-files/folder', methods=['POST'])
@api_login_required
def api_shared_files_create_folder():
    data = request.get_json(force=True)
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Folder name is required'}), 400
    parent_id = data.get('parent_id')
    conn = get_db()
    # Check duplicate name in same parent
    if parent_id:
        dup = conn.execute(
            'SELECT id FROM shared_files WHERE parent_id = ? AND name = ? AND is_folder = 1',
            (parent_id, name)
        ).fetchone()
    else:
        dup = conn.execute(
            'SELECT id FROM shared_files WHERE parent_id IS NULL AND name = ? AND is_folder = 1',
            (name,)
        ).fetchone()
    if dup:
        conn.close()
        return jsonify({'error': 'A folder with that name already exists here'}), 400
    conn.execute(
        'INSERT INTO shared_files (parent_id, name, is_folder, uploaded_by) VALUES (?, ?, 1, ?)',
        (parent_id, name, session['user_id'])
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/shared-files/upload', methods=['POST'])
@api_login_required
def api_shared_files_upload():
    from werkzeug.utils import secure_filename
    import mimetypes
    parent_id = request.form.get('parent_id') or None
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'No files provided'}), 400
    conn = get_db()
    for f in files:
        if not f.filename:
            continue
        fname = secure_filename(f.filename)
        ts_fname = f"{int(datetime.now().timestamp())}_{fname}"
        fpath = os.path.join(SHARED_FILES_DIR, ts_fname)
        f.save(fpath)
        fsize = os.path.getsize(fpath)
        mime = mimetypes.guess_type(fname)[0] or 'application/octet-stream'
        conn.execute(
            '''INSERT INTO shared_files (parent_id, name, is_folder, file_path, file_size, mime_type, uploaded_by)
               VALUES (?, ?, 0, ?, ?, ?, ?)''',
            (parent_id, fname, ts_fname, fsize, mime, session['user_id'])
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/shared-files/<int:fid>/rename', methods=['PUT'])
@api_login_required
def api_shared_files_rename(fid):
    data = request.get_json(force=True)
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Name is required'}), 400
    conn = get_db()
    conn.execute(
        "UPDATE shared_files SET name = ?, updated_at = datetime('now','localtime') WHERE id = ?",
        (name, fid)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/shared-files/<int:fid>', methods=['DELETE'])
@api_login_required
def api_shared_files_delete(fid):
    conn = get_db()
    row = conn.execute('SELECT * FROM shared_files WHERE id = ?', (fid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    # Recursively collect all file_paths in subtree
    def collect_files(item_id):
        paths = []
        item = conn.execute('SELECT file_path, is_folder FROM shared_files WHERE id = ?', (item_id,)).fetchone()
        if not item:
            return paths
        if item['is_folder']:
            children = conn.execute('SELECT id FROM shared_files WHERE parent_id = ?', (item_id,)).fetchall()
            for child in children:
                paths.extend(collect_files(child['id']))
        else:
            if item['file_path']:
                paths.append(item['file_path'])
        return paths

    file_paths = collect_files(fid)
    # Delete physical files
    for fp in file_paths:
        full = os.path.join(SHARED_FILES_DIR, fp)
        if os.path.exists(full):
            os.remove(full)

    # Delete DB row (cascade handles children)
    conn.execute('DELETE FROM shared_files WHERE id = ?', (fid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/shared-files/<int:fid>/download')
@api_login_required
def api_shared_files_download(fid):
    conn = get_db()
    row = conn.execute('SELECT * FROM shared_files WHERE id = ?', (fid,)).fetchone()
    conn.close()
    if not row or row['is_folder'] or not row['file_path']:
        return 'Not found', 404
    fpath = os.path.join(SHARED_FILES_DIR, row['file_path'])
    if not os.path.exists(fpath):
        return 'File not found', 404
    as_attachment = request.args.get('download') == '1'
    return send_file(fpath, mimetype=row['mime_type'] or 'application/octet-stream',
                     as_attachment=as_attachment, download_name=row['name'])

# ─── Service Invoices ────────────────────────────────────────────

@app.route('/service-invoices')
@login_required
def service_invoices_page():
    return render_template('service_invoices/list.html')


@app.route('/api/service-invoices', methods=['GET'])
@api_role_required('owner', 'admin', 'project_manager')
def api_service_invoices_list():
    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    rows = conn.execute('''
        SELECT ci.*, j.name as job_name,
               (SELECT COUNT(*) FROM service_invoice_items WHERE invoice_id = ci.id) as item_count
        FROM client_invoices ci
        LEFT JOIN jobs j ON ci.job_id = j.id
        ORDER BY ci.issue_date DESC, ci.id DESC
    ''').fetchall()
    result = []
    for r in rows:
        d = dict(r)
        # Auto-mark overdue
        if d['status'] not in ('Paid',) and d['due_date'] and d['due_date'] < today:
            if d['status'] != 'Overdue' and d['amount_paid'] == 0:
                d['status'] = 'Overdue'
            elif d['status'] != 'Overdue' and d['amount_paid'] > 0 and d['amount_paid'] < d['amount']:
                d['status'] = 'Partial'
        result.append(d)
    conn.close()
    return jsonify(result)


@app.route('/api/service-invoices/<int:sid>', methods=['GET'])
@api_role_required('owner', 'admin', 'project_manager')
def api_service_invoice_detail(sid):
    conn = get_db()
    inv = conn.execute('SELECT ci.*, j.name as job_name FROM client_invoices ci LEFT JOIN jobs j ON ci.job_id = j.id WHERE ci.id = ?', (sid,)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    d = dict(inv)
    d['items'] = [dict(i) for i in conn.execute('SELECT * FROM service_invoice_items WHERE invoice_id = ? ORDER BY id', (sid,)).fetchall()]
    d['payments'] = [dict(p) for p in conn.execute('''
        SELECT sip.*, u.display_name as recorded_by_name
        FROM service_invoice_payments sip
        LEFT JOIN users u ON sip.created_by = u.id
        WHERE sip.invoice_id = ? ORDER BY sip.payment_date DESC, sip.id DESC
    ''', (sid,)).fetchall()]
    conn.close()
    return jsonify(d)


@app.route('/api/service-invoices', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_service_invoice_create():
    data = request.get_json()
    conn = get_db()

    # Auto-generate invoice number if not provided
    inv_num = data.get('invoice_number', '').strip()
    if not inv_num:
        last = conn.execute("SELECT invoice_number FROM client_invoices WHERE invoice_number LIKE 'INV-%' ORDER BY id DESC LIMIT 1").fetchone()
        if last and last['invoice_number']:
            try:
                num = int(last['invoice_number'].replace('INV-', '')) + 1
            except ValueError:
                num = 1
        else:
            num = 1
        inv_num = f'INV-{num:04d}'

    # Calculate totals from line items
    items = data.get('items', [])
    subtotal = sum(float(it.get('quantity', 0)) * float(it.get('unit_price', 0)) for it in items)
    tax_rate = float(data.get('tax_rate', 0))
    tax_amount = round(subtotal * tax_rate / 100, 2)
    total = round(subtotal + tax_amount, 2)

    job_id = data.get('job_id') or None

    cursor = conn.execute(
        '''INSERT INTO client_invoices (job_id, invoice_number, amount, status, description,
           issue_date, due_date, created_by, customer_name, customer_email, customer_address,
           notes, terms, tax_rate, subtotal, tax_amount, balance_due, amount_paid)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0)''',
        (job_id, inv_num, total, data.get('status', 'Draft'),
         data.get('description', ''), data.get('issue_date', datetime.now().strftime('%Y-%m-%d')),
         data.get('due_date', ''), session.get('user_id'),
         data.get('customer_name', ''), data.get('customer_email', ''),
         data.get('customer_address', ''), data.get('notes', ''),
         data.get('terms', 'Net 30'), tax_rate, subtotal, tax_amount, total)
    )
    inv_id = cursor.lastrowid

    for it in items:
        qty = float(it.get('quantity', 1))
        up = float(it.get('unit_price', 0))
        amt = round(qty * up, 2)
        conn.execute(
            'INSERT INTO service_invoice_items (invoice_id, description, quantity, unit_price, amount) VALUES (?,?,?,?,?)',
            (inv_id, it.get('description', ''), qty, up, amt)
        )

    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': inv_id, 'invoice_number': inv_num}), 201


@app.route('/api/service-invoices/<int:sid>', methods=['PUT'])
@api_role_required('owner', 'admin', 'project_manager')
def api_service_invoice_update(sid):
    data = request.get_json()
    conn = get_db()

    inv = conn.execute('SELECT * FROM client_invoices WHERE id = ?', (sid,)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    items = data.get('items', [])
    subtotal = sum(float(it.get('quantity', 0)) * float(it.get('unit_price', 0)) for it in items)
    tax_rate = float(data.get('tax_rate', 0))
    tax_amount = round(subtotal * tax_rate / 100, 2)
    total = round(subtotal + tax_amount, 2)
    amount_paid = float(inv['amount_paid'] or 0)
    balance_due = round(total - amount_paid, 2)

    job_id = data.get('job_id') or None

    conn.execute(
        '''UPDATE client_invoices SET job_id=?, invoice_number=?, amount=?, status=?,
           description=?, issue_date=?, due_date=?, customer_name=?, customer_email=?,
           customer_address=?, notes=?, terms=?, tax_rate=?, subtotal=?, tax_amount=?,
           balance_due=?
           WHERE id=?''',
        (job_id, data.get('invoice_number', inv['invoice_number']),
         total, data.get('status', inv['status']),
         data.get('description', inv['description']),
         data.get('issue_date', inv['issue_date']),
         data.get('due_date', inv['due_date']),
         data.get('customer_name', inv['customer_name']),
         data.get('customer_email', inv['customer_email']),
         data.get('customer_address', inv['customer_address']),
         data.get('notes', inv['notes']),
         data.get('terms', inv['terms']),
         tax_rate, subtotal, tax_amount, balance_due, sid)
    )

    # Replace line items
    conn.execute('DELETE FROM service_invoice_items WHERE invoice_id = ?', (sid,))
    for it in items:
        qty = float(it.get('quantity', 1))
        up = float(it.get('unit_price', 0))
        amt = round(qty * up, 2)
        conn.execute(
            'INSERT INTO service_invoice_items (invoice_id, description, quantity, unit_price, amount) VALUES (?,?,?,?,?)',
            (sid, it.get('description', ''), qty, up, amt)
        )

    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/service-invoices/<int:sid>', methods=['DELETE'])
@api_role_required('owner', 'admin', 'project_manager')
def api_service_invoice_delete(sid):
    conn = get_db()
    inv = conn.execute('SELECT status FROM client_invoices WHERE id = ?', (sid,)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if inv['status'] not in ('Draft',):
        conn.close()
        return jsonify({'error': 'Only draft invoices can be deleted'}), 400
    conn.execute('DELETE FROM service_invoice_items WHERE invoice_id = ?', (sid,))
    conn.execute('DELETE FROM service_invoice_payments WHERE invoice_id = ?', (sid,))
    conn.execute('DELETE FROM client_invoices WHERE id = ?', (sid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/service-invoices/<int:sid>/generate-pdf', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_service_invoice_generate_pdf(sid):
    conn = get_db()
    inv = conn.execute('SELECT ci.*, j.name as job_name FROM client_invoices ci LEFT JOIN jobs j ON ci.job_id = j.id WHERE ci.id = ?', (sid,)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    items = conn.execute('SELECT * FROM service_invoice_items WHERE invoice_id = ? ORDER BY id', (sid,)).fetchall()
    payments = conn.execute('SELECT * FROM service_invoice_payments WHERE invoice_id = ? ORDER BY payment_date', (sid,)).fetchall()

    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'icons', 'sidebar-logo.png')
    if not os.path.exists(logo_path):
        logo_path = ''
    else:
        logo_path = 'file://' + os.path.abspath(logo_path)

    def fmt_money(val):
        try:
            v = float(val or 0)
            return f'{v:,.2f}'
        except (ValueError, TypeError):
            return '0.00'

    html = render_template('service_invoices/invoice_pdf.html',
        invoice=dict(inv),
        items=[dict(i) for i in items],
        payments=[dict(p) for p in payments],
        logo_path=logo_path,
        fmt=fmt_money
    )

    pdf_dir = os.path.join(os.path.dirname(__file__), 'data', 'service_invoices')
    os.makedirs(pdf_dir, exist_ok=True)

    safe_num = ''.join(c if c.isalnum() or c in '-_' else '' for c in (inv['invoice_number'] or 'invoice'))
    filename = f"ServiceInvoice_{safe_num}.pdf"
    filepath = os.path.join(pdf_dir, filename)

    if weasyprint is None:
        conn.close()
        return jsonify({'error': 'PDF generation not available. WeasyPrint is not installed.'}), 500

    try:
        wp = weasyprint.HTML(string=html, base_url=os.path.dirname(__file__))
        wp.write_pdf(filepath)
    except Exception as e:
        conn.close()
        return jsonify({'error': f'PDF generation failed: {str(e)[:200]}'}), 500

    conn.execute('UPDATE client_invoices SET pdf_file = ? WHERE id = ?', (filename, sid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'filename': filename, 'path': f'/api/service-invoices/{sid}/pdf/{filename}'})


@app.route('/api/service-invoices/<int:sid>/pdf/<filename>')
@api_role_required('owner', 'admin', 'project_manager')
def api_service_invoice_download_pdf(sid, filename):
    pdf_dir = os.path.join(os.path.dirname(__file__), 'data', 'service_invoices')
    filepath = os.path.join(pdf_dir, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    if request.args.get('download'):
        return send_file(filepath, as_attachment=True, download_name=filename)
    return send_file(filepath, mimetype='application/pdf')


@app.route('/api/service-invoices/<int:sid>/email', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_service_invoice_email(sid):
    data = request.get_json()
    recipients = [e.strip() for e in data.get('recipients', []) if e.strip()]
    subject = data.get('subject', 'Invoice from LGHVAC LLC')
    body_text = data.get('body', '')

    if not recipients:
        return jsonify({'error': 'No recipients specified'}), 400

    saved_settings = {}
    settings_path = os.path.join(os.path.dirname(__file__), 'data', 'email_settings.json')
    if os.path.exists(settings_path):
        with open(settings_path) as f:
            saved_settings = json.load(f)

    smtp_host = saved_settings.get('smtp_host', '')
    smtp_port = int(saved_settings.get('smtp_port', 587) or 587)
    smtp_user = saved_settings.get('smtp_user', '')
    smtp_pass = saved_settings.get('smtp_pass', '')
    from_email = saved_settings.get('from_email', '') or smtp_user

    if not smtp_host or not smtp_user:
        return jsonify({'error': 'SMTP not configured. Go to Settings to set up email.'}), 400

    conn = get_db()
    inv = conn.execute('SELECT pdf_file, invoice_number, status FROM client_invoices WHERE id = ?', (sid,)).fetchone()
    if not inv or not inv['pdf_file']:
        conn.close()
        return jsonify({'error': 'No PDF generated yet. Generate the PDF first.'}), 404

    pdf_dir = os.path.join(os.path.dirname(__file__), 'data', 'service_invoices')
    pdf_path = os.path.join(pdf_dir, inv['pdf_file'])
    if not os.path.exists(pdf_path):
        conn.close()
        return jsonify({'error': 'PDF file not found on disk.'}), 404

    try:
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = subject
        msg.attach(MIMEText(body_text, 'plain'))

        with open(pdf_path, 'rb') as f:
            part = MIMEBase('application', 'pdf')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(pdf_path)}"')
            msg.attach(part)

        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)

        # Update status to Sent if still Draft
        if inv['status'] == 'Draft':
            conn.execute("UPDATE client_invoices SET status = 'Sent' WHERE id = ?", (sid,))
            conn.commit()

        conn.close()
        save_email_autocomplete(recipients)
        return jsonify({'ok': True, 'sent_to': recipients})
    except Exception as e:
        conn.close()
        return jsonify({'error': f'Email failed: {str(e)[:200]}'}), 500


@app.route('/api/service-invoices/<int:sid>/record-payment', methods=['POST'])
@api_role_required('owner', 'admin', 'project_manager')
def api_service_invoice_record_payment(sid):
    data = request.get_json()
    conn = get_db()

    inv = conn.execute('SELECT * FROM client_invoices WHERE id = ?', (sid,)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    pay_amount = float(data.get('amount', 0))
    if pay_amount <= 0:
        conn.close()
        return jsonify({'error': 'Payment amount must be positive'}), 400

    conn.execute(
        '''INSERT INTO service_invoice_payments (invoice_id, amount, payment_date, payment_method, reference_number, notes, created_by)
           VALUES (?,?,?,?,?,?,?)''',
        (sid, pay_amount, data.get('payment_date', datetime.now().strftime('%Y-%m-%d')),
         data.get('payment_method', ''), data.get('reference_number', ''),
         data.get('notes', ''), session.get('user_id'))
    )

    # Recalculate totals
    total_paid = conn.execute('SELECT COALESCE(SUM(amount), 0) FROM service_invoice_payments WHERE invoice_id = ?', (sid,)).fetchone()[0]
    balance = round(float(inv['amount']) - total_paid, 2)

    if balance <= 0:
        new_status = 'Paid'
        paid_date = datetime.now().strftime('%Y-%m-%d')
        # Calculate days_to_pay
        days = None
        if inv['issue_date']:
            try:
                issue = datetime.strptime(inv['issue_date'], '%Y-%m-%d')
                days = (datetime.now() - issue).days
            except ValueError:
                pass
        conn.execute('UPDATE client_invoices SET status=?, amount_paid=?, balance_due=?, paid_date=?, days_to_pay=? WHERE id=?',
                     (new_status, total_paid, 0, paid_date, days, sid))
    else:
        conn.execute('UPDATE client_invoices SET status=?, amount_paid=?, balance_due=? WHERE id=?',
                     ('Partial', total_paid, balance, sid))

    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ─── Startup ────────────────────────────────────────────────────

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return '127.0.0.1'

if __name__ == '__main__':
    init_db()
    local_ip = get_local_ip()
    print(f'\n  Construction Management')
    print(f'  ────────────────────────────────')
    print(f'  Local:   http://localhost:5001')
    print(f'  Network: http://{local_ip}:5001')
    print(f'  ────────────────────────────────')
    print(f'  Default login: admin / admin')
    print(f'  ────────────────────────────────\n')
    app.run(host='0.0.0.0', port=5001, debug=True)
